"""
Propel Health MCP Server
========================
Connects Claude to the Propel Health toolkits.

Available Tool Categories:

CONFIGURATIONS TOOLKIT:
- User Management: list_users, get_user, add_user, export_annual_review
- Access Management: list_access, get_reviews_due
- Training: get_training_status, get_expired_training
- Compliance: get_compliance_report
- Configuration: list_programs, get_config

REQUIREMENTS TOOLKIT:
- Client/Program: list_clients, get_client_programs, get_program_by_prefix, create_program, create_requirement, link_requirement_to_story
- User Stories: list_stories, get_story, create_story, update_story, get_approval_pipeline
- Test Cases: list_test_cases, get_test_summary, create_test_case, update_test_result
- Reporting: get_program_health, get_client_tree, search_stories, get_coverage_gaps

UAT CYCLE MANAGEMENT:
- Cycle Management: create_uat_cycle, get_uat_cycle, list_uat_cycles, update_uat_cycle_status
- Test Assignment: assign_test_case, bulk_assign_by_profile
- Test Execution: update_test_execution
- Gate Management: update_pre_uat_gate
- Reporting: get_cycle_dashboard, get_tester_workload
- Decisions: record_go_nogo_decision

UAT TOOLKIT INTEGRATION:
- Cycle Setup: setup_uat_cycle_with_testers - Create cycle + assign testers with overlap
- Tester Assignment: assign_uat_testers - Add/modify tester assignments
- Tracker Generation: generate_uat_trackers - Generate HTML files + push to GitHub
- Results Import: import_uat_results_json - Paste JSON from Formspree email
- Dashboard: regenerate_uat_dashboard, get_uat_progress - View/update progress

ONBOARDING PROJECT MANAGEMENT:
- Projects: create_onboarding_project, get_onboarding_project, list_onboarding_projects
- Milestones: update_milestone
- Dependencies: add_onboarding_dependency, resolve_dependency
- Readiness: get_go_live_readiness

ONBOARDING FORM TOOLKIT:
- Question Management: list_form_questions, add_form_question, update_form_question, remove_form_question, reorder_form_questions

DASHBOARD DATA GENERATION:
- generate_dashboard_data: Generate JSON for clinic configuration dashboard (GitHub Pages)

REQUIREMENTS DASHBOARD:
- generate_requirements_dashboard: Generate requirements dashboard HTML + push to GitHub Pages
"""

# ============================================================
# IMPORTS
# ============================================================
import os
import sys
import csv
import json
import logging
import importlib.util
from datetime import datetime, date, timedelta
from typing import Optional
from logging.handlers import RotatingFileHandler

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# MCP library
from mcp.server.fastmcp import FastMCP

# ============================================================
# LOGGING CONFIGURATION
# ============================================================
# Structured logging for audit trail and debugging.
# Logs are written to a rotating file (max 5MB, 3 backups) for compliance.
#
# Log levels used:
#   - INFO: Successful operations, tool invocations
#   - WARNING: Non-critical issues, missing optional data
#   - ERROR: Failed operations, exceptions
#   - DEBUG: Detailed execution info (disabled by default)

LOG_DIR = os.environ.get("PROPEL_LOG_DIR", os.path.expanduser("~/projects/logs"))
LOG_FILE = os.path.join(LOG_DIR, "propel_mcp.log")
LOG_LEVEL = os.environ.get("PROPEL_LOG_LEVEL", "INFO").upper()

# Ensure log directory exists
os.makedirs(LOG_DIR, exist_ok=True)

# Create logger
logger = logging.getLogger("propel_mcp")
logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))

# Prevent duplicate handlers if module reloaded
if not logger.handlers:
    # File handler with rotation (5MB max, keep 3 backups)
    file_handler = RotatingFileHandler(
        LOG_FILE,
        maxBytes=5 * 1024 * 1024,  # 5MB
        backupCount=3,
        encoding="utf-8"
    )
    file_handler.setLevel(logging.DEBUG)

    # Console handler for errors only (doesn't clutter MCP output)
    console_handler = logging.StreamHandler(sys.stderr)
    console_handler.setLevel(logging.ERROR)

    # Format: timestamp | level | message
    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

logger.info("=" * 60)
logger.info("Propel MCP Server starting")
logger.info(f"Log file: {LOG_FILE}")
logger.info(f"Log level: {LOG_LEVEL}")

# Import from configurations_toolkit (installed as editable package)
from configurations_toolkit import (
    AccessManager,
    ConfigurationManager,
    InheritanceManager,
    ComplianceReports,
)


# Load requirements_toolkit modules using importlib to avoid name conflicts
# (both toolkits have a 'database' module)
def _import_from_path(name: str, path: str):
    """Import a module from file path without polluting sys.modules."""
    spec = importlib.util.spec_from_file_location(name, path)
    if spec and spec.loader:
        module = importlib.util.module_from_spec(spec)
        # Temporarily add parent path for internal imports
        parent = os.path.dirname(os.path.dirname(path))
        sys.path.insert(0, parent)
        try:
            spec.loader.exec_module(module)
        finally:
            sys.path.remove(parent)
        return module
    raise ImportError(f"Could not load {path}")


_REQ_PATH = os.path.expanduser("~/projects/requirements_toolkit")
_req_db_manager = _import_from_path(
    "req_db_manager",
    os.path.join(_REQ_PATH, "database", "db_manager.py")
)
_req_queries_mod = _import_from_path(
    "req_queries_mod",
    os.path.join(_REQ_PATH, "database", "queries.py")
)

# Expose with cleaner names
ReqClientProductDatabase = _req_db_manager.ClientProductDatabase
req_queries = _req_queries_mod

# ============================================================
# SERVER SETUP
# ============================================================
mcp = FastMCP("propel-health")

# Unified database path - all toolkits share a single database
# The database contains: requirements, configurations, UAT, and access management
DB_PATH = os.environ.get(
    "PROPEL_DB_PATH",
    os.path.expanduser("~/projects/data/client_product_database.db")
)

# Legacy alias for backwards compatibility - both paths now resolve to the same unified DB
# Other toolkits use symlinks: ~/projects/requirements_toolkit/data/ and ~/projects/uat_toolkit/data/
REQ_DB_PATH = os.environ.get(
    "REQUIREMENTS_DB_PATH",
    DB_PATH  # Now points to the same unified database
)

# Notion Dashboard Page ID (created by Claude)
NOTION_DASHBOARD_PAGE_ID = "2dab5d1d-1631-81bb-8eeb-c4f4397de747"


# ============================================================
# HELPER FUNCTIONS
# ============================================================


def validate_choice(value: str, valid_options: list, field_name: str, descriptions: dict = None) -> Optional[str]:
    """
    Validate that a value is one of the valid options.

    Args:
        value: The value to validate
        valid_options: List of valid option values
        field_name: Name of the field for error message
        descriptions: Optional dict mapping option values to descriptions

    Returns:
        Error message string if invalid, None if valid
    """
    if value not in valid_options:
        error = f"Invalid {field_name}: '{value}'\n\nValid {field_name}s:\n"
        for opt in valid_options:
            desc = descriptions.get(opt, "") if descriptions else ""
            if desc:
                error += f"  • {opt} - {desc}\n"
            else:
                error += f"  • {opt}\n"
        return error.rstrip()
    return None


def validate_optional_choice(value: Optional[str], valid_options: list, field_name: str,
                             descriptions: dict = None) -> Optional[str]:
    """
    Validate an optional field - only validates if value is not None.

    Returns:
        Error message string if invalid, None if valid or not provided
    """
    if value is None:
        return None
    return validate_choice(value, valid_options, field_name, descriptions)


def log_audit(cursor, record_type: str, record_id: str, action: str,
              field_changed: str, old_value, new_value, changed_by: str,
              changed_date: str, change_reason: str) -> None:
    """
    Insert an audit history record.

    Args:
        cursor: Database cursor
        record_type: Type of record (client, story, test, etc.)
        record_id: ID of the record being audited
        action: Action performed (Created, Updated, Deleted)
        field_changed: Name of field changed or record type for creates
        old_value: Previous value (None for creates)
        new_value: New value (usually JSON for creates/deletes)
        changed_by: Who made the change (e.g., 'MCP:tool_name')
        changed_date: Timestamp of the change
        change_reason: Description of why the change was made
    """
    cursor.execute(
        """
        INSERT INTO audit_history (
            record_type, record_id, action, field_changed,
            old_value, new_value, changed_by, changed_date, change_reason
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (record_type, record_id, action, field_changed, old_value,
         new_value, changed_by, changed_date, change_reason)
    )


def get_access_manager() -> AccessManager:
    """Create AccessManager instance with configured DB path."""
    return AccessManager(db_path=DB_PATH)


def get_config_manager() -> ConfigurationManager:
    """Create ConfigurationManager instance with configured DB path."""
    return ConfigurationManager(db_path=DB_PATH)


def get_req_database() -> ReqClientProductDatabase:
    """Create ClientProductDatabase instance with configured DB path."""
    return ReqClientProductDatabase(db_path=REQ_DB_PATH)


# ============================================================
# USER MANAGEMENT TOOLS
# ============================================================

@mcp.tool()
def hello_propel() -> str:
    """
    Test function to verify MCP server is working.
    """
    logger.info("hello_propel() called - connection test")
    return "Hello from Propel Health MCP Server! Connection successful."


@mcp.tool()
def list_users(
    program: Optional[str] = None,
    status: Optional[str] = None,
    organization: Optional[str] = None,
    clinic: Optional[str] = None,
    location: Optional[str] = None
) -> str:
    """
    List users from the Propel Health database.

    Args:
        program: Filter by program name (e.g., "Prevention4ME")
        status: Filter by status ("Active", "Inactive", "Terminated")
        organization: Filter by organization name
        clinic: Filter by clinic name or code (e.g., "Franz", "FRANZ")
        location: Filter by location name or code (e.g., "Richland")

    Returns:
        Formatted list of users with their status and access count
    """
    logger.info(f"list_users() called - program={program}, status={status}, clinic={clinic}")
    manager = None
    try:
        manager = get_access_manager()

        # If clinic or location filter is provided, use get_access_by_scope
        if clinic or location:
            # Resolve program_id if provided
            program_id = None
            if program:
                try:
                    program_id = manager._resolve_program_id(program)
                except ValueError:
                    return f"Program not found: {program}"

            # Resolve clinic_id
            clinic_id = None
            if clinic:
                if not program_id:
                    # Need to find which program this clinic belongs to
                    conn = manager.conn
                    cursor = conn.execute(
                        "SELECT clinic_id, program_id FROM clinics WHERE name LIKE ? OR code LIKE ?",
                        (f"%{clinic}%", f"%{clinic.upper()}%")
                    )
                    row = cursor.fetchone()
                    if not row:
                        return f"Clinic not found: {clinic}"
                    clinic_id = row['clinic_id']
                    program_id = row['program_id']
                else:
                    try:
                        clinic_id = manager._resolve_clinic_id(clinic, program_id)
                    except ValueError:
                        return f"Clinic not found: {clinic}"

            # Resolve location_id
            location_id = None
            if location:
                if not clinic_id:
                    return "Location filter requires a clinic to be specified."
                try:
                    location_id = manager._resolve_location_id(location, clinic_id)
                except ValueError:
                    return f"Location not found: {location}"

            # Get users by scope
            access_list = manager.get_access_by_scope(
                program_id=program_id,
                clinic_id=clinic_id,
                location_id=location_id,
                active_only=True
            )

            # get_access_by_scope doesn't include user status, so fetch it separately
            if access_list:
                conn = manager.conn
                user_ids = list(set(a.get('user_id') for a in access_list if a.get('user_id')))
                if user_ids:
                    placeholders = ','.join(['?' for _ in user_ids])
                    cursor = conn.execute(
                        f"SELECT user_id, status FROM users WHERE user_id IN ({placeholders})",
                        user_ids
                    )
                    status_map = {row['user_id']: row['status'] for row in cursor.fetchall()}
                    for access in access_list:
                        access['status'] = status_map.get(access.get('user_id'))

            # Filter by status and organization if provided
            if status:
                access_list = [a for a in access_list if a.get('status') == status]
            if organization:
                access_list = [a for a in access_list
                              if organization.lower() in (a.get('organization') or '').lower()]

            if not access_list:
                return "No users found matching the criteria."

            # Deduplicate by user (a user might have multiple access grants)
            seen_users = {}
            for access in access_list:
                user_id = access.get('user_id')
                if user_id not in seen_users:
                    seen_users[user_id] = {
                        'name': access.get('user_name'),
                        'email': access.get('email'),
                        'status': access.get('status'),
                        'organization': access.get('organization'),
                        'roles': [access.get('role')],
                        'clinics': [access.get('clinic_name')] if access.get('clinic_name') else [],
                        'locations': [access.get('location_name')] if access.get('location_name') else []
                    }
                else:
                    if access.get('role') and access.get('role') not in seen_users[user_id]['roles']:
                        seen_users[user_id]['roles'].append(access.get('role'))
                    if access.get('clinic_name') and access.get('clinic_name') not in seen_users[user_id]['clinics']:
                        seen_users[user_id]['clinics'].append(access.get('clinic_name'))
                    if access.get('location_name') and access.get('location_name') not in seen_users[user_id]['locations']:
                        seen_users[user_id]['locations'].append(access.get('location_name'))

            users = list(seen_users.values())

            # Build filter description
            filters = []
            if program:
                filters.append(f"program={program}")
            if clinic:
                filters.append(f"clinic={clinic}")
            if location:
                filters.append(f"location={location}")
            if status:
                filters.append(f"status={status}")
            if organization:
                filters.append(f"org={organization}")
            filter_str = f" [{', '.join(filters)}]" if filters else ""

            result = f"Found {len(users)} user(s){filter_str}:\n\n"
            for user in users:
                result += f"• {user.get('name', 'Unknown')} ({user.get('email', 'No email')})\n"
                result += f"  Status: {user.get('status', 'Unknown')}"
                if user.get('organization'):
                    result += f" | Org: {user.get('organization')}"
                result += "\n"
                if user.get('roles'):
                    result += f"  Roles: {', '.join(user['roles'])}\n"
                if user.get('clinics'):
                    result += f"  Clinics: {', '.join(user['clinics'])}\n"

            return result

        else:
            # Use original list_users for non-clinic/location filters
            users = manager.list_users(
                program_filter=program,
                status_filter=status,
                organization_filter=organization,
                include_access_count=True
            )

            if not users:
                return "No users found matching the criteria."

            result = f"Found {len(users)} user(s):\n\n"
            for user in users:
                result += f"• {user.get('name', 'Unknown')} ({user.get('email', 'No email')})\n"
                result += f"  Status: {user.get('status', 'Unknown')}"
                if user.get('organization'):
                    result += f" | Org: {user.get('organization')}"
                if user.get('active_access_count') is not None:
                    result += f" | Access grants: {user.get('active_access_count')}"
                result += "\n"

            return result

    except Exception as e:
        logger.error(f"list_users() failed: {str(e)}", exc_info=True)
        return f"Error listing users: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def get_user(email: str) -> str:
    """
    Get detailed information about a specific user.

    Args:
        email: User's email address

    Returns:
        User details including status, organization, and access grants
    """
    manager = None
    try:
        manager = get_access_manager()
        user = manager.get_user(email=email)

        if not user:
            return f"User not found: {email}"

        # Get user's access grants
        access_list = manager.get_user_access(user['user_id'], active_only=False)

        # Get user's training status
        training = manager.get_training_status(user['user_id'])

        result = f"User: {user['name']}\n"
        result += f"Email: {user['email']}\n"
        result += f"Status: {user['status']}\n"
        result += f"Organization: {user.get('organization', 'N/A')}\n"
        result += f"Business Associate: {'Yes' if user.get('is_business_associate') else 'No'}\n"
        result += f"User ID: {user['user_id']}\n"

        if access_list:
            result += f"\nAccess Grants ({len(access_list)}):\n"
            for access in access_list:
                if access.get('is_active'):
                    result += f"  • {access.get('program_name', 'Unknown')} - {access.get('role')}"
                    if access.get('clinic_name'):
                        result += f" ({access.get('clinic_name')})"
                    result += "\n"

        if training:
            result += f"\nTraining Records ({len(training)}):\n"
            for t in training:
                result += f"  • {t.get('training_type')}: {t.get('status')}"
                if t.get('expires_date'):
                    result += f" (expires: {t.get('expires_date')})"
                result += "\n"

        return result

    except Exception as e:
        return f"Error getting user: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def add_user(
    name: str,
    email: str,
    organization: str = "Internal",
    is_business_associate: bool = False
) -> str:
    """
    Create a new user in the system.

    Args:
        name: Full name (e.g., "John Smith")
        email: Email address (must be unique)
        organization: Organization name (default: "Internal")
        is_business_associate: True if external party requiring HIPAA BAA

    Returns:
        Confirmation with new user's ID
    """
    manager = None
    try:
        manager = get_access_manager()

        # Check if user already exists
        existing = manager.get_user(email=email)
        if existing:
            return f"User already exists with email: {email}"

        user_id = manager.create_user(
            name=name,
            email=email,
            organization=organization,
            is_business_associate=is_business_associate
        )

        result = f"User created successfully!\n"
        result += f"  Name: {name}\n"
        result += f"  Email: {email}\n"
        result += f"  User ID: {user_id}\n"
        result += f"  Organization: {organization}\n"
        if is_business_associate:
            result += f"  Business Associate: Yes (HIPAA BAA required)\n"

        return result

    except Exception as e:
        return f"Error creating user: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def export_annual_review(
    program: Optional[str] = None,
    clinic: Optional[str] = None,
    location: Optional[str] = None,
    output_format: str = "csv",
    output_dir: Optional[str] = None
) -> str:
    """
    Generate an annual access review export for clinic managers.

    Creates a formatted spreadsheet for HIPAA annual access review compliance.
    Clinic managers can use this to review and confirm user access.

    Args:
        program: Filter by program name (e.g., "Prevention4ME")
        clinic: Filter by clinic name or code (e.g., "Franz", "FRANZ")
        location: Filter by location name or code (e.g., "Richland")
        output_format: Output format - "csv" or "xlsx" (default: "csv")
        output_dir: Directory to save the file (default: ~/Downloads)

    Returns:
        Path to the generated file and summary statistics
    """
    manager = None
    try:
        manager = get_access_manager()

        # Resolve filters
        program_id = None
        clinic_id = None
        location_id = None
        clinic_name = None

        if program:
            try:
                program_id = manager._resolve_program_id(program)
            except ValueError:
                return f"Program not found: {program}"

        if clinic:
            if not program_id:
                # Find which program this clinic belongs to
                conn = manager.conn
                cursor = conn.execute(
                    "SELECT clinic_id, program_id, name FROM clinics WHERE name LIKE ? OR code LIKE ?",
                    (f"%{clinic}%", f"%{clinic.upper()}%")
                )
                row = cursor.fetchone()
                if not row:
                    return f"Clinic not found: {clinic}"
                clinic_id = row['clinic_id']
                program_id = row['program_id']
                clinic_name = row['name']
            else:
                try:
                    clinic_id = manager._resolve_clinic_id(clinic, program_id)
                    # Get clinic name for file naming
                    conn = manager.conn
                    cursor = conn.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic_id,))
                    row = cursor.fetchone()
                    clinic_name = row['name'] if row else clinic
                except ValueError:
                    return f"Clinic not found: {clinic}"

        if location:
            if not clinic_id:
                return "Location filter requires a clinic to be specified."
            try:
                location_id = manager._resolve_location_id(location, clinic_id)
            except ValueError:
                return f"Location not found: {location}"

        # Must have at least a program or clinic filter
        if not program_id and not clinic_id:
            return "Please specify at least a program or clinic filter."

        # Get access data
        access_list = manager.get_access_by_scope(
            program_id=program_id,
            clinic_id=clinic_id,
            location_id=location_id,
            active_only=True
        )

        if not access_list:
            return "No users found matching the criteria."

        # get_access_by_scope doesn't include user status, so fetch it separately
        conn = manager.conn
        user_ids = list(set(a.get('user_id') for a in access_list if a.get('user_id')))
        if user_ids:
            placeholders = ','.join(['?' for _ in user_ids])
            cursor = conn.execute(
                f"SELECT user_id, status FROM users WHERE user_id IN ({placeholders})",
                user_ids
            )
            status_map = {row['user_id']: row['status'] for row in cursor.fetchall()}
            for access in access_list:
                access['status'] = status_map.get(access.get('user_id'))

        # Get program name for file
        program_name = None
        if program_id:
            conn = manager.conn
            cursor = conn.execute("SELECT name FROM programs WHERE program_id = ?", (program_id,))
            row = cursor.fetchone()
            program_name = row['name'] if row else "Unknown"

        # Prepare export data
        export_rows = []
        for access in access_list:
            # Get last review date if available
            last_review = None
            try:
                conn = manager.conn
                cursor = conn.execute("""
                    SELECT MAX(review_date) as last_review
                    FROM access_reviews
                    WHERE access_id = ?
                """, (access.get('access_id'),))
                row = cursor.fetchone()
                if row and row['last_review']:
                    last_review = row['last_review']
            except Exception:
                pass

            export_rows.append({
                'Name': access.get('user_name', ''),
                'Email': access.get('email', ''),
                'Status': access.get('status', ''),
                'Role': access.get('role', ''),
                'Program': access.get('program_name', ''),
                'Clinic': access.get('clinic_name', ''),
                'Location': access.get('location_name', ''),
                'Access Granted': access.get('granted_date', ''),
                'Last Review Date': last_review or '',
                'Next Review Due': access.get('next_review_due', ''),
                'Review Action': '',  # For manager to fill in: Keep/Remove/Modify
                'Notes': ''  # For manager comments
            })

        # Generate filename
        today = datetime.now().strftime('%Y-%m-%d')
        if clinic_name:
            safe_name = clinic_name.replace(' ', '_').replace('/', '-')
            filename = f"{safe_name}_access_review_{today}"
        elif program_name:
            safe_name = program_name.replace(' ', '_').replace('/', '-')
            filename = f"{safe_name}_access_review_{today}"
        else:
            filename = f"access_review_{today}"

        # Determine output directory
        if output_dir:
            out_path = os.path.expanduser(output_dir)
        else:
            out_path = os.path.expanduser("~/Downloads")

        os.makedirs(out_path, exist_ok=True)

        # Generate the file
        if output_format.lower() == "xlsx":
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill

                filepath = os.path.join(out_path, f"{filename}.xlsx")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Access Review"

                # Header row
                headers = list(export_rows[0].keys()) if export_rows else []
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Data rows
                for row_num, row_data in enumerate(export_rows, 2):
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col, value=row_data.get(header, ''))

                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

                # Add review action dropdown hint
                if export_rows:
                    ws.cell(row=1, column=len(headers) + 1, value="Review Actions: Keep / Remove / Modify")

                wb.save(filepath)

            except ImportError:
                return ("Excel export requires openpyxl. Install with: pip install openpyxl\n"
                        "Alternatively, use output_format='csv'")
        else:
            # CSV export
            filepath = os.path.join(out_path, f"{filename}.csv")
            headers = list(export_rows[0].keys()) if export_rows else []

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(export_rows)

        # Build summary
        result = f"Annual Access Review Export Generated\n"
        result += f"{'=' * 50}\n\n"
        result += f"File: {filepath}\n"
        result += f"Format: {output_format.upper()}\n"
        result += f"Generated: {today}\n\n"

        result += f"Scope:\n"
        if program_name:
            result += f"  Program: {program_name}\n"
        if clinic_name:
            result += f"  Clinic: {clinic_name}\n"
        if location:
            result += f"  Location: {location}\n"

        result += f"\nStatistics:\n"
        result += f"  Total Users: {len(export_rows)}\n"

        # Count by status
        status_counts = {}
        role_counts = {}
        for row in export_rows:
            status = row.get('Status', 'Unknown')
            role = row.get('Role', 'Unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
            role_counts[role] = role_counts.get(role, 0) + 1

        result += f"  By Status:\n"
        for status, count in status_counts.items():
            result += f"    • {status}: {count}\n"

        result += f"  By Role:\n"
        for role, count in role_counts.items():
            result += f"    • {role}: {count}\n"

        result += f"\nWorkflow:\n"
        result += f"  1. Share this file with the clinic manager\n"
        result += f"  2. Manager reviews each user and fills in 'Review Action' column\n"
        result += f"  3. Collect marked-up spreadsheet and process changes\n"

        return result

    except Exception as e:
        return f"Error generating export: {str(e)}"
    finally:
        if manager:
            manager.close()


# ============================================================
# ACCESS MANAGEMENT TOOLS
# ============================================================

@mcp.tool()
def list_access(
    user_email: Optional[str] = None,
    program: Optional[str] = None
) -> str:
    """
    List access grants, optionally filtered by user or program.

    Args:
        user_email: Filter by user's email address
        program: Filter by program name

    Returns:
        List of access grants with roles and scope
    """
    manager = None
    try:
        manager = get_access_manager()

        user_id = None
        user_info = None
        if user_email:
            user = manager.get_user(email=user_email)
            if not user:
                return f"User not found: {user_email}"
            user_id = user['user_id']
            user_info = user  # Store for display

        # Get access based on filters provided
        if user_id:
            # User-specific access
            access_list = manager.get_user_access(user_id, active_only=True)
            # Filter by program if provided
            if program:
                try:
                    program_id = manager._resolve_program_id(program)
                    access_list = [a for a in access_list if a.get('program_id') == program_id]
                except ValueError:
                    return f"Program not found: {program}"
        elif program:
            # Scope-specific access
            try:
                program_id = manager._resolve_program_id(program)
                access_list = manager.get_access_by_scope(program_id=program_id, active_only=True)
            except ValueError:
                return f"Program not found: {program}"
        else:
            # No filters - get all access (scope-based with no filter)
            access_list = manager.get_access_by_scope(active_only=True)

        if not access_list:
            return "No access grants found matching the criteria."

        # Format results
        result = f"Found {len(access_list)} access grant(s):\n\n"

        # If user-specific query, show user header once
        if user_info:
            result += f"User: {user_info['name']} ({user_info['email']})\n\n"

        for access in access_list:
            status = "Active" if access.get('is_active') else "Revoked"
            # For scope-based queries, show user per access grant
            if not user_info:
                result += f"• {access.get('user_name', 'Unknown')} ({access.get('email', 'N/A')})\n"
            result += f"  Program: {access.get('program_name')} | Role: {access.get('role')} | Status: {status}\n"
            if access.get('clinic_name'):
                result += f"  Scope: {access.get('clinic_name')}"
                if access.get('location_name'):
                    result += f" > {access.get('location_name')}"
                result += "\n"
            if access.get('next_review_due'):
                result += f"  Next review: {access.get('next_review_due')}\n"
            result += "\n"

        return result

    except Exception as e:
        return f"Error listing access: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def get_reviews_due(program: Optional[str] = None) -> str:
    """
    Get access reviews that are overdue or due soon.

    Args:
        program: Filter by program name (optional)

    Returns:
        List of access grants needing review
    """
    manager = None
    try:
        manager = get_access_manager()

        program_id = None
        if program:
            try:
                program_id = manager._resolve_program_id(program)
            except ValueError:
                return f"Program not found: {program}"

        reviews = manager.get_reviews_due(program_id=program_id)

        if not reviews:
            return "No reviews are currently due. All access reviews are up to date!"

        overdue = [r for r in reviews if r.get('is_overdue')]
        due_soon = [r for r in reviews if not r.get('is_overdue')]

        result = f"Access Reviews Due: {len(reviews)} total\n\n"

        if overdue:
            result += f"OVERDUE ({len(overdue)}):\n"
            for r in overdue:
                result += f"  • {r.get('user_name')} - {r.get('program_name')} ({r.get('role')})\n"
                result += f"    Due: {r.get('next_review_due')} | Days overdue: {r.get('days_overdue')}\n"
            result += "\n"

        if due_soon:
            result += f"Due Soon ({len(due_soon)}):\n"
            for r in due_soon:
                result += f"  • {r.get('user_name')} - {r.get('program_name')} ({r.get('role')})\n"
                result += f"    Due: {r.get('next_review_due')}\n"

        return result

    except Exception as e:
        return f"Error getting reviews: {str(e)}"
    finally:
        if manager:
            manager.close()


# ============================================================
# TRAINING MANAGEMENT TOOLS
# ============================================================

@mcp.tool()
def get_training_status(user_email: str) -> str:
    """
    Get training status for a specific user.

    Args:
        user_email: User's email address

    Returns:
        List of training records with completion status
    """
    manager = None
    try:
        manager = get_access_manager()

        user = manager.get_user(email=user_email)
        if not user:
            return f"User not found: {user_email}"

        training = manager.get_training_status(user['user_id'])

        if not training:
            return f"No training records found for {user['name']}."

        result = f"Training Status for {user['name']}:\n\n"

        # Group by status
        current = [t for t in training if t.get('status') == 'Current']
        pending = [t for t in training if t.get('status') == 'Pending']
        expired = [t for t in training if t.get('status') == 'Expired']

        if current:
            result += f"Current ({len(current)}):\n"
            for t in current:
                result += f"  • {t.get('training_type')}"
                if t.get('expires_date'):
                    result += f" (expires: {t.get('expires_date')})"
                result += "\n"
            result += "\n"

        if pending:
            result += f"Pending ({len(pending)}):\n"
            for t in pending:
                result += f"  • {t.get('training_type')} (assigned: {t.get('assigned_date')})\n"
            result += "\n"

        if expired:
            result += f"Expired ({len(expired)}):\n"
            for t in expired:
                result += f"  • {t.get('training_type')} (expired: {t.get('expires_date')})\n"

        return result

    except Exception as e:
        return f"Error getting training status: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def get_expired_training() -> str:
    """
    Get all users with expired training that needs renewal.

    Returns:
        List of users and their expired training records
    """
    manager = None
    try:
        manager = get_access_manager()
        expired = manager.get_expired_training()

        if not expired:
            return "No expired training found. All training is current!"

        result = f"Expired Training: {len(expired)} record(s)\n\n"

        for record in expired:
            result += f"• {record.get('user_name')} ({record.get('user_email')})\n"
            result += f"  Training: {record.get('training_type')}\n"
            result += f"  Expired: {record.get('expires_date')}\n\n"

        return result

    except Exception as e:
        return f"Error getting expired training: {str(e)}"
    finally:
        if manager:
            manager.close()


# ============================================================
# COMPLIANCE REPORTING TOOLS
# ============================================================

@mcp.tool()
def get_compliance_report(
    report_type: str,
    program: Optional[str] = None
) -> str:
    """
    Generate a compliance report.

    Args:
        report_type: Type of report:
            - "access_list": Who has access to what
            - "review_status": Are access reviews current?
            - "training_compliance": Training completion status
            - "terminated_audit": Check for terminated users with access
            - "business_associates": List of business associates
        program: Filter by program name (optional)

    Returns:
        Formatted compliance report
    """
    valid_reports = [
        'access_list', 'review_status', 'training_compliance',
        'terminated_audit', 'business_associates'
    ]

    if report_type not in valid_reports:
        return f"Invalid report type. Choose from: {', '.join(valid_reports)}"

    manager = None
    try:
        manager = get_access_manager()
        reports = ComplianceReports(manager)

        program_id = None
        if program:
            try:
                program_id = manager._resolve_program_id(program)
            except ValueError:
                return f"Program not found: {program}"

        result = ""

        if report_type == 'access_list':
            data = reports.access_list_report(program_id=program_id)
            result = f"Access List Report\n"
            result += f"==================\n\n"
            result += f"Total Users: {data['summary']['total_users']}\n"
            result += f"Total Access Grants: {data['summary']['total_access_grants']}\n\n"

            # Group by user for cleaner display
            users_seen = set()
            for access in data.get('access_list', [])[:50]:  # Limit to 50 entries
                user_key = access.get('user_id')
                if user_key not in users_seen:
                    result += f"• {access.get('user_name')} ({access.get('email')})\n"
                    users_seen.add(user_key)
                result += f"  - {access.get('program_name')}: {access.get('role')}\n"

            if len(data.get('access_list', [])) > 50:
                result += f"\n... and {len(data['access_list']) - 50} more access grants\n"

        elif report_type == 'review_status':
            data = reports.review_status_report(program_id=program_id)
            result = f"Access Review Status\n"
            result += f"====================\n\n"
            result += f"Current: {data['summary']['current']}\n"
            result += f"Due Soon: {data['summary']['due_soon']}\n"
            result += f"Overdue: {data['summary']['overdue']}\n"

            if data['summary']['overdue'] > 0:
                result += f"\nAction Required: {data['summary']['overdue']} overdue reviews\n"

        elif report_type == 'training_compliance':
            data = reports.training_compliance_report()
            result = f"Training Compliance Report\n"
            result += f"==========================\n\n"
            result += f"Total Users: {data['summary']['total_users']}\n"
            result += f"Fully Compliant: {data['summary']['compliant']}\n"
            result += f"Missing Training: {data['summary']['missing_training']}\n"
            result += f"Expired Training: {data['summary']['expired_training']}\n"

        elif report_type == 'terminated_audit':
            data = reports.terminated_user_audit()
            result = f"Terminated User Audit\n"
            result += f"=====================\n\n"

            if not data.get('violations'):
                result += "No issues found. All terminated users have had access revoked.\n"
            else:
                result += f"VIOLATIONS FOUND: {len(data['violations'])}\n\n"
                for v in data['violations']:
                    result += f"• {v.get('user_name')} - still has {v.get('active_grants')} active grant(s)\n"

        elif report_type == 'business_associates':
            data = reports.business_associate_report()
            result = f"Business Associates Report\n"
            result += f"==========================\n\n"
            result += f"Total External Users: {data['summary']['total_external_users']}\n"
            result += f"Organizations: {data['summary']['organizations']}\n"
            result += f"Total Access Grants: {data['summary']['total_access_grants']}\n\n"

            for ba in data.get('all_external_users', []):
                result += f"• {ba.get('name')} ({ba.get('organization')})\n"
                result += f"  Email: {ba.get('email')} | Programs: {ba.get('programs')}\n\n"

        return result

    except Exception as e:
        return f"Error generating report: {str(e)}"
    finally:
        if manager:
            manager.close()


# ============================================================
# CONFIGURATION MANAGEMENT TOOLS
# ============================================================

@mcp.tool()
def list_programs() -> str:
    """
    List all programs with their clinic and location hierarchy.

    Returns:
        Formatted hierarchy of programs, clinics, and locations
    """
    cm = None
    try:
        cm = get_config_manager()
        programs = cm.list_programs()

        if not programs:
            return "No programs found in the database."

        result = "Programs:\n"
        result += "=========\n\n"

        for program in programs:
            result += f"[{program.get('prefix')}] {program.get('name')}\n"
            result += f"   Type: {program.get('program_type')} | Status: {program.get('status')}\n"

            clinics = program.get('clinics', [])
            if clinics:
                for clinic in clinics:
                    result += f"   +-- {clinic.get('name')}"
                    if clinic.get('code'):
                        result += f" [{clinic.get('code')}]"
                    result += "\n"

                    locations = clinic.get('locations', [])
                    for loc in locations:
                        result += f"       +-- {loc.get('name')}"
                        if loc.get('code'):
                            result += f" [{loc.get('code')}]"
                        result += "\n"

            result += "\n"

        return result

    except Exception as e:
        return f"Error listing programs: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def get_config(
    config_key: str,
    program: str,
    clinic: Optional[str] = None,
    location: Optional[str] = None
) -> str:
    """
    Get a configuration value with inheritance information.

    Args:
        config_key: Configuration key (e.g., "helpdesk_phone", "hours_open")
        program: Program prefix (e.g., "P4M")
        clinic: Clinic name (optional)
        location: Location name (optional, requires clinic)

    Returns:
        Configuration value with source and inheritance chain
    """
    cm = None
    try:
        cm = get_config_manager()
        im = InheritanceManager(cm)

        # Resolve IDs
        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        clinic_id = None
        if clinic:
            clinic_id = cm.get_clinic_id(program_id, clinic)
            if not clinic_id:
                return f"Clinic not found: {clinic}"

        location_id = None
        if location:
            if not clinic_id:
                return "Location requires clinic to be specified"
            location_id = cm.get_location_id(clinic_id, location)
            if not location_id:
                return f"Location not found: {location}"

        # Get config with inheritance
        config = im.resolve_with_inheritance(
            config_key, program_id, clinic_id, location_id
        )

        if not config:
            return f"Configuration not found: {config_key}"

        result = f"Configuration: {config_key}\n"
        result += f"{'=' * 40}\n\n"
        result += f"Effective Value: {config.get('value', '(not set)')}\n"
        result += f"Set At: {config.get('effective_level', 'default')}\n"
        result += f"Is Override: {'Yes' if config.get('is_override') else 'No'}\n"

        chain = config.get('inheritance_chain', [])
        if chain:
            result += f"\nInheritance Chain:\n"
            for level in chain:
                marker = ">" if level.get('is_effective') else " "
                override = "*" if level.get('is_override') else ""
                result += f"  {marker} {level.get('level')}: {level.get('value', '(inherited)')}{override}\n"

        return result

    except Exception as e:
        return f"Error getting config: {str(e)}"
    finally:
        if cm:
            cm.close()


# ============================================================
# REQUIREMENTS TOOLKIT - CLIENT/PROGRAM TOOLS
# ============================================================

@mcp.tool()
def list_clients(status: Optional[str] = "Active") -> str:
    """
    List all clients in the requirements database.

    Args:
        status: Filter by status ("Active", "Inactive", or None for all)

    Returns:
        Formatted list of clients
    """
    db = None
    try:
        db = get_req_database()
        clients = db.list_clients(status_filter=status)

        if not clients:
            return "No clients found."

        result = f"Found {len(clients)} client(s):\n\n"
        for client in clients:
            result += f"• {client['name']}\n"
            result += f"  ID: {client['client_id']} | Status: {client['status']}\n"
            if client.get('description'):
                result += f"  Description: {client['description']}\n"
            if client.get('primary_contact'):
                result += f"  Contact: {client['primary_contact']}"
                if client.get('contact_email'):
                    result += f" ({client['contact_email']})"
                result += "\n"
            result += "\n"

        return result

    except Exception as e:
        return f"Error listing clients: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def get_client_programs(client_name: str) -> str:
    """
    Get all programs for a specific client.

    Args:
        client_name: Name of the client

    Returns:
        List of programs with story/test counts
    """
    db = None
    try:
        db = get_req_database()
        client = db.get_client_by_name(client_name)
        if not client:
            return f"Client not found: {client_name}"

        programs = db.list_programs(client_id=client['client_id'], status_filter=None)

        if not programs:
            return f"No programs found for {client_name}."

        result = f"Programs for {client_name}:\n\n"
        for prog in programs:
            summary = db.get_program_summary(prog['program_id'])
            result += f"[{prog['prefix']}] {prog['name']}\n"
            result += f"  Status: {prog['status']}"
            if prog.get('program_type'):
                result += f" | Type: {prog['program_type']}"
            result += "\n"
            result += f"  Stories: {summary['story_count']} | Tests: {summary['test_count']} | Requirements: {summary['requirement_count']}\n"
            if summary.get('stories_by_status'):
                approved = summary['stories_by_status'].get('Approved', 0)
                result += f"  Approved stories: {approved}\n"
            result += "\n"

        return result

    except Exception as e:
        return f"Error getting programs: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def get_program_by_prefix(prefix: str) -> str:
    """
    Get detailed information about a program by its prefix.

    Args:
        prefix: Program prefix (e.g., "PROP", "GRX")

    Returns:
        Program details with summary statistics
    """
    db = None
    try:
        db = get_req_database()
        program = db.get_program_by_prefix(prefix.upper())

        if not program:
            return f"Program not found with prefix: {prefix}"

        summary = db.get_program_summary(program['program_id'])

        result = f"Program: {program['name']} [{program['prefix']}]\n"
        result += f"{'=' * 50}\n\n"
        result += f"ID: {program['program_id']}\n"
        result += f"Status: {program['status']}\n"
        if program.get('program_type'):
            result += f"Type: {program['program_type']}\n"
        if program.get('description'):
            result += f"Description: {program['description']}\n"
        if program.get('source_file'):
            result += f"Source: {program['source_file']}\n"

        result += f"\nStatistics:\n"
        result += f"  Requirements: {summary['requirement_count']}\n"
        result += f"  User Stories: {summary['story_count']}\n"
        result += f"  Test Cases: {summary['test_count']}\n"

        if summary.get('stories_by_status'):
            result += f"\nStories by Status:\n"
            for status, count in summary['stories_by_status'].items():
                result += f"  • {status}: {count}\n"

        if summary.get('tests_by_status'):
            result += f"\nTests by Status:\n"
            for status, count in summary['tests_by_status'].items():
                result += f"  • {status}: {count}\n"

        if summary.get('coverage'):
            cov = summary['coverage']
            result += f"\nCoverage:\n"
            result += f"  Full: {cov.get('full_pct', 0)}% | Partial: {cov.get('partial_pct', 0)}% | None: {cov.get('none_pct', 0)}%\n"

        return result

    except Exception as e:
        return f"Error getting program: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def create_client(
    client_name: str,
    status: str = "Active",
    short_name: Optional[str] = None,
    client_type: Optional[str] = None,
    description: Optional[str] = None,
    primary_contact_name: Optional[str] = None,
    primary_contact_email: Optional[str] = None,
    primary_contact_phone: Optional[str] = None,
    contract_reference: Optional[str] = None,
    contract_start_date: Optional[str] = None,
    contract_end_date: Optional[str] = None,
    source_document: Optional[str] = None
) -> str:
    """
    Create a new client organization in the requirements database.

    PURPOSE:
        Creates a top-level client record. Clients are the root of the hierarchy:
        Client → Program → Story → Test Case. Each client can have multiple
        programs, and all requirements traceability flows from the client.

    R EQUIVALENT:
        Like adding a row to a clients data.frame, with auto-generated UUID
        and schema validation. Similar to tibble::add_row() with constraints.

    Args:
        client_name: Full client name (e.g., "Providence Health & Services", "Propel Health")
        status: "Active" or "Inactive" (default: "Active")
        short_name: Abbreviated name for display (e.g., "Providence", "Propel")
        client_type: "External" (customer) or "Internal" (Propel Health internal)
        description: Description of the client or relationship
        primary_contact_name: Main point of contact
        primary_contact_email: Contact email
        primary_contact_phone: Contact phone
        contract_reference: Contract or agreement reference (e.g., "Providence PSA 2026")
        contract_start_date: Contract start date (YYYY-MM-DD format)
        contract_end_date: Contract end date (YYYY-MM-DD format)
        source_document: Source document for requirements (e.g., "Exhibit E - KPIs")

    Returns:
        Confirmation with client_id, name, and next steps

    WHY THIS APPROACH:
        - Case-insensitive uniqueness check prevents duplicate clients
        - UUID generation ensures globally unique identifiers
        - Self-healing schema: adds missing columns if they don't exist
        - Comprehensive audit logging for Part 11 compliance
        - Stores all metadata needed for contract and contact tracking

    Example (minimal - internal):
        create_client(
            client_name="Propel Health",
            client_type="Internal",
            description="Platform-wide features and internal capabilities"
        )

    Example (full - external customer):
        create_client(
            client_name="Providence Health & Services",
            short_name="Providence",
            client_type="External",
            description="Health system partner - breast cancer prevention programs",
            primary_contact_name="Kim Smith",
            primary_contact_email="kim.smith@providence.org",
            contract_reference="Providence PSA 2026",
            contract_start_date="2026-01-01",
            contract_end_date="2026-12-31",
            source_document="Exhibit E - Key Performance Indicators"
        )
    """
    import sqlite3
    import uuid
    import json
    import re

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate status
        # ----------------------------------------------------------------
        status_descriptions = {
            "Active": "Client is currently active",
            "Inactive": "Client relationship is paused or ended"
        }
        if error := validate_choice(status, list(status_descriptions.keys()), "status", status_descriptions):
            return error

        # ----------------------------------------------------------------
        # STEP 2: Validate client_type if provided
        # ----------------------------------------------------------------
        client_type_descriptions = {
            "External": "Customer organization",
            "Internal": "Propel Health internal project"
        }
        if error := validate_optional_choice(client_type, list(client_type_descriptions.keys()),
                                              "client_type", client_type_descriptions):
            return error

        # ----------------------------------------------------------------
        # STEP 3: Validate date formats if provided
        # ----------------------------------------------------------------
        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')

        if contract_start_date is not None:
            if not date_pattern.match(contract_start_date):
                return (
                    f"Invalid contract_start_date format: '{contract_start_date}'\n\n"
                    f"Please use YYYY-MM-DD format (e.g., '2026-01-01')"
                )

        if contract_end_date is not None:
            if not date_pattern.match(contract_end_date):
                return (
                    f"Invalid contract_end_date format: '{contract_end_date}'\n\n"
                    f"Please use YYYY-MM-DD format (e.g., '2026-12-31')"
                )

        # ----------------------------------------------------------------
        # STEP 4: Validate email format if provided
        # ----------------------------------------------------------------
        if primary_contact_email is not None:
            email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
            if not email_pattern.match(primary_contact_email):
                return (
                    f"Invalid email format: '{primary_contact_email}'\n\n"
                    f"Please provide a valid email address."
                )

        # ----------------------------------------------------------------
        # STEP 5: Connect to database
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # ----------------------------------------------------------------
        # STEP 6: Self-healing schema - add missing columns if needed
        # This ensures the tool works even if the schema hasn't been updated
        # ----------------------------------------------------------------
        additional_columns = [
            ("short_name", "TEXT"),
            ("client_type", "TEXT"),
            ("primary_contact_name", "TEXT"),
            ("primary_contact_email", "TEXT"),
            ("primary_contact_phone", "TEXT"),
            ("contract_reference", "TEXT"),
            ("contract_start_date", "TEXT"),
            ("contract_end_date", "TEXT"),
            ("source_document", "TEXT")
        ]

        # Get existing columns
        cursor.execute("PRAGMA table_info(clients)")
        existing_columns = {row['name'] for row in cursor.fetchall()}

        # Add missing columns
        for col_name, col_type in additional_columns:
            if col_name not in existing_columns:
                cursor.execute(f"ALTER TABLE clients ADD COLUMN {col_name} {col_type}")

        # ----------------------------------------------------------------
        # STEP 7: Check for duplicate client name (case-insensitive)
        # ----------------------------------------------------------------
        cursor.execute(
            "SELECT client_id, name FROM clients WHERE LOWER(name) = LOWER(?)",
            (client_name,)
        )
        existing = cursor.fetchone()

        if existing:
            return (
                f"Client already exists: '{existing['name']}'\n\n"
                f"Client ID: {existing['client_id']}\n\n"
                f"If you need to update this client, use update_client() instead.\n"
                f"If this is a different client, please use a unique name."
            )

        # ----------------------------------------------------------------
        # STEP 8: Generate client_id and prepare data
        # ----------------------------------------------------------------
        client_id = str(uuid.uuid4())
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ----------------------------------------------------------------
        # STEP 9: Insert the client record
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO clients (
                client_id, name, short_name, client_type, description, status,
                primary_contact_name, primary_contact_email, primary_contact_phone,
                contract_reference, contract_start_date, contract_end_date,
                source_document, created_date, updated_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_id,
                client_name,
                short_name,
                client_type,
                description,
                status,
                primary_contact_name,
                primary_contact_email,
                primary_contact_phone,
                contract_reference,
                contract_start_date,
                contract_end_date,
                source_document,
                now,
                now
            )
        )

        # ----------------------------------------------------------------
        # STEP 10: Log to audit_history for Part 11 compliance
        # ----------------------------------------------------------------
        # Create a JSON snapshot of all provided values for the audit trail
        audit_data = {
            'client_id': client_id,
            'name': client_name,
            'short_name': short_name,
            'client_type': client_type,
            'description': description,
            'status': status,
            'primary_contact_name': primary_contact_name,
            'primary_contact_email': primary_contact_email,
            'primary_contact_phone': primary_contact_phone,
            'contract_reference': contract_reference,
            'contract_start_date': contract_start_date,
            'contract_end_date': contract_end_date,
            'source_document': source_document
        }
        # Remove None values for cleaner audit log
        audit_data = {k: v for k, v in audit_data.items() if v is not None}

        log_audit(
            cursor,
            record_type='client',
            record_id=client_id,
            action='Created',
            field_changed='client',
            old_value=None,
            new_value=json.dumps(audit_data),
            changed_by='MCP:create_client',
            changed_date=now,
            change_reason=f'New client created: {client_name}'
        )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 11: Build success response
        # ----------------------------------------------------------------
        display_name = short_name or client_name
        result = f"""Client created successfully!

Client Details:
  Client ID: {client_id}
  Name: {client_name}
"""
        if short_name:
            result += f"  Short Name: {short_name}\n"
        if client_type:
            result += f"  Type: {client_type}\n"
        result += f"  Status: {status}\n"
        if description:
            result += f"  Description: {description[:80]}{'...' if len(description) > 80 else ''}\n"

        if primary_contact_name or primary_contact_email:
            result += "\nContact Information:\n"
            if primary_contact_name:
                result += f"  Contact: {primary_contact_name}\n"
            if primary_contact_email:
                result += f"  Email: {primary_contact_email}\n"
            if primary_contact_phone:
                result += f"  Phone: {primary_contact_phone}\n"

        if contract_reference or contract_start_date:
            result += "\nContract Information:\n"
            if contract_reference:
                result += f"  Reference: {contract_reference}\n"
            if contract_start_date:
                result += f"  Start Date: {contract_start_date}\n"
            if contract_end_date:
                result += f"  End Date: {contract_end_date}\n"
            if source_document:
                result += f"  Source Document: {source_document}\n"

        result += f"""
Created At: {now}

Next Steps:
  • create_program(client_name="{display_name}", prefix="XXX", program_name="...") - Add a program
  • get_client_tree() - View client/program hierarchy
"""
        return result

    except Exception as e:
        return f"Error creating client: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def create_program(
    client_name: str,
    prefix: str,
    program_name: str,
    description: Optional[str] = None,
    status: str = "Active"
) -> str:
    """
    Create a new program for tracking user stories, acceptance criteria, and test cases.

    PURPOSE:
        Creates a program under an existing client. Programs group related user stories
        and test cases. Each program has a unique prefix used to generate story IDs
        (e.g., prefix "P4M" creates stories like P4M-AUTH-001).

    R EQUIVALENT:
        Like creating a new project folder in RStudio, but with database tracking
        and unique ID generation.

    Args:
        client_name: Parent client name (must exist, case-insensitive match)
        prefix: Short prefix for story IDs (2-6 alphanumeric chars, auto-converted to uppercase)
        program_name: Full program name (e.g., "Prevention4ME")
        description: Optional program description
        status: Program status (default: "Active")

    Returns:
        Confirmation with program ID and example story IDs, or helpful error message

    WHY THIS APPROACH:
        - Validates inputs thoroughly to prevent data integrity issues
        - Case-insensitive client lookup for user convenience
        - Prefix uniqueness across ALL programs prevents story ID collisions
        - Audit logging for FDA 21 CFR Part 11 compliance
    """
    # Import here to keep dependency local (matches pattern in other tools)
    import sqlite3
    import uuid
    import re

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate prefix format
        # Prefix must be 2-6 alphanumeric characters (letters and numbers only)
        # Auto-convert to uppercase for consistency
        # ----------------------------------------------------------------
        prefix_upper = prefix.upper().strip()

        # Check length
        if len(prefix_upper) < 2 or len(prefix_upper) > 6:
            return (
                f"Invalid prefix length: '{prefix}'\n\n"
                f"Prefix must be 2-6 characters. Examples:\n"
                f"  • P4M (3 chars) - good\n"
                f"  • DISCOVER (8 chars) - too long, try 'DISC'\n"
                f"  • X (1 char) - too short, try 'XX' or longer"
            )

        # Check alphanumeric only (no special characters, spaces, or dashes)
        if not re.match(r'^[A-Z0-9]+$', prefix_upper):
            return (
                f"Invalid prefix format: '{prefix}'\n\n"
                f"Prefix must contain only letters and numbers (no spaces, dashes, or special characters).\n"
                f"Examples:\n"
                f"  • P4M - good\n"
                f"  • P-4M - bad (contains dash)\n"
                f"  • P 4M - bad (contains space)"
            )

        # ----------------------------------------------------------------
        # STEP 2: Connect to requirements database
        # Using direct sqlite3 connection (not the db manager class) for
        # more control over the transaction and audit logging
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row  # Allows dict-style access to rows
        cursor = conn.cursor()

        # ----------------------------------------------------------------
        # STEP 3: Validate client exists (case-insensitive search)
        # Show available clients if not found to help the user
        # ----------------------------------------------------------------
        cursor.execute(
            "SELECT client_id, name FROM clients WHERE LOWER(name) = LOWER(?)",
            (client_name,)
        )
        client_row = cursor.fetchone()

        if not client_row:
            # Client not found - show available clients to help user
            cursor.execute("SELECT name FROM clients WHERE status = 'Active' ORDER BY name")
            available = cursor.fetchall()

            if available:
                client_list = "\n".join(f"  • {row['name']}" for row in available)
                return (
                    f"Client not found: '{client_name}'\n\n"
                    f"Available clients:\n{client_list}\n\n"
                    f"Hint: Client names are matched case-insensitively."
                )
            else:
                return (
                    f"Client not found: '{client_name}'\n\n"
                    f"No active clients exist in the database. "
                    f"Please create a client first."
                )

        client_id = client_row['client_id']
        client_name_actual = client_row['name']  # Use actual casing from database

        # ----------------------------------------------------------------
        # STEP 4: Check prefix uniqueness across ALL programs
        # This is critical - prevents story ID collisions like P4M-AUTH-001
        # appearing in multiple programs
        # ----------------------------------------------------------------
        cursor.execute(
            "SELECT name, prefix FROM programs WHERE UPPER(prefix) = ?",
            (prefix_upper,)
        )
        existing_prefix = cursor.fetchone()

        if existing_prefix:
            return (
                f"Prefix already in use: '{prefix_upper}'\n\n"
                f"The prefix '{existing_prefix['prefix']}' is already assigned to "
                f"program '{existing_prefix['name']}'.\n\n"
                f"Each prefix must be unique across ALL programs to prevent story ID collisions.\n"
                f"Suggestion: Try a variation like '{prefix_upper}2' or choose a different abbreviation."
            )

        # ----------------------------------------------------------------
        # STEP 5: Check program name uniqueness within the client
        # Same client can't have two programs with identical names
        # ----------------------------------------------------------------
        cursor.execute(
            "SELECT name FROM programs WHERE client_id = ? AND LOWER(name) = LOWER(?)",
            (client_id, program_name)
        )
        existing_name = cursor.fetchone()

        if existing_name:
            return (
                f"Program name already exists for this client: '{existing_name['name']}'\n\n"
                f"Client '{client_name_actual}' already has a program with this name.\n"
                f"Please choose a different program name."
            )

        # ----------------------------------------------------------------
        # STEP 6: Generate program_id and insert
        # Using UUID4 for globally unique identifier
        # ----------------------------------------------------------------
        program_id = str(uuid.uuid4())
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute(
            """
            INSERT INTO programs (
                program_id, client_id, name, prefix, description, status,
                created_date, updated_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (program_id, client_id, program_name, prefix_upper, description, status, now, now)
        )

        # ----------------------------------------------------------------
        # STEP 7: Log to audit_history for Part 11 compliance
        # This creates a permanent record of who created what and when
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                'program',                          # record_type
                program_id,                         # record_id
                'Created',                          # action
                'program',                          # field_changed
                None,                               # old_value (none for create)
                f'{program_name} [{prefix_upper}]', # new_value
                'MCP:create_program',               # changed_by - identifies the tool
                now,                                # changed_date
                f'New program created under client {client_name_actual}'  # change_reason
            )
        )

        # Commit both inserts together (atomic transaction)
        conn.commit()

        # ----------------------------------------------------------------
        # STEP 8: Build success response with next steps
        # Show example story IDs so user understands the prefix usage
        # ----------------------------------------------------------------
        result = f"""Program created successfully!

Program Details:
  Name: {program_name}
  Prefix: {prefix_upper}
  Client: {client_name_actual}
  Status: {status}
  Program ID: {program_id}
"""
        if description:
            result += f"  Description: {description}\n"

        result += f"""
Next Steps - Example Story IDs:
  When you create user stories for this program, they'll be named like:
  • {prefix_upper}-AUTH-001 (Authentication story)
  • {prefix_upper}-DASH-001 (Dashboard story)
  • {prefix_upper}-DATA-001 (Data management story)

Use these commands to add content:
  • list_stories(program_prefix="{prefix_upper}") - View all stories
  • get_program_by_prefix(prefix="{prefix_upper}") - View program stats
"""
        return result

    except Exception as e:
        return f"Error creating program: {str(e)}"

    finally:
        # Always close the connection, even if an error occurred
        if conn:
            conn.close()


@mcp.tool()
def create_requirement(
    program_prefix: str,
    title: str,
    description: str,
    source_file: Optional[str] = None,
    priority: str = "Medium",
    requirement_type: str = "Technical"
) -> str:
    """
    Create a raw requirement entry that can later be converted to user stories.

    PURPOSE:
        Captures raw requirements from various sources (meetings, documents,
        stakeholder requests) before they're refined into user stories.
        This preserves the original requirement for traceability.

    R EQUIVALENT:
        Like adding a row to a requirements data.frame, with auto-generated
        IDs and source tracking.

    REQUIREMENT ID FORMAT:
        Auto-generated as {PREFIX}-REQ-{NNN}
        Example: P4M-REQ-001, P4M-REQ-002

    Args:
        program_prefix: Program prefix (e.g., "P4M", "PROP")
        title: Short requirement title
        description: Full requirement description
        source_file: Where this came from (e.g., "Client Meeting 2025-01-06", "SRS v1.2")
        priority: "High", "Medium", "Low" (default: "Medium")
        requirement_type: "Technical", "Workflow", "Process", "Integration" (default: "Technical")

    Returns:
        Confirmation with generated requirement_id

    WHY THIS APPROACH:
        - Preserves raw requirements before story conversion
        - Tracks source for traceability
        - Auto-increments requirement number within program
        - Audit logging for Part 11 compliance
    """
    import sqlite3
    import uuid
    import re

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate priority
        # ----------------------------------------------------------------
        valid_priorities = ["High", "Medium", "Low"]
        if priority not in valid_priorities:
            return (
                f"Invalid priority: '{priority}'\n\n"
                f"Valid priorities:\n"
                f"  • High - Critical, must be addressed first\n"
                f"  • Medium - Important, normal priority\n"
                f"  • Low - Nice to have, can be deferred"
            )

        # ----------------------------------------------------------------
        # STEP 2: Validate requirement_type
        # ----------------------------------------------------------------
        valid_types = ["Technical", "Workflow", "Process", "Integration"]
        if requirement_type not in valid_types:
            type_list = "\n".join(f"  • {t}" for t in valid_types)
            return (
                f"Invalid requirement_type: '{requirement_type}'\n\n"
                f"Valid types:\n{type_list}\n\n"
                f"Descriptions:\n"
                f"  • Technical - System/software functionality\n"
                f"  • Workflow - Business process steps\n"
                f"  • Process - Operational procedures\n"
                f"  • Integration - External system connections"
            )

        # ----------------------------------------------------------------
        # STEP 3: Connect to database and validate program exists
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            "SELECT program_id, name, prefix FROM programs WHERE UPPER(prefix) = ?",
            (program_prefix.upper(),)
        )
        program_row = cursor.fetchone()

        if not program_row:
            # Program not found - show available programs
            cursor.execute(
                "SELECT prefix, name FROM programs WHERE status = 'Active' ORDER BY prefix"
            )
            available = cursor.fetchall()

            if available:
                prog_list = "\n".join(f"  • [{row['prefix']}] {row['name']}" for row in available)
                return (
                    f"Program not found with prefix: '{program_prefix}'\n\n"
                    f"Available programs:\n{prog_list}\n\n"
                    f"Hint: Use the prefix (e.g., 'P4M'), not the full name."
                )
            else:
                return (
                    f"Program not found: '{program_prefix}'\n\n"
                    f"No active programs exist. Create a program first with create_program()."
                )

        program_id = program_row['program_id']
        program_name = program_row['name']
        prefix = program_row['prefix']

        # ----------------------------------------------------------------
        # STEP 4: Generate requirement_id by finding max existing number
        # Format: {PREFIX}-REQ-{NNN}
        # ----------------------------------------------------------------
        cursor.execute(
            """
            SELECT requirement_id FROM requirements
            WHERE requirement_id LIKE ?
            ORDER BY requirement_id DESC
            LIMIT 1
            """,
            (f"{prefix}-REQ-%",)
        )
        last_req = cursor.fetchone()

        if last_req:
            # Extract the number from the last requirement ID
            last_id = last_req['requirement_id']
            match = re.search(r'-REQ-(\d+)$', last_id)
            if match:
                next_num = int(match.group(1)) + 1
            else:
                next_num = 1
        else:
            next_num = 1

        # Format with leading zeros (3 digits)
        requirement_id = f"{prefix}-REQ-{next_num:03d}"

        # ----------------------------------------------------------------
        # STEP 5: Insert the requirement
        # ----------------------------------------------------------------
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute(
            """
            INSERT INTO requirements (
                requirement_id, program_id, title, description, raw_text,
                source_file, priority, requirement_type,
                created_date, updated_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                requirement_id,
                program_id,
                title,
                description,
                description,  # raw_text = original description
                source_file,
                priority,
                requirement_type,
                now,
                now
            )
        )

        # ----------------------------------------------------------------
        # STEP 6: Log to audit_history for Part 11 compliance
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                'requirement',                      # record_type
                requirement_id,                     # record_id
                'Created',                          # action
                'requirement',                      # field_changed
                None,                               # old_value (none for create)
                title,                              # new_value
                'MCP:create_requirement',           # changed_by
                now,                                # changed_date
                f'New requirement created in {program_name} [{prefix}]'  # change_reason
            )
        )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 7: Build success response
        # ----------------------------------------------------------------
        result = f"""Requirement created successfully!

Requirement Details:
  Requirement ID: {requirement_id}
  Title: {title}
  Program: {program_name} [{prefix}]
  Priority: {priority}
  Type: {requirement_type}
"""
        if source_file:
            result += f"  Source: {source_file}\n"

        result += f"""
Description:
  {description[:200]}{'...' if len(description) > 200 else ''}

Next Steps:
  • Convert to user story: create_story(program_prefix="{prefix}", ...)
  • View program requirements (use search_stories with program filter)
  • Link stories to this requirement using requirement_id

Traceability:
  This requirement ID ({requirement_id}) can be linked to user stories
  for full requirements traceability in compliance reporting.
"""
        return result

    except sqlite3.IntegrityError as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        if "UNIQUE constraint" in str(e):
            return (
                f"Requirement ID conflict: '{requirement_id}' already exists.\n\n"
                f"This shouldn't happen with auto-increment. Please report this bug."
            )
        return f"Database integrity error: {str(e)}"

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        return f"Error creating requirement: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def link_requirement_to_story(
    requirement_id: str,
    story_id: str,
    coverage_status: str = "Full",
    gap_notes: Optional[str] = None
) -> str:
    """
    Create traceability link between a requirement and a user story.

    PURPOSE:
        Establishes formal traceability between raw requirements and user stories.
        Essential for compliance reporting and gap analysis.

    R EQUIVALENT:
        Like creating a join table row that links two data.frames by foreign key,
        with metadata about the relationship quality.

    Args:
        requirement_id: Requirement ID (e.g., "P4M-REQ-001")
        story_id: Story ID (e.g., "P4M-AUTH-001")
        coverage_status: How well the story covers the requirement
                        "Full" - Story completely addresses requirement
                        "Partial" - Story partially addresses requirement
                        "None" - Placeholder link, no coverage yet
        gap_notes: Notes explaining gaps if coverage is Partial or None

    Returns:
        Confirmation with traceability link details

    WHY THIS APPROACH:
        - Creates formal audit trail between requirements and implementation
        - Supports compliance gap analysis reporting
        - Validates both entities exist and belong to same program
        - Audit logging for Part 11 compliance
    """
    import sqlite3

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate coverage_status
        # ----------------------------------------------------------------
        valid_coverage = ["Full", "Partial", "None"]
        if coverage_status not in valid_coverage:
            return (
                f"Invalid coverage_status: '{coverage_status}'\n\n"
                f"Valid values:\n"
                f"  • Full - Story completely addresses the requirement\n"
                f"  • Partial - Story partially addresses the requirement\n"
                f"  • None - Placeholder link, requirement not yet covered"
            )

        # Warn if Partial/None without gap_notes
        if coverage_status in ["Partial", "None"] and not gap_notes:
            missing_notes_warning = True
        else:
            missing_notes_warning = False

        # ----------------------------------------------------------------
        # STEP 2: Connect to database
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # ----------------------------------------------------------------
        # STEP 3: Validate requirement exists and get its program_id
        # ----------------------------------------------------------------
        cursor.execute(
            """
            SELECT r.requirement_id, r.title as req_title, r.program_id,
                   p.name as program_name, p.prefix
            FROM requirements r
            JOIN programs p ON r.program_id = p.program_id
            WHERE r.requirement_id = ?
            """,
            (requirement_id,)
        )
        req_row = cursor.fetchone()

        if not req_row:
            # Requirement not found - provide helpful suggestions
            parts = requirement_id.split('-')
            if len(parts) >= 2:
                prefix = parts[0]
                cursor.execute(
                    "SELECT requirement_id, title FROM requirements WHERE requirement_id LIKE ? ORDER BY requirement_id LIMIT 5",
                    (f"{prefix}-REQ-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {r['requirement_id']}: {r['title'][:40]}" for r in similar)
                    return (
                        f"Requirement not found: '{requirement_id}'\n\n"
                        f"Requirements in {prefix}:\n{similar_list}"
                    )

            return f"Requirement not found: '{requirement_id}'"

        req_program_id = req_row['program_id']
        req_title = req_row['req_title']
        program_name = req_row['program_name']
        prefix = req_row['prefix']

        # ----------------------------------------------------------------
        # STEP 4: Validate story exists and get its program_id
        # ----------------------------------------------------------------
        cursor.execute(
            """
            SELECT s.story_id, s.title as story_title, s.program_id
            FROM user_stories s
            WHERE s.story_id = ?
            """,
            (story_id,)
        )
        story_row = cursor.fetchone()

        if not story_row:
            # Story not found - provide helpful suggestions
            parts = story_id.split('-')
            if len(parts) >= 2:
                prefix_part = parts[0]
                cursor.execute(
                    "SELECT story_id, title FROM user_stories WHERE story_id LIKE ? ORDER BY story_id LIMIT 5",
                    (f"{prefix_part}-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {s['story_id']}: {s['title'][:40]}" for s in similar)
                    return (
                        f"Story not found: '{story_id}'\n\n"
                        f"Stories in {prefix_part}:\n{similar_list}"
                    )

            return f"Story not found: '{story_id}'"

        story_program_id = story_row['program_id']
        story_title = story_row['story_title']

        # ----------------------------------------------------------------
        # STEP 5: Validate both belong to the same program
        # ----------------------------------------------------------------
        if req_program_id != story_program_id:
            return (
                f"Program mismatch!\n\n"
                f"Requirement '{requirement_id}' belongs to program: {program_name}\n"
                f"Story '{story_id}' belongs to a different program.\n\n"
                f"Both must belong to the same program for traceability."
            )

        # ----------------------------------------------------------------
        # STEP 6: Check if link already exists
        # ----------------------------------------------------------------
        cursor.execute(
            """
            SELECT trace_id, coverage_status FROM traceability
            WHERE requirement_id = ? AND story_id = ?
            """,
            (requirement_id, story_id)
        )
        existing_link = cursor.fetchone()

        if existing_link:
            return (
                f"Link already exists!\n\n"
                f"Requirement '{requirement_id}' is already linked to story '{story_id}'\n"
                f"Current coverage: {existing_link['coverage_status']}\n"
                f"Trace ID: {existing_link['trace_id']}\n\n"
                f"To update coverage, use a direct database update or create a new tool."
            )

        # ----------------------------------------------------------------
        # STEP 7: Insert the traceability link
        # ----------------------------------------------------------------
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute(
            """
            INSERT INTO traceability (
                program_id, requirement_id, story_id,
                coverage_status, gap_notes,
                created_date, updated_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                req_program_id,
                requirement_id,
                story_id,
                coverage_status,
                gap_notes,
                now,
                now
            )
        )

        trace_id = cursor.lastrowid

        # ----------------------------------------------------------------
        # STEP 8: Log to audit_history for Part 11 compliance
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                'traceability',                     # record_type
                str(trace_id),                      # record_id
                'Created',                          # action
                'link',                             # field_changed
                None,                               # old_value
                f'{requirement_id} -> {story_id}', # new_value
                'MCP:link_requirement_to_story',    # changed_by
                now,                                # changed_date
                f'Traceability link created with {coverage_status} coverage'
            )
        )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 9: Build success response
        # ----------------------------------------------------------------
        # Coverage indicator
        coverage_indicators = {
            "Full": "✓",
            "Partial": "◐",
            "None": "○"
        }
        indicator = coverage_indicators.get(coverage_status, "?")

        result = f"""Traceability link created!

{indicator} {requirement_id} → {story_id}

Link Details:
  Trace ID: {trace_id}
  Program: {program_name} [{prefix}]
  Coverage: {coverage_status}

Requirement:
  {requirement_id}: {req_title[:50]}{'...' if len(req_title) > 50 else ''}

Story:
  {story_id}: {story_title[:50]}{'...' if len(story_title) > 50 else ''}
"""
        if gap_notes:
            result += f"""
Gap Notes:
  {gap_notes}
"""

        if missing_notes_warning:
            result += f"""
⚠ Warning: Coverage is '{coverage_status}' but no gap_notes provided.
  Consider adding notes explaining what's missing or why coverage is incomplete.
"""

        result += f"""
Next Steps:
  • get_coverage_gaps(program_prefix="{prefix}") - View all coverage gaps
  • Create additional links for other requirements/stories
  • Add test cases to the story for full traceability
"""
        return result

    except Exception as e:
        return f"Error creating traceability link: {str(e)}"

    finally:
        if conn:
            conn.close()


# ============================================================
# REQUIREMENTS TOOLKIT - USER STORY TOOLS
# ============================================================

@mcp.tool()
def list_stories(
    program_prefix: str,
    status: Optional[str] = None,
    category: Optional[str] = None,
    roadmap_target: Optional[str] = None
) -> str:
    """
    List user stories for a program.

    PURPOSE:
        Lists user stories for a program with optional filtering by status,
        category, and/or roadmap target quarter.

    Args:
        program_prefix: Program prefix (e.g., "PROP")
        status: Filter by status (Draft, Approved, etc.)
        category: Filter by category
        roadmap_target: Filter by roadmap quarter/year - "Q1 2026", "Q2 2026",
                        "Q3 2026", "Q4 2026", "2027", or "Backlog"

    Returns:
        List of user stories with status, priority, and roadmap target
    """
    # ----------------------------------------------------------------
    # Import sqlite3 for direct query with roadmap_target filter
    # The db.get_stories() method doesn't support roadmap_target yet
    # ----------------------------------------------------------------
    import sqlite3

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Connect and validate program exists
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            "SELECT program_id, name, prefix FROM programs WHERE UPPER(prefix) = ?",
            (program_prefix.upper(),)
        )
        program = cursor.fetchone()

        if not program:
            return f"Program not found: {program_prefix}"

        # ----------------------------------------------------------------
        # STEP 2: Build query with optional filters
        # ----------------------------------------------------------------
        query = "SELECT * FROM user_stories WHERE program_id = ?"
        params = [program['program_id']]

        if status:
            query += " AND status = ?"
            params.append(status)

        if category:
            query += " AND UPPER(category) = ?"
            params.append(category.upper())

        if roadmap_target:
            query += " AND roadmap_target = ?"
            params.append(roadmap_target)

        query += " ORDER BY category, story_id"

        cursor.execute(query, params)
        stories = [dict(row) for row in cursor.fetchall()]

        # ----------------------------------------------------------------
        # STEP 3: Handle no results
        # ----------------------------------------------------------------
        if not stories:
            filters = []
            if status:
                filters.append(f"status={status}")
            if category:
                filters.append(f"category={category}")
            if roadmap_target:
                filters.append(f"roadmap_target={roadmap_target}")
            filter_str = f" (filters: {', '.join(filters)})" if filters else ""
            return f"No stories found for {program_prefix}{filter_str}."

        # ----------------------------------------------------------------
        # STEP 4: Build output with roadmap_target included
        # ----------------------------------------------------------------
        result = f"User Stories for {program['name']} [{program['prefix']}]:\n"
        result += f"Found {len(stories)} story(ies)\n\n"

        for story in stories[:50]:  # Limit display
            result += f"[{story['story_id']}] {story.get('title', 'Untitled')}\n"
            result += f"  Status: {story['status']} | Priority: {story.get('priority', 'N/A')}"
            if story.get('category'):
                result += f" | Category: {story['category']}"
            if story.get('roadmap_target'):
                result += f" | Roadmap: {story['roadmap_target']}"
            result += "\n"

        if len(stories) > 50:
            result += f"\n... and {len(stories) - 50} more stories\n"

        return result

    except Exception as e:
        return f"Error listing stories: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_story(story_id: str) -> str:
    """
    Get detailed information about a specific user story.

    PURPOSE:
        Retrieves full story details including roadmap planning info and
        status transition timeline for audit/tracking purposes.

    Args:
        story_id: The story ID (e.g., "PROP-AUTH-001")

    Returns:
        Full story details including acceptance criteria, roadmap target,
        and status history timeline
    """
    db = None
    try:
        db = get_req_database()
        conn = db.get_connection()

        cursor = conn.execute(
            "SELECT * FROM user_stories WHERE story_id = ?",
            (story_id,)
        )
        row = cursor.fetchone()

        if not row:
            return f"Story not found: {story_id}"

        story = dict(row)

        result = f"Story: {story['story_id']}\n"
        result += f"{'=' * 50}\n\n"
        result += f"Title: {story.get('title', 'N/A')}\n"
        result += f"Status: {story['status']} | Version: {story.get('version', 1)}\n"
        result += f"Priority: {story.get('priority', 'N/A')}\n"

        # ----------------------------------------------------------------
        # Show roadmap target (annual planning timeline)
        # ----------------------------------------------------------------
        if story.get('roadmap_target'):
            result += f"Roadmap Target: {story['roadmap_target']}\n"

        if story.get('category'):
            result += f"Category: {story['category']}"
            if story.get('category_full'):
                result += f" ({story['category_full']})"
            result += "\n"

        if story.get('user_story'):
            result += f"\nUser Story:\n{story['user_story']}\n"

        if story.get('acceptance_criteria'):
            result += f"\nAcceptance Criteria:\n"
            criteria = story['acceptance_criteria'].split('\n')
            for i, ac in enumerate(criteria, 1):
                if ac.strip():
                    result += f"  {i}. {ac.strip()}\n"

        if story.get('success_metrics'):
            result += f"\nSuccess Metrics: {story['success_metrics']}\n"

        # ----------------------------------------------------------------
        # Status History Timeline (audit trail for status transitions)
        # Shows when the story moved through each workflow stage
        # ----------------------------------------------------------------
        timeline_entries = []
        if story.get('draft_date'):
            timeline_entries.append(('Draft', story['draft_date']))
        if story.get('internal_review_date'):
            timeline_entries.append(('Internal Review', story['internal_review_date']))
        if story.get('client_review_date'):
            timeline_entries.append(('Pending Client Review', story['client_review_date']))
        if story.get('needs_discussion_date'):
            timeline_entries.append(('Needs Discussion', story['needs_discussion_date']))
        if story.get('approved_date'):
            timeline_entries.append(('Approved', story['approved_date']))

        if timeline_entries:
            result += f"\nStatus Timeline:\n"
            # Sort by date to show chronological order
            timeline_entries.sort(key=lambda x: x[1] if x[1] else '')
            for status_name, date_val in timeline_entries:
                if date_val:
                    result += f"  • {status_name}: {date_val}\n"

        # Show approved_by separately if available
        if story.get('approved_date') and story.get('approved_by'):
            result += f"  (Approved by: {story['approved_by']})\n"

        return result

    except Exception as e:
        return f"Error getting story: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def create_story(
    program_prefix: str,
    category: str,
    title: str,
    user_story: str,
    acceptance_criteria: str,
    priority: str = "Should Have",
    status: str = "Draft",
    role: Optional[str] = None,
    internal_notes: Optional[str] = None,
    roadmap_target: Optional[str] = None
) -> str:
    """
    Create a new user story for tracking requirements, acceptance criteria, and test coverage.

    PURPOSE:
        Creates a user story under an existing program. Stories track requirements
        in a structured format with clear acceptance criteria for testing.

    R EQUIVALENT:
        Like creating a row in a data.frame, but with auto-generated IDs
        and database persistence.

    STORY ID FORMAT:
        Auto-generated as {PREFIX}-{CATEGORY}-{NNN}
        Example: P4M-AUTH-001, P4M-AUTH-002, P4M-DASH-001

    Args:
        program_prefix: Program prefix (e.g., "P4M", "PROP")
        category: Story category (e.g., "AUTH", "DASH", "RECRUIT", "MSG")
        title: Short descriptive title for the story
        user_story: Full story in "As a [role], I want [feature], so that [benefit]" format
        acceptance_criteria: Acceptance criteria (can be multi-line, use \\n for line breaks)
        priority: MoSCoW priority - "Must Have", "Should Have", "Could Have", "Won't Have"
        status: Workflow status (default: "Draft")
        role: User role if not extractable from user_story
        internal_notes: Internal team notes (not shared with client)
        roadmap_target: Annual planning timeline - "Q1 2026", "Q2 2026", "Q3 2026",
                        "Q4 2026", "2027", or "Backlog" (separate from priority)

    Returns:
        Confirmation with generated story_id and link to view it

    WHY THIS APPROACH:
        - Auto-increments story number within prefix+category for unique IDs
        - Validates program exists before creating
        - Parses user_story to extract role if not provided
        - Sets draft_date automatically for status timeline tracking
        - Audit logging for FDA 21 CFR Part 11 compliance
    """
    logger.info(f"create_story() called - program={program_prefix}, category={category}, title={title[:50]}...")

    # Import here to keep dependency local (matches pattern in other tools)
    import sqlite3
    import re

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate and normalize category
        # Category should be uppercase for consistency in story IDs
        # ----------------------------------------------------------------
        category_upper = category.upper().strip()

        # Validate category format (letters only, 2-10 chars)
        if not re.match(r'^[A-Z]{2,10}$', category_upper):
            return (
                f"Invalid category format: '{category}'\n\n"
                f"Category must be 2-10 letters (no numbers, spaces, or special characters).\n"
                f"Common categories:\n"
                f"  • AUTH - Authentication/Authorization\n"
                f"  • DASH - Dashboard\n"
                f"  • RECRUIT - Recruitment/Enrollment\n"
                f"  • MSG - Messaging/Communications\n"
                f"  • DATA - Data Management\n"
                f"  • REPORT - Reporting\n"
                f"  • ADMIN - Administration"
            )

        # ----------------------------------------------------------------
        # STEP 2: Validate priority is a valid MoSCoW value
        # ----------------------------------------------------------------
        priority_descriptions = {
            "Must Have": "Critical requirement",
            "Should Have": "Important but not critical",
            "Could Have": "Nice to have",
            "Won't Have": "Out of scope for now"
        }
        if error := validate_choice(priority, list(priority_descriptions.keys()), "priority", priority_descriptions):
            return error

        # ----------------------------------------------------------------
        # STEP 3: Connect to requirements database
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # ----------------------------------------------------------------
        # STEP 4: Validate program exists (lookup by prefix)
        # ----------------------------------------------------------------
        cursor.execute(
            "SELECT program_id, name, prefix FROM programs WHERE UPPER(prefix) = ?",
            (program_prefix.upper(),)
        )
        program_row = cursor.fetchone()

        if not program_row:
            # Program not found - show available programs to help user
            cursor.execute(
                "SELECT prefix, name FROM programs WHERE status = 'Active' ORDER BY prefix"
            )
            available = cursor.fetchall()

            if available:
                prog_list = "\n".join(f"  • [{row['prefix']}] {row['name']}" for row in available)
                return (
                    f"Program not found with prefix: '{program_prefix}'\n\n"
                    f"Available programs:\n{prog_list}\n\n"
                    f"Hint: Use the prefix (e.g., 'P4M'), not the full name."
                )
            else:
                return (
                    f"Program not found: '{program_prefix}'\n\n"
                    f"No active programs exist. Create a program first with create_program()."
                )

        program_id = program_row['program_id']
        program_name = program_row['name']
        prefix = program_row['prefix']  # Use stored prefix for consistent casing

        # ----------------------------------------------------------------
        # STEP 5: Generate story_id by finding max existing number
        # Format: {PREFIX}-{CATEGORY}-{NNN}
        # Query: Find highest NNN for this prefix+category, add 1
        # ----------------------------------------------------------------
        cursor.execute(
            """
            SELECT story_id FROM user_stories
            WHERE story_id LIKE ?
            ORDER BY story_id DESC
            LIMIT 1
            """,
            (f"{prefix}-{category_upper}-%",)
        )
        last_story = cursor.fetchone()

        if last_story:
            # Extract the number from the last story ID (e.g., "P4M-AUTH-005" -> 5)
            last_id = last_story['story_id']
            match = re.search(r'-(\d+)$', last_id)
            if match:
                next_num = int(match.group(1)) + 1
            else:
                next_num = 1
        else:
            next_num = 1

        # Format with leading zeros (3 digits)
        story_id = f"{prefix}-{category_upper}-{next_num:03d}"

        # ----------------------------------------------------------------
        # STEP 6: Try to extract role from user_story if not provided
        # Looks for "As a [role]," pattern
        # ----------------------------------------------------------------
        extracted_role = role
        if not extracted_role and user_story:
            role_match = re.search(r'[Aa]s\s+(?:a|an)\s+([^,]+),', user_story)
            if role_match:
                extracted_role = role_match.group(1).strip()

        # ----------------------------------------------------------------
        # STEP 7: Insert the user story
        # Sets draft_date automatically since all stories start as Draft
        # ----------------------------------------------------------------
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute(
            """
            INSERT INTO user_stories (
                story_id, program_id, title, user_story, role,
                acceptance_criteria, priority, category, status,
                internal_notes, version, created_date, updated_date,
                roadmap_target, draft_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                story_id,
                program_id,
                title,
                user_story,
                extracted_role,
                acceptance_criteria,
                priority,
                category_upper,
                status,
                internal_notes,
                1,  # version starts at 1
                now,
                now,
                roadmap_target,  # Annual roadmap planning target
                now              # draft_date = now since all stories start as Draft
            )
        )

        # ----------------------------------------------------------------
        # STEP 8: Log to audit_history for Part 11 compliance
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                'user_story',                       # record_type
                story_id,                           # record_id
                'Created',                          # action
                'user_story',                       # field_changed
                None,                               # old_value (none for create)
                title,                              # new_value
                'MCP:create_story',                 # changed_by
                now,                                # changed_date
                f'New story created in {program_name} [{prefix}]'  # change_reason
            )
        )

        # Commit both inserts together
        conn.commit()

        # ----------------------------------------------------------------
        # STEP 9: Build success response
        # ----------------------------------------------------------------
        result = f"""Story created successfully!

Story Details:
  Story ID: {story_id}
  Title: {title}
  Program: {program_name} [{prefix}]
  Category: {category_upper}
  Priority: {priority}
  Status: {status}
"""
        if roadmap_target:
            result += f"  Roadmap Target: {roadmap_target}\n"
        if extracted_role:
            result += f"  Role: {extracted_role}\n"

        result += f"""
User Story:
  {user_story}

Acceptance Criteria:
"""
        # Format acceptance criteria with line numbers
        for i, line in enumerate(acceptance_criteria.split('\n'), 1):
            if line.strip():
                result += f"  {i}. {line.strip()}\n"

        result += f"""
Next Steps:
  • get_story(story_id="{story_id}") - View full story details
  • list_stories(program_prefix="{prefix}") - View all stories in program
  • Create test cases linked to this story
"""
        return result

    except sqlite3.IntegrityError as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        # Handle unique constraint violations
        if "UNIQUE constraint" in str(e):
            return (
                f"Story ID conflict: '{story_id}' already exists.\n\n"
                f"This shouldn't happen with auto-increment. Please report this bug."
            )
        return f"Database integrity error: {str(e)}"

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        logger.error(f"create_story() failed: {str(e)}", exc_info=True)
        return f"Error creating story: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_story(
    story_id: str,
    title: Optional[str] = None,
    user_story: Optional[str] = None,
    acceptance_criteria: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    client_feedback: Optional[str] = None,
    internal_notes: Optional[str] = None,
    change_reason: Optional[str] = None,
    roadmap_target: Optional[str] = None
) -> str:
    """
    Update an existing user story's content, status, or add feedback.

    PURPOSE:
        Modifies an existing user story. Only fields that are provided (not None)
        will be updated. Each change is logged to the audit trail for compliance.

    R EQUIVALENT:
        Like dplyr::mutate() on a single row, but with version tracking
        and audit logging.

    Args:
        story_id: Story ID to update (e.g., "P4M-AUTH-001")
        title: Updated title (optional)
        user_story: Updated story text (optional)
        acceptance_criteria: Updated acceptance criteria (optional)
        status: Workflow status - "Draft", "Internal Review", "Pending Client Review",
                "Approved", "Needs Discussion", "Out of Scope"
        priority: MoSCoW priority - "Must Have", "Should Have", "Could Have", "Won't Have"
        client_feedback: Feedback from client review (optional)
        internal_notes: Internal team notes (optional)
        change_reason: Why this change was made - recorded in audit trail (optional)
        roadmap_target: Annual planning timeline - "Q1 2026", "Q2 2026", "Q3 2026",
                        "Q4 2026", "2027", or "Backlog" (separate from priority)

    Returns:
        Confirmation with updated fields and new version number

    WHY THIS APPROACH:
        - Only updates provided fields (sparse update pattern)
        - Auto-increments version on any change for tracking
        - Status changes automatically set corresponding date field for audit trail
        - Per-field audit logging for regulatory compliance
    """
    import sqlite3

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate status if provided
        # ----------------------------------------------------------------
        valid_statuses = [
            "Draft",
            "Internal Review",
            "Pending Client Review",
            "Approved",
            "Needs Discussion",
            "Out of Scope"
        ]
        if status is not None and status not in valid_statuses:
            status_list = "\n".join(f"  • {s}" for s in valid_statuses)
            return (
                f"Invalid status: '{status}'\n\n"
                f"Valid statuses:\n{status_list}"
            )

        # ----------------------------------------------------------------
        # STEP 2: Validate priority if provided
        # ----------------------------------------------------------------
        valid_priorities = ["Must Have", "Should Have", "Could Have", "Won't Have"]
        if priority is not None and priority not in valid_priorities:
            return (
                f"Invalid priority: '{priority}'\n\n"
                f"Priority must be one of (MoSCoW):\n"
                f"  • Must Have - Critical requirement\n"
                f"  • Should Have - Important but not critical\n"
                f"  • Could Have - Nice to have\n"
                f"  • Won't Have - Out of scope for now"
            )

        # ----------------------------------------------------------------
        # STEP 2b: Validate roadmap_target if provided
        # Roadmap target is separate from priority - it indicates WHEN on the
        # annual roadmap, while priority indicates importance for THIS release
        # ----------------------------------------------------------------
        valid_roadmap_targets = ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026", "2027", "Backlog"]
        if roadmap_target is not None and roadmap_target not in valid_roadmap_targets:
            target_list = "\n".join(f"  • {t}" for t in valid_roadmap_targets)
            return (
                f"Invalid roadmap_target: '{roadmap_target}'\n\n"
                f"Valid roadmap targets:\n{target_list}"
            )

        # ----------------------------------------------------------------
        # STEP 3: Connect to database and fetch existing story
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT s.*, p.name as program_name, p.prefix
            FROM user_stories s
            JOIN programs p ON s.program_id = p.program_id
            WHERE s.story_id = ?
            """,
            (story_id,)
        )
        existing = cursor.fetchone()

        if not existing:
            # Story not found - provide helpful error
            # Try to extract prefix to suggest similar stories
            parts = story_id.split('-')
            if len(parts) >= 2:
                prefix = parts[0]
                cursor.execute(
                    "SELECT story_id, title FROM user_stories WHERE story_id LIKE ? LIMIT 5",
                    (f"{prefix}-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {s['story_id']}: {s['title'][:40]}" for s in similar)
                    return (
                        f"Story not found: '{story_id}'\n\n"
                        f"Similar stories in {prefix}:\n{similar_list}"
                    )

            return f"Story not found: '{story_id}'"

        # Convert to dict for easier access
        old_story = dict(existing)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ----------------------------------------------------------------
        # STEP 4: Build update fields and track changes
        # Only include fields that were provided (not None)
        # ----------------------------------------------------------------
        updates = {}
        changes = []  # List of (field_name, old_value, new_value) for audit

        # Check each optional field
        if title is not None and title != old_story.get('title'):
            updates['title'] = title
            changes.append(('title', old_story.get('title'), title))

        if user_story is not None and user_story != old_story.get('user_story'):
            updates['user_story'] = user_story
            changes.append(('user_story', old_story.get('user_story'), user_story))

        if acceptance_criteria is not None and acceptance_criteria != old_story.get('acceptance_criteria'):
            updates['acceptance_criteria'] = acceptance_criteria
            changes.append(('acceptance_criteria', old_story.get('acceptance_criteria'), acceptance_criteria))

        if status is not None and status != old_story.get('status'):
            updates['status'] = status
            changes.append(('status', old_story.get('status'), status))

            # ----------------------------------------------------------------
            # Set the corresponding date field for each status transition
            # This creates an audit trail showing when the story moved through
            # each stage of the workflow
            # ----------------------------------------------------------------
            status_date_map = {
                "Draft": "draft_date",
                "Internal Review": "internal_review_date",
                "Pending Client Review": "client_review_date",
                "Approved": "approved_date",
                "Needs Discussion": "needs_discussion_date"
            }

            if status in status_date_map:
                date_field = status_date_map[status]
                updates[date_field] = now
                changes.append((date_field, old_story.get(date_field), now))

            # Special handling for Approved status - also set approved_by
            if status == "Approved":
                updates['approved_by'] = 'MCP:update_story'

        if priority is not None and priority != old_story.get('priority'):
            updates['priority'] = priority
            changes.append(('priority', old_story.get('priority'), priority))

        if client_feedback is not None:
            # For feedback, we append rather than replace (if existing)
            old_feedback = old_story.get('client_feedback') or ''
            if old_feedback:
                new_feedback = f"{old_feedback}\n\n[{now}]\n{client_feedback}"
            else:
                new_feedback = f"[{now}]\n{client_feedback}"
            updates['client_feedback'] = new_feedback
            changes.append(('client_feedback', old_feedback, new_feedback))

        if internal_notes is not None:
            # For notes, we append rather than replace (if existing)
            old_notes = old_story.get('internal_notes') or ''
            if old_notes:
                new_notes = f"{old_notes}\n\n[{now}]\n{internal_notes}"
            else:
                new_notes = f"[{now}]\n{internal_notes}"
            updates['internal_notes'] = new_notes
            changes.append(('internal_notes', old_notes, new_notes))

        # ----------------------------------------------------------------
        # Handle roadmap_target updates (annual planning timeline)
        # This is separate from priority - priority = what's in this release,
        # roadmap_target = when on the annual roadmap
        # ----------------------------------------------------------------
        if roadmap_target is not None and roadmap_target != old_story.get('roadmap_target'):
            updates['roadmap_target'] = roadmap_target
            changes.append(('roadmap_target', old_story.get('roadmap_target'), roadmap_target))

        # ----------------------------------------------------------------
        # STEP 5: Check if there are any changes to make
        # ----------------------------------------------------------------
        if not updates:
            return (
                f"No changes to make for {story_id}\n\n"
                f"All provided values match the current values, or no update fields were provided.\n"
                f"Current story:\n"
                f"  Title: {old_story.get('title')}\n"
                f"  Status: {old_story.get('status')}\n"
                f"  Priority: {old_story.get('priority')}\n"
                f"  Version: {old_story.get('version', 1)}"
            )

        # ----------------------------------------------------------------
        # STEP 6: Always update version and updated_date when changes occur
        # ----------------------------------------------------------------
        new_version = (old_story.get('version') or 1) + 1
        updates['version'] = new_version
        updates['updated_date'] = now

        # ----------------------------------------------------------------
        # STEP 7: Build and execute UPDATE statement
        # ----------------------------------------------------------------
        set_clause = ", ".join(f"{field} = ?" for field in updates.keys())
        values = list(updates.values()) + [story_id]

        cursor.execute(
            f"UPDATE user_stories SET {set_clause} WHERE story_id = ?",
            values
        )

        # ----------------------------------------------------------------
        # STEP 8: Log each change to audit_history
        # ----------------------------------------------------------------
        audit_reason = change_reason or "Story updated via MCP"

        for field_name, old_value, new_value in changes:
            # Truncate long values for audit log readability
            old_display = str(old_value)[:200] if old_value else None
            new_display = str(new_value)[:200] if new_value else None

            cursor.execute(
                """
                INSERT INTO audit_history (
                    record_type, record_id, action, field_changed,
                    old_value, new_value, changed_by, changed_date, change_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    'user_story',
                    story_id,
                    'Updated',
                    field_name,
                    old_display,
                    new_display,
                    'MCP:update_story',
                    now,
                    audit_reason
                )
            )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 9: Build success response
        # ----------------------------------------------------------------
        result = f"""Story updated successfully!

Story: {story_id}
Program: {old_story['program_name']} [{old_story['prefix']}]
Version: {old_story.get('version', 1)} → {new_version}

Changes made ({len(changes)} field(s)):
"""
        for field_name, old_val, new_val in changes:
            # Format display based on field type
            if field_name in ('client_feedback', 'internal_notes', 'user_story', 'acceptance_criteria'):
                # For long text fields, just show that it was updated
                result += f"  • {field_name}: (updated)\n"
            else:
                result += f"  • {field_name}: '{old_val}' → '{new_val}'\n"

        if status == "Approved":
            result += f"\n✓ Story marked as APPROVED at {now}\n"

        result += f"""
Next Steps:
  • get_story(story_id="{story_id}") - View updated story
  • get_approval_pipeline(program_prefix="{old_story['prefix']}") - View workflow status
"""
        return result

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        return f"Error updating story: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_approval_pipeline(program_prefix: str) -> str:
    """
    Get Kanban-style view of story workflow status.

    Args:
        program_prefix: Program prefix (e.g., "PROP")

    Returns:
        Stories grouped by workflow status
    """
    db = None
    try:
        db = get_req_database()
        program = db.get_program_by_prefix(program_prefix.upper())

        if not program:
            return f"Program not found: {program_prefix}"

        pipeline = req_queries.get_approval_pipeline(db, program['program_id'])

        result = f"Approval Pipeline for {program['name']}:\n"
        result += f"{'=' * 50}\n\n"

        for status, stories in pipeline.items():
            result += f"{status} ({len(stories)}):\n"
            if stories:
                for story in stories[:5]:
                    result += f"  • [{story['story_id']}] {story['title'][:40]}\n"
                if len(stories) > 5:
                    result += f"  ... and {len(stories) - 5} more\n"
            else:
                result += "  (none)\n"
            result += "\n"

        return result

    except Exception as e:
        return f"Error getting pipeline: {str(e)}"
    finally:
        if db:
            db.close()


# ============================================================
# REQUIREMENTS TOOLKIT - TEST CASE TOOLS
# ============================================================

@mcp.tool()
def list_test_cases(
    program_prefix: str,
    status: Optional[str] = None,
    test_type: Optional[str] = None
) -> str:
    """
    List UAT test cases for a program.

    Args:
        program_prefix: Program prefix (e.g., "PROP")
        status: Filter by status (Not Run, Pass, Fail, Blocked, Skipped)
        test_type: Filter by type (happy_path, negative, validation, edge_case)

    Returns:
        List of test cases
    """
    db = None
    try:
        db = get_req_database()
        program = db.get_program_by_prefix(program_prefix.upper())

        if not program:
            return f"Program not found: {program_prefix}"

        tests = db.get_test_cases(
            program_id=program['program_id'],
            status_filter=status,
            test_type=test_type
        )

        if not tests:
            return f"No test cases found for {program_prefix}."

        result = f"Test Cases for {program['name']} [{program_prefix}]:\n"
        result += f"Found {len(tests)} test(s)\n\n"

        for test in tests[:30]:
            result += f"[{test['test_id']}] {test.get('title', 'Untitled')}\n"
            result += f"  Status: {test.get('test_status', 'Not Run')} | Type: {test.get('test_type', 'N/A')}\n"

        if len(tests) > 30:
            result += f"\n... and {len(tests) - 30} more tests\n"

        return result

    except Exception as e:
        return f"Error listing tests: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def get_test_summary(program_prefix: str) -> str:
    """
    Get test execution summary for a program.

    Args:
        program_prefix: Program prefix (e.g., "PROP")

    Returns:
        Test execution statistics including pass/fail rates
    """
    db = None
    try:
        db = get_req_database()
        program = db.get_program_by_prefix(program_prefix.upper())

        if not program:
            return f"Program not found: {program_prefix}"

        summary = req_queries.get_test_execution_summary(db, program['program_id'])

        result = f"Test Execution Summary for {program['name']}:\n"
        result += f"{'=' * 50}\n\n"
        result += f"Total Tests: {summary['total']}\n"
        result += f"Execution Rate: {summary['execution_rate']}%\n"
        result += f"Pass Rate: {summary['pass_rate']}%\n\n"

        result += "By Status:\n"
        for status, count in summary['by_status'].items():
            result += f"  • {status}: {count}\n"

        if summary['by_type']:
            result += "\nBy Test Type:\n"
            for test_type, data in summary['by_type'].items():
                result += f"  {test_type}: {data['total']} total"
                if data.get('Pass'):
                    result += f", {data['Pass']} passed"
                if data.get('Fail'):
                    result += f", {data['Fail']} failed"
                result += "\n"

        return result

    except Exception as e:
        return f"Error getting test summary: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def create_test_case(
    story_id: str,
    title: str,
    test_type: str,
    test_steps: str,
    expected_result: str,
    preconditions: Optional[str] = None,
    priority: str = "Should Have",
    compliance_framework: Optional[str] = None
) -> str:
    """
    Create a UAT test case linked to a user story.

    PURPOSE:
        Creates a test case that validates a user story's acceptance criteria.
        Test cases track execution status and can be linked to compliance frameworks.

    R EQUIVALENT:
        Like adding a row to a test plan data.frame, with automatic ID generation
        and linkage to the parent story.

    TEST ID FORMAT:
        Auto-generated as {STORY_ID}-TC{NN}
        Example: P4M-AUTH-001-TC01, P4M-AUTH-001-TC02

    Args:
        story_id: Parent story ID (e.g., "P4M-AUTH-001")
        title: Test case title (short description)
        test_type: Type of test - "happy_path", "negative", "validation", "edge_case"
        test_steps: Numbered test steps (can be multi-line)
        expected_result: Expected outcome when test passes
        preconditions: Setup required before running test (optional)
        priority: MoSCoW priority (default: "Should Have")
        compliance_framework: "Part11", "HIPAA", "SOC2", or None

    Returns:
        Confirmation with generated test_id

    WHY THIS APPROACH:
        - Links test to parent story for traceability
        - Auto-increments test number within the story
        - Supports compliance framework tagging
        - Audit logging for Part 11 compliance
    """
    import sqlite3
    import re

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate test_type
        # ----------------------------------------------------------------
        valid_test_types = ["happy_path", "negative", "validation", "edge_case"]
        test_type_lower = test_type.lower().strip()

        if test_type_lower not in valid_test_types:
            type_list = "\n".join(f"  • {t}" for t in valid_test_types)
            return (
                f"Invalid test_type: '{test_type}'\n\n"
                f"Valid test types:\n{type_list}\n\n"
                f"Descriptions:\n"
                f"  • happy_path - Normal successful scenario\n"
                f"  • negative - Error handling, invalid inputs\n"
                f"  • validation - Input/data validation rules\n"
                f"  • edge_case - Boundary conditions, unusual scenarios"
            )

        # ----------------------------------------------------------------
        # STEP 2: Validate priority
        # ----------------------------------------------------------------
        valid_priorities = ["Must Have", "Should Have", "Could Have", "Won't Have"]
        if priority not in valid_priorities:
            return (
                f"Invalid priority: '{priority}'\n\n"
                f"Priority must be one of (MoSCoW):\n"
                f"  • Must Have - Critical test\n"
                f"  • Should Have - Important but not critical\n"
                f"  • Could Have - Nice to have\n"
                f"  • Won't Have - Out of scope for now"
            )

        # ----------------------------------------------------------------
        # STEP 3: Validate compliance_framework if provided
        # ----------------------------------------------------------------
        valid_frameworks = ["Part11", "HIPAA", "SOC2"]
        if compliance_framework is not None:
            # Normalize common variations
            framework_normalized = compliance_framework.strip()
            framework_upper = framework_normalized.upper().replace(" ", "").replace("-", "")

            # Map common variations to standard names
            framework_map = {
                "PART11": "Part11",
                "21CFRPART11": "Part11",
                "CFR11": "Part11",
                "HIPAA": "HIPAA",
                "SOC2": "SOC2",
                "SOCII": "SOC2",
                "SOC": "SOC2"
            }

            if framework_upper in framework_map:
                compliance_framework = framework_map[framework_upper]
            elif framework_normalized not in valid_frameworks:
                return (
                    f"Invalid compliance_framework: '{compliance_framework}'\n\n"
                    f"Valid frameworks:\n"
                    f"  • Part11 - FDA 21 CFR Part 11\n"
                    f"  • HIPAA - Health Insurance Portability and Accountability\n"
                    f"  • SOC2 - Service Organization Control 2\n"
                    f"  • None - No compliance requirement"
                )

        # ----------------------------------------------------------------
        # STEP 4: Connect to database and validate parent story exists
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT s.story_id, s.title as story_title, s.program_id,
                   p.name as program_name, p.prefix
            FROM user_stories s
            JOIN programs p ON s.program_id = p.program_id
            WHERE s.story_id = ?
            """,
            (story_id,)
        )
        parent_story = cursor.fetchone()

        if not parent_story:
            # Story not found - provide helpful suggestions
            parts = story_id.split('-')
            if len(parts) >= 2:
                prefix = parts[0]
                cursor.execute(
                    "SELECT story_id, title FROM user_stories WHERE story_id LIKE ? ORDER BY story_id LIMIT 5",
                    (f"{prefix}-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {s['story_id']}: {s['title'][:40]}" for s in similar)
                    return (
                        f"Parent story not found: '{story_id}'\n\n"
                        f"Similar stories in {prefix}:\n{similar_list}\n\n"
                        f"Hint: Use get_story() to verify the story exists."
                    )

            return f"Parent story not found: '{story_id}'"

        program_id = parent_story['program_id']
        program_name = parent_story['program_name']
        prefix = parent_story['prefix']

        # ----------------------------------------------------------------
        # STEP 5: Generate test_id by finding max existing number
        # Format: {STORY_ID}-TC{NN}
        # Query: Find highest NN for this story, add 1
        # ----------------------------------------------------------------
        cursor.execute(
            """
            SELECT test_id FROM uat_test_cases
            WHERE test_id LIKE ?
            ORDER BY test_id DESC
            LIMIT 1
            """,
            (f"{story_id}-TC%",)
        )
        last_test = cursor.fetchone()

        if last_test:
            # Extract the number from the last test ID (e.g., "P4M-AUTH-001-TC05" -> 5)
            last_id = last_test['test_id']
            match = re.search(r'-TC(\d+)$', last_id)
            if match:
                next_num = int(match.group(1)) + 1
            else:
                next_num = 1
        else:
            next_num = 1

        # Format with leading zeros (2 digits)
        test_id = f"{story_id}-TC{next_num:02d}"

        # ----------------------------------------------------------------
        # STEP 6: Insert the test case
        # ----------------------------------------------------------------
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute(
            """
            INSERT INTO uat_test_cases (
                test_id, story_id, program_id, title, test_type,
                prerequisites, test_steps, expected_results,
                priority, compliance_framework, test_status,
                created_date, updated_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                test_id,
                story_id,
                program_id,
                title,
                test_type_lower,
                preconditions,
                test_steps,
                expected_result,
                priority,
                compliance_framework,
                "Not Run",  # Initial status
                now,
                now
            )
        )

        # ----------------------------------------------------------------
        # STEP 7: Log to audit_history for Part 11 compliance
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                'test_case',                        # record_type
                test_id,                            # record_id
                'Created',                          # action
                'test_case',                        # field_changed
                None,                               # old_value (none for create)
                title,                              # new_value
                'MCP:create_test_case',             # changed_by
                now,                                # changed_date
                f'New test case for story {story_id}'  # change_reason
            )
        )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 8: Build success response
        # ----------------------------------------------------------------
        result = f"""Test case created successfully!

Test Case Details:
  Test ID: {test_id}
  Title: {title}
  Parent Story: {story_id} - {parent_story['story_title'][:40]}
  Program: {program_name} [{prefix}]
  Test Type: {test_type_lower}
  Priority: {priority}
  Status: Not Run
"""
        if compliance_framework:
            result += f"  Compliance: {compliance_framework}\n"

        if preconditions:
            result += f"""
Preconditions:
  {preconditions}
"""

        result += f"""
Test Steps:
"""
        for line in test_steps.split('\n'):
            if line.strip():
                result += f"  {line.strip()}\n"

        result += f"""
Expected Result:
  {expected_result}

Next Steps:
  • list_test_cases(program_prefix="{prefix}") - View all test cases
  • get_test_summary(program_prefix="{prefix}") - View test execution stats
  • Create more test cases for this story with different test_types
"""
        return result

    except sqlite3.IntegrityError as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        if "UNIQUE constraint" in str(e):
            return (
                f"Test ID conflict: '{test_id}' already exists.\n\n"
                f"This shouldn't happen with auto-increment. Please report this bug."
            )
        return f"Database integrity error: {str(e)}"

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        return f"Error creating test case: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_test_result(
    test_id: str,
    status: str,
    tested_by: str,
    notes: Optional[str] = None,
    defect_id: Optional[str] = None
) -> str:
    """
    Record the result of executing a test case.

    PURPOSE:
        Records test execution results including pass/fail status, who ran
        the test, and any defects found. Essential for test coverage tracking
        and compliance reporting.

    R EQUIVALENT:
        Like updating a row in a test results data.frame with execution data
        and timestamp.

    Args:
        test_id: Test case ID (e.g., "P4M-AUTH-001-TC01")
        status: Execution result - "Pass", "Fail", "Blocked", "Skipped"
        tested_by: Name or ID of person who executed the test
        notes: Execution notes or observations (optional)
        defect_id: Link to defect/bug ticket if test failed (optional)

    Returns:
        Confirmation with updated test status and summary

    WHY THIS APPROACH:
        - Captures who ran the test and when for audit trail
        - Links failed tests to defect tracking
        - Per-field audit logging for Part 11 compliance
    """
    import sqlite3

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate status
        # ----------------------------------------------------------------
        valid_statuses = ["Pass", "Fail", "Blocked", "Skipped"]
        if status not in valid_statuses:
            status_list = "\n".join(f"  • {s}" for s in valid_statuses)
            return (
                f"Invalid status: '{status}'\n\n"
                f"Valid statuses:\n{status_list}\n\n"
                f"Descriptions:\n"
                f"  • Pass - Test completed successfully\n"
                f"  • Fail - Test found a defect\n"
                f"  • Blocked - Cannot run due to environment/dependency issue\n"
                f"  • Skipped - Intentionally not run (with justification)"
            )

        # Warn if Fail status without defect_id
        if status == "Fail" and not defect_id:
            # Don't block, just note it in the response later
            missing_defect_warning = True
        else:
            missing_defect_warning = False

        # ----------------------------------------------------------------
        # STEP 2: Connect to database and fetch existing test case
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT t.*, s.title as story_title, p.name as program_name, p.prefix
            FROM uat_test_cases t
            JOIN user_stories s ON t.story_id = s.story_id
            JOIN programs p ON t.program_id = p.program_id
            WHERE t.test_id = ?
            """,
            (test_id,)
        )
        existing = cursor.fetchone()

        if not existing:
            # Test not found - provide helpful suggestions
            parts = test_id.split('-')
            if len(parts) >= 3:
                # Try to find the parent story's tests
                story_prefix = '-'.join(parts[:-1])  # e.g., "P4M-AUTH-001" from "P4M-AUTH-001-TC01"
                cursor.execute(
                    "SELECT test_id, title FROM uat_test_cases WHERE test_id LIKE ? ORDER BY test_id LIMIT 5",
                    (f"{story_prefix}-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {t['test_id']}: {t['title'][:40]}" for t in similar)
                    return (
                        f"Test case not found: '{test_id}'\n\n"
                        f"Test cases for {story_prefix}:\n{similar_list}"
                    )

            return f"Test case not found: '{test_id}'"

        # Convert to dict for easier access
        old_test = dict(existing)
        old_status = old_test.get('test_status', 'Not Run')
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ----------------------------------------------------------------
        # STEP 3: Build update fields
        # ----------------------------------------------------------------
        updates = {
            'test_status': status,
            'tested_by': tested_by,
            'tested_date': now,
            'updated_date': now
        }

        # Track changes for audit
        changes = [('test_status', old_status, status)]

        if notes is not None:
            updates['execution_notes'] = notes
            changes.append(('execution_notes', old_test.get('execution_notes'), notes))

        if defect_id is not None:
            updates['defect_id'] = defect_id
            changes.append(('defect_id', old_test.get('defect_id'), defect_id))

        # ----------------------------------------------------------------
        # STEP 4: Execute UPDATE
        # ----------------------------------------------------------------
        set_clause = ", ".join(f"{field} = ?" for field in updates.keys())
        values = list(updates.values()) + [test_id]

        cursor.execute(
            f"UPDATE uat_test_cases SET {set_clause} WHERE test_id = ?",
            values
        )

        # ----------------------------------------------------------------
        # STEP 5: Log to audit_history
        # ----------------------------------------------------------------
        for field_name, old_value, new_value in changes:
            cursor.execute(
                """
                INSERT INTO audit_history (
                    record_type, record_id, action, field_changed,
                    old_value, new_value, changed_by, changed_date, change_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    'test_case',
                    test_id,
                    'Executed',
                    field_name,
                    str(old_value) if old_value else None,
                    str(new_value) if new_value else None,
                    f'MCP:update_test_result ({tested_by})',
                    now,
                    f'Test executed with result: {status}'
                )
            )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 6: Build success response
        # ----------------------------------------------------------------
        # Determine status emoji/indicator
        status_indicators = {
            "Pass": "✓",
            "Fail": "✗",
            "Blocked": "⊘",
            "Skipped": "—"
        }
        indicator = status_indicators.get(status, "?")

        result = f"""Test result recorded!

{indicator} {test_id}: {status}

Test Details:
  Title: {old_test['title']}
  Parent Story: {old_test['story_id']} - {old_test['story_title'][:40]}
  Program: {old_test['program_name']} [{old_test['prefix']}]
  Test Type: {old_test.get('test_type', 'N/A')}

Execution:
  Previous Status: {old_status}
  New Status: {status}
  Tested By: {tested_by}
  Tested Date: {now}
"""
        if notes:
            result += f"  Notes: {notes[:100]}{'...' if len(notes) > 100 else ''}\n"

        if defect_id:
            result += f"  Defect ID: {defect_id}\n"

        if missing_defect_warning:
            result += f"""
⚠ Warning: Test failed but no defect_id provided.
  Consider linking to a defect ticket for tracking.
"""

        result += f"""
Next Steps:
  • get_test_summary(program_prefix="{old_test['prefix']}") - View overall test stats
  • list_test_cases(program_prefix="{old_test['prefix']}", status="{status}") - View all {status} tests
"""
        return result

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        return f"Error updating test result: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def delete_test_case(
    test_id: str,
    reason: str
) -> str:
    """
    Delete a test case from the database with full audit trail.

    PURPOSE:
        Permanently removes a test case that is no longer needed (out of scope,
        duplicate, or superseded). The full test case data is captured in the
        audit trail before deletion for compliance.

    R EQUIVALENT:
        Like dplyr::filter(!test_id == id) but with audit logging and
        validation. The deleted record is preserved in audit_history.

    Args:
        test_id: Test case ID to delete (e.g., "P4M-CONFIG-001-TC02")
        reason: Reason for deletion - recorded in audit trail for compliance

    Returns:
        Confirmation with deleted test_id, parent story, and timestamp

    WHY THIS APPROACH:
        - Hard delete (not soft delete) since test cases can be recreated if needed
        - Full record snapshot captured in audit_history before deletion
        - Reason field ensures audit trail explains why the deletion occurred
        - Part 11 compliant audit trail

    Example:
        delete_test_case(
            test_id="P4M-CONFIG-001-TC02",
            reason="Out of scope - patient assessment moved to future story"
        )
    """
    import sqlite3
    import json

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Connect to database and fetch existing test case
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT t.*, s.title as story_title, p.name as program_name, p.prefix
            FROM uat_test_cases t
            JOIN user_stories s ON t.story_id = s.story_id
            JOIN programs p ON t.program_id = p.program_id
            WHERE t.test_id = ?
            """,
            (test_id,)
        )
        existing = cursor.fetchone()

        if not existing:
            # Test not found - provide helpful suggestions
            parts = test_id.split('-')
            if len(parts) >= 3:
                # Try to find similar test cases
                story_prefix = '-'.join(parts[:-1])  # e.g., "P4M-CONFIG-001" from "P4M-CONFIG-001-TC02"
                cursor.execute(
                    "SELECT test_id, title FROM uat_test_cases WHERE test_id LIKE ? ORDER BY test_id LIMIT 5",
                    (f"{story_prefix}-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {t['test_id']}: {t['title'][:40]}" for t in similar)
                    return (
                        f"Test case not found: '{test_id}'\n\n"
                        f"Test cases for {story_prefix}:\n{similar_list}"
                    )

            return f"Test case not found: '{test_id}'"

        # Convert to dict for easier access and JSON serialization
        test_data = dict(existing)
        story_id = test_data['story_id']
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ----------------------------------------------------------------
        # STEP 2: Prepare audit record with full test case data
        # Remove non-serializable or redundant joined fields
        # ----------------------------------------------------------------
        audit_record = {
            'test_id': test_data.get('test_id'),
            'story_id': test_data.get('story_id'),
            'program_id': test_data.get('program_id'),
            'title': test_data.get('title'),
            'test_type': test_data.get('test_type'),
            'prerequisites': test_data.get('prerequisites'),
            'test_steps': test_data.get('test_steps'),
            'expected_results': test_data.get('expected_results'),
            'priority': test_data.get('priority'),
            'compliance_framework': test_data.get('compliance_framework'),
            'test_status': test_data.get('test_status'),
            'tested_by': test_data.get('tested_by'),
            'tested_date': test_data.get('tested_date'),
            'execution_notes': test_data.get('execution_notes'),
            'defect_id': test_data.get('defect_id'),
            'created_date': test_data.get('created_date'),
            'updated_date': test_data.get('updated_date')
        }

        # ----------------------------------------------------------------
        # STEP 3: Log deletion to audit_history BEFORE deleting
        # ----------------------------------------------------------------
        cursor.execute(
            """
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                'test_case',                    # record_type
                test_id,                        # record_id
                'Deleted',                      # action
                'record',                       # field_changed
                json.dumps(audit_record),       # old_value (full record)
                None,                           # new_value (null for delete)
                'MCP:delete_test_case',         # changed_by
                now,                            # changed_date
                reason                          # change_reason
            )
        )

        # ----------------------------------------------------------------
        # STEP 4: Delete the test case
        # ----------------------------------------------------------------
        cursor.execute(
            "DELETE FROM uat_test_cases WHERE test_id = ?",
            (test_id,)
        )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 5: Build success response
        # ----------------------------------------------------------------
        result = f"""Test case deleted successfully!

Deleted Test Case:
  Test ID: {test_id}
  Title: {test_data['title']}
  Parent Story: {story_id} - {test_data['story_title'][:40]}
  Program: {test_data['program_name']} [{test_data['prefix']}]
  Previous Status: {test_data.get('test_status', 'Not Run')}

Deletion Details:
  Reason: {reason}
  Deleted At: {now}
  Deleted By: MCP:delete_test_case

Audit Trail:
  Full test case record has been preserved in audit_history.
  The deletion is traceable for compliance purposes.

Next Steps:
  • list_test_cases(story_id="{story_id}") - View remaining test cases
  • create_test_case(...) - Create a replacement test case if needed
"""
        return result

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        return f"Error deleting test case: {str(e)}"

    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_test_case(
    test_id: str,
    title: Optional[str] = None,
    test_type: Optional[str] = None,
    test_steps: Optional[str] = None,
    expected_result: Optional[str] = None,
    preconditions: Optional[str] = None,
    priority: Optional[str] = None,
    compliance_framework: Optional[str] = None,
    story_id: Optional[str] = None,
    change_reason: Optional[str] = None
) -> str:
    """
    Update test case content (NOT execution results — use update_test_result for that).

    PURPOSE:
        Modifies test case definition: title, steps, expected results, etc.
        Only provided fields are updated (sparse update pattern).
        Each change is logged to the audit trail for compliance.

    R EQUIVALENT:
        Like dplyr::mutate() on a single row with coalesce() for optional fields.
        Only non-NULL parameters trigger updates.

    Args:
        test_id: Test case ID to update (e.g., "P4M-CONFIG-001-TC05")
        title: Updated test case title (optional)
        test_type: Type of test - "happy_path", "negative", "validation", "edge_case" (optional)
        test_steps: Updated test steps (optional)
        expected_result: Updated expected result (optional)
        preconditions: Updated preconditions/prerequisites (optional)
        priority: MoSCoW priority - "Must Have", "Should Have", "Could Have", "Won't Have" (optional)
        compliance_framework: "Part11", "HIPAA", "SOC2", or None (optional)
        story_id: Reassign to different parent story (must be same program) (optional)
        change_reason: Why this change was made - recorded in audit trail (optional)

    Returns:
        Confirmation with list of updated fields showing old → new values

    WHY THIS APPROACH:
        - Sparse update: Only updates fields that are explicitly provided
        - Distinguishes from update_test_result which records EXECUTION outcomes
        - This tool is for changing WHAT the test does, not HOW it performed
        - Cross-program reassignment is blocked to maintain data integrity
        - Per-field audit logging for Part 11 compliance

    Example:
        update_test_case(
            test_id="P4M-CONFIG-001-TC05",
            title="TC Enabled Clinic Clinical Summary Functions Correctly",
            expected_result="Clinical summary generates with TC score, lifetime risk %, TC recommendations",
            change_reason="Narrowed scope to clinical summary only per story update"
        )
    """
    import sqlite3

    conn = None
    try:
        # ----------------------------------------------------------------
        # STEP 1: Validate test_type if provided
        # ----------------------------------------------------------------
        valid_test_types = ["happy_path", "negative", "validation", "edge_case"]
        if test_type is not None:
            test_type_lower = test_type.lower().strip()
            if test_type_lower not in valid_test_types:
                type_list = "\n".join(f"  • {t}" for t in valid_test_types)
                return (
                    f"Invalid test_type: '{test_type}'\n\n"
                    f"Valid test types:\n{type_list}\n\n"
                    f"Descriptions:\n"
                    f"  • happy_path - Normal successful scenario\n"
                    f"  • negative - Error handling, invalid inputs\n"
                    f"  • validation - Input/data validation rules\n"
                    f"  • edge_case - Boundary conditions, unusual scenarios"
                )
            test_type = test_type_lower  # Normalize to lowercase

        # ----------------------------------------------------------------
        # STEP 2: Validate priority if provided
        # ----------------------------------------------------------------
        valid_priorities = ["Must Have", "Should Have", "Could Have", "Won't Have"]
        if priority is not None and priority not in valid_priorities:
            return (
                f"Invalid priority: '{priority}'\n\n"
                f"Priority must be one of (MoSCoW):\n"
                f"  • Must Have - Critical test\n"
                f"  • Should Have - Important but not critical\n"
                f"  • Could Have - Nice to have\n"
                f"  • Won't Have - Out of scope for now"
            )

        # ----------------------------------------------------------------
        # STEP 3: Validate compliance_framework if provided
        # ----------------------------------------------------------------
        valid_frameworks = ["Part11", "HIPAA", "SOC2"]
        if compliance_framework is not None:
            # Normalize common variations
            framework_normalized = compliance_framework.strip()
            framework_upper = framework_normalized.upper().replace(" ", "").replace("-", "")

            # Map common variations to standard names
            framework_map = {
                "PART11": "Part11",
                "21CFRPART11": "Part11",
                "CFR11": "Part11",
                "HIPAA": "HIPAA",
                "SOC2": "SOC2",
                "SOCII": "SOC2",
                "SOC": "SOC2",
                "NONE": None  # Allow clearing the framework
            }

            if framework_upper in framework_map:
                compliance_framework = framework_map[framework_upper]
            elif framework_normalized not in valid_frameworks:
                return (
                    f"Invalid compliance_framework: '{compliance_framework}'\n\n"
                    f"Valid frameworks:\n"
                    f"  • Part11 - FDA 21 CFR Part 11\n"
                    f"  • HIPAA - Health Insurance Portability and Accountability\n"
                    f"  • SOC2 - Service Organization Control 2\n"
                    f"  • None - No compliance requirement (or use 'NONE' to clear)"
                )

        # ----------------------------------------------------------------
        # STEP 4: Connect to database and fetch existing test case
        # ----------------------------------------------------------------
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT t.*, s.title as story_title, p.name as program_name, p.prefix, p.program_id as parent_program_id
            FROM uat_test_cases t
            JOIN user_stories s ON t.story_id = s.story_id
            JOIN programs p ON t.program_id = p.program_id
            WHERE t.test_id = ?
            """,
            (test_id,)
        )
        existing = cursor.fetchone()

        if not existing:
            # Test not found - provide helpful suggestions
            parts = test_id.split('-')
            if len(parts) >= 3:
                story_prefix = '-'.join(parts[:-1])
                cursor.execute(
                    "SELECT test_id, title FROM uat_test_cases WHERE test_id LIKE ? ORDER BY test_id LIMIT 5",
                    (f"{story_prefix}-%",)
                )
                similar = cursor.fetchall()
                if similar:
                    similar_list = "\n".join(f"  • {t['test_id']}: {t['title'][:40]}" for t in similar)
                    return (
                        f"Test case not found: '{test_id}'\n\n"
                        f"Test cases for {story_prefix}:\n{similar_list}"
                    )

            return f"Test case not found: '{test_id}'"

        # Convert to dict for easier access
        old_test = dict(existing)
        current_program_id = old_test['program_id']
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ----------------------------------------------------------------
        # STEP 5: Validate story_id reassignment if provided
        # ----------------------------------------------------------------
        new_program_id = None
        if story_id is not None:
            # Validate new story exists
            cursor.execute(
                """
                SELECT s.story_id, s.title, s.program_id, p.prefix
                FROM user_stories s
                JOIN programs p ON s.program_id = p.program_id
                WHERE s.story_id = ?
                """,
                (story_id,)
            )
            new_story = cursor.fetchone()

            if not new_story:
                return (
                    f"New parent story not found: '{story_id}'\n\n"
                    f"Please verify the story_id exists before reassigning."
                )

            new_program_id = new_story['program_id']

            # Cross-program check: parse program prefix from test_id
            # Test ID format: {STORY_ID}-TC{NN} where STORY_ID = {PREFIX}-{CATEGORY}-{NUM}
            # e.g., "P4M-CONFIG-001-TC02" -> program prefix is "P4M"
            test_parts = test_id.split('-')
            if len(test_parts) >= 1:
                test_program_prefix = test_parts[0]
                new_story_prefix = new_story['prefix']

                if test_program_prefix.upper() != new_story_prefix.upper():
                    return (
                        f"Cross-program reassignment not allowed!\n\n"
                        f"Test case {test_id} belongs to program '{test_program_prefix}'\n"
                        f"but the new story {story_id} belongs to program '{new_story_prefix}'.\n\n"
                        f"Test cases cannot be moved between programs.\n"
                        f"If needed, delete this test case and create a new one in the target program."
                    )

        # ----------------------------------------------------------------
        # STEP 6: Build update fields and track changes
        # Only include fields that were provided (not None)
        # ----------------------------------------------------------------
        updates = {}
        changes = []  # List of (field_name, old_value, new_value) for audit

        # Check each optional field
        if title is not None and title != old_test.get('title'):
            updates['title'] = title
            changes.append(('title', old_test.get('title'), title))

        if test_type is not None and test_type != old_test.get('test_type'):
            updates['test_type'] = test_type
            changes.append(('test_type', old_test.get('test_type'), test_type))

        if test_steps is not None and test_steps != old_test.get('test_steps'):
            updates['test_steps'] = test_steps
            changes.append(('test_steps', old_test.get('test_steps'), test_steps))

        if expected_result is not None and expected_result != old_test.get('expected_results'):
            updates['expected_results'] = expected_result
            changes.append(('expected_results', old_test.get('expected_results'), expected_result))

        if preconditions is not None and preconditions != old_test.get('prerequisites'):
            updates['prerequisites'] = preconditions
            changes.append(('prerequisites', old_test.get('prerequisites'), preconditions))

        if priority is not None and priority != old_test.get('priority'):
            updates['priority'] = priority
            changes.append(('priority', old_test.get('priority'), priority))

        if compliance_framework is not None:
            # Handle clearing framework (set to None) vs updating
            old_framework = old_test.get('compliance_framework')
            if compliance_framework != old_framework:
                updates['compliance_framework'] = compliance_framework
                changes.append(('compliance_framework', old_framework, compliance_framework))

        if story_id is not None and story_id != old_test.get('story_id'):
            updates['story_id'] = story_id
            if new_program_id:
                updates['program_id'] = new_program_id
            changes.append(('story_id', old_test.get('story_id'), story_id))

        # ----------------------------------------------------------------
        # STEP 7: Check if there are any changes to make
        # ----------------------------------------------------------------
        if not updates:
            return (
                f"No changes to make for {test_id}\n\n"
                f"All provided values match the current values, or no update fields were provided.\n"
                f"Current test case:\n"
                f"  Title: {old_test.get('title')}\n"
                f"  Test Type: {old_test.get('test_type')}\n"
                f"  Priority: {old_test.get('priority')}\n"
                f"  Parent Story: {old_test.get('story_id')}"
            )

        # ----------------------------------------------------------------
        # STEP 8: Always update updated_date when changes occur
        # ----------------------------------------------------------------
        updates['updated_date'] = now

        # ----------------------------------------------------------------
        # STEP 9: Build and execute UPDATE statement
        # ----------------------------------------------------------------
        set_clause = ", ".join(f"{field} = ?" for field in updates.keys())
        values = list(updates.values()) + [test_id]

        cursor.execute(
            f"UPDATE uat_test_cases SET {set_clause} WHERE test_id = ?",
            values
        )

        # ----------------------------------------------------------------
        # STEP 10: Log each change to audit_history
        # ----------------------------------------------------------------
        audit_reason = change_reason or "Test case updated via MCP"

        for field_name, old_value, new_value in changes:
            # Truncate long values for audit log readability
            old_display = str(old_value)[:500] if old_value else None
            new_display = str(new_value)[:500] if new_value else None

            cursor.execute(
                """
                INSERT INTO audit_history (
                    record_type, record_id, action, field_changed,
                    old_value, new_value, changed_by, changed_date, change_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    'test_case',
                    test_id,
                    'Updated',
                    field_name,
                    old_display,
                    new_display,
                    'MCP:update_test_case',
                    now,
                    audit_reason
                )
            )

        conn.commit()

        # ----------------------------------------------------------------
        # STEP 11: Build success response
        # ----------------------------------------------------------------
        result = f"""Test case updated successfully!

Test Case: {test_id}
Parent Story: {old_test['story_id']} - {old_test['story_title'][:40]}
Program: {old_test['program_name']} [{old_test['prefix']}]

Changes made ({len(changes)} field(s)):
"""
        for field_name, old_val, new_val in changes:
            # Format display based on field type
            if field_name in ('test_steps', 'expected_results', 'prerequisites'):
                # For long text fields, just show that it was updated
                result += f"  • {field_name}: (updated)\n"
            else:
                old_str = str(old_val)[:40] if old_val else '(none)'
                new_str = str(new_val)[:40] if new_val else '(none)'
                result += f"  • {field_name}: '{old_str}' → '{new_str}'\n"

        if change_reason:
            result += f"\nChange Reason: {change_reason}\n"

        result += f"""
Updated At: {now}

Next Steps:
  • list_test_cases(story_id="{updates.get('story_id', old_test['story_id'])}") - View test cases
  • update_test_result(test_id="{test_id}", ...) - Record execution results
"""
        return result

    except Exception as e:
        # Explicit rollback for compliance audit trail clarity
        if conn:
            conn.rollback()
        return f"Error updating test case: {str(e)}"

    finally:
        if conn:
            conn.close()


# ============================================================
# REQUIREMENTS TOOLKIT - REPORTING TOOLS
# ============================================================

@mcp.tool()
def get_program_health(program_prefix: str) -> str:
    """
    Get health score and recommendations for a program.

    Args:
        program_prefix: Program prefix (e.g., "PROP")

    Returns:
        Health score (0-100), grade, and recommendations
    """
    db = None
    try:
        db = get_req_database()
        program = db.get_program_by_prefix(program_prefix.upper())

        if not program:
            return f"Program not found: {program_prefix}"

        health = req_queries.get_program_health_score(db, program['program_id'])

        result = f"Program Health: {program['name']}\n"
        result += f"{'=' * 50}\n\n"
        result += f"Overall Score: {health['score']}/100 (Grade: {health['grade']})\n\n"

        result += "Component Scores:\n"
        for name, data in health['components'].items():
            bar = "█" * (data['score'] // 10) + "░" * (10 - data['score'] // 10)
            result += f"  {name}: {bar} {data['score']}% (weight: {int(data['weight']*100)}%)\n"

        if health['recommendations']:
            result += f"\nRecommendations:\n"
            for rec in health['recommendations']:
                result += f"  • {rec}\n"

        return result

    except Exception as e:
        return f"Error getting health score: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def get_client_tree() -> str:
    """
    Get hierarchical view of all clients and their programs.

    Returns:
        Tree view: Clients → Programs with story/test counts
    """
    db = None
    try:
        db = get_req_database()
        tree = req_queries.get_client_program_tree(db)

        if not tree:
            return "No clients found in the database."

        result = "Client/Program Hierarchy:\n"
        result += "=" * 50 + "\n\n"

        for client in tree:
            result += f"📁 {client['name']}\n"
            programs = client.get('programs', [])
            if programs:
                for i, prog in enumerate(programs):
                    prefix = "└──" if i == len(programs) - 1 else "├──"
                    result += f"   {prefix} [{prog['prefix']}] {prog['name']}\n"
                    result += f"       Stories: {prog['story_count']} ({prog['approved_count']} approved) | Tests: {prog['test_count']}\n"
            else:
                result += "   (no programs)\n"
            result += "\n"

        return result

    except Exception as e:
        return f"Error getting client tree: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def search_stories(keyword: str) -> str:
    """
    Search for stories across all programs by keyword.

    Args:
        keyword: Search term to find in title, story, or acceptance criteria

    Returns:
        Matching stories with program context
    """
    db = None
    try:
        db = get_req_database()
        results = req_queries.search_stories_global(db, keyword)

        if not results:
            return f"No stories found matching: {keyword}"

        result = f"Search Results for '{keyword}':\n"
        result += f"Found {len(results)} matching story(ies)\n\n"

        for story in results:
            result += f"[{story['story_id']}] {story['title']}\n"
            result += f"  Program: {story['program_name']} [{story['prefix']}] | Client: {story['client_name']}\n"
            result += f"  Status: {story['status']} | Priority: {story.get('priority', 'N/A')}\n\n"

        return result

    except Exception as e:
        return f"Error searching stories: {str(e)}"
    finally:
        if db:
            db.close()


@mcp.tool()
def get_coverage_gaps(program_prefix: str) -> str:
    """
    Find requirements without stories and stories without tests.

    Args:
        program_prefix: Program prefix (e.g., "PROP")

    Returns:
        List of coverage gaps
    """
    db = None
    try:
        db = get_req_database()
        program = db.get_program_by_prefix(program_prefix.upper())

        if not program:
            return f"Program not found: {program_prefix}"

        orphan_reqs = req_queries.get_orphan_requirements(db, program['program_id'])
        untested_stories = req_queries.get_stories_without_tests(db, program['program_id'])

        result = f"Coverage Gaps for {program['name']}:\n"
        result += "=" * 50 + "\n\n"

        result += f"Requirements without Stories ({len(orphan_reqs)}):\n"
        if orphan_reqs:
            for req in orphan_reqs[:10]:
                result += f"  • [{req['requirement_id']}] {req.get('title', req.get('description', '')[:50])}\n"
            if len(orphan_reqs) > 10:
                result += f"  ... and {len(orphan_reqs) - 10} more\n"
        else:
            result += "  None - all requirements have stories!\n"

        result += f"\nApproved Stories without Tests ({len(untested_stories)}):\n"
        if untested_stories:
            for story in untested_stories[:10]:
                result += f"  • [{story['story_id']}] {story.get('title', 'Untitled')}\n"
            if len(untested_stories) > 10:
                result += f"  ... and {len(untested_stories) - 10} more\n"
        else:
            result += "  None - all approved stories have tests!\n"

        return result

    except Exception as e:
        return f"Error getting coverage gaps: {str(e)}"
    finally:
        if db:
            db.close()


# ============================================================
# EXCEL EXPORT HELPER FUNCTIONS
# ============================================================

def _create_review_status_excel(rows: list, output_dir: str, filename: str) -> str:
    """Create formatted Excel file for review status report with manager response columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Status"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    due_soon_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    response_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for response columns
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if rows:
        # Add manager response columns to each row
        for row in rows:
            row['Action'] = ''  # Blank = Keep, or "Terminate" or "Update"
            row['New Role'] = ''  # Only if Action = Update
            row['Manager Notes'] = ''  # Reason for change

        headers = list(rows[0].keys())

        # Find which columns are response columns
        response_cols = ['Action', 'New Role', 'Manager Notes']

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Write data rows
        for row_num, row_data in enumerate(rows, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=row_data.get(header, ''))
                cell.border = thin_border

                # Apply status-based highlighting to non-response columns
                if header not in response_cols:
                    if row_data.get('Review Status') == 'OVERDUE':
                        cell.fill = overdue_fill
                    elif row_data.get('Review Status') == 'Due Soon':
                        cell.fill = due_soon_fill
                else:
                    # Light green background for response columns (to fill in)
                    cell.fill = response_fill

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(rows) + 2):
                try:
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Make response columns wider for easier data entry
        for col, header in enumerate(headers, 1):
            if header in response_cols:
                ws.column_dimensions[get_column_letter(col)].width = 20

    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions_ws['A1'] = "Annual Access Review Instructions"
    instructions_ws['A1'].font = Font(bold=True, size=14)
    instructions_ws['A3'] = "For each user, review their access and fill in the green columns:"
    instructions_ws['A5'] = "ACTION Column:"
    instructions_ws['B5'] = "Leave BLANK to keep current access (recertify)"
    instructions_ws['B6'] = "Enter 'Terminate' to revoke access"
    instructions_ws['B7'] = "Enter 'Update' to change their role"
    instructions_ws['A9'] = "NEW ROLE Column (only if Action = Update):"
    instructions_ws['B9'] = "Read-Write-Order"
    instructions_ws['B10'] = "Read-Write"
    instructions_ws['B11'] = "Read-Only"
    instructions_ws['A13'] = "MANAGER NOTES Column:"
    instructions_ws['B13'] = "Required for Terminate or Update actions"
    instructions_ws['B14'] = "Example: 'Left organization', 'Role change per manager request'"
    instructions_ws['A16'] = "When complete, save the file and send back for import."
    instructions_ws.column_dimensions['A'].width = 35
    instructions_ws.column_dimensions['B'].width = 50

    # Freeze header row on main sheet
    ws.freeze_panes = 'A2'

    filepath = os.path.join(output_dir, f"{filename}.xlsx")
    wb.save(filepath)
    return filepath


def _create_terminated_audit_excel(rows: list, output_dir: str, filename: str, violations_count: int) -> str:
    """Create formatted Excel file for terminated user audit."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Terminated Audit"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    violation_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if rows:
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_num, row_data in enumerate(rows, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=row_data.get(header, ''))
                cell.border = thin_border
                compliance_status = row_data.get('Compliance Status', '')
                if 'VIOLATION' in compliance_status:
                    cell.fill = violation_fill
                elif 'Compliant' in compliance_status:
                    cell.fill = compliant_fill
                if header == 'Access Still Active' and 'VIOLATION' in str(row_data.get(header, '')):
                    cell.fill = violation_fill
                    cell.font = Font(bold=True, color="9C0006")

        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(rows) + 2):
                try:
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    ws.freeze_panes = 'A2'

    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = "Terminated User Audit Summary"
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A3'] = "Generated:"
    summary_ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    summary_ws['A4'] = "Total Users:"
    summary_ws['B4'] = len(set(r.get('Email') for r in rows if r.get('Email')))
    summary_ws['A5'] = "Violations:"
    summary_ws['B5'] = violations_count
    if violations_count > 0:
        summary_ws['B5'].fill = violation_fill
        summary_ws['C5'] = "ACTION REQUIRED"
        summary_ws['C5'].font = Font(bold=True, color="9C0006")
    else:
        summary_ws['B5'].fill = compliant_fill
        summary_ws['C5'] = "All access properly revoked"
        summary_ws['C5'].font = Font(color="006100")

    filepath = os.path.join(output_dir, f"{filename}.xlsx")
    wb.save(filepath)
    return filepath


# ============================================================
# COMPLIANCE EXCEL EXPORT TOOLS
# ============================================================

@mcp.tool()
def export_review_status(
    program: Optional[str] = None,
    include_current: bool = False,
    output_format: str = "xlsx",
    output_dir: Optional[str] = None
) -> str:
    """
    Export access review status to Excel for compliance tracking.

    Args:
        program: Filter by program name or prefix
        include_current: If True, include reviews that are current (not just issues)
        output_format: "xlsx" or "csv"
        output_dir: Directory to save file (default: ~/Downloads)

    Returns:
        Path to generated file and summary statistics
    """
    manager = None
    try:
        manager = get_access_manager()

        program_id = None
        program_name = None
        if program:
            try:
                program_id = manager._resolve_program_id(program)
                cursor = manager.conn.cursor()
                cursor.execute("SELECT name FROM programs WHERE program_id = ?", (program_id,))
                row = cursor.fetchone()
                if row:
                    program_name = row['name']
            except ValueError:
                return f"Program not found: {program}"

        try:
            status_data = manager.get_review_status_detail(
                program_id=program_id,
                include_current=include_current
            )
        except AttributeError:
            reviews_due = manager.get_reviews_due(program_id=program_id)
            today = date.today()
            due_soon_cutoff = (today + timedelta(days=30)).isoformat()
            today_str = today.isoformat()

            status_data = {
                'overdue': [],
                'due_soon': [],
                'current': [],
                'summary': {'overdue_count': 0, 'due_soon_count': 0, 'current_count': 0, 'as_of_date': today_str}
            }

            for review in reviews_due:
                next_due = review.get('next_review_due')
                if next_due and next_due <= today_str:
                    status_data['overdue'].append(review)
                elif next_due and next_due <= due_soon_cutoff:
                    status_data['due_soon'].append(review)

            status_data['summary']['overdue_count'] = len(status_data['overdue'])
            status_data['summary']['due_soon_count'] = len(status_data['due_soon'])

        has_issues = (status_data['summary']['overdue_count'] > 0 or
                      status_data['summary']['due_soon_count'] > 0)

        if not has_issues and not include_current:
            return (
                "All access reviews are current!\n\n"
                f"Summary (as of {status_data['summary']['as_of_date']}):\n"
                f"  Overdue: 0\n"
                f"  Due Soon: 0\n"
                f"  Current: {status_data['summary'].get('current_count', 'N/A')}\n\n"
                "No export generated - use include_current=True for full export."
            )

        if output_dir:
            out_path = os.path.expanduser(output_dir)
        else:
            out_path = os.path.expanduser("~/Downloads")
        os.makedirs(out_path, exist_ok=True)

        today_str = datetime.now().strftime('%Y-%m-%d')
        if program_name:
            safe_name = program_name.replace(' ', '_').replace('/', '-')
            filename = f"{safe_name}_review_status_{today_str}"
        else:
            filename = f"access_review_status_{today_str}"

        export_rows = []

        for item in status_data['overdue']:
            export_rows.append({
                'Review Status': 'OVERDUE',
                'Days Overdue': item.get('days_overdue', ''),
                'User Name': item.get('user_name', ''),
                'Email': item.get('email', ''),
                'Role': item.get('role', ''),
                'Program': item.get('program_name', ''),
                'Clinic': item.get('clinic_name', ''),
                'Location': item.get('location_name', ''),
                'Granted Date': item.get('granted_date', ''),
                'Last Review': item.get('last_review_date', ''),
                'Next Review Due': item.get('next_review_due', ''),
                'Review Cycle': item.get('review_cycle', ''),
                'Action Needed': ''
            })

        for item in status_data['due_soon']:
            export_rows.append({
                'Review Status': 'Due Soon',
                'Days Overdue': '',
                'User Name': item.get('user_name', ''),
                'Email': item.get('email', ''),
                'Role': item.get('role', ''),
                'Program': item.get('program_name', ''),
                'Clinic': item.get('clinic_name', ''),
                'Location': item.get('location_name', ''),
                'Granted Date': item.get('granted_date', ''),
                'Last Review': item.get('last_review_date', ''),
                'Next Review Due': item.get('next_review_due', ''),
                'Review Cycle': item.get('review_cycle', ''),
                'Action Needed': ''
            })

        if include_current:
            for item in status_data.get('current', []):
                export_rows.append({
                    'Review Status': 'Current',
                    'Days Overdue': '',
                    'User Name': item.get('user_name', ''),
                    'Email': item.get('email', ''),
                    'Role': item.get('role', ''),
                    'Program': item.get('program_name', ''),
                    'Clinic': item.get('clinic_name', ''),
                    'Location': item.get('location_name', ''),
                    'Granted Date': item.get('granted_date', ''),
                    'Last Review': item.get('last_review_date', ''),
                    'Next Review Due': item.get('next_review_due', ''),
                    'Review Cycle': item.get('review_cycle', ''),
                    'Action Needed': ''
                })

        if not export_rows:
            return "No data to export."

        if output_format.lower() == "xlsx":
            filepath = _create_review_status_excel(export_rows, out_path, filename)
        else:
            filepath = os.path.join(out_path, f"{filename}.csv")
            headers = list(export_rows[0].keys())
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(export_rows)

        result = f"Access Review Status Export\n"
        result += f"{'=' * 50}\n\n"
        result += f"File: {filepath}\n"
        result += f"Generated: {today_str}\n\n"
        if program_name:
            result += f"Program: {program_name}\n\n"
        result += f"Summary:\n"
        result += f"  OVERDUE: {status_data['summary']['overdue_count']}"
        if status_data['summary']['overdue_count'] > 0:
            result += " - ACTION REQUIRED"
        result += "\n"
        result += f"  Due Soon: {status_data['summary']['due_soon_count']}\n"
        result += f"  Total exported: {len(export_rows)}\n"

        return result

    except Exception as e:
        return f"Error exporting review status: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def import_review_response(
    file_path: str,
    reviewed_by: str,
    preview_only: bool = True
) -> str:
    """
    Import completed annual access review responses from clinic manager.

    Processes the Excel file that was exported with export_review_status
    and filled in by the clinic manager.

    Action column values:
    - Blank or empty: Recertify (confirm access, set next review date)
    - "Terminate": Revoke access (Manager Notes required)
    - "Update": Change role to New Role value (New Role required)

    Args:
        file_path: Path to completed Excel review file
        reviewed_by: Name of clinic manager who completed the review
        preview_only: If True (default), show what WOULD happen without making changes.
                     Set to False to actually process the review.

    Returns:
        Summary of review actions taken or planned
    """
    manager = None
    try:
        manager = get_access_manager()

        result = manager.process_review_response(
            file_path=file_path,
            reviewed_by=reviewed_by,
            preview_only=preview_only
        )

        # Build response
        mode = "PREVIEW MODE (no changes made)" if preview_only else "REVIEW IMPORT COMPLETE"

        output = f"Annual Access Review Import - {mode}\n"
        output += f"{'=' * 50}\n\n"
        output += f"File: {file_path}\n"
        output += f"Reviewed By: {reviewed_by}\n\n"

        # Summary
        summary = result['summary']
        output += f"Summary:\n"
        output += f"  Total rows: {summary['total_rows']}\n"
        output += f"  Recertified (keep): {summary['recertified']}\n"
        output += f"  Updated (role change): {summary['updated']}\n"
        output += f"  Terminated (revoked): {summary['terminated']}\n"
        output += f"  Skipped: {summary['skipped']}\n"
        output += f"  Errors: {summary['errors']}\n\n"

        # Show preview actions
        if result['preview']:
            output += f"Actions:\n"
            for action in result['preview'][:30]:
                icon = {
                    'RECERTIFY': '[KEEP]',
                    'UPDATE': '[UPDATE]',
                    'TERMINATE': '[REVOKE]'
                }.get(action['action'], '-')

                output += f"  {icon} {action['name']} ({action['email']})\n"
                output += f"      {action['action']}: {action['details']}\n"

            if len(result['preview']) > 30:
                remaining = len(result['preview']) - 30
                output += f"\n  ... and {remaining} more actions\n"

        # Show errors
        if result['errors']:
            output += f"\nErrors:\n"
            for err in result['errors'][:10]:
                output += f"  Row {err['row']}: {err['email']} - {err['error']}\n"
            if len(result['errors']) > 10:
                output += f"  ... and {len(result['errors']) - 10} more errors\n"

        # Next step hint
        if preview_only:
            total_actions = summary['recertified'] + summary['updated'] + summary['terminated']
            if total_actions > 0:
                output += f"\nTo apply these changes, run again with preview_only=False"

        return output

    except FileNotFoundError as e:
        return f"File not found: {file_path}\n\nMake sure the file exists and the path is correct."
    except Exception as e:
        return f"Error processing review: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def export_terminated_audit(
    since_date: Optional[str] = None,
    output_format: str = "xlsx",
    output_dir: Optional[str] = None
) -> str:
    """
    Export full terminated user audit to Excel.

    Args:
        since_date: Only include users terminated on/after this date (YYYY-MM-DD)
        output_format: "xlsx" or "csv"
        output_dir: Directory to save file (default: ~/Downloads)

    Returns:
        Path to generated file and compliance summary
    """
    manager = None
    try:
        manager = get_access_manager()

        try:
            terminated_users = manager.get_all_terminated_users(
                include_access_history=True,
                since_date=since_date
            )
        except AttributeError:
            cursor = manager.conn.cursor()
            query = """
                SELECT u.user_id, u.name, u.email, u.organization, u.updated_date as termination_date
                FROM users u WHERE u.status = 'Terminated'
            """
            params = []
            if since_date:
                query += " AND u.updated_date >= ?"
                params.append(since_date)
            query += " ORDER BY u.updated_date DESC"

            cursor.execute(query, params)
            terminated_users = []

            for row in cursor.fetchall():
                user = dict(row)
                user['terminated_by'] = 'See audit log'
                user['termination_reason'] = 'See audit log'

                cursor.execute("""
                    SELECT ua.access_id, ua.role,
                           CASE WHEN ua.status = 'Active' THEN 1 ELSE 0 END as is_active,
                           ua.granted_date, ua.revoked_date, ua.revoked_by, ua.revoke_reason,
                           p.name as program_name, c.name as clinic_name, l.name as location_name
                    FROM user_access ua
                    JOIN programs p ON ua.program_id = p.program_id
                    LEFT JOIN clinics c ON ua.clinic_id = c.clinic_id
                    LEFT JOIN locations l ON ua.location_id = l.location_id
                    WHERE ua.user_id = ?
                """, (user['user_id'],))

                access_history = [dict(r) for r in cursor.fetchall()]
                user['access_history'] = access_history
                active_count = sum(1 for a in access_history if a['is_active'])
                user['active_access_count'] = active_count
                user['compliance_status'] = 'VIOLATION' if active_count > 0 else 'Compliant'
                terminated_users.append(user)

        if not terminated_users:
            period = f" since {since_date}" if since_date else ""
            return f"No terminated users found{period}."

        if output_dir:
            out_path = os.path.expanduser(output_dir)
        else:
            out_path = os.path.expanduser("~/Downloads")
        os.makedirs(out_path, exist_ok=True)

        today_str = datetime.now().strftime('%Y-%m-%d')
        if since_date:
            filename = f"terminated_audit_{since_date}_to_{today_str}"
        else:
            filename = f"terminated_audit_full_{today_str}"

        export_rows = []
        violations_count = 0

        for user in terminated_users:
            is_violation = user['compliance_status'] == 'VIOLATION'
            if is_violation:
                violations_count += 1

            access_history = user.get('access_history', [])

            if not access_history:
                export_rows.append({
                    'Compliance Status': 'Compliant' if not is_violation else 'VIOLATION',
                    'User Name': user.get('name', ''),
                    'Email': user.get('email', ''),
                    'Organization': user.get('organization', ''),
                    'Termination Date': user.get('termination_date', ''),
                    'Terminated By': user.get('terminated_by', ''),
                    'Termination Reason': user.get('termination_reason', ''),
                    'Previous Role': '(No access grants)',
                    'Previous Program': '',
                    'Previous Clinic': '',
                    'Access Granted': '',
                    'Access Revoked': '',
                    'Revoked By': '',
                    'Revocation Reason': '',
                    'Access Still Active': 'No'
                })
            else:
                for i, access in enumerate(access_history):
                    export_rows.append({
                        'Compliance Status': ('Compliant' if not is_violation else 'VIOLATION') if i == 0 else '',
                        'User Name': user.get('name', '') if i == 0 else '',
                        'Email': user.get('email', '') if i == 0 else '',
                        'Organization': user.get('organization', '') if i == 0 else '',
                        'Termination Date': user.get('termination_date', '') if i == 0 else '',
                        'Terminated By': user.get('terminated_by', '') if i == 0 else '',
                        'Termination Reason': user.get('termination_reason', '') if i == 0 else '',
                        'Previous Role': access.get('role', ''),
                        'Previous Program': access.get('program_name', ''),
                        'Previous Clinic': access.get('clinic_name', ''),
                        'Access Granted': access.get('granted_date', ''),
                        'Access Revoked': access.get('revoked_date', '') if not access.get('is_active') else '',
                        'Revoked By': access.get('revoked_by', '') if not access.get('is_active') else '',
                        'Revocation Reason': access.get('revoke_reason', '') if not access.get('is_active') else '',
                        'Access Still Active': 'YES - VIOLATION!' if access.get('is_active') else 'No'
                    })

        if output_format.lower() == "xlsx":
            filepath = _create_terminated_audit_excel(export_rows, out_path, filename, violations_count)
        else:
            filepath = os.path.join(out_path, f"{filename}.csv")
            headers = list(export_rows[0].keys())
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(export_rows)

        result = f"Terminated User Audit Export\n"
        result += f"{'=' * 50}\n\n"
        result += f"File: {filepath}\n"
        result += f"Generated: {today_str}\n"
        if since_date:
            result += f"Period: {since_date} to present\n"
        result += "\n"
        result += f"Summary:\n"
        result += f"  Total terminated users: {len(terminated_users)}\n"
        result += f"  Compliant: {len(terminated_users) - violations_count}\n"
        result += f"  VIOLATIONS: {violations_count}"
        if violations_count > 0:
            result += " - IMMEDIATE ACTION REQUIRED"
        result += "\n"

        return result

    except Exception as e:
        return f"Error exporting terminated audit: {str(e)}"
    finally:
        if manager:
            manager.close()


# ============================================================
# DASHBOARD TOOLS
# ============================================================

def _generate_dashboard_react(data: dict) -> str:
    """Generate React component code for visual dashboard."""

    # Prepare data for the component
    programs_data = json.dumps([
        {"name": p["program_name"], "users": p["user_count"]}
        for p in data["users_by_program"]
    ])

    roles_data = json.dumps([
        {"name": r["role"], "count": r["count"], "percentage": r["percentage"]}
        for r in data["role_distribution"]
    ])

    imm = data["immediate_attention"]
    upcoming = data["upcoming"]
    totals = data["totals"]

    return f'''
import React from 'react';
import {{ BarChart, Bar, XAxis, YAxis, Tooltip, ResponsiveContainer, PieChart, Pie, Cell }} from 'recharts';

export default function ComplianceDashboard() {{
  const programData = {programs_data};
  const roleData = {roles_data};
  const COLORS = ['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000', '#5B9BD5'];

  return (
    <div className="p-6 bg-gray-50 min-h-screen">
      <h1 className="text-2xl font-bold text-gray-800 mb-6">Compliance Dashboard</h1>

      {{/* Alert Cards */}}
      <div className="grid grid-cols-1 md:grid-cols-4 gap-4 mb-6">
        <div className={{"p-4 rounded-lg " + ({imm['reviews_overdue']} > 0 ? "bg-red-100 border-2 border-red-400" : "bg-green-100")}}>
          <div className="text-3xl font-bold">{imm['reviews_overdue']}</div>
          <div className="text-sm text-gray-600">Reviews Overdue</div>
        </div>
        <div className={{"p-4 rounded-lg " + ({imm['terminated_violations']} > 0 ? "bg-red-100 border-2 border-red-400" : "bg-green-100")}}>
          <div className="text-3xl font-bold">{imm['terminated_violations']}</div>
          <div className="text-sm text-gray-600">Access Violations</div>
        </div>
        <div className={{"p-4 rounded-lg " + ({imm['training_expired']} > 0 ? "bg-red-100 border-2 border-red-400" : "bg-green-100")}}>
          <div className="text-3xl font-bold">{imm['training_expired']}</div>
          <div className="text-sm text-gray-600">Training Expired</div>
        </div>
        <div className="p-4 rounded-lg bg-yellow-100">
          <div className="text-3xl font-bold">{upcoming['reviews_due_soon']}</div>
          <div className="text-sm text-gray-600">Reviews Due (30 days)</div>
        </div>
      </div>

      {{/* Stats Row */}}
      <div className="grid grid-cols-4 gap-4 mb-6">
        <div className="bg-white p-4 rounded-lg shadow">
          <div className="text-2xl font-bold text-blue-600">{totals['active_users']}</div>
          <div className="text-sm text-gray-500">Active Users</div>
        </div>
        <div className="bg-white p-4 rounded-lg shadow">
          <div className="text-2xl font-bold text-blue-600">{totals['total_programs']}</div>
          <div className="text-sm text-gray-500">Programs</div>
        </div>
        <div className="bg-white p-4 rounded-lg shadow">
          <div className="text-2xl font-bold text-blue-600">{totals['total_clinics']}</div>
          <div className="text-sm text-gray-500">Clinics</div>
        </div>
        <div className="bg-white p-4 rounded-lg shadow">
          <div className="text-2xl font-bold text-blue-600">{totals['recent_grants_count']}</div>
          <div className="text-sm text-gray-500">Grants (30 days)</div>
        </div>
      </div>

      {{/* Charts Row */}}
      <div className="grid grid-cols-1 md:grid-cols-2 gap-6">
        <div className="bg-white p-4 rounded-lg shadow">
          <h3 className="font-semibold mb-4">Users by Program</h3>
          <ResponsiveContainer width="100%" height={{200}}>
            <BarChart data={{programData}}>
              <XAxis dataKey="name" tick={{{{fontSize: 12}}}} />
              <YAxis />
              <Tooltip />
              <Bar dataKey="users" fill="#4472C4" />
            </BarChart>
          </ResponsiveContainer>
        </div>

        <div className="bg-white p-4 rounded-lg shadow">
          <h3 className="font-semibold mb-4">Role Distribution</h3>
          <ResponsiveContainer width="100%" height={{200}}>
            <PieChart>
              <Pie
                data={{roleData}}
                dataKey="count"
                nameKey="name"
                cx="50%"
                cy="50%"
                outerRadius={{70}}
                label={{({{name, percentage}}) => `${{name}} (${{percentage}}%)`}}
              >
                {{roleData.map((entry, index) => (
                  <Cell key={{index}} fill={{COLORS[index % COLORS.length]}} />
                ))}}
              </Pie>
              <Tooltip />
            </PieChart>
          </ResponsiveContainer>
        </div>
      </div>
    </div>
  );
}}
'''


@mcp.tool()
def get_compliance_dashboard() -> str:
    """
    Get compliance dashboard with text summary and visual artifact.

    Shows:
    - Immediate attention items (overdue reviews, violations, expired training)
    - Upcoming reviews (next 30 days)
    - Users by program and clinic
    - Role distribution
    - Recent activity

    Returns:
        Text summary followed by React artifact code for visual dashboard
    """
    manager = None
    try:
        manager = get_access_manager()
        data = manager.get_dashboard_data()

        # Build text summary
        output = "COMPLIANCE DASHBOARD\n"
        output += f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}\n"
        output += "=" * 50 + "\n\n"

        # Immediate attention
        imm = data['immediate_attention']
        has_issues = imm['reviews_overdue'] > 0 or imm['terminated_violations'] > 0 or imm['training_expired'] > 0

        if has_issues:
            output += "IMMEDIATE ATTENTION\n"
            if imm['reviews_overdue'] > 0:
                output += f"   - {imm['reviews_overdue']} access reviews OVERDUE\n"
            if imm['terminated_violations'] > 0:
                output += f"   - {imm['terminated_violations']} terminated user(s) with ACTIVE ACCESS\n"
            if imm['training_expired'] > 0:
                output += f"   - {imm['training_expired']} training record(s) EXPIRED\n"
            output += "\n"
        else:
            output += "NO IMMEDIATE ISSUES\n"
            output += "   All reviews current, no violations, training up to date\n\n"

        # Upcoming
        upcoming = data['upcoming']['reviews_due_soon']
        if upcoming > 0:
            output += f"UPCOMING (Next 30 Days)\n"
            output += f"   - {upcoming} review(s) due soon\n\n"

        # Totals
        totals = data['totals']
        output += "TOTALS\n"
        output += f"   - {totals['active_users']} active users\n"
        output += f"   - {totals['total_programs']} programs\n"
        output += f"   - {totals['total_clinics']} clinics\n"
        output += f"   - {totals['recent_grants_count']} access grants in last 30 days\n\n"

        # Users by program
        output += "USERS BY PROGRAM\n"
        for prog in data['users_by_program']:
            output += f"   - {prog['program_name']}: {prog['user_count']} users, {prog['clinic_count']} clinics\n"
        output += "\n"

        # Role distribution
        output += "ROLE DISTRIBUTION\n"
        for role in data['role_distribution']:
            output += f"   - {role['role']}: {role['count']} ({role['percentage']}%)\n"
        output += "\n"

        # Recent activity (top 5)
        if data['recent_activity']:
            output += "RECENT ACTIVITY (Last 30 Days)\n"
            for i, activity in enumerate(data['recent_activity'][:5]):
                output += f"   - {activity['granted_date']}: {activity['user_name']} -> {activity['program_name']}"
                if activity['clinic_name'] != '(Program-wide)':
                    output += f" / {activity['clinic_name']}"
                output += f" ({activity['role']})\n"
            if len(data['recent_activity']) > 5:
                output += f"   ... and {len(data['recent_activity']) - 5} more\n"

        # Add React artifact instructions
        output += "\n" + "=" * 50 + "\n"
        output += "VISUAL DASHBOARD\n"
        output += "Creating React artifact with charts...\n\n"

        # Generate React artifact code
        react_code = _generate_dashboard_react(data)
        output += "```jsx\n" + react_code + "\n```"

        return output

    except Exception as e:
        return f"Error generating dashboard: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def push_dashboard_to_notion() -> str:
    """
    Update the Notion compliance dashboard page with current data.

    Updates the page at: https://www.notion.so/2dab5d1d163181bb8eebc4f4397de747

    Returns:
        Confirmation of update with summary
    """
    manager = None
    try:
        manager = get_access_manager()
        data = manager.get_dashboard_data()

        # Build Notion-flavored markdown content
        now = datetime.now().strftime('%B %d, %Y at %I:%M %p')

        content = f"""# Client Access Compliance Dashboard

**Auto-Updated by Claude MCP**
This dashboard is automatically updated when you run "push compliance dashboard to Notion" in Claude.

**Last Updated:** {now}

---

## Immediate Attention

| Metric | Count | Status |
|--------|-------|--------|
| Access Reviews Overdue | {data['immediate_attention']['reviews_overdue']} | {'ACTION REQUIRED' if data['immediate_attention']['reviews_overdue'] > 0 else 'OK'} |
| Terminated with Active Access | {data['immediate_attention']['terminated_violations']} | {'VIOLATION' if data['immediate_attention']['terminated_violations'] > 0 else 'OK'} |
| Training Expired | {data['immediate_attention']['training_expired']} | {'ACTION REQUIRED' if data['immediate_attention']['training_expired'] > 0 else 'OK'} |

---

## Upcoming (Next 30 Days)

| Metric | Count |
|--------|-------|
| Reviews Due Soon | {data['upcoming']['reviews_due_soon']} |

---

## Program Summary

| Program | Active Users | Clinics |
|---------|--------------|---------|
"""
        for prog in data['users_by_program']:
            content += f"| {prog['program_name']} | {prog['user_count']} | {prog['clinic_count']} |\n"

        content += """
---

## Users by Clinic

| Clinic | Program | Users |
|--------|---------|-------|
"""
        for clinic in data['users_by_clinic'][:15]:  # Top 15
            content += f"| {clinic['clinic_name']} | {clinic['program_name']} | {clinic['user_count']} |\n"

        content += """
---

## Role Distribution

| Role | Count | % |
|------|-------|---|
"""
        for role in data['role_distribution']:
            content += f"| {role['role']} | {role['count']} | {role['percentage']}% |\n"

        content += """
---

## Recent Activity (Last 30 Days)

| Date | User | Program | Clinic | Role |
|------|------|---------|--------|------|
"""
        for activity in data['recent_activity'][:10]:  # Top 10
            content += f"| {activity['granted_date']} | {activity['user_name']} | {activity['program_name']} | {activity['clinic_name']} | {activity['role']} |\n"

        content += f"""
---

## Totals

| Metric | Value |
|--------|-------|
| Total Active Users | {data['totals']['active_users']} |
| Total Programs | {data['totals']['total_programs']} |
| Total Clinics | {data['totals']['total_clinics']} |
| Access Grants (Last 30 Days) | {data['totals']['recent_grants_count']} |
"""

        # Return the content for Notion update
        # The actual Notion update will be handled by the Notion MCP tools
        return f"""Dashboard data ready for Notion update.

**Page ID:** {NOTION_DASHBOARD_PAGE_ID}
**URL:** https://www.notion.so/{NOTION_DASHBOARD_PAGE_ID.replace('-', '')}
**Last Updated:** {now}

**Summary:**
- Reviews Overdue: {data['immediate_attention']['reviews_overdue']}
- Violations: {data['immediate_attention']['terminated_violations']}
- Training Expired: {data['immediate_attention']['training_expired']}
- Active Users: {data['totals']['active_users']}

To update Notion, I need to call the Notion update-page tool with this content.

---NOTION_CONTENT_START---
{content}
---NOTION_CONTENT_END---
"""

    except Exception as e:
        return f"Error preparing dashboard for Notion: {str(e)}"
    finally:
        if manager:
            manager.close()


@mcp.tool()
def export_compliance_dashboard(
    output_dir: Optional[str] = None
) -> str:
    """
    Export compliance dashboard to Excel for stakeholders.

    Creates a formatted Excel file with multiple sheets:
    - Summary (all metrics)
    - Reviews Overdue (detail)
    - Users by Program
    - Users by Clinic
    - Role Distribution
    - Recent Activity

    Args:
        output_dir: Directory to save file (default: ~/Downloads)

    Returns:
        Path to generated file
    """
    manager = None
    try:
        manager = get_access_manager()
        data = manager.get_dashboard_data()

        if output_dir:
            out_path = os.path.expanduser(output_dir)
        else:
            out_path = os.path.expanduser("~/Downloads")
        os.makedirs(out_path, exist_ok=True)

        today_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"compliance_dashboard_{today_str}.xlsx"
        filepath = os.path.join(out_path, filename)

        wb = Workbook()

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # --- SUMMARY SHEET ---
        ws = wb.active
        ws.title = "Summary"

        ws['A1'] = "Compliance Dashboard Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"

        ws['A4'] = "IMMEDIATE ATTENTION"
        ws['A4'].font = Font(bold=True, size=12)

        metrics = [
            ('Access Reviews Overdue', data['immediate_attention']['reviews_overdue']),
            ('Terminated with Active Access', data['immediate_attention']['terminated_violations']),
            ('Training Expired', data['immediate_attention']['training_expired']),
        ]

        row = 5
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            cell = ws.cell(row=row, column=2, value=value)
            cell.fill = alert_fill if value > 0 else ok_fill
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="UPCOMING (Next 30 Days)").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Reviews Due Soon")
        ws.cell(row=row, column=2, value=data['upcoming']['reviews_due_soon'])

        row += 2
        ws.cell(row=row, column=1, value="TOTALS").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Active Users")
        ws.cell(row=row, column=2, value=data['totals']['active_users'])
        row += 1
        ws.cell(row=row, column=1, value="Total Programs")
        ws.cell(row=row, column=2, value=data['totals']['total_programs'])
        row += 1
        ws.cell(row=row, column=1, value="Total Clinics")
        ws.cell(row=row, column=2, value=data['totals']['total_clinics'])
        row += 1
        ws.cell(row=row, column=1, value="Access Grants (Last 30 Days)")
        ws.cell(row=row, column=2, value=data['totals']['recent_grants_count'])

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

        # --- USERS BY PROGRAM SHEET ---
        ws2 = wb.create_sheet("Users by Program")
        headers = ['Program', 'Prefix', 'Active Users', 'Clinics']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, prog in enumerate(data['users_by_program'], 2):
            ws2.cell(row=row_num, column=1, value=prog['program_name']).border = thin_border
            ws2.cell(row=row_num, column=2, value=prog.get('program_prefix', '')).border = thin_border
            ws2.cell(row=row_num, column=3, value=prog['user_count']).border = thin_border
            ws2.cell(row=row_num, column=4, value=prog['clinic_count']).border = thin_border

        ws2.column_dimensions['A'].width = 25
        ws2.freeze_panes = 'A2'

        # --- USERS BY CLINIC SHEET ---
        ws3 = wb.create_sheet("Users by Clinic")
        headers = ['Clinic', 'Program', 'Active Users']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, clinic in enumerate(data['users_by_clinic'], 2):
            ws3.cell(row=row_num, column=1, value=clinic['clinic_name']).border = thin_border
            ws3.cell(row=row_num, column=2, value=clinic['program_name']).border = thin_border
            ws3.cell(row=row_num, column=3, value=clinic['user_count']).border = thin_border

        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 25
        ws3.freeze_panes = 'A2'

        # --- ROLE DISTRIBUTION SHEET ---
        ws4 = wb.create_sheet("Role Distribution")
        headers = ['Role', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, role in enumerate(data['role_distribution'], 2):
            ws4.cell(row=row_num, column=1, value=role['role']).border = thin_border
            ws4.cell(row=row_num, column=2, value=role['count']).border = thin_border
            ws4.cell(row=row_num, column=3, value=f"{role['percentage']}%").border = thin_border

        ws4.column_dimensions['A'].width = 20
        ws4.freeze_panes = 'A2'

        # --- RECENT ACTIVITY SHEET ---
        ws5 = wb.create_sheet("Recent Activity")
        headers = ['Date', 'User', 'Email', 'Role', 'Program', 'Clinic', 'Granted By']
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, activity in enumerate(data['recent_activity'], 2):
            ws5.cell(row=row_num, column=1, value=activity['granted_date']).border = thin_border
            ws5.cell(row=row_num, column=2, value=activity['user_name']).border = thin_border
            ws5.cell(row=row_num, column=3, value=activity['email']).border = thin_border
            ws5.cell(row=row_num, column=4, value=activity['role']).border = thin_border
            ws5.cell(row=row_num, column=5, value=activity['program_name']).border = thin_border
            ws5.cell(row=row_num, column=6, value=activity['clinic_name']).border = thin_border
            ws5.cell(row=row_num, column=7, value=activity['granted_by']).border = thin_border

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws5.column_dimensions[col].width = 20
        ws5.freeze_panes = 'A2'

        wb.save(filepath)

        # Build summary
        output = f"Compliance Dashboard Export\n"
        output += f"{'=' * 50}\n\n"
        output += f"File: {filepath}\n"
        output += f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}\n\n"
        output += f"Sheets included:\n"
        output += f"  - Summary\n"
        output += f"  - Users by Program ({len(data['users_by_program'])} programs)\n"
        output += f"  - Users by Clinic ({len(data['users_by_clinic'])} clinics)\n"
        output += f"  - Role Distribution ({len(data['role_distribution'])} roles)\n"
        output += f"  - Recent Activity ({len(data['recent_activity'])} records)\n"

        return output

    except Exception as e:
        return f"Error exporting dashboard: {str(e)}"
    finally:
        if manager:
            manager.close()


# ============================================================
# CONFIGURATION VIEWING TOOLS
# ============================================================

@mcp.tool()
def get_program_overview(program: str) -> str:
    """
    Get comprehensive overview of a program including all clinics, locations, and config summary.

    Args:
        program: Program prefix (e.g., "P4M") or name

    Returns:
        Program hierarchy with config counts and override summary
    """
    cm = None
    try:
        cm = get_config_manager()

        # Resolve program
        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        # Get program details
        cursor = cm.conn.cursor()
        cursor.execute("""
            SELECT p.*,
                   (SELECT COUNT(*) FROM clinics c WHERE c.program_id = p.program_id) as clinic_count,
                   (SELECT COUNT(*) FROM locations l
                    JOIN clinics c ON l.clinic_id = c.clinic_id
                    WHERE c.program_id = p.program_id) as location_count,
                   (SELECT COUNT(*) FROM config_values cv WHERE cv.program_id = p.program_id AND cv.clinic_id IS NULL) as program_config_count
            FROM programs p WHERE p.program_id = ?
        """, (program_id,))
        program_data = dict(cursor.fetchone())

        result = f"Program: {program_data['name']} [{program_data.get('prefix', '')}]\n"
        result += "=" * 50 + "\n\n"
        result += f"Type: {program_data.get('program_type', 'N/A')}\n"
        result += f"Status: {program_data.get('status', 'Active')}\n\n"

        result += f"Hierarchy:\n"
        result += f"  Clinics: {program_data['clinic_count']}\n"
        result += f"  Locations: {program_data['location_count']}\n"
        result += f"  Program-level configs: {program_data['program_config_count']}\n\n"

        # Get clinics with their locations
        cursor.execute("""
            SELECT c.clinic_id, c.name, c.code,
                   (SELECT COUNT(*) FROM locations l WHERE l.clinic_id = c.clinic_id) as location_count,
                   (SELECT COUNT(*) FROM config_values cv WHERE cv.clinic_id = c.clinic_id AND cv.location_id IS NULL) as clinic_config_count
            FROM clinics c
            WHERE c.program_id = ?
            ORDER BY c.name
        """, (program_id,))

        clinics = cursor.fetchall()

        if clinics:
            result += "Clinics:\n"
            for clinic in clinics:
                clinic_dict = dict(clinic)
                result += f"\n  [{clinic_dict.get('code', '')}] {clinic_dict['name']}\n"
                result += f"      Locations: {clinic_dict['location_count']} | Clinic-level configs: {clinic_dict['clinic_config_count']}\n"

                # Get locations for this clinic
                cursor.execute("""
                    SELECT l.name, l.code,
                           (SELECT COUNT(*) FROM config_values cv WHERE cv.location_id = l.location_id) as location_config_count
                    FROM locations l
                    WHERE l.clinic_id = ?
                    ORDER BY l.name
                """, (clinic_dict['clinic_id'],))

                locations = cursor.fetchall()
                for loc in locations:
                    loc_dict = dict(loc)
                    result += f"      +-- {loc_dict['name']}"
                    if loc_dict.get('code'):
                        result += f" [{loc_dict['code']}]"
                    result += f" ({loc_dict['location_config_count']} overrides)\n"

        return result

    except Exception as e:
        return f"Error getting program overview: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def get_clinic_config(
    program: str,
    clinic: str,
    category: Optional[str] = None
) -> str:
    """
    Get all configuration values for a specific clinic with inheritance info.

    Args:
        program: Program prefix (e.g., "P4M")
        clinic: Clinic name or code
        category: Optional category filter (e.g., "helpdesk", "operations", "lab_order")

    Returns:
        All configs for the clinic showing effective values and sources
    """
    cm = None
    try:
        cm = get_config_manager()
        im = InheritanceManager(cm)

        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        clinic_id = cm.get_clinic_id(program_id, clinic)
        if not clinic_id:
            return f"Clinic not found: {clinic}"

        # Get clinic name
        cursor = cm.conn.cursor()
        cursor.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic_id,))
        clinic_name = cursor.fetchone()['name']

        # Get all config definitions
        cursor.execute("""
            SELECT config_key, display_name, category, default_value
            FROM config_definitions
            ORDER BY category, display_name
        """)
        definitions = cursor.fetchall()

        result = f"Configuration: {clinic_name} ({program})\n"
        result += "=" * 50 + "\n\n"

        current_category = None
        for defn in definitions:
            defn_dict = dict(defn)

            # Filter by category if specified
            if category and category.lower() not in defn_dict['category'].lower():
                continue

            # Print category header
            if defn_dict['category'] != current_category:
                current_category = defn_dict['category']
                result += f"\n[{current_category.upper()}]\n"

            # Get effective value with inheritance
            config = im.resolve_with_inheritance(
                defn_dict['config_key'], program_id, clinic_id, None
            )

            effective_value = config.get('value') if config else defn_dict['default_value']
            effective_level = config.get('effective_level', 'default') if config else 'default'
            is_override = config.get('is_override', False) if config else False

            # Format display
            display = defn_dict.get('display_name') or defn_dict['config_key']
            override_marker = "*" if is_override else ""
            source_indicator = f" [{effective_level}]" if effective_level != 'clinic' else ""

            result += f"  {display}: {effective_value or '(not set)'}{override_marker}{source_indicator}\n"

        result += "\n* = overridden at this level\n"
        result += "[level] = inherited from that level\n"

        return result

    except Exception as e:
        return f"Error getting clinic config: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def compare_clinic_configs(
    program: str,
    clinic1: str,
    clinic2: str
) -> str:
    """
    Compare configuration values between two clinics.

    Args:
        program: Program prefix (e.g., "P4M")
        clinic1: First clinic name or code
        clinic2: Second clinic name or code

    Returns:
        Side-by-side comparison highlighting differences
    """
    cm = None
    try:
        cm = get_config_manager()
        im = InheritanceManager(cm)

        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        clinic1_id = cm.get_clinic_id(program_id, clinic1)
        if not clinic1_id:
            return f"Clinic not found: {clinic1}"

        clinic2_id = cm.get_clinic_id(program_id, clinic2)
        if not clinic2_id:
            return f"Clinic not found: {clinic2}"

        # Get clinic names
        cursor = cm.conn.cursor()
        cursor.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic1_id,))
        clinic1_name = cursor.fetchone()['name']
        cursor.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic2_id,))
        clinic2_name = cursor.fetchone()['name']

        # Get all config definitions
        cursor.execute("""
            SELECT config_key, display_name, category
            FROM config_definitions
            ORDER BY category, display_name
        """)
        definitions = cursor.fetchall()

        result = f"Config Comparison: {clinic1_name} vs {clinic2_name}\n"
        result += "=" * 60 + "\n\n"

        differences = []
        same_count = 0

        current_category = None
        for defn in definitions:
            defn_dict = dict(defn)

            # Get values for both clinics
            config1 = im.resolve_with_inheritance(defn_dict['config_key'], program_id, clinic1_id, None)
            config2 = im.resolve_with_inheritance(defn_dict['config_key'], program_id, clinic2_id, None)

            value1 = config1.get('value') if config1 else None
            value2 = config2.get('value') if config2 else None

            if value1 != value2:
                differences.append({
                    'category': defn_dict['category'],
                    'key': defn_dict.get('display_name') or defn_dict['config_key'],
                    'value1': value1 or '(not set)',
                    'value2': value2 or '(not set)'
                })
            else:
                same_count += 1

        # Display differences grouped by category
        if differences:
            result += f"DIFFERENCES ({len(differences)}):\n"
            current_category = None
            for diff in differences:
                if diff['category'] != current_category:
                    current_category = diff['category']
                    result += f"\n[{current_category.upper()}]\n"
                result += f"  {diff['key']}:\n"
                result += f"    {clinic1_name}: {diff['value1']}\n"
                result += f"    {clinic2_name}: {diff['value2']}\n"
        else:
            result += "No differences found.\n"

        result += f"\n\nSummary: {len(differences)} differences, {same_count} same\n"

        return result

    except Exception as e:
        return f"Error comparing clinics: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def get_config_overrides(
    program: str,
    clinic: Optional[str] = None
) -> str:
    """
    Get all configuration overrides (non-inherited values) for a program or clinic.

    Args:
        program: Program prefix (e.g., "P4M")
        clinic: Optional clinic name to filter to specific clinic

    Returns:
        List of all overridden configs with their values
    """
    cm = None
    try:
        cm = get_config_manager()

        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        cursor = cm.conn.cursor()

        # Get program name
        cursor.execute("SELECT name FROM programs WHERE program_id = ?", (program_id,))
        program_name = cursor.fetchone()['name']

        result = f"Configuration Overrides: {program_name}\n"
        result += "=" * 50 + "\n"

        # Build query based on scope
        if clinic:
            clinic_id = cm.get_clinic_id(program_id, clinic)
            if not clinic_id:
                return f"Clinic not found: {clinic}"

            cursor.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic_id,))
            clinic_name = cursor.fetchone()['name']
            result += f"Clinic: {clinic_name}\n\n"

            # Get clinic and location overrides
            cursor.execute("""
                SELECT cv.config_key, cv.value, cv.source, cv.updated_date,
                       cd.display_name, cd.category,
                       c.name as clinic_name, l.name as location_name
                FROM config_values cv
                JOIN config_definitions cd ON cv.config_key = cd.config_key
                LEFT JOIN clinics c ON cv.clinic_id = c.clinic_id
                LEFT JOIN locations l ON cv.location_id = l.location_id
                WHERE cv.program_id = ? AND cv.clinic_id = ?
                ORDER BY cv.location_id IS NULL DESC, cd.category, cd.display_name
            """, (program_id, clinic_id))
        else:
            result += "\n"
            # Get all overrides for the program
            cursor.execute("""
                SELECT cv.config_key, cv.value, cv.source, cv.updated_date,
                       cd.display_name, cd.category,
                       c.name as clinic_name, l.name as location_name
                FROM config_values cv
                JOIN config_definitions cd ON cv.config_key = cd.config_key
                LEFT JOIN clinics c ON cv.clinic_id = c.clinic_id
                LEFT JOIN locations l ON cv.location_id = l.location_id
                WHERE cv.program_id = ?
                ORDER BY cv.clinic_id IS NULL DESC, cv.location_id IS NULL DESC,
                         c.name, l.name, cd.category, cd.display_name
            """, (program_id,))

        overrides = cursor.fetchall()

        if not overrides:
            result += "No overrides found. All values are inherited from defaults.\n"
            return result

        # Group and display
        current_scope = None
        for override in overrides:
            ov = dict(override)

            # Determine scope
            if ov['location_name']:
                scope = f"Location: {ov['clinic_name']} > {ov['location_name']}"
            elif ov['clinic_name']:
                scope = f"Clinic: {ov['clinic_name']}"
            else:
                scope = "Program-level"

            if scope != current_scope:
                current_scope = scope
                result += f"\n[{scope}]\n"

            display = ov.get('display_name') or ov['config_key']
            result += f"  {display}: {ov['value']}\n"
            result += f"    Source: {ov['source']} | Updated: {ov['updated_date']}\n"

        result += f"\nTotal overrides: {len(overrides)}\n"

        return result

    except Exception as e:
        return f"Error getting overrides: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def get_clinic_providers(
    program: str,
    clinic: str
) -> str:
    """
    Get all providers (healthcare staff) for a clinic with their NPIs.

    Args:
        program: Program prefix (e.g., "P4M")
        clinic: Clinic name or code

    Returns:
        List of providers with NPI, location, and specialty info
    """
    cm = None
    try:
        cm = get_config_manager()

        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        clinic_id = cm.get_clinic_id(program_id, clinic)
        if not clinic_id:
            return f"Clinic not found: {clinic}"

        cursor = cm.conn.cursor()

        # Get clinic name
        cursor.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic_id,))
        clinic_name = cursor.fetchone()['name']

        # Get providers
        cursor.execute("""
            SELECT p.name, p.npi, p.specialty, p.status,
                   l.name as location_name
            FROM providers p
            LEFT JOIN locations l ON p.location_id = l.location_id
            JOIN clinics c ON (p.location_id IN (SELECT location_id FROM locations WHERE clinic_id = c.clinic_id)
                              OR p.clinic_id = c.clinic_id)
            WHERE c.clinic_id = ?
            ORDER BY l.name, p.name
        """, (clinic_id,))

        providers = cursor.fetchall()

        result = f"Providers: {clinic_name} ({program})\n"
        result += "=" * 50 + "\n\n"

        if not providers:
            result += "No providers found for this clinic.\n"
            return result

        current_location = None
        active_count = 0
        for prov in providers:
            prov_dict = dict(prov)

            location = prov_dict.get('location_name') or '(Clinic-wide)'
            if location != current_location:
                current_location = location
                result += f"\n[{location}]\n"

            status = prov_dict.get('status', 'Active')
            status_marker = "" if status == 'Active' else f" ({status})"
            if status == 'Active':
                active_count += 1

            result += f"  {prov_dict['name']}{status_marker}\n"
            result += f"    NPI: {prov_dict.get('npi', 'Not set')}"
            if prov_dict.get('specialty'):
                result += f" | Specialty: {prov_dict['specialty']}"
            result += "\n"

        result += f"\nTotal: {len(providers)} providers ({active_count} active)\n"

        return result

    except Exception as e:
        return f"Error getting providers: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def get_clinic_appointment_types(
    program: str,
    clinic: str
) -> str:
    """
    Get appointment type filters for a clinic and its locations.

    Args:
        program: Program prefix (e.g., "P4M")
        clinic: Clinic name or code

    Returns:
        List of appointment types included/excluded for extract filtering
    """
    cm = None
    try:
        cm = get_config_manager()

        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        clinic_id = cm.get_clinic_id(program_id, clinic)
        if not clinic_id:
            return f"Clinic not found: {clinic}"

        cursor = cm.conn.cursor()

        # Get clinic name
        cursor.execute("SELECT name FROM clinics WHERE clinic_id = ?", (clinic_id,))
        clinic_name = cursor.fetchone()['name']

        # Get appointment types
        cursor.execute("""
            SELECT at.name, at.code, at.include, at.exclude_reason,
                   l.name as location_name
            FROM appointment_types at
            LEFT JOIN locations l ON at.location_id = l.location_id
            WHERE at.clinic_id = ? OR at.location_id IN (
                SELECT location_id FROM locations WHERE clinic_id = ?
            )
            ORDER BY l.name NULLS FIRST, at.include DESC, at.name
        """, (clinic_id, clinic_id))

        appt_types = cursor.fetchall()

        result = f"Appointment Types: {clinic_name} ({program})\n"
        result += "=" * 50 + "\n\n"

        if not appt_types:
            result += "No appointment type filters configured.\n"
            result += "All appointment types will be included in extracts.\n"
            return result

        current_location = None
        included_count = 0
        excluded_count = 0

        for at in appt_types:
            at_dict = dict(at)

            location = at_dict.get('location_name') or '(Clinic-wide)'
            if location != current_location:
                current_location = location
                result += f"\n[{location}]\n"

            included = at_dict.get('include', True)
            if included:
                included_count += 1
                result += f"  [INCLUDE] {at_dict['name']}"
            else:
                excluded_count += 1
                result += f"  [EXCLUDE] {at_dict['name']}"

            if at_dict.get('code'):
                result += f" ({at_dict['code']})"
            result += "\n"

            if not included and at_dict.get('exclude_reason'):
                result += f"           Reason: {at_dict['exclude_reason']}\n"

        result += f"\nSummary: {included_count} included, {excluded_count} excluded\n"

        return result

    except Exception as e:
        return f"Error getting appointment types: {str(e)}"
    finally:
        if cm:
            cm.close()


@mcp.tool()
def export_program_configs(
    program: str,
    output_dir: Optional[str] = None,
    include_audit: bool = False
) -> str:
    """
    Export all configurations for a program to Excel (Configuration Matrix format).

    Args:
        program: Program prefix (e.g., "P4M")
        output_dir: Directory to save file (default: ~/Downloads)
        include_audit: If True, include audit history sheet

    Returns:
        Path to generated Excel file
    """
    cm = None
    try:
        cm = get_config_manager()

        program_id = cm.get_program_id(program)
        if not program_id:
            return f"Program not found: {program}"

        cursor = cm.conn.cursor()

        # Get program info
        cursor.execute("SELECT name, prefix FROM programs WHERE program_id = ?", (program_id,))
        program_data = dict(cursor.fetchone())
        program_name = program_data['name']
        program_prefix = program_data.get('prefix', '')

        if output_dir:
            out_path = os.path.expanduser(output_dir)
        else:
            out_path = os.path.expanduser("~/Downloads")
        os.makedirs(out_path, exist_ok=True)

        today_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"{program_prefix}_configurations_{today_str}.xlsx"
        filepath = os.path.join(out_path, filename)

        wb = Workbook()

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        category_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
        override_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # --- CONFIGURATION MATRIX SHEET ---
        ws = wb.active
        ws.title = "Configuration Matrix"

        # Get clinics and locations
        cursor.execute("""
            SELECT c.clinic_id, c.name as clinic_name
            FROM clinics c WHERE c.program_id = ?
            ORDER BY c.name
        """, (program_id,))
        clinics = [dict(row) for row in cursor.fetchall()]

        # Build location map
        location_columns = []
        for clinic in clinics:
            cursor.execute("""
                SELECT l.location_id, l.name
                FROM locations l WHERE l.clinic_id = ?
                ORDER BY l.name
            """, (clinic['clinic_id'],))
            locations = [dict(row) for row in cursor.fetchall()]

            if locations:
                for loc in locations:
                    location_columns.append({
                        'clinic_id': clinic['clinic_id'],
                        'clinic_name': clinic['clinic_name'],
                        'location_id': loc['location_id'],
                        'location_name': loc['name'],
                        'header': f"{clinic['clinic_name']}\n{loc['name']}"
                    })
            else:
                location_columns.append({
                    'clinic_id': clinic['clinic_id'],
                    'clinic_name': clinic['clinic_name'],
                    'location_id': None,
                    'location_name': None,
                    'header': clinic['clinic_name']
                })

        # Headers
        headers = ['Config Key', 'Display Name', 'Program Default'] + [lc['header'] for lc in location_columns]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Get config definitions
        cursor.execute("""
            SELECT config_key, display_name, category, default_value
            FROM config_definitions
            ORDER BY category, display_name
        """)
        definitions = [dict(row) for row in cursor.fetchall()]

        im = InheritanceManager(cm)

        row = 2
        current_category = None

        for defn in definitions:
            # Category header row
            if defn['category'] != current_category:
                current_category = defn['category']
                cell = ws.cell(row=row, column=1, value=current_category.upper())
                cell.fill = category_fill
                cell.font = Font(bold=True)
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = category_fill
                    ws.cell(row=row, column=col).border = thin_border
                row += 1

            # Config key
            ws.cell(row=row, column=1, value=defn['config_key']).border = thin_border
            ws.cell(row=row, column=2, value=defn.get('display_name', '')).border = thin_border

            # Program default
            prog_config = im.resolve_with_inheritance(defn['config_key'], program_id, None, None)
            prog_value = prog_config.get('value') if prog_config else defn.get('default_value')
            cell = ws.cell(row=row, column=3, value=prog_value or "—")
            cell.border = thin_border
            if prog_config and prog_config.get('is_override'):
                cell.fill = override_fill

            # Location columns
            for col_idx, loc_col in enumerate(location_columns, 4):
                config = im.resolve_with_inheritance(
                    defn['config_key'],
                    program_id,
                    loc_col['clinic_id'],
                    loc_col['location_id']
                )

                effective_value = config.get('value') if config else None
                effective_level = config.get('effective_level') if config else None

                # Determine if this is inherited or set at this level
                if loc_col['location_id']:
                    is_set_here = effective_level == 'location'
                else:
                    is_set_here = effective_level == 'clinic'

                if is_set_here:
                    cell = ws.cell(row=row, column=col_idx, value=effective_value or "—")
                    cell.fill = override_fill
                else:
                    cell = ws.cell(row=row, column=col_idx, value="—")

                cell.border = thin_border

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        for col in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.freeze_panes = 'D2'

        # --- AUDIT HISTORY SHEET (if requested) ---
        if include_audit:
            ws2 = wb.create_sheet("Audit History")
            cursor.execute("""
                SELECT ch.config_key, ch.old_value, ch.new_value,
                       ch.changed_by, ch.changed_date, ch.change_reason,
                       c.name as clinic_name, l.name as location_name
                FROM config_history ch
                LEFT JOIN clinics c ON ch.clinic_id = c.clinic_id
                LEFT JOIN locations l ON ch.location_id = l.location_id
                WHERE ch.program_id = ?
                ORDER BY ch.changed_date DESC
                LIMIT 500
            """, (program_id,))

            audit_rows = cursor.fetchall()

            audit_headers = ['Date', 'Config Key', 'Clinic', 'Location', 'Old Value', 'New Value', 'Changed By', 'Reason']
            for col, header in enumerate(audit_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

            for row_num, audit in enumerate(audit_rows, 2):
                audit_dict = dict(audit)
                ws2.cell(row=row_num, column=1, value=audit_dict['changed_date']).border = thin_border
                ws2.cell(row=row_num, column=2, value=audit_dict['config_key']).border = thin_border
                ws2.cell(row=row_num, column=3, value=audit_dict.get('clinic_name', '')).border = thin_border
                ws2.cell(row=row_num, column=4, value=audit_dict.get('location_name', '')).border = thin_border
                ws2.cell(row=row_num, column=5, value=audit_dict.get('old_value', '')).border = thin_border
                ws2.cell(row=row_num, column=6, value=audit_dict.get('new_value', '')).border = thin_border
                ws2.cell(row=row_num, column=7, value=audit_dict.get('changed_by', '')).border = thin_border
                ws2.cell(row=row_num, column=8, value=audit_dict.get('change_reason', '')).border = thin_border

            ws2.freeze_panes = 'A2'

        wb.save(filepath)

        result = f"Configuration Export: {program_name}\n"
        result += "=" * 50 + "\n\n"
        result += f"File: {filepath}\n"
        result += f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}\n\n"
        result += f"Contents:\n"
        result += f"  - Configuration Matrix ({len(definitions)} configs x {len(location_columns)} columns)\n"
        if include_audit:
            result += f"  - Audit History (up to 500 recent changes)\n"
        result += f"\nNote: Yellow cells indicate overrides (non-inherited values)\n"

        return result

    except Exception as e:
        return f"Error exporting configs: {str(e)}"
    finally:
        if cm:
            cm.close()


# ============================================================
# AUDIT COMPLETION TOOLS
# ============================================================

@mcp.tool()
def update_clinic_manager(
    clinic: str,
    manager_name: str,
    manager_email: str,
    program: Optional[str] = None
) -> str:
    """
    Update the clinic manager contact information.

    Args:
        clinic: Clinic name or code
        manager_name: Manager's full name
        manager_email: Manager's email address
        program: Program name/prefix if clinic name is ambiguous

    Returns:
        Confirmation of update
    """
    manager = None
    try:
        import sqlite3
        db_path = os.path.expanduser("~/projects/data/client_product_database.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Find clinic
        query = """
            SELECT c.clinic_id, c.name, p.name as program_name
            FROM clinics c
            JOIN programs p ON c.program_id = p.program_id
            WHERE LOWER(c.name) LIKE LOWER(?) OR LOWER(c.code) = LOWER(?)
        """
        params = [f"%{clinic}%", clinic]

        if program:
            query += " AND (LOWER(p.name) LIKE LOWER(?) OR LOWER(p.prefix) = LOWER(?))"
            params.extend([f"%{program}%", program])

        cursor.execute(query, params)
        clinic_row = cursor.fetchone()

        if not clinic_row:
            conn.close()
            return f"Clinic not found: {clinic}"

        # Update manager info
        cursor.execute("""
            UPDATE clinics
            SET manager_name = ?, manager_email = ?, updated_date = CURRENT_TIMESTAMP
            WHERE clinic_id = ?
        """, (manager_name, manager_email, clinic_row['clinic_id']))

        conn.commit()
        conn.close()

        return f"""Clinic manager updated

Clinic: {clinic_row['name']}
Program: {clinic_row['program_name']}
Manager: {manager_name}
Email: {manager_email}
"""

    except Exception as e:
        return f"Error updating clinic manager: {str(e)}"


@mcp.tool()
def record_audit_completion(
    clinic: str,
    audit_year: int,
    date_initiated: Optional[str] = None,
    date_reviewed: Optional[str] = None,
    date_finalized: Optional[str] = None,
    date_tickets_submitted: Optional[str] = None,
    date_confirmed: Optional[str] = None,
    ticket_number: Optional[str] = None,
    audit_type: str = "Annual",
    document_version: str = "1.0",
    notes: Optional[str] = None,
    program: Optional[str] = None
) -> str:
    """
    Record or update audit completion milestone dates.

    All dates should be in YYYY-MM-DD format.
    You can update individual dates without affecting others.

    Args:
        clinic: Clinic name or code
        audit_year: Year of the audit (e.g., 2025)
        date_initiated: Date roster sent to clinic manager
        date_reviewed: Date client returned validation
        date_finalized: Date internal review finalized
        date_tickets_submitted: Date change tickets sent to dev
        date_confirmed: Date dev team confirmed complete
        ticket_number: Zendesk ticket number (without #)
        audit_type: 'Annual' or 'Quarterly' (default: Annual)
        document_version: Document version (default: 1.0)
        notes: Optional notes
        program: Program name/prefix if clinic name is ambiguous

    Returns:
        Confirmation of recorded data
    """
    try:
        import sqlite3
        db_path = os.path.expanduser("~/projects/data/client_product_database.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Find clinic
        query = """
            SELECT c.clinic_id, c.name, c.program_id, p.name as program_name
            FROM clinics c
            JOIN programs p ON c.program_id = p.program_id
            WHERE LOWER(c.name) LIKE LOWER(?) OR LOWER(c.code) = LOWER(?)
        """
        params = [f"%{clinic}%", clinic]

        if program:
            query += " AND (LOWER(p.name) LIKE LOWER(?) OR LOWER(p.prefix) = LOWER(?))"
            params.extend([f"%{program}%", program])

        cursor.execute(query, params)
        clinic_row = cursor.fetchone()

        if not clinic_row:
            conn.close()
            return f"Clinic not found: {clinic}"

        clinic_id = clinic_row['clinic_id']
        program_id = clinic_row['program_id']

        # Check if record exists
        cursor.execute("""
            SELECT completion_id FROM audit_completions
            WHERE clinic_id = ? AND audit_year = ? AND audit_type = ?
        """, (clinic_id, audit_year, audit_type))

        existing = cursor.fetchone()

        if existing:
            # Update existing
            cursor.execute("""
                UPDATE audit_completions SET
                    date_initiated = COALESCE(?, date_initiated),
                    date_reviewed = COALESCE(?, date_reviewed),
                    date_finalized = COALESCE(?, date_finalized),
                    date_tickets_submitted = COALESCE(?, date_tickets_submitted),
                    date_confirmed = COALESCE(?, date_confirmed),
                    ticket_number = COALESCE(?, ticket_number),
                    document_version = COALESCE(?, document_version),
                    notes = COALESCE(?, notes),
                    updated_date = CURRENT_TIMESTAMP
                WHERE completion_id = ?
            """, (
                date_initiated, date_reviewed, date_finalized,
                date_tickets_submitted, date_confirmed, ticket_number,
                document_version, notes, existing['completion_id']
            ))
            action = "Updated"
        else:
            # Insert new
            cursor.execute("""
                INSERT INTO audit_completions (
                    clinic_id, program_id, audit_year, audit_type,
                    date_initiated, date_reviewed, date_finalized,
                    date_tickets_submitted, date_confirmed,
                    ticket_number, document_version, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                clinic_id, program_id, audit_year, audit_type,
                date_initiated, date_reviewed, date_finalized,
                date_tickets_submitted, date_confirmed,
                ticket_number, document_version, notes
            ))
            action = "Created"

        conn.commit()

        # Fetch the current state
        cursor.execute("""
            SELECT * FROM audit_completions
            WHERE clinic_id = ? AND audit_year = ? AND audit_type = ?
        """, (clinic_id, audit_year, audit_type))
        record = cursor.fetchone()

        conn.close()

        output = f"""{action} audit completion record

Clinic: {clinic_row['name']}
Program: {clinic_row['program_name']}
Audit: {audit_year} {audit_type}

TIMELINE
---------
Initiated:          {record['date_initiated'] or '(not set)'}
Reviewed by Client: {record['date_reviewed'] or '(not set)'}
Finalized:          {record['date_finalized'] or '(not set)'}
Tickets Submitted:  {record['date_tickets_submitted'] or '(not set)'}
Confirmed:          {record['date_confirmed'] or '(not set)'}

DETAILS
-------
Ticket #: {record['ticket_number'] or '(not set)'}
Version:  {record['document_version']}
Notes:    {record['notes'] or '(none)'}
"""
        return output

    except Exception as e:
        return f"Error recording audit completion: {str(e)}"


@mcp.tool()
def generate_audit_memo(
    clinic: str,
    audit_year: int,
    audit_type: str = "Annual",
    program: Optional[str] = None,
    output_dir: Optional[str] = None
) -> str:
    """
    Generate a filled audit completion memo Word document.

    Pulls data from the audit_completions table and calculates
    metrics from access_reviews. Outputs a ready-to-sign document.

    Args:
        clinic: Clinic name or code
        audit_year: Year of the audit (e.g., 2025)
        audit_type: 'Annual' or 'Quarterly' (default: Annual)
        program: Program name/prefix if clinic name is ambiguous
        output_dir: Directory to save file (default: ~/Downloads)

    Returns:
        Path to generated document and summary
    """
    try:
        from docxtpl import DocxTemplate
        from datetime import datetime
        import sqlite3

        db_path = os.path.expanduser("~/projects/data/client_product_database.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Find clinic
        query = """
            SELECT c.clinic_id, c.name, c.manager_name, c.manager_email,
                   c.program_id, p.name as program_name
            FROM clinics c
            JOIN programs p ON c.program_id = p.program_id
            WHERE LOWER(c.name) LIKE LOWER(?) OR LOWER(c.code) = LOWER(?)
        """
        params = [f"%{clinic}%", clinic]

        if program:
            query += " AND (LOWER(p.name) LIKE LOWER(?) OR LOWER(p.prefix) = LOWER(?))"
            params.extend([f"%{program}%", program])

        cursor.execute(query, params)
        clinic_row = cursor.fetchone()

        if not clinic_row:
            conn.close()
            return f"Clinic not found: {clinic}"

        clinic_id = clinic_row['clinic_id']

        # Get audit completion record
        cursor.execute("""
            SELECT * FROM audit_completions
            WHERE clinic_id = ? AND audit_year = ? AND audit_type = ?
        """, (clinic_id, audit_year, audit_type))

        record = cursor.fetchone()
        if not record:
            conn.close()
            return f"No audit completion record found for {clinic_row['name']} {audit_year} {audit_type}. Use record_audit_completion first."

        # Calculate metrics
        cursor.execute("""
            SELECT
                COUNT(*) as total_reviewed,
                SUM(CASE WHEN status = 'Revoked' THEN 1 ELSE 0 END) as revocations,
                SUM(CASE WHEN status = 'Modified' THEN 1 ELSE 0 END) as modifications
            FROM access_reviews ar
            JOIN user_access ua ON ar.access_id = ua.access_id
            WHERE ua.clinic_id = ?
            AND strftime('%Y', ar.review_date) = ?
        """, (clinic_id, str(audit_year)))
        metrics = cursor.fetchone()

        cursor.execute("""
            SELECT COUNT(*) as new_requests
            FROM user_access
            WHERE clinic_id = ? AND strftime('%Y', granted_date) = ?
        """, (clinic_id, str(audit_year)))
        new_grants = cursor.fetchone()

        conn.close()

        # Format dates
        def format_date(date_str):
            if not date_str:
                return '(Not Set)'
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                return dt.strftime('%B %d, %Y')
            except (ValueError, TypeError):
                return date_str

        # Prepare context
        context = {
            'clinic_name': clinic_row['name'],
            'audit_year': audit_year,
            'audit_type': audit_type,
            'document_version': record['document_version'] or '1.0',
            'date_initiated': format_date(record['date_initiated']),
            'date_reviewed': format_date(record['date_reviewed']),
            'date_finalized': format_date(record['date_finalized']),
            'date_tickets_submitted': format_date(record['date_tickets_submitted']),
            'date_confirmed': format_date(record['date_confirmed']),
            'ticket_number': record['ticket_number'] or '(Not Set)',
            'manager_name': clinic_row['manager_name'] or '(Not Set)',
            'manager_email': clinic_row['manager_email'] or '(Not Set)',
            'total_reviewed': metrics['total_reviewed'] if metrics else 0,
            'revocations': metrics['revocations'] if metrics else 0,
            'modifications': metrics['modifications'] if metrics else 0,
            'new_requests': new_grants['new_requests'] if new_grants else 0
        }

        # Load and render template
        template_path = os.path.expanduser(
            '~/projects/configurations_toolkit/templates/audit_completion_memo.docx'
        )

        if not os.path.exists(template_path):
            return f"Template not found: {template_path}\nRun the template creation script first."

        doc = DocxTemplate(template_path)
        doc.render(context)

        # Save output
        if output_dir:
            out_path = os.path.expanduser(output_dir)
        else:
            out_path = os.path.expanduser('~/Downloads')
        os.makedirs(out_path, exist_ok=True)

        safe_clinic = clinic_row['name'].replace(' ', '_').replace('/', '-')
        filename = f"{safe_clinic}_Audit_Memo_{audit_year}_{audit_type}.docx"
        filepath = os.path.join(out_path, filename)

        doc.save(filepath)

        return f"""Audit memo generated

File: {filepath}

DOCUMENT CONTENTS
-----------------
Clinic: {clinic_row['name']}
Audit: {audit_year} {audit_type}
Manager: {context['manager_name']} ({context['manager_email']})

Metrics:
- Total Reviewed: {context['total_reviewed']}
- Revocations: {context['revocations']}
- Modifications: {context['modifications']}
- New Requests: {context['new_requests']}

Timeline:
- Initiated: {context['date_initiated']}
- Reviewed: {context['date_reviewed']}
- Finalized: {context['date_finalized']}
- Tickets: {context['date_tickets_submitted']}
- Confirmed: {context['date_confirmed']}

Ticket #: {context['ticket_number']}
"""

    except ImportError:
        return "Error: docxtpl not installed. Run: pip install docxtpl --break-system-packages"
    except Exception as e:
        return f"Error generating audit memo: {str(e)}"


# ============================================================
# FORM DEFINITION MANAGEMENT TOOLS
# ============================================================
# Tools for managing the onboarding form questions in form-definition.json

# Path to form definition file
FORM_DEF_PATH = os.path.expanduser("~/projects/propel-onboarding-form/src/data/form-definition.json")


def _log_form_audit(
    action: str,
    entity_type: str,
    entity_id: str,
    details: str,
    old_value: str = None,
    new_value: str = None
):
    """Log form definition changes to audit_history table."""
    import sqlite3

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO audit_history (
                timestamp, action, entity_type, entity_id,
                field_name, old_value, new_value, changed_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datetime.now().isoformat(),
            action,
            entity_type,
            entity_id,
            details,
            old_value,
            new_value,
            "MCP:form_admin"
        ))

        conn.commit()
    except Exception as e:
        print(f"Audit log warning: {e}")
    finally:
        conn.close()


@mcp.tool()
def list_form_questions(
    step_id: str = None,
    show_details: bool = False
) -> str:
    """
    List questions in the onboarding form.

    Args:
        step_id: Optional step ID to filter (e.g., "clinic_info"). If None, shows all steps.
        show_details: If True, show full question properties. If False, show summary only.

    Returns:
        Formatted list of questions organized by step

    Example:
        list_form_questions()  # All steps, summary view
        list_form_questions(step_id="lab_config", show_details=True)  # One step, full details
    """
    import json

    try:
        with open(FORM_DEF_PATH, 'r') as f:
            form_def = json.load(f)
    except FileNotFoundError:
        return "Error: form-definition.json not found. Is the repo cloned?"
    except json.JSONDecodeError as e:
        return f"Error: Invalid JSON in form-definition.json: {e}"

    output = []
    output.append(f"Form: {form_def.get('title', 'Unknown')}")
    output.append(f"Version: {form_def.get('version', 'Unknown')}")
    output.append("=" * 60)

    for step in form_def.get('steps', []):
        # Filter by step_id if provided
        if step_id and step.get('step_id') != step_id:
            continue

        output.append(f"\nStep {step.get('order', '?')}: {step.get('title', 'Unknown')}")
        output.append(f"   ID: {step.get('step_id', 'Unknown')}")

        if step.get('repeatable'):
            output.append(f"   Repeatable (min: {step.get('repeatable_config', {}).get('min_items', 0)}, max: {step.get('repeatable_config', {}).get('max_items', 'unlimited')})")

        if step.get('is_review_step'):
            output.append("   Review/Download step (no questions)")
            continue

        questions = step.get('questions', [])
        output.append(f"   Questions: {len(questions)}")

        for q in questions:
            req = "*" if q.get('required') else ""
            q_type = q.get('type', 'unknown')

            if show_details:
                output.append(f"\n   [{q.get('question_id')}] {q.get('label', 'No label')}{req}")
                output.append(f"      Type: {q_type}")
                if q.get('options_ref'):
                    output.append(f"      Options: ref:{q.get('options_ref')}")
                if q.get('show_when'):
                    sw = q.get('show_when')
                    output.append(f"      Conditional: {sw.get('question_id')} {sw.get('operator')} {sw.get('value')}")
                if q.get('help_text'):
                    help_preview = q.get('help_text')[:50] + "..." if len(q.get('help_text', '')) > 50 else q.get('help_text')
                    output.append(f"      Help: {help_preview}")
                if q.get('placeholder'):
                    output.append(f"      Placeholder: {q.get('placeholder')}")
                if q.get('pattern'):
                    output.append(f"      Pattern: {q.get('pattern')}")
            else:
                output.append(f"   - {q.get('question_id')}: {q.get('label', 'No label')}{req} ({q_type})")

    if step_id and not any(s.get('step_id') == step_id for s in form_def.get('steps', [])):
        return f"Error: Step '{step_id}' not found. Use list_form_questions() to see available steps."

    return "\n".join(output)


@mcp.tool()
def add_form_question(
    step_id: str,
    question_id: str,
    question_type: str,
    label: str,
    required: bool = False,
    options_ref: str = None,
    help_text: str = None,
    placeholder: str = None,
    pattern: str = None,
    max_length: int = None,
    show_when_question: str = None,
    show_when_operator: str = None,
    show_when_value: str = None,
    insert_after: str = None
) -> str:
    """
    Add a new question to a step in the onboarding form.

    Args:
        step_id: Step to add question to (e.g., "clinic_info", "lab_config")
        question_id: Unique ID for the question (e.g., "clinic_fax")
        question_type: Type of question: text, textarea, select, radio, checkbox,
                       address, contact_group, stakeholder_group, select_with_alternates
        label: Display label for the question
        required: Whether the field is required (default: False)
        options_ref: Reference to options in reference-data.json (for select/radio types)
        help_text: Help text displayed below the field
        placeholder: Placeholder text for text inputs
        pattern: Regex pattern for validation (e.g., "^[0-9]{10}$" for NPI)
        max_length: Maximum character length for text fields
        show_when_question: Question ID that controls visibility (conditional field)
        show_when_operator: Operator for condition: "equals", "in", "not_equals"
        show_when_value: Value(s) that trigger showing this field
        insert_after: Question ID to insert after. If None, adds to end of step.

    Returns:
        Confirmation message with the added question details

    Example:
        add_form_question(
            step_id="clinic_info",
            question_id="clinic_fax",
            question_type="text",
            label="Clinic Fax Number",
            required=False,
            placeholder="406-555-1234",
            insert_after="clinic_phone"
        )
    """
    import json

    # Load current form definition
    try:
        with open(FORM_DEF_PATH, 'r') as f:
            form_def = json.load(f)
    except FileNotFoundError:
        return "Error: form-definition.json not found."
    except json.JSONDecodeError as e:
        return f"Error: Invalid JSON: {e}"

    # Find the step
    step = None
    for s in form_def.get('steps', []):
        if s.get('step_id') == step_id:
            step = s
            break

    if not step:
        available = [s.get('step_id') for s in form_def.get('steps', [])]
        return f"Error: Step '{step_id}' not found. Available steps: {', '.join(available)}"

    # Check for duplicate question_id across ALL steps
    for s in form_def.get('steps', []):
        for q in s.get('questions', []):
            if q.get('question_id') == question_id:
                return f"Error: question_id '{question_id}' already exists in step '{s.get('step_id')}'"

    # Build the question object
    new_question = {
        "question_id": question_id,
        "type": question_type,
        "label": label,
        "required": required
    }

    # Add optional fields
    if options_ref:
        new_question["options_ref"] = options_ref
    if help_text:
        new_question["help_text"] = help_text
    if placeholder:
        new_question["placeholder"] = placeholder
    if pattern:
        new_question["pattern"] = pattern
    if max_length:
        new_question["max_length"] = max_length

    # Add conditional visibility
    if show_when_question and show_when_operator and show_when_value:
        # Parse value - could be a list for "in" operator
        value = show_when_value
        if show_when_operator == "in" and "," in show_when_value:
            value = [v.strip() for v in show_when_value.split(",")]

        new_question["show_when"] = {
            "question_id": show_when_question,
            "operator": show_when_operator,
            "value": value
        }

    # Ensure questions list exists
    if 'questions' not in step:
        step['questions'] = []

    # Insert at position
    if insert_after:
        # Find the index of insert_after question
        insert_idx = None
        for idx, q in enumerate(step['questions']):
            if q.get('question_id') == insert_after:
                insert_idx = idx + 1
                break

        if insert_idx is None:
            return f"Error: insert_after question '{insert_after}' not found in step '{step_id}'"

        step['questions'].insert(insert_idx, new_question)
        position_msg = f"after '{insert_after}'"
    else:
        step['questions'].append(new_question)
        position_msg = "at end of step"

    # Save the updated form definition
    with open(FORM_DEF_PATH, 'w') as f:
        json.dump(form_def, f, indent=2)

    # Audit log
    _log_form_audit(
        action="ADD_QUESTION",
        entity_type="form_question",
        entity_id=question_id,
        details=f"Added to step '{step_id}' {position_msg}",
        new_value=json.dumps(new_question)
    )

    return f"""Question added successfully!

Step: {step_id}
Question ID: {question_id}
Label: {label}
Type: {question_type}
Required: {required}
Position: {position_msg}

Next steps:
  - Run list_form_questions(step_id="{step_id}") to verify
  - Push changes: cd ~/projects/propel-onboarding-form && git add . && git commit -m "Add {question_id} question" && git push
"""


@mcp.tool()
def update_form_question(
    question_id: str,
    label: str = None,
    required: bool = None,
    help_text: str = None,
    placeholder: str = None,
    pattern: str = None,
    max_length: int = None,
    options_ref: str = None,
    show_when_question: str = None,
    show_when_operator: str = None,
    show_when_value: str = None,
    clear_show_when: bool = False
) -> str:
    """
    Update an existing question's properties. Only provided fields are updated.

    Args:
        question_id: ID of question to update
        label: New display label
        required: Change required status
        help_text: New help text (use empty string "" to clear)
        placeholder: New placeholder text
        pattern: New regex validation pattern
        max_length: New max length
        options_ref: New options reference
        show_when_question: Question ID for conditional visibility
        show_when_operator: Operator for condition
        show_when_value: Value(s) for condition
        clear_show_when: If True, removes conditional visibility

    Returns:
        Confirmation with old and new values

    Example:
        update_form_question(
            question_id="clinic_name",
            label="Primary Clinic Name",
            help_text="Enter the official clinic name as it appears on your license"
        )
    """
    import json

    with open(FORM_DEF_PATH, 'r') as f:
        form_def = json.load(f)

    # Find the question
    found_step = None
    found_question = None
    for step in form_def.get('steps', []):
        for q in step.get('questions', []):
            if q.get('question_id') == question_id:
                found_step = step
                found_question = q
                break
        if found_question:
            break

    if not found_question:
        return f"Error: Question '{question_id}' not found in any step."

    # Track changes
    changes = []
    old_values = {}

    # Update provided fields
    if label is not None:
        old_values['label'] = found_question.get('label')
        found_question['label'] = label
        changes.append(f"label: '{old_values['label']}' -> '{label}'")

    if required is not None:
        old_values['required'] = found_question.get('required')
        found_question['required'] = required
        changes.append(f"required: {old_values['required']} -> {required}")

    if help_text is not None:
        old_values['help_text'] = found_question.get('help_text')
        if help_text == "":
            found_question.pop('help_text', None)
            changes.append("help_text: removed")
        else:
            found_question['help_text'] = help_text
            changes.append("help_text: updated")

    if placeholder is not None:
        old_values['placeholder'] = found_question.get('placeholder')
        found_question['placeholder'] = placeholder
        changes.append(f"placeholder: '{old_values['placeholder']}' -> '{placeholder}'")

    if pattern is not None:
        old_values['pattern'] = found_question.get('pattern')
        found_question['pattern'] = pattern
        changes.append(f"pattern: '{old_values['pattern']}' -> '{pattern}'")

    if max_length is not None:
        old_values['max_length'] = found_question.get('max_length')
        found_question['max_length'] = max_length
        changes.append(f"max_length: {old_values['max_length']} -> {max_length}")

    if options_ref is not None:
        old_values['options_ref'] = found_question.get('options_ref')
        found_question['options_ref'] = options_ref
        changes.append(f"options_ref: '{old_values['options_ref']}' -> '{options_ref}'")

    if clear_show_when:
        if 'show_when' in found_question:
            old_values['show_when'] = found_question.pop('show_when')
            changes.append("show_when: removed (question now always visible)")
    elif show_when_question and show_when_operator and show_when_value:
        old_values['show_when'] = found_question.get('show_when')
        value = show_when_value
        if show_when_operator == "in" and "," in show_when_value:
            value = [v.strip() for v in show_when_value.split(",")]
        found_question['show_when'] = {
            "question_id": show_when_question,
            "operator": show_when_operator,
            "value": value
        }
        changes.append(f"show_when: updated to depend on '{show_when_question}'")

    if not changes:
        return "No changes provided. Specify at least one field to update."

    # Save
    with open(FORM_DEF_PATH, 'w') as f:
        json.dump(form_def, f, indent=2)

    # Audit log
    _log_form_audit(
        action="UPDATE_QUESTION",
        entity_type="form_question",
        entity_id=question_id,
        details=f"Updated in step '{found_step.get('step_id')}'",
        old_value=json.dumps(old_values),
        new_value=json.dumps({k: found_question.get(k) for k in old_values.keys()})
    )

    changes_formatted = "\n  - ".join(changes)
    return f"""Question updated successfully!

Question ID: {question_id}
Step: {found_step.get('step_id')}

Changes:
  - {changes_formatted}

Next steps:
  - Run list_form_questions(step_id="{found_step.get('step_id')}", show_details=True) to verify
  - Push changes to GitHub
"""


@mcp.tool()
def remove_form_question(
    question_id: str,
    reason: str
) -> str:
    """
    Remove a question from the onboarding form.

    Args:
        question_id: ID of question to remove
        reason: Reason for removal (required for audit trail)

    Returns:
        Confirmation with removed question details

    Example:
        remove_form_question(
            question_id="clinic_fax",
            reason="Fax numbers no longer collected - moved to optional notes"
        )
    """
    import json

    with open(FORM_DEF_PATH, 'r') as f:
        form_def = json.load(f)

    # Find and remove the question
    found_step = None
    removed_question = None
    for step in form_def.get('steps', []):
        for idx, q in enumerate(step.get('questions', [])):
            if q.get('question_id') == question_id:
                found_step = step
                removed_question = step['questions'].pop(idx)
                break
        if removed_question:
            break

    if not removed_question:
        return f"Error: Question '{question_id}' not found in any step."

    # Check if any other questions depend on this one (show_when)
    dependencies = []
    for step in form_def.get('steps', []):
        for q in step.get('questions', []):
            show_when = q.get('show_when', {})
            if show_when.get('question_id') == question_id:
                dependencies.append(f"{q.get('question_id')} (in {step.get('step_id')})")

    if dependencies:
        # Re-add the question since we can't remove it
        found_step['questions'].append(removed_question)
        deps_formatted = "\n  - ".join(dependencies)
        return f"""Error: Cannot remove '{question_id}' - other questions depend on it:

Dependent questions:
  - {deps_formatted}

Remove or update the dependent questions first, then try again."""

    # Save
    with open(FORM_DEF_PATH, 'w') as f:
        json.dump(form_def, f, indent=2)

    # Audit log
    _log_form_audit(
        action="REMOVE_QUESTION",
        entity_type="form_question",
        entity_id=question_id,
        details=f"Removed from step '{found_step.get('step_id')}'. Reason: {reason}",
        old_value=json.dumps(removed_question)
    )

    return f"""Question removed successfully!

Question ID: {question_id}
Label: {removed_question.get('label')}
Step: {found_step.get('step_id')}
Reason: {reason}

The full question definition has been saved to the audit log.

Next steps:
  - Push changes to GitHub
"""


@mcp.tool()
def reorder_form_questions(
    step_id: str,
    question_id: str,
    move_to: str,
    target_question_id: str = None
) -> str:
    """
    Reorder questions within a step.

    Args:
        step_id: Step containing the question
        question_id: Question to move
        move_to: Where to move: "first", "last", "before", "after"
        target_question_id: Required if move_to is "before" or "after"

    Returns:
        Confirmation with new question order

    Example:
        reorder_form_questions(
            step_id="clinic_info",
            question_id="clinic_phone",
            move_to="after",
            target_question_id="clinic_name"
        )
    """
    import json

    with open(FORM_DEF_PATH, 'r') as f:
        form_def = json.load(f)

    # Find the step
    step = None
    for s in form_def.get('steps', []):
        if s.get('step_id') == step_id:
            step = s
            break

    if not step:
        return f"Error: Step '{step_id}' not found."

    questions = step.get('questions', [])

    # Find the question to move
    move_idx = None
    for idx, q in enumerate(questions):
        if q.get('question_id') == question_id:
            move_idx = idx
            break

    if move_idx is None:
        return f"Error: Question '{question_id}' not found in step '{step_id}'."

    # Remove the question from current position
    question_to_move = questions.pop(move_idx)

    # Determine new position
    if move_to == "first":
        new_idx = 0
    elif move_to == "last":
        new_idx = len(questions)
    elif move_to in ("before", "after"):
        if not target_question_id:
            questions.insert(move_idx, question_to_move)  # restore
            return f"Error: target_question_id required when move_to is '{move_to}'"

        target_idx = None
        for idx, q in enumerate(questions):
            if q.get('question_id') == target_question_id:
                target_idx = idx
                break

        if target_idx is None:
            questions.insert(move_idx, question_to_move)  # restore
            return f"Error: Target question '{target_question_id}' not found in step '{step_id}'."

        new_idx = target_idx if move_to == "before" else target_idx + 1
    else:
        questions.insert(move_idx, question_to_move)  # restore
        return f"Error: Invalid move_to value '{move_to}'. Use: first, last, before, after"

    # Insert at new position
    questions.insert(new_idx, question_to_move)

    # Save
    with open(FORM_DEF_PATH, 'w') as f:
        json.dump(form_def, f, indent=2)

    # Build new order display
    new_order = [q.get('question_id') for q in questions]
    order_formatted = "\n  ".join([f"{i+1}. {qid}" for i, qid in enumerate(new_order)])

    # Audit log
    _log_form_audit(
        action="REORDER_QUESTION",
        entity_type="form_question",
        entity_id=question_id,
        details=f"Moved to position {new_idx + 1} in step '{step_id}'"
    )

    target_msg = f" {target_question_id}" if target_question_id else ""
    return f"""Question reordered successfully!

Step: {step_id}
Moved: {question_id}
Position: {move_to}{target_msg}

New order:
  {order_formatted}

Next steps:
  - Push changes to GitHub
"""


# ============================================================
# DASHBOARD DATA GENERATION TOOLS
# ============================================================
# Tools for generating static JSON data files for the clinic dashboard
# Dashboard location: ~/projects/propel-clinic-dashboard/

# Default output path for dashboard data
DASHBOARD_DATA_PATH = os.path.expanduser(
    "~/projects/propel-clinic-dashboard/src/data/dashboard-data.json"
)

# Required configuration fields for a clinic to be considered "configured"
# A clinic is fully configured only when ALL of these fields have values
REQUIRED_CONFIG_FIELDS = [
    "clinic_phone",        # Clinic's helpdesk phone number
    "default_test",        # Default genetic test panel
    "default_specimen",    # Default specimen type (Blood, Saliva, etc.)
    "ordering_provider_npi"  # At least one ordering provider NPI
]


@mcp.tool()
def generate_dashboard_data(
    show_audit_trail: bool = False,
    output_path: str = None,
    days_of_audit: int = 30
) -> str:
    """
    Generate dashboard JSON data for the clinic configuration dashboard.

    Args:
        show_audit_trail: If True, includes audit_trail in output (default: False)
        output_path: Where to write JSON file (default: ~/projects/propel-clinic-dashboard/src/data/dashboard-data.json)
        days_of_audit: How many days of audit history to include (default: 30)

    Returns:
        Confirmation message with file path and summary stats
    """
    logger.info(f"generate_dashboard_data() called - audit={show_audit_trail}, days={days_of_audit}")

    import json
    import sqlite3

    # =========================================================================
    # SETUP
    # =========================================================================
    if output_path is None:
        output_path = DASHBOARD_DATA_PATH
    else:
        output_path = os.path.expanduser(output_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    try:
        cursor = conn.cursor()

        # =====================================================================
        # QUERY 1: PROGRAMS LIST
        # =====================================================================
        cursor.execute("""
            SELECT DISTINCT
                prefix as id,
                name as name
            FROM programs
            WHERE status = 'Active'
              AND prefix IS NOT NULL
            ORDER BY name
        """)
        programs = [dict(row) for row in cursor.fetchall()]

        # =====================================================================
        # QUERY 2: USERS WITH ACCESS
        # =====================================================================
        cursor.execute("""
            SELECT DISTINCT
                u.name,
                u.email,
                p.prefix as program,
                c.name as clinic,
                ua.role,
                u.status,
                ua.granted_date as last_access
            FROM users u
            JOIN user_access ua ON u.user_id = ua.user_id
            JOIN clinics c ON ua.clinic_id = c.clinic_id
            JOIN programs p ON ua.program_id = p.program_id
            WHERE ua.status = 'Active'
            ORDER BY u.name
        """)
        users = [dict(row) for row in cursor.fetchall()]

        # =====================================================================
        # QUERY 3: CLINICS WITH CONFIGURATION STATUS
        # =====================================================================
        # Get all clinics and their configuration values to determine
        # which are fully configured vs missing required fields.
        # Required fields: clinic_phone, default_test
        # Optional fields: default_specimen, optional_tests
        cursor.execute("""
            SELECT
                p.prefix as program,
                c.name as clinic,
                c.clinic_id,
                c.phone as clinic_phone
            FROM clinics c
            JOIN programs p ON c.program_id = p.program_id
            WHERE p.status = 'Active'
            ORDER BY p.prefix, c.name
        """)
        clinic_rows = cursor.fetchall()

        # Build configurations list with config values for each clinic
        configurations = []
        clinics_configured = []
        config_requests_pending = []
        missing_configurations = []

        for clinic_row in clinic_rows:
            clinic_data = dict(clinic_row)
            clinic_id = clinic_data['clinic_id']
            program = clinic_data['program']
            clinic_name = clinic_data['clinic']

            # Get config values for this clinic
            cursor.execute("""
                SELECT config_key, config_value as value
                FROM config_values
                WHERE clinic_id = ?
            """, (clinic_id,))
            config_rows = cursor.fetchall()
            config_dict = {row['config_key']: row['value'] for row in config_rows}

            # Extract specific config values
            clinic_phone = clinic_data.get('clinic_phone')
            default_test = config_dict.get('default_test')
            default_specimen = config_dict.get('default_specimen')

            # Parse optional_tests as JSON array if it exists
            optional_tests_raw = config_dict.get('optional_tests')
            optional_tests = []
            if optional_tests_raw:
                try:
                    optional_tests = json.loads(optional_tests_raw)
                except (json.JSONDecodeError, TypeError):
                    # If not JSON, treat as single value
                    optional_tests = [optional_tests_raw] if optional_tests_raw else []

            # Check if clinic has at least one ordering provider
            # Query providers linked to this clinic through locations
            cursor.execute("""
                SELECT prov.npi
                FROM providers prov
                JOIN locations l ON prov.location_id = l.location_id
                WHERE l.clinic_id = ? AND prov.is_active = 1
                LIMIT 1
            """, (clinic_id,))
            provider_row = cursor.fetchone()
            has_ordering_provider = provider_row is not None
            ordering_provider_npi = provider_row['npi'] if provider_row else None

            # Determine configuration status
            # config_submitted = True if any config values exist for this clinic
            config_submitted = len(config_dict) > 0 or clinic_phone is not None

            # Check all required fields from REQUIRED_CONFIG_FIELDS
            # A clinic is only "configured" when ALL required fields have values
            missing_fields = []
            if not clinic_phone:
                missing_fields.append('clinic_phone')
            if not default_test:
                missing_fields.append('default_test')
            if not default_specimen:
                missing_fields.append('default_specimen')
            if not has_ordering_provider:
                missing_fields.append('ordering_provider_npi')

            is_configured = len(missing_fields) == 0

            # Build configuration entry
            config_entry = {
                "program": program,
                "clinic": clinic_name,
                "clinic_phone": clinic_phone,
                "default_test": default_test,
                "optional_tests": optional_tests,
                "default_specimen": default_specimen,
                "ordering_provider_npi": ordering_provider_npi,
                "config_submitted": config_submitted,
                "is_configured": is_configured,
                "missing_fields": missing_fields
            }
            configurations.append(config_entry)

            # Track onboarding status
            if is_configured:
                clinics_configured.append({
                    "clinic": clinic_name,
                    "program": program
                })
            elif config_submitted:
                # Config was submitted but missing required fields
                missing_configurations.append({
                    "clinic": clinic_name,
                    "program": program,
                    "missing": missing_fields
                })
            else:
                # No config submitted yet - request is pending
                config_requests_pending.append({
                    "clinic": clinic_name,
                    "program": program
                })

        # =====================================================================
        # QUERY 4: PROVIDERS (SIMPLIFIED)
        # =====================================================================
        cursor.execute("""
            SELECT
                prov.name as name,
                prov.npi,
                p.prefix as program,
                c.name as clinic
            FROM providers prov
            JOIN locations l ON prov.location_id = l.location_id
            JOIN clinics c ON l.clinic_id = c.clinic_id
            JOIN programs p ON c.program_id = p.program_id
            WHERE prov.is_active = 1
              AND p.status = 'Active'
            ORDER BY prov.name
        """)
        providers = [dict(row) for row in cursor.fetchall()]

        # =====================================================================
        # QUERY 5: AUDIT TRAIL (CONDITIONAL)
        # =====================================================================
        audit_trail = []
        if show_audit_trail:
            cursor.execute("""
                SELECT
                    changed_date as timestamp,
                    action,
                    record_type as entity_type,
                    record_id as entity_id,
                    changed_by,
                    change_reason as details
                FROM audit_history
                WHERE changed_date >= date('now', '-' || ? || ' days')
                ORDER BY changed_date DESC
                LIMIT 500
            """, (days_of_audit,))
            audit_trail = [dict(row) for row in cursor.fetchall()]

        # =====================================================================
        # CALCULATE SUMMARY STATISTICS
        # =====================================================================
        total_users = len(users)
        active_users = len([u for u in users if u.get('status') == 'Active'])
        total_clinics = len(configurations)
        num_configured = len(clinics_configured)
        num_pending = len(config_requests_pending)
        num_missing = len(missing_configurations)

        summary = {
            "total_users": total_users,
            "active_users": active_users,
            "total_clinics": total_clinics,
            "clinics_configured": num_configured,
            "config_requests_pending": num_pending,
            "clinics_missing_configs": num_missing
        }

        # =====================================================================
        # BUILD ONBOARDING STATUS
        # =====================================================================
        onboarding_status = {
            "clinics_configured": clinics_configured,
            "config_requests_pending": config_requests_pending,
            "missing_configurations": missing_configurations
        }

        # =====================================================================
        # BUILD OUTPUT STRUCTURE
        # =====================================================================
        generated_at = datetime.now().isoformat() + "Z"

        output = {
            "generated_at": generated_at,
            "show_audit_trail": show_audit_trail,
            "summary": summary,
            "programs": programs,
            "users": users,
            "configurations": configurations,
            "providers": providers,
            "onboarding_status": onboarding_status,
            "audit_trail": audit_trail
        }

        # =====================================================================
        # WRITE JSON FILE
        # =====================================================================
        with open(output_path, 'w') as f:
            json.dump(output, f, indent=2, default=str)

        # =====================================================================
        # LOG AUDIT ENTRY
        # =====================================================================
        cursor.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "dashboard_data",
            "dashboard-data.json",
            "Generated",
            None,
            json.dumps(summary),
            "MCP:generate_dashboard_data",
            generated_at,
            f"Users: {total_users}, Clinics: {total_clinics}, Configured: {num_configured}"
        ))
        conn.commit()

        # =====================================================================
        # BUILD RETURN MESSAGE
        # =====================================================================
        audit_status = f"Included ({len(audit_trail)} entries)" if show_audit_trail else "Excluded"

        return f"""Dashboard data generated successfully!

File: {output_path}
Generated: {generated_at}

Summary:
  Users: {total_users} ({active_users} active)
  Clinics: {total_clinics} total
    - Configured: {num_configured}
    - Pending: {num_pending}
    - Missing configs: {num_missing}
  Providers: {len(providers)}
  Audit Trail: {audit_status}

Data sections:
  - programs: {len(programs)} programs
  - users: {len(users)} user access records
  - configurations: {len(configurations)} clinic configs
  - providers: {len(providers)} providers
  - onboarding_status: tracked
  - audit_trail: {len(audit_trail)} entries

Next steps:
  1. Review the generated JSON file
  2. Push to GitHub Pages: cd ~/projects/propel-clinic-dashboard && git add -A && git commit -m "Update dashboard data" && git push
  3. Dashboard will update automatically at your GitHub Pages URL
"""

    except sqlite3.Error as e:
        return f"Database error: {str(e)}\n\nMake sure the database exists at: {DB_PATH}"

    except Exception as e:
        logger.error(f"generate_dashboard_data() failed: {str(e)}", exc_info=True)
        return f"Error generating dashboard data: {str(e)}"

    finally:
        conn.close()


# ============================================================
# CLINIC MANAGEMENT TOOLS
# ============================================================
# Tools for creating, updating, and managing clinics, providers,
# configurations, and importing onboarding data.
#
# All tools support preview_only mode (default=True) for safe testing.

import uuid
import json as json_lib


def fuzzy_match_clinic(conn, search_term: str, program_id: str = None):
    """
    Find clinics matching a search term (fuzzy match on name, code, or epic_id).

    Args:
        conn: Database connection
        search_term: Clinic name, code, or ID to search for
        program_id: Optional program_id to narrow search

    Returns:
        List of matching clinic rows (dict format)
    """
    search_lower = search_term.lower().strip()

    query = """
        SELECT c.*, p.name as program_name, p.prefix as program_prefix
        FROM clinics c
        JOIN programs p ON c.program_id = p.program_id
        WHERE (
            LOWER(c.name) LIKE ?
            OR LOWER(c.code) LIKE ?
            OR c.clinic_id = ?
            OR c.epic_id = ?
        )
    """
    params = [f"%{search_lower}%", f"%{search_lower}%", search_term, search_term]

    if program_id:
        query += " AND c.program_id = ?"
        params.append(program_id)

    cursor = conn.execute(query, params)
    return [dict(row) for row in cursor.fetchall()]


def resolve_program_id_by_prefix(conn, prefix: str) -> str:
    """
    Get program_id from prefix (P4M, PR4M, GRX).

    Args:
        conn: Database connection
        prefix: Program prefix

    Returns:
        program_id string

    Raises:
        ValueError if program not found
    """
    cursor = conn.execute(
        "SELECT program_id FROM programs WHERE UPPER(prefix) = ?",
        (prefix.upper(),)
    )
    row = cursor.fetchone()
    if not row:
        raise ValueError(f"Program not found with prefix: {prefix}")
    return row['program_id']


def get_nested_value(data: dict, path: str):
    """
    Get nested dict value by dot-notation path.

    Example: get_nested_value(data, 'lab_config.default_sample_type')

    Args:
        data: Dictionary to traverse
        path: Dot-separated path (e.g., 'a.b.0.c')

    Returns:
        Value at path, or None if not found
    """
    keys = path.split('.')
    value = data
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        elif isinstance(value, list) and key.isdigit():
            idx = int(key)
            value = value[idx] if idx < len(value) else None
        else:
            return None
        if value is None:
            return None
    return value


@mcp.tool()
def list_clinics(
    program: Optional[str] = None,
    status: Optional[str] = None,
    include_config_status: bool = True
) -> str:
    """
    List all clinics with their key information.

    Args:
        program: Filter by program prefix (P4M, PR4M, GRX)
        status: Filter by status (Active, Inactive, Onboarding)
        include_config_status: Include configuration completeness info

    Returns:
        Formatted list of clinics grouped by program

    Example:
        list_clinics()                    # All clinics
        list_clinics(program="P4M")       # Only P4M clinics
        list_clinics(status="Active")     # Only active clinics
    """
    logger.info(f"list_clinics() called - program={program}, status={status}")

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Build query
        query = """
            SELECT
                c.*,
                p.name as program_name,
                p.prefix as program_prefix,
                (SELECT COUNT(*) FROM providers pr
                 JOIN locations l ON pr.location_id = l.location_id
                 WHERE l.clinic_id = c.clinic_id AND pr.is_active = 1) as provider_count,
                (SELECT COUNT(*) FROM config_values cv
                 WHERE cv.clinic_id = c.clinic_id) as config_count
            FROM clinics c
            JOIN programs p ON c.program_id = p.program_id
            WHERE 1=1
        """
        params = []

        if program:
            query += " AND UPPER(p.prefix) = ?"
            params.append(program.upper())

        if status:
            query += " AND c.status = ?"
            params.append(status)

        query += " ORDER BY p.prefix, c.name"

        cursor = conn.execute(query, params)
        clinics = cursor.fetchall()

        if not clinics:
            return "No clinics found matching the criteria."

        # Group by program
        by_program = {}
        for clinic in clinics:
            prefix = clinic['program_prefix']
            if prefix not in by_program:
                by_program[prefix] = {
                    'name': clinic['program_name'],
                    'clinics': []
                }
            by_program[prefix]['clinics'].append(clinic)

        # Format output
        lines = [f"Clinics ({len(clinics)} total)", "=" * 40, ""]

        for prefix, data in sorted(by_program.items()):
            lines.append(f"{prefix} - {data['name']} ({len(data['clinics'])} clinics):")

            for c in data['clinics']:
                epic_str = f"(EPIC: {c['epic_id']})" if c['epic_id'] else "(no EPIC ID)"
                status_str = c['status'] or 'Unknown'
                phone_str = c['phone'] or '(no phone)'
                provider_str = f"Providers: {c['provider_count']}"

                # Config status
                config_status = "✓" if c['config_submitted_at'] else "○"
                if c['config_count'] > 0:
                    config_str = f"Config: {config_status} ({c['config_count']} values)"
                else:
                    config_str = "Config: ✗ Not configured"

                lines.append(f"  • {c['name']} {epic_str} - {status_str}")
                lines.append(f"    Phone: {phone_str} | {provider_str} | {config_str}")

            lines.append("")

        conn.close()
        return "\n".join(lines)

    except sqlite3.Error as e:
        logger.error(f"list_clinics() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"list_clinics() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def create_clinic(
    clinic_name: str,
    program: str,
    epic_id: str = None,
    code: str = None,
    address_street: str = None,
    address_city: str = None,
    address_state: str = None,
    address_zip: str = None,
    phone: str = None,
    hours_of_operation: str = None,
    preview_only: bool = True
) -> str:
    """
    Create a new clinic in the database.

    Args:
        clinic_name: Full name of the clinic (e.g., "Franz Clinic")
        program: Program prefix (P4M, PR4M, GRX)
        epic_id: Clinic's EPIC system identifier
        code: Short code for the clinic (e.g., "FRANZ")
        address_street: Street address
        address_city: City
        address_state: State (2-letter code)
        address_zip: ZIP code
        phone: Main clinic phone number
        hours_of_operation: Operating hours (e.g., "08:00 - 17:00")
        preview_only: If True (default), shows what WOULD happen without making changes

    Returns:
        Confirmation with clinic_id or preview summary

    Example:
        create_clinic(
            clinic_name="Franz Clinic",
            program="P4M",
            epic_id="12345",
            phone="555-123-4567",
            preview_only=False
        )
    """
    logger.info(f"create_clinic() called - name={clinic_name}, program={program}, preview={preview_only}")

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # 1. Validate program exists
        try:
            program_id = resolve_program_id_by_prefix(conn, program)
        except ValueError as e:
            conn.close()
            return str(e)

        # Get program info for display
        cursor = conn.execute(
            "SELECT name, prefix FROM programs WHERE program_id = ?",
            (program_id,)
        )
        program_info = cursor.fetchone()

        # 2. Check for existing clinic with similar name or same EPIC ID
        similar = fuzzy_match_clinic(conn, clinic_name, program_id)

        # Also check for exact EPIC ID match
        if epic_id:
            cursor = conn.execute(
                "SELECT * FROM clinics WHERE epic_id = ?",
                (epic_id,)
            )
            epic_match = cursor.fetchone()
            if epic_match and dict(epic_match) not in similar:
                similar.append(dict(epic_match))

        # 3. If similar clinics found, return warning
        if similar:
            lines = ["⚠️  POTENTIAL DUPLICATES FOUND", ""]
            for s in similar:
                lines.append(f"  • {s['name']} (ID: {s['clinic_id'][:8]}...)")
                if s.get('epic_id'):
                    lines.append(f"    EPIC ID: {s['epic_id']}")
            lines.extend([
                "",
                "If this is intentional, you can still create the clinic.",
                "Otherwise, use update_clinic() to modify the existing record."
            ])
            # Don't block, just warn

        # 4. Generate clinic ID
        clinic_id = str(uuid.uuid4())

        # Auto-generate code if not provided
        if not code:
            # Take first 4-6 letters of clinic name
            code = ''.join(c for c in clinic_name.upper() if c.isalpha())[:6]

        # Format address
        address_parts = []
        if address_street:
            address_parts.append(address_street)
        if address_city or address_state or address_zip:
            city_state_zip = f"{address_city or ''}, {address_state or ''} {address_zip or ''}".strip(', ')
            address_parts.append(city_state_zip)
        address_display = ", ".join(address_parts) if address_parts else "(not provided)"

        # 5. Preview mode
        if preview_only:
            lines = [
                "=== CREATE CLINIC PREVIEW ===",
                "",
                f"Clinic: {clinic_name}",
                f"Code: {code}",
                f"Program: {program} ({program_info['name']})",
                f"EPIC ID: {epic_id or '(not provided)'}",
                f"Address: {address_display}",
                f"Phone: {phone or '(not provided)'}",
                f"Hours: {hours_of_operation or '(not provided)'}",
                ""
            ]

            if similar:
                lines.extend(["⚠️  Similar clinics exist (see above)", ""])

            lines.extend([
                "To create this clinic, run:",
                f'  create_clinic("{clinic_name}", "{program}", preview_only=False, ...)'
            ])

            conn.close()
            return "\n".join(lines)

        # 6. Execute INSERT
        cursor = conn.execute("""
            INSERT INTO clinics (
                clinic_id, program_id, name, code, status,
                epic_id, address_street, address_city, address_state, address_zip,
                phone, hours_of_operation, created_by
            ) VALUES (?, ?, ?, ?, 'Onboarding', ?, ?, ?, ?, ?, ?, ?, 'MCP:create_clinic')
        """, (
            clinic_id, program_id, clinic_name, code,
            epic_id, address_street, address_city, address_state, address_zip,
            phone, hours_of_operation
        ))

        # Log to audit
        log_audit(
            cursor, 'clinic', clinic_id, 'Created', 'clinic',
            None, clinic_name, 'MCP:create_clinic',
            datetime.now().isoformat(), f"Created clinic {clinic_name} for {program}"
        )

        conn.commit()
        conn.close()

        logger.info(f"create_clinic() SUCCESS - created {clinic_id}")

        return f"""✓ Clinic created successfully!

Clinic ID: {clinic_id}
Name: {clinic_name}
Code: {code}
Program: {program}
Status: Onboarding

Next steps:
  • Set configurations: set_clinic_config("{clinic_name}", "{program}", "helpdesk_phone", "555-999-9999")
  • Add providers: create_provider("{clinic_name}", "{program}", "Dr. Jane Smith", "1234567890")
  • Or import full config: import_onboarding_json("~/imports/file.json", "{program}")
"""

    except sqlite3.Error as e:
        logger.error(f"create_clinic() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"create_clinic() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def update_clinic(
    clinic: str,
    program: str = None,
    clinic_name: str = None,
    epic_id: str = None,
    code: str = None,
    address_street: str = None,
    address_city: str = None,
    address_state: str = None,
    address_zip: str = None,
    phone: str = None,
    hours_of_operation: str = None,
    status: str = None,
    preview_only: bool = True
) -> str:
    """
    Update an existing clinic's information.

    Args:
        clinic: Clinic name, code, or ID to update (fuzzy match supported)
        program: Program prefix to help identify clinic if name is ambiguous
        clinic_name: Updated clinic name
        epic_id: Updated EPIC ID
        code: Updated clinic code
        address_street: Updated street address
        address_city: Updated city
        address_state: Updated state
        address_zip: Updated ZIP
        phone: Updated phone
        hours_of_operation: Updated hours
        status: Updated status (Active, Inactive, Onboarding, Archived)
        preview_only: If True (default), shows what WOULD change

    Returns:
        Summary of changes (preview or committed)

    Example:
        update_clinic("Franz", "P4M", phone="555-999-8888", preview_only=False)
    """
    logger.info(f"update_clinic() called - clinic={clinic}, program={program}, preview={preview_only}")

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Resolve program_id if provided
        program_id = None
        if program:
            try:
                program_id = resolve_program_id_by_prefix(conn, program)
            except ValueError as e:
                conn.close()
                return str(e)

        # Find clinic
        matches = fuzzy_match_clinic(conn, clinic, program_id)

        if not matches:
            conn.close()
            return f"Clinic not found: {clinic}" + (f" in program {program}" if program else "")

        if len(matches) > 1:
            lines = ["Multiple clinics found. Please be more specific:", ""]
            for m in matches:
                lines.append(f"  • {m['name']} ({m['program_prefix']}) - ID: {m['clinic_id'][:8]}...")
            lines.extend(["", "Specify the program parameter to narrow down."])
            conn.close()
            return "\n".join(lines)

        current = matches[0]

        # Build update dict with only provided (non-None) values
        updates = {}
        field_labels = {
            'name': 'Clinic Name',
            'epic_id': 'EPIC ID',
            'code': 'Code',
            'address_street': 'Street',
            'address_city': 'City',
            'address_state': 'State',
            'address_zip': 'ZIP',
            'phone': 'Phone',
            'hours_of_operation': 'Hours',
            'status': 'Status'
        }

        if clinic_name is not None:
            updates['name'] = clinic_name
        if epic_id is not None:
            updates['epic_id'] = epic_id
        if code is not None:
            updates['code'] = code
        if address_street is not None:
            updates['address_street'] = address_street
        if address_city is not None:
            updates['address_city'] = address_city
        if address_state is not None:
            updates['address_state'] = address_state
        if address_zip is not None:
            updates['address_zip'] = address_zip
        if phone is not None:
            updates['phone'] = phone
        if hours_of_operation is not None:
            updates['hours_of_operation'] = hours_of_operation
        if status is not None:
            # Validate status
            valid_statuses = ['Active', 'Inactive', 'Onboarding', 'Archived']
            if status not in valid_statuses:
                conn.close()
                return f"Invalid status: {status}. Valid values: {', '.join(valid_statuses)}"
            updates['status'] = status

        if not updates:
            conn.close()
            return "No changes specified. Provide at least one field to update."

        # Preview mode
        if preview_only:
            lines = [
                "=== UPDATE CLINIC PREVIEW ===",
                "",
                f"Clinic: {current['name']} (ID: {current['clinic_id'][:8]}...)",
                f"Program: {current['program_prefix']}",
                "",
                "Changes:"
            ]

            for field, new_val in updates.items():
                old_val = current.get(field) or "(not set)"
                label = field_labels.get(field, field)
                lines.append(f'  {label}: "{old_val}" → "{new_val}"')

            # List unchanged fields
            unchanged = [f for f in field_labels if f not in updates]
            if unchanged:
                lines.extend(["", f"No changes to: {', '.join(field_labels[f] for f in unchanged)}"])

            lines.extend([
                "",
                "To apply these changes, run:",
                f'  update_clinic("{clinic}", "{current["program_prefix"]}", ..., preview_only=False)'
            ])

            conn.close()
            return "\n".join(lines)

        # Execute UPDATE
        set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
        set_clause += ", updated_date = CURRENT_TIMESTAMP"

        conn.execute(
            f"UPDATE clinics SET {set_clause} WHERE clinic_id = ?",
            list(updates.values()) + [current['clinic_id']]
        )

        # Log to audit
        cursor = conn.cursor()
        for field, new_val in updates.items():
            old_val = current.get(field)
            log_audit(
                cursor, 'clinic', current['clinic_id'], 'Updated', field,
                str(old_val), str(new_val), 'MCP:update_clinic',
                datetime.now().isoformat(), f"Updated {field}"
            )

        conn.commit()
        conn.close()

        logger.info(f"update_clinic() SUCCESS - updated {current['clinic_id']}")

        changes_summary = ", ".join(f"{field_labels.get(f, f)}" for f in updates.keys())
        return f"""✓ Clinic updated successfully!

Clinic: {current['name']}
Updated fields: {changes_summary}
"""

    except sqlite3.Error as e:
        logger.error(f"update_clinic() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"update_clinic() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def set_clinic_config(
    clinic: str,
    program: str,
    config_key: str,
    config_value: str,
    preview_only: bool = True
) -> str:
    """
    Set a configuration value for a clinic.

    Args:
        clinic: Clinic name (fuzzy match supported)
        program: Program prefix (P4M, PR4M, GRX)
        config_key: Configuration key (e.g., "helpdesk_phone")
        config_value: Value to set
        preview_only: If True (default), shows what WOULD change

    Returns:
        Confirmation of config change

    Common config keys:
        helpdesk_phone, helpdesk_email, helpdesk_hours,
        default_test, default_specimen, lab_account_number,
        extract_patient_status, extract_procedure_type

    Example:
        set_clinic_config("Franz", "P4M", "helpdesk_phone", "555-999-9999", preview_only=False)
    """
    logger.info(f"set_clinic_config() called - clinic={clinic}, key={config_key}, preview={preview_only}")

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Resolve program
        try:
            program_id = resolve_program_id_by_prefix(conn, program)
        except ValueError as e:
            conn.close()
            return str(e)

        # Find clinic
        matches = fuzzy_match_clinic(conn, clinic, program_id)

        if not matches:
            conn.close()
            return f"Clinic not found: {clinic} in program {program}"

        if len(matches) > 1:
            lines = ["Multiple clinics found:", ""]
            for m in matches:
                lines.append(f"  • {m['name']}")
            conn.close()
            return "\n".join(lines)

        clinic_info = matches[0]
        clinic_id = clinic_info['clinic_id']

        # Check if config key exists in definitions
        cursor = conn.execute(
            "SELECT * FROM config_definitions WHERE config_key = ?",
            (config_key,)
        )
        config_def = cursor.fetchone()

        if not config_def:
            # Get list of valid keys
            cursor = conn.execute(
                "SELECT config_key, display_name, category FROM config_definitions ORDER BY category, display_name"
            )
            all_keys = cursor.fetchall()

            lines = [f"Unknown config key: {config_key}", "", "Valid config keys:"]
            current_cat = None
            for k in all_keys:
                if k['category'] != current_cat:
                    current_cat = k['category']
                    lines.append(f"\n  {current_cat}:")
                lines.append(f"    • {k['config_key']} - {k['display_name']}")

            conn.close()
            return "\n".join(lines)

        # Check current value at clinic level
        cursor = conn.execute(
            "SELECT value FROM config_values WHERE config_key = ? AND clinic_id = ? AND location_id IS NULL",
            (config_key, clinic_id)
        )
        current_clinic_val = cursor.fetchone()

        # Check program-level value (inherited)
        cursor = conn.execute(
            "SELECT value FROM config_values WHERE config_key = ? AND program_id = ? AND clinic_id IS NULL",
            (config_key, program_id)
        )
        program_val = cursor.fetchone()

        # Determine current effective value
        if current_clinic_val:
            current_str = f'"{current_clinic_val["value"]}" [Clinic Override]'
            is_update = True
        elif program_val:
            current_str = f'"{program_val["value"]}" [Inherited from Program]'
            is_update = False
        else:
            current_str = f'"{config_def["default_value"] or "(none)"}" [Default]'
            is_update = False

        # Preview mode
        if preview_only:
            lines = [
                "=== SET CONFIG PREVIEW ===",
                "",
                f"Clinic: {clinic_info['name']} ({program})",
                f"Config: {config_key} ({config_def['display_name']})",
                "",
                f"Current: {current_str}",
                f'New: "{config_value}" [Clinic Override]',
                "",
                "To apply, run:",
                f'  set_clinic_config("{clinic}", "{program}", "{config_key}", "{config_value}", preview_only=False)'
            ]
            conn.close()
            return "\n".join(lines)

        # Execute INSERT or UPDATE
        if is_update:
            conn.execute("""
                UPDATE config_values
                SET value = ?, is_override = TRUE, updated_date = CURRENT_TIMESTAMP
                WHERE config_key = ? AND clinic_id = ? AND location_id IS NULL
            """, (config_value, config_key, clinic_id))
        else:
            conn.execute("""
                INSERT INTO config_values (config_key, program_id, clinic_id, value, is_override, source, created_by)
                VALUES (?, ?, ?, ?, TRUE, 'manual', 'MCP:set_clinic_config')
            """, (config_key, program_id, clinic_id, config_value))

        # Log to audit
        cursor = conn.cursor()
        log_audit(
            cursor, 'config', f"{clinic_id}:{config_key}",
            'Updated' if is_update else 'Created',
            config_key,
            current_clinic_val['value'] if current_clinic_val else None,
            config_value, 'MCP:set_clinic_config',
            datetime.now().isoformat(), f"Set {config_key} for {clinic_info['name']}"
        )

        conn.commit()
        conn.close()

        logger.info(f"set_clinic_config() SUCCESS - {config_key}={config_value} for {clinic_id}")

        return f"""✓ Configuration set successfully!

Clinic: {clinic_info['name']}
Config: {config_key}
Value: {config_value}
"""

    except sqlite3.Error as e:
        logger.error(f"set_clinic_config() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"set_clinic_config() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def set_clinic_configs_batch(
    clinic: str,
    program: str,
    configs: str,
    preview_only: bool = True
) -> str:
    """
    Set multiple configuration values for a clinic at once.

    Args:
        clinic: Clinic name (fuzzy match supported)
        program: Program prefix (P4M, PR4M, GRX)
        configs: JSON string of key-value pairs
        preview_only: If True (default), shows what WOULD change

    Returns:
        Summary of all config changes

    Example:
        set_clinic_configs_batch(
            clinic="Franz",
            program="P4M",
            configs='{"helpdesk_phone": "555-999-9999", "default_specimen": "Saliva"}',
            preview_only=False
        )
    """
    logger.info(f"set_clinic_configs_batch() called - clinic={clinic}, preview={preview_only}")

    try:
        # Parse JSON configs
        try:
            config_dict = json_lib.loads(configs)
        except json_lib.JSONDecodeError as e:
            return f"Invalid JSON in configs parameter: {e}"

        if not isinstance(config_dict, dict):
            return "configs must be a JSON object (dict)"

        results = []
        for key, value in config_dict.items():
            result = set_clinic_config(clinic, program, key, str(value), preview_only)
            results.append(f"--- {key} ---\n{result}")

        return "\n\n".join(results)

    except Exception as e:
        logger.error(f"set_clinic_configs_batch() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def create_provider(
    clinic: str,
    program: str,
    provider_name: str,
    npi: str,
    phone: str = None,
    email: str = None,
    specialty: str = None,
    office_street: str = None,
    office_city: str = None,
    office_state: str = None,
    office_zip: str = None,
    preview_only: bool = True
) -> str:
    """
    Add an ordering provider to a clinic.

    Args:
        clinic: Clinic name (fuzzy match supported)
        program: Program prefix (P4M, PR4M, GRX)
        provider_name: Full name (e.g., "Dr. Jane Smith")
        npi: 10-digit NPI number
        phone: Provider's phone number
        email: Provider's email address
        specialty: Medical specialty (optional)
        office_street: Office address street
        office_city: Office city
        office_state: Office state (2-letter)
        office_zip: Office ZIP code
        preview_only: If True (default), shows what WOULD be created

    Returns:
        Confirmation with provider_id

    Example:
        create_provider(
            clinic="Franz",
            program="P4M",
            provider_name="Dr. Jane Smith",
            npi="1234567890",
            phone="555-123-4567",
            preview_only=False
        )
    """
    logger.info(f"create_provider() called - clinic={clinic}, name={provider_name}, preview={preview_only}")

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Resolve program
        try:
            program_id = resolve_program_id_by_prefix(conn, program)
        except ValueError as e:
            conn.close()
            return str(e)

        # Find clinic
        matches = fuzzy_match_clinic(conn, clinic, program_id)

        if not matches:
            conn.close()
            return f"Clinic not found: {clinic} in program {program}"

        if len(matches) > 1:
            lines = ["Multiple clinics found:", ""]
            for m in matches:
                lines.append(f"  • {m['name']}")
            conn.close()
            return "\n".join(lines)

        clinic_info = matches[0]
        clinic_id = clinic_info['clinic_id']

        # Validate NPI (10 digits)
        npi_clean = ''.join(c for c in npi if c.isdigit())
        if len(npi_clean) != 10:
            conn.close()
            return f"Invalid NPI: {npi}. Must be exactly 10 digits."

        # Get the default location for this clinic (or create one if none exists)
        cursor = conn.execute(
            "SELECT location_id, name FROM locations WHERE clinic_id = ? LIMIT 1",
            (clinic_id,)
        )
        location = cursor.fetchone()

        if not location:
            # Create a default location
            location_id = str(uuid.uuid4())
            conn.execute("""
                INSERT INTO locations (location_id, clinic_id, name, code, is_primary, created_by)
                VALUES (?, ?, ?, 'MAIN', TRUE, 'MCP:create_provider')
            """, (location_id, clinic_id, f"{clinic_info['name']} - Main"))
            location_name = f"{clinic_info['name']} - Main"
        else:
            location_id = location['location_id']
            location_name = location['name']

        # Check for duplicate NPI in this clinic
        cursor = conn.execute("""
            SELECT pr.name FROM providers pr
            JOIN locations l ON pr.location_id = l.location_id
            WHERE l.clinic_id = ? AND pr.npi = ? AND pr.is_active = TRUE
        """, (clinic_id, npi_clean))
        existing = cursor.fetchone()

        if existing:
            conn.close()
            return f"Provider with NPI {npi_clean} already exists in this clinic: {existing['name']}"

        # Format office address
        office_parts = []
        if office_street:
            office_parts.append(office_street)
        if office_city or office_state or office_zip:
            csz = f"{office_city or ''}, {office_state or ''} {office_zip or ''}".strip(', ')
            office_parts.append(csz)
        office_display = ", ".join(office_parts) if office_parts else "(not provided)"

        # Preview mode
        if preview_only:
            lines = [
                "=== CREATE PROVIDER PREVIEW ===",
                "",
                f"Clinic: {clinic_info['name']} ({program})",
                f"Location: {location_name}",
                "",
                f"Provider: {provider_name}",
                f"NPI: {npi_clean}",
                f"Phone: {phone or '(not provided)'}",
                f"Email: {email or '(not provided)'}",
                f"Specialty: {specialty or '(not provided)'}",
                f"Office Address: {office_display}",
                "",
                "To create this provider, run:",
                f'  create_provider("{clinic}", "{program}", "{provider_name}", "{npi}", preview_only=False, ...)'
            ]
            conn.close()
            return "\n".join(lines)

        # Execute INSERT
        cursor = conn.execute("""
            INSERT INTO providers (
                location_id, name, npi, phone, email, specialty,
                office_street, office_city, office_state, office_zip,
                is_active, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, TRUE, 'MCP:create_provider')
        """, (
            location_id, provider_name, npi_clean, phone, email, specialty,
            office_street, office_city, office_state, office_zip
        ))

        provider_id = cursor.lastrowid

        # Log to audit
        log_audit(
            cursor, 'provider', str(provider_id), 'Created', 'provider',
            None, provider_name, 'MCP:create_provider',
            datetime.now().isoformat(), f"Created provider {provider_name} (NPI: {npi_clean})"
        )

        conn.commit()
        conn.close()

        logger.info(f"create_provider() SUCCESS - created {provider_id}")

        return f"""✓ Provider created successfully!

Provider ID: {provider_id}
Name: {provider_name}
NPI: {npi_clean}
Clinic: {clinic_info['name']}
Location: {location_name}
"""

    except sqlite3.Error as e:
        logger.error(f"create_provider() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"create_provider() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def import_onboarding_json(
    file_path: str,
    program: str,
    preview_only: bool = True,
    update_clinic_id: str = None,
    force_create: bool = False
) -> str:
    """
    Import clinic onboarding data from a JSON file into the database.

    This is the main import tool that processes the JSON output from the
    onboarding form and creates/updates all related records.

    Args:
        file_path: Path to the onboarding JSON file (e.g., ~/imports/franz.json)
        program: Program prefix (P4M, PR4M, GRX)
        preview_only: If True (default), shows what WOULD happen without making changes
        update_clinic_id: If set, update this existing clinic instead of creating new
        force_create: If True, create new clinic even if similar names exist

    Returns:
        Summary of import actions (preview or committed)

    Example:
        # Preview first
        import_onboarding_json("~/imports/franz.json", "P4M")

        # Then commit
        import_onboarding_json("~/imports/franz.json", "P4M", preview_only=False)
    """
    logger.info(f"import_onboarding_json() called - file={file_path}, program={program}, preview={preview_only}")

    try:
        # Expand ~ in path
        file_path = os.path.expanduser(file_path)

        # 1. Load and validate JSON
        if not os.path.exists(file_path):
            return f"Error: File not found at {file_path}"

        try:
            with open(file_path, 'r') as f:
                data = json_lib.load(f)
        except json_lib.JSONDecodeError as e:
            return f"Error: Invalid JSON in file. Details: {e}"

        # 2. Validate required fields
        # Support both old format (clinic_info) and new format (clinic_information)
        clinic_info = data.get('clinic_information', data.get('clinic_info', {}))
        clinic_name = clinic_info.get('clinic_name')
        # Support both epic_department_id (new) and clinic_epic_id (old)
        epic_id = clinic_info.get('epic_department_id') or clinic_info.get('clinic_epic_id')

        if not clinic_name:
            return "Error: Missing clinic_information.clinic_name in JSON file"

        # Connect to database
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Resolve program
        try:
            program_id = resolve_program_id_by_prefix(conn, program)
        except ValueError as e:
            conn.close()
            return str(e)

        # Get program info
        cursor = conn.execute("SELECT name FROM programs WHERE program_id = ?", (program_id,))
        program_info = cursor.fetchone()

        # 3. Check for existing clinic
        existing_clinic = None
        if update_clinic_id:
            cursor = conn.execute("SELECT * FROM clinics WHERE clinic_id = ?", (update_clinic_id,))
            existing_clinic = cursor.fetchone()
            if not existing_clinic:
                conn.close()
                return f"Error: Clinic not found with ID: {update_clinic_id}"
        elif not force_create:
            matches = fuzzy_match_clinic(conn, clinic_name, program_id)
            if epic_id:
                cursor = conn.execute("SELECT * FROM clinics WHERE epic_id = ?", (epic_id,))
                epic_match = cursor.fetchone()
                if epic_match:
                    matches.append(dict(epic_match))

            if matches:
                lines = [
                    "⚠️  SIMILAR CLINICS FOUND",
                    "",
                    "Existing clinics that may match:"
                ]
                for m in matches:
                    lines.append(f"  • {m['name']} (ID: {m['clinic_id'][:8]}...)")
                    if m.get('epic_id'):
                        lines.append(f"    EPIC ID: {m['epic_id']}")
                lines.extend([
                    "",
                    "Options:",
                    f"  • Update existing: import_onboarding_json(..., update_clinic_id=\"{matches[0]['clinic_id']}\")",
                    "  • Force new: import_onboarding_json(..., force_create=True)"
                ])
                conn.close()
                return "\n".join(lines)

        # 4. Build import plan
        plan = {
            'clinic': {
                'action': 'UPDATE' if existing_clinic else 'CREATE',
                'clinic_id': existing_clinic['clinic_id'] if existing_clinic else str(uuid.uuid4()),
                'name': clinic_name,
                'epic_id': epic_id,
                # Support both 'address' (new) and 'clinic_address' (old)
                'address': clinic_info.get('address') or clinic_info.get('clinic_address'),
                # Support both 'clinic_phone' (form field) and direct 'phone'
                'phone': clinic_info.get('clinic_phone') or clinic_info.get('phone'),
                'hours': clinic_info.get('hours_of_operation')
            },
            'locations': [],
            'providers': [],
            'configs': []
        }

        # Parse satellite locations (nested under clinic_information in new format)
        satellite_locations = clinic_info.get('satellite_locations', data.get('satellite_locations', []))
        for loc in satellite_locations:
            # Support both 'name' (new format) and 'location_name' (old format)
            loc_name = loc.get('name') or loc.get('location_name')
            if loc_name:
                plan['locations'].append({
                    'action': 'CREATE',
                    'name': loc_name,
                    # Support both 'address' (new) and 'location_address' (old)
                    'address': loc.get('address') or loc.get('location_address'),
                    'phone': loc.get('phone') or loc.get('location_phone'),
                    'hours': loc.get('hours_of_operation') or loc.get('location_hours')
                })

        # Parse ordering providers
        for prov in data.get('ordering_providers', []):
            if prov.get('name') and prov.get('npi'):
                plan['providers'].append({
                    'action': 'CREATE',
                    'name': prov.get('name'),
                    'npi': prov.get('npi'),
                    'phone': prov.get('phone'),
                    'email': prov.get('email'),
                    'specialty': prov.get('specialty'),
                    'office_address': prov.get('office_address')
                })

        # Parse configurations
        config_mappings = [
            ('lab_order_configuration.billing_method', 'billing_method'),
            ('lab_order_configuration.send_kit_to_patient', 'send_kit_to_patient'),
            ('lab_order_configuration.indication', 'default_indication'),
            ('lab_order_configuration.criteria_for_testing', 'criteria_for_testing'),
            ('lab_order_configuration.specimen_collection.default', 'default_specimen'),
            ('lab_order_configuration.test_panel.test_name', 'default_test'),
            ('lab_order_configuration.test_panel.test_code', 'default_test_code'),
            ('helpdesk.phone', 'helpdesk_phone'),
            ('helpdesk.include_in_emails', 'helpdesk_phone_in_emails'),
            ('extract_filtering.patient_status', 'extract_patient_status'),
            ('extract_filtering.procedure_type', 'extract_procedure_type'),
            ('extract_filtering.filter_by_provider', 'extract_filter_by_provider'),
        ]

        for json_path, config_key in config_mappings:
            value = get_nested_value(data, json_path)
            if value is not None:
                plan['configs'].append({
                    'action': 'SET',
                    'key': config_key,
                    'value': str(value) if not isinstance(value, str) else value
                })

        # Handle provider_list specially - can be array of objects or comma-separated string
        provider_list = get_nested_value(data, 'extract_filtering.provider_list')
        if provider_list:
            # New format: array of {first_name, last_name} objects
            if isinstance(provider_list, list) and len(provider_list) > 0:
                if isinstance(provider_list[0], dict):
                    # Convert to comma-separated names: "Jane Smith, John Doe"
                    names = [f"{p.get('first_name', '')} {p.get('last_name', '')}".strip()
                             for p in provider_list if p.get('first_name') or p.get('last_name')]
                    provider_list = ', '.join(names)
                else:
                    # Already a list of strings
                    provider_list = ', '.join(str(p) for p in provider_list)
            plan['configs'].append({
                'action': 'SET',
                'key': 'extract_providers',
                'value': provider_list
            })

        # 5. Preview mode
        if preview_only:
            lines = [
                "=== IMPORT PREVIEW ===",
                f"File: {file_path}",
                f"Program: {program} ({program_info['name']})",
                "Mode: PREVIEW (no changes will be made)",
                "",
                "CLINIC:",
                f"  {'✎' if plan['clinic']['action'] == 'UPDATE' else '✓'} {plan['clinic']['action']} clinic \"{plan['clinic']['name']}\""
            ]

            if plan['clinic']['epic_id']:
                lines.append(f"    EPIC ID: {plan['clinic']['epic_id']}")
            if plan['clinic']['phone']:
                lines.append(f"    Phone: {plan['clinic']['phone']}")

            if plan['locations']:
                lines.append(f"\nSATELLITE LOCATIONS ({len(plan['locations'])}):")
                for loc in plan['locations']:
                    lines.append(f"  ✓ CREATE location \"{loc['name']}\"")

            if plan['providers']:
                lines.append(f"\nORDERING PROVIDERS ({len(plan['providers'])}):")
                for prov in plan['providers']:
                    lines.append(f"  ✓ CREATE provider \"{prov['name']}\" (NPI: {prov['npi']})")

            if plan['configs']:
                lines.append(f"\nCONFIGURATIONS ({len(plan['configs'])}):")
                for cfg in plan['configs']:
                    lines.append(f"  ✓ SET {cfg['key']} = \"{cfg['value']}\"")

            lines.extend([
                "",
                "=== SUMMARY ===",
                f"Ready to {'update' if plan['clinic']['action'] == 'UPDATE' else 'create'}:",
                f"  • 1 clinic",
                f"  • {len(plan['locations'])} satellite location(s)",
                f"  • {len(plan['providers'])} ordering provider(s)",
                f"  • {len(plan['configs'])} configuration value(s)",
                "",
                "To execute this import, run:",
                f'  import_onboarding_json("{file_path}", "{program}", preview_only=False)'
            ])

            conn.close()
            return "\n".join(lines)

        # 6. Execute import
        clinic_id = plan['clinic']['clinic_id']

        # Create or update clinic
        if plan['clinic']['action'] == 'CREATE':
            address = plan['clinic'].get('address') or {}
            conn.execute("""
                INSERT INTO clinics (
                    clinic_id, program_id, name, epic_id, status,
                    address_street, address_city, address_state, address_zip,
                    phone, hours_of_operation, config_submitted_at, created_by
                ) VALUES (?, ?, ?, ?, 'Onboarding', ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 'MCP:import_onboarding_json')
            """, (
                clinic_id, program_id, plan['clinic']['name'], plan['clinic']['epic_id'],
                address.get('street'), address.get('city'), address.get('state'), address.get('zip'),
                plan['clinic']['phone'], plan['clinic']['hours']
            ))
        else:
            address = plan['clinic'].get('address') or {}
            conn.execute("""
                UPDATE clinics SET
                    name = ?, epic_id = ?,
                    address_street = ?, address_city = ?, address_state = ?, address_zip = ?,
                    phone = ?, hours_of_operation = ?,
                    config_submitted_at = CURRENT_TIMESTAMP, updated_date = CURRENT_TIMESTAMP
                WHERE clinic_id = ?
            """, (
                plan['clinic']['name'], plan['clinic']['epic_id'],
                address.get('street'), address.get('city'), address.get('state'), address.get('zip'),
                plan['clinic']['phone'], plan['clinic']['hours'],
                clinic_id
            ))

        # Get or create default location
        cursor = conn.execute(
            "SELECT location_id FROM locations WHERE clinic_id = ? LIMIT 1",
            (clinic_id,)
        )
        loc_row = cursor.fetchone()
        if loc_row:
            main_location_id = loc_row['location_id']
        else:
            main_location_id = str(uuid.uuid4())
            conn.execute("""
                INSERT INTO locations (location_id, clinic_id, name, code, is_primary, created_by)
                VALUES (?, ?, ?, 'MAIN', TRUE, 'MCP:import_onboarding_json')
            """, (main_location_id, clinic_id, f"{plan['clinic']['name']} - Main"))

        # Create satellite locations
        for loc in plan['locations']:
            loc_id = str(uuid.uuid4())
            loc_addr = loc.get('address') or {}
            conn.execute("""
                INSERT INTO locations (location_id, clinic_id, name, is_primary, created_by)
                VALUES (?, ?, ?, FALSE, 'MCP:import_onboarding_json')
            """, (loc_id, clinic_id, loc['name']))

        # Create providers
        for prov in plan['providers']:
            npi_clean = ''.join(c for c in prov['npi'] if c.isdigit())
            office = prov.get('office_address') or {}
            conn.execute("""
                INSERT INTO providers (
                    location_id, name, npi, phone, email, specialty,
                    office_street, office_city, office_state, office_zip,
                    is_active, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, TRUE, 'MCP:import_onboarding_json')
            """, (
                main_location_id, prov['name'], npi_clean,
                prov.get('phone'), prov.get('email'), prov.get('specialty'),
                office.get('street'), office.get('city'), office.get('state'), office.get('zip')
            ))

        # Set configurations
        for cfg in plan['configs']:
            # Check if exists
            cursor = conn.execute(
                "SELECT value_id FROM config_values WHERE config_key = ? AND clinic_id = ? AND location_id IS NULL",
                (cfg['key'], clinic_id)
            )
            existing = cursor.fetchone()

            if existing:
                conn.execute("""
                    UPDATE config_values SET value = ?, is_override = TRUE, updated_date = CURRENT_TIMESTAMP
                    WHERE value_id = ?
                """, (cfg['value'], existing['value_id']))
            else:
                # Check if config_key exists in definitions
                cursor = conn.execute("SELECT 1 FROM config_definitions WHERE config_key = ?", (cfg['key'],))
                if cursor.fetchone():
                    conn.execute("""
                        INSERT INTO config_values (config_key, program_id, clinic_id, value, is_override, source, created_by)
                        VALUES (?, ?, ?, ?, TRUE, 'import', 'MCP:import_onboarding_json')
                    """, (cfg['key'], program_id, clinic_id, cfg['value']))

        # Log to audit
        cursor = conn.cursor()
        log_audit(
            cursor, 'clinic', clinic_id,
            'Updated' if plan['clinic']['action'] == 'UPDATE' else 'Created',
            'import', None,
            json_lib.dumps({
                'file': file_path,
                'locations': len(plan['locations']),
                'providers': len(plan['providers']),
                'configs': len(plan['configs'])
            }),
            'MCP:import_onboarding_json',
            datetime.now().isoformat(),
            f"Imported from {os.path.basename(file_path)}"
        )

        conn.commit()
        conn.close()

        logger.info(f"import_onboarding_json() SUCCESS - imported {clinic_id}")

        return f"""✓ Import completed successfully!

Clinic: {plan['clinic']['name']}
Clinic ID: {clinic_id}
Action: {'Updated' if plan['clinic']['action'] == 'UPDATE' else 'Created'}

Imported:
  • {len(plan['locations'])} satellite location(s)
  • {len(plan['providers'])} ordering provider(s)
  • {len(plan['configs'])} configuration value(s)

Next steps:
  • View clinic: list_clinics(program="{program}")
  • Add more configs: set_clinic_config("{plan['clinic']['name']}", "{program}", ...)
  • Regenerate dashboard: generate_dashboard_data()
"""

    except sqlite3.Error as e:
        logger.error(f"import_onboarding_json() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"import_onboarding_json() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


# ============================================================
# REQUIREMENTS TOOLKIT - UAT CYCLE MANAGEMENT
# ============================================================
# These tools support the UAT process including cycle creation,
# test assignment, execution tracking, and reporting.


@mcp.tool()
def create_uat_cycle(
    program_prefix: str,
    name: str,
    uat_type: str,
    description: str = None,
    target_launch_date: str = None,
    clinical_pm: str = None,
    clinical_pm_email: str = None
) -> str:
    """
    Create a new UAT cycle for a program.

    Args:
        program_prefix: Program prefix (e.g., "PROP", "GRX")
        name: Cycle name (e.g., "NCCN Q4 2025", "GenoRx e-Consent v1")
        uat_type: Type of UAT - 'feature', 'rule_validation', or 'regression'
        description: Optional description
        target_launch_date: Target launch date (YYYY-MM-DD)
        clinical_pm: Clinical PM name
        clinical_pm_email: Clinical PM email

    Returns:
        Created cycle details

    Example:
        create_uat_cycle("PROP", "NCCN Q4 2025", "rule_validation",
                        target_launch_date="2025-01-15", clinical_pm="Jane Smith")
    """
    logger.info(f"create_uat_cycle() called - program={program_prefix}, name={name}")

    # Validate uat_type
    valid_types = ['feature', 'rule_validation', 'regression']
    if uat_type not in valid_types:
        return f"Error: Invalid uat_type '{uat_type}'. Valid types: {', '.join(valid_types)}"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Resolve program
        cursor = conn.execute(
            "SELECT program_id, name, prefix FROM programs WHERE UPPER(prefix) = UPPER(?)",
            (program_prefix,)
        )
        program = cursor.fetchone()
        if not program:
            return f"Error: Program not found with prefix '{program_prefix}'"

        # Generate cycle_id
        cycle_id = f"UAT-{program['prefix']}-{str(uuid.uuid4())[:8].upper()}"

        # Insert cycle
        conn.execute("""
            INSERT INTO uat_cycles (
                cycle_id, program_id, name, description, uat_type,
                target_launch_date, clinical_pm, clinical_pm_email,
                status, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'planning', 'MCP:create_uat_cycle')
        """, (
            cycle_id, program['program_id'], name, description, uat_type,
            target_launch_date, clinical_pm, clinical_pm_email
        ))

        # Create default pre-UAT gate items based on uat_type
        gate_items = []
        if uat_type == 'rule_validation':
            gate_items = [
                ('feature_deployment', 1, 'NCCN rules deployed to QA environment', 1),
                ('feature_deployment', 2, 'All rule IDs verified in system', 1),
                ('critical_path', 1, 'Patient registration flow tested', 1),
                ('critical_path', 2, 'Test profiles created and validated', 1),
                ('environment', 1, 'QA environment stable and accessible', 1),
                ('blocker_check', 1, 'No critical defects in backlog', 1),
                ('sign_off', 1, 'Clinical PM approval for test start', 1),
            ]
        elif uat_type == 'feature':
            gate_items = [
                ('feature_deployment', 1, 'Feature deployed to QA environment', 1),
                ('feature_deployment', 2, 'Feature flags configured correctly', 1),
                ('critical_path', 1, 'Core happy path verified', 1),
                ('environment', 1, 'QA environment stable and accessible', 1),
                ('blocker_check', 1, 'No critical defects blocking feature', 1),
                ('sign_off', 1, 'Product Owner approval for test start', 1),
            ]
        else:  # regression
            gate_items = [
                ('feature_deployment', 1, 'Release candidate deployed to QA', 1),
                ('critical_path', 1, 'Smoke tests passing', 1),
                ('environment', 1, 'QA environment mirrors production', 1),
                ('blocker_check', 1, 'No P0/P1 defects open', 1),
                ('sign_off', 1, 'Release Manager approval', 1),
            ]

        for category, seq, item_text, is_required in gate_items:
            conn.execute("""
                INSERT INTO pre_uat_gate_items (cycle_id, category, sequence, item_text, is_required)
                VALUES (?, ?, ?, ?, ?)
            """, (cycle_id, category, seq, item_text, is_required))

        conn.commit()

        logger.info(f"create_uat_cycle() SUCCESS - created {cycle_id}")

        result = f"""✓ UAT Cycle created successfully!

Cycle ID: {cycle_id}
Program: {program['name']} [{program['prefix']}]
Name: {name}
Type: {uat_type}
Status: planning
"""
        if target_launch_date:
            result += f"Target Launch: {target_launch_date}\n"
        if clinical_pm:
            result += f"Clinical PM: {clinical_pm}\n"

        result += f"""
Pre-UAT Gate Items: {len(gate_items)} items created

Next Steps:
  • get_uat_cycle("{cycle_id}") - View cycle details
  • assign_test_case(test_id, "{cycle_id}", ...) - Assign tests to cycle
  • bulk_assign_by_profile("{cycle_id}", profile_id, tester) - Bulk assign
"""
        return result

    except sqlite3.Error as e:
        logger.error(f"create_uat_cycle() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"create_uat_cycle() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_uat_cycle(cycle_id: str) -> str:
    """
    Get detailed information about a UAT cycle.

    Args:
        cycle_id: The cycle ID (e.g., "UAT-PROP-12345678")

    Returns:
        Cycle details including status, dates, and test summary

    Example:
        get_uat_cycle("UAT-PROP-12345678")
    """
    logger.info(f"get_uat_cycle() called - cycle_id={cycle_id}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get cycle with program info
        cursor = conn.execute("""
            SELECT c.*, p.name as program_name, p.prefix as program_prefix
            FROM uat_cycles c
            JOIN programs p ON c.program_id = p.program_id
            WHERE c.cycle_id = ?
        """, (cycle_id,))
        cycle = cursor.fetchone()

        if not cycle:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        cycle = dict(cycle)

        # Get test case summary
        cursor = conn.execute("""
            SELECT
                test_status,
                COUNT(*) as count
            FROM uat_test_cases
            WHERE uat_cycle_id = ?
            GROUP BY test_status
        """, (cycle_id,))
        test_summary = {row['test_status']: row['count'] for row in cursor.fetchall()}
        total_tests = sum(test_summary.values())

        # Get pre-UAT gate status
        cursor = conn.execute("""
            SELECT
                COUNT(*) as total,
                SUM(is_complete) as completed,
                SUM(CASE WHEN is_required = 1 AND is_complete = 0 THEN 1 ELSE 0 END) as required_pending
            FROM pre_uat_gate_items
            WHERE cycle_id = ?
        """, (cycle_id,))
        gate_status = dict(cursor.fetchone())

        # Get tester assignments
        cursor = conn.execute("""
            SELECT assigned_to, COUNT(*) as count
            FROM uat_test_cases
            WHERE uat_cycle_id = ? AND assigned_to IS NOT NULL
            GROUP BY assigned_to
        """, (cycle_id,))
        testers = {row['assigned_to']: row['count'] for row in cursor.fetchall()}

        # Build response
        result = f"""UAT Cycle Details
================

Cycle ID: {cycle['cycle_id']}
Program: {cycle['program_name']} [{cycle['program_prefix']}]
Name: {cycle['name']}
Type: {cycle['uat_type']}
Status: {cycle['status'].upper()}
"""
        if cycle['description']:
            result += f"Description: {cycle['description']}\n"
        if cycle['target_launch_date']:
            result += f"Target Launch: {cycle['target_launch_date']}\n"
        if cycle['clinical_pm']:
            result += f"Clinical PM: {cycle['clinical_pm']}"
            if cycle['clinical_pm_email']:
                result += f" ({cycle['clinical_pm_email']})"
            result += "\n"

        # Phase dates
        phase_dates = []
        if cycle['kickoff_date']:
            phase_dates.append(f"Kickoff: {cycle['kickoff_date']}")
        if cycle['testing_start']:
            phase_dates.append(f"Testing: {cycle['testing_start']} to {cycle['testing_end'] or 'TBD'}")
        if cycle['review_date']:
            phase_dates.append(f"Review: {cycle['review_date']}")
        if cycle['go_nogo_date']:
            phase_dates.append(f"Go/No-Go: {cycle['go_nogo_date']}")

        if phase_dates:
            result += "\nPhase Dates:\n"
            for pd in phase_dates:
                result += f"  • {pd}\n"

        # Pre-UAT Gate
        result += f"""
Pre-UAT Gate:
  Items: {gate_status['completed'] or 0}/{gate_status['total'] or 0} complete
  Required Pending: {gate_status['required_pending'] or 0}
  Gate Passed: {'Yes' if cycle['pre_uat_gate_passed'] else 'No'}
"""

        # Test summary
        result += f"""
Test Cases: {total_tests} total
"""
        if test_summary:
            for status, count in sorted(test_summary.items()):
                pct = round(100 * count / total_tests) if total_tests > 0 else 0
                result += f"  • {status}: {count} ({pct}%)\n"

        # Testers
        if testers:
            result += "\nAssigned Testers:\n"
            for tester, count in sorted(testers.items()):
                result += f"  • {tester}: {count} tests\n"

        # Go/No-Go decision
        if cycle['go_nogo_decision']:
            result += f"""
Go/No-Go Decision: {cycle['go_nogo_decision'].upper()}
  Signed by: {cycle['go_nogo_signed_by']}
  Date: {cycle['go_nogo_signed_date']}
"""
            if cycle['go_nogo_notes']:
                result += f"  Notes: {cycle['go_nogo_notes']}\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"get_uat_cycle() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"get_uat_cycle() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def list_uat_cycles(
    program_prefix: str = None,
    status: str = None
) -> str:
    """
    List UAT cycles, optionally filtered by program and/or status.

    Args:
        program_prefix: Filter by program prefix (e.g., "PROP")
        status: Filter by status (planning, validation, kickoff, testing, review, retesting, decision, complete, cancelled)

    Returns:
        List of UAT cycles

    Example:
        list_uat_cycles("PROP")
        list_uat_cycles(status="testing")
    """
    logger.info(f"list_uat_cycles() called - program={program_prefix}, status={status}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        query = """
            SELECT c.*, p.name as program_name, p.prefix as program_prefix,
                   (SELECT COUNT(*) FROM uat_test_cases WHERE uat_cycle_id = c.cycle_id) as test_count
            FROM uat_cycles c
            JOIN programs p ON c.program_id = p.program_id
            WHERE 1=1
        """
        params = []

        if program_prefix:
            query += " AND UPPER(p.prefix) = UPPER(?)"
            params.append(program_prefix)

        if status:
            query += " AND c.status = ?"
            params.append(status)

        query += " ORDER BY c.created_at DESC"

        cursor = conn.execute(query, params)
        cycles = cursor.fetchall()

        if not cycles:
            filters = []
            if program_prefix:
                filters.append(f"program={program_prefix}")
            if status:
                filters.append(f"status={status}")
            filter_str = f" (filters: {', '.join(filters)})" if filters else ""
            return f"No UAT cycles found{filter_str}."

        result = f"UAT Cycles ({len(cycles)} found)\n"
        result += "=" * 40 + "\n\n"

        for cycle in cycles:
            status_icon = {
                'planning': '📋',
                'validation': '🔍',
                'kickoff': '🚀',
                'testing': '🧪',
                'review': '📊',
                'retesting': '🔄',
                'decision': '⚖️',
                'complete': '✅',
                'cancelled': '❌'
            }.get(cycle['status'], '•')

            result += f"{status_icon} [{cycle['cycle_id']}]\n"
            result += f"   {cycle['name']} ({cycle['program_prefix']})\n"
            result += f"   Type: {cycle['uat_type']} | Status: {cycle['status']}\n"
            result += f"   Tests: {cycle['test_count']}"
            if cycle['target_launch_date']:
                result += f" | Launch: {cycle['target_launch_date']}"
            result += "\n\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"list_uat_cycles() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"list_uat_cycles() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def assign_test_case(
    test_id: str,
    cycle_id: str,
    assigned_to: str,
    assignment_type: str = "primary",
    profile_id: str = None,
    platform: str = None,
    persona: str = None,
    target_rule: str = None,
    patient_conditions: str = None
) -> str:
    """
    Assign a test case to a tester within a UAT cycle.

    Args:
        test_id: Test case ID to assign
        cycle_id: UAT cycle ID
        assigned_to: Tester name or email
        assignment_type: 'primary', 'secondary', or 'cross_check'
        profile_id: NCCN profile ID if applicable
        platform: Testing platform (e.g., "iOS", "Android", "Web")
        persona: Tester persona (e.g., "Patient", "Provider", "Admin")
        target_rule: Target NCCN rule being validated
        patient_conditions: Patient conditions for the test (JSON or comma-separated)

    Returns:
        Assignment confirmation

    Example:
        assign_test_case("PROP-AUTH-001-TC01", "UAT-PROP-12345678", "john.doe@example.com",
                        profile_id="NCCN-PROF-001", platform="iOS", persona="Patient")
    """
    logger.info(f"assign_test_case() called - test={test_id}, cycle={cycle_id}, tester={assigned_to}")

    # Validate assignment_type
    valid_types = ['primary', 'secondary', 'cross_check']
    if assignment_type not in valid_types:
        return f"Error: Invalid assignment_type '{assignment_type}'. Valid types: {', '.join(valid_types)}"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Verify cycle exists
        cursor = conn.execute("SELECT cycle_id, status FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        cycle = cursor.fetchone()
        if not cycle:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        # Verify test case exists
        cursor = conn.execute("SELECT test_id, title, uat_cycle_id FROM uat_test_cases WHERE test_id = ?", (test_id,))
        test = cursor.fetchone()
        if not test:
            return f"Error: Test case not found with ID '{test_id}'"

        # Update test case assignment
        conn.execute("""
            UPDATE uat_test_cases SET
                uat_cycle_id = ?,
                assigned_to = ?,
                assignment_type = ?,
                profile_id = COALESCE(?, profile_id),
                platform = COALESCE(?, platform),
                persona = COALESCE(?, persona),
                target_rule = COALESCE(?, target_rule),
                patient_conditions = COALESCE(?, patient_conditions),
                updated_date = CURRENT_TIMESTAMP
            WHERE test_id = ?
        """, (
            cycle_id, assigned_to, assignment_type,
            profile_id, platform, persona, target_rule, patient_conditions,
            test_id
        ))

        conn.commit()

        logger.info(f"assign_test_case() SUCCESS - {test_id} assigned to {assigned_to}")

        result = f"""✓ Test case assigned successfully!

Test: {test_id}
  "{test['title'][:60]}{'...' if len(test['title']) > 60 else ''}"

Assigned to: {assigned_to}
Assignment Type: {assignment_type}
Cycle: {cycle_id}
"""
        if profile_id:
            result += f"Profile: {profile_id}\n"
        if platform:
            result += f"Platform: {platform}\n"
        if persona:
            result += f"Persona: {persona}\n"
        if target_rule:
            result += f"Target Rule: {target_rule}\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"assign_test_case() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"assign_test_case() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_test_execution(
    test_id: str,
    status: str,
    tested_by: str,
    execution_notes: str = None,
    defect_id: str = None,
    defect_description: str = None,
    dev_status: str = None,
    dev_notes: str = None
) -> str:
    """
    Update test execution results.

    Args:
        test_id: Test case ID
        status: New status - 'Pass', 'Fail', 'Blocked', 'Skipped', 'Not Run'
        tested_by: Who executed the test
        execution_notes: Notes from test execution
        defect_id: Defect/bug ID if test failed
        defect_description: Description of the defect
        dev_status: Dev response status ('acknowledged', 'investigating', 'fixed', 'wont_fix')
        dev_notes: Developer notes

    Returns:
        Update confirmation

    Example:
        update_test_execution("PROP-AUTH-001-TC01", "Fail", "john.doe@example.com",
                             execution_notes="Login button unresponsive",
                             defect_id="BUG-123")
    """
    logger.info(f"update_test_execution() called - test={test_id}, status={status}")

    # Validate status
    valid_statuses = ['Pass', 'Fail', 'Blocked', 'Skipped', 'Not Run']
    if status not in valid_statuses:
        return f"Error: Invalid status '{status}'. Valid statuses: {', '.join(valid_statuses)}"

    # Validate dev_status if provided
    if dev_status:
        valid_dev = ['acknowledged', 'investigating', 'fixed', 'wont_fix']
        if dev_status not in valid_dev:
            return f"Error: Invalid dev_status '{dev_status}'. Valid: {', '.join(valid_dev)}"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Verify test case exists and get current state
        cursor = conn.execute(
            "SELECT test_id, title, test_status, uat_cycle_id FROM uat_test_cases WHERE test_id = ?",
            (test_id,)
        )
        test = cursor.fetchone()
        if not test:
            return f"Error: Test case not found with ID '{test_id}'"

        old_status = test['test_status']

        # Update test case
        conn.execute("""
            UPDATE uat_test_cases SET
                test_status = ?,
                tested_by = ?,
                tested_date = CURRENT_TIMESTAMP,
                execution_notes = COALESCE(?, execution_notes),
                defect_id = COALESCE(?, defect_id),
                defect_description = COALESCE(?, defect_description),
                dev_status = COALESCE(?, dev_status),
                dev_notes = COALESCE(?, dev_notes),
                updated_date = CURRENT_TIMESTAMP
            WHERE test_id = ?
        """, (
            status, tested_by, execution_notes,
            defect_id, defect_description, dev_status, dev_notes,
            test_id
        ))

        # Log to audit history
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            'uat_test_case', test_id, 'Test Executed', 'test_status',
            old_status, status, f'MCP:update_test_execution:{tested_by}',
            now, execution_notes or f'Status changed to {status}'
        ))

        conn.commit()

        logger.info(f"update_test_execution() SUCCESS - {test_id} now {status}")

        status_icon = {'Pass': '✅', 'Fail': '❌', 'Blocked': '🚧', 'Skipped': '⏭️', 'Not Run': '⚪'}.get(status, '•')

        result = f"""{status_icon} Test execution recorded!

Test: {test_id}
  "{test['title'][:60]}{'...' if len(test['title']) > 60 else ''}"

Status: {old_status} → {status}
Tested by: {tested_by}
Tested date: {now}
"""
        if execution_notes:
            result += f"\nNotes: {execution_notes}\n"
        if defect_id:
            result += f"\nDefect: {defect_id}"
            if defect_description:
                result += f"\n  {defect_description}"
            result += "\n"
        if dev_status:
            result += f"\nDev Status: {dev_status}\n"
            if dev_notes:
                result += f"Dev Notes: {dev_notes}\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"update_test_execution() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"update_test_execution() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def bulk_assign_by_profile(
    cycle_id: str,
    profile_id: str,
    assigned_to: str,
    assignment_type: str = "primary",
    platform: str = None
) -> str:
    """
    Bulk assign all test cases for a profile to a tester.

    Args:
        cycle_id: UAT cycle ID
        profile_id: Profile ID to match
        assigned_to: Tester name or email
        assignment_type: 'primary', 'secondary', or 'cross_check'
        platform: Testing platform

    Returns:
        Summary of assignments made

    Example:
        bulk_assign_by_profile("UAT-PROP-12345678", "NCCN-PROF-001", "jane.doe@example.com")
    """
    logger.info(f"bulk_assign_by_profile() called - cycle={cycle_id}, profile={profile_id}, tester={assigned_to}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Verify cycle exists
        cursor = conn.execute("SELECT cycle_id, name FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        cycle = cursor.fetchone()
        if not cycle:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        # Find test cases with matching profile_id
        cursor = conn.execute("""
            SELECT test_id, title FROM uat_test_cases
            WHERE profile_id = ? AND (uat_cycle_id IS NULL OR uat_cycle_id = ?)
        """, (profile_id, cycle_id))
        tests = cursor.fetchall()

        if not tests:
            return f"No test cases found with profile_id '{profile_id}' for cycle '{cycle_id}'"

        # Update all matching test cases
        conn.execute("""
            UPDATE uat_test_cases SET
                uat_cycle_id = ?,
                assigned_to = ?,
                assignment_type = ?,
                platform = COALESCE(?, platform),
                updated_date = CURRENT_TIMESTAMP
            WHERE profile_id = ? AND (uat_cycle_id IS NULL OR uat_cycle_id = ?)
        """, (cycle_id, assigned_to, assignment_type, platform, profile_id, cycle_id))

        conn.commit()

        logger.info(f"bulk_assign_by_profile() SUCCESS - {len(tests)} tests assigned to {assigned_to}")

        result = f"""✓ Bulk assignment complete!

Profile: {profile_id}
Cycle: {cycle['name']} ({cycle_id})
Assigned to: {assigned_to}
Assignment Type: {assignment_type}
Tests Assigned: {len(tests)}
"""
        if platform:
            result += f"Platform: {platform}\n"

        result += "\nTest Cases:\n"
        for test in tests[:10]:
            result += f"  • {test['test_id']}: {test['title'][:50]}...\n"
        if len(tests) > 10:
            result += f"  ... and {len(tests) - 10} more\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"bulk_assign_by_profile() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"bulk_assign_by_profile() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_uat_cycle_status(
    cycle_id: str,
    status: str,
    phase_date: str = None,
    notes: str = None
) -> str:
    """
    Update UAT cycle status and phase dates.

    Args:
        cycle_id: UAT cycle ID
        status: New status (planning, validation, kickoff, testing, review, retesting, decision, complete, cancelled)
        phase_date: Date for the phase transition (YYYY-MM-DD), defaults to today
        notes: Optional notes about the status change

    Returns:
        Status update confirmation

    Example:
        update_uat_cycle_status("UAT-PROP-12345678", "testing", "2025-01-15")
    """
    logger.info(f"update_uat_cycle_status() called - cycle={cycle_id}, status={status}")

    valid_statuses = ['planning', 'validation', 'kickoff', 'testing', 'review', 'retesting', 'decision', 'complete', 'cancelled']
    if status not in valid_statuses:
        return f"Error: Invalid status '{status}'. Valid statuses: {', '.join(valid_statuses)}"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get current cycle
        cursor = conn.execute("SELECT * FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        cycle = cursor.fetchone()
        if not cycle:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        old_status = cycle['status']
        phase_date = phase_date or date.today().isoformat()

        # Map status to phase date column
        date_column_map = {
            'validation': 'validation_start',
            'kickoff': 'kickoff_date',
            'testing': 'testing_start',
            'review': 'review_date',
            'retesting': 'retest_start',
            'decision': 'go_nogo_date',
        }

        # Update status and phase date
        update_sql = "UPDATE uat_cycles SET status = ?, updated_at = CURRENT_TIMESTAMP, updated_by = 'MCP:update_uat_cycle_status'"
        params = [status]

        if status in date_column_map:
            date_col = date_column_map[status]
            update_sql += f", {date_col} = ?"
            params.append(phase_date)

        # Handle end dates for previous phase
        if status == 'testing' and old_status == 'validation':
            update_sql += ", validation_end = ?"
            params.append(phase_date)
        elif status == 'review' and old_status == 'testing':
            update_sql += ", testing_end = ?"
            params.append(phase_date)
        elif status == 'decision' and old_status == 'retesting':
            update_sql += ", retest_end = ?"
            params.append(phase_date)

        update_sql += " WHERE cycle_id = ?"
        params.append(cycle_id)

        conn.execute(update_sql, params)

        # Log to audit
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            'uat_cycle', cycle_id, 'Status Changed', 'status',
            old_status, status, 'MCP:update_uat_cycle_status',
            now, notes or f'Status changed to {status}'
        ))

        conn.commit()

        logger.info(f"update_uat_cycle_status() SUCCESS - {cycle_id} now {status}")

        status_icon = {
            'planning': '📋', 'validation': '🔍', 'kickoff': '🚀',
            'testing': '🧪', 'review': '📊', 'retesting': '🔄',
            'decision': '⚖️', 'complete': '✅', 'cancelled': '❌'
        }.get(status, '•')

        return f"""{status_icon} UAT Cycle status updated!

Cycle: {cycle_id}
  {cycle['name']}

Status: {old_status} → {status}
Date: {phase_date}
"""

    except sqlite3.Error as e:
        logger.error(f"update_uat_cycle_status() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"update_uat_cycle_status() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_pre_uat_gate(
    cycle_id: str,
    item_id: int = None,
    category: str = None,
    is_complete: bool = None,
    completed_by: str = None,
    notes: str = None,
    sign_off: bool = False,
    signed_by: str = None
) -> str:
    """
    Update pre-UAT gate items or sign off on the gate.

    Args:
        cycle_id: UAT cycle ID
        item_id: Specific gate item ID to update
        category: Update all items in a category
        is_complete: Mark item(s) as complete/incomplete
        completed_by: Who completed the item
        notes: Notes about completion
        sign_off: If True, attempt to sign off on the entire gate
        signed_by: Who is signing off (required if sign_off=True)

    Returns:
        Gate update confirmation

    Example:
        # Complete a specific item
        update_pre_uat_gate("UAT-PROP-12345678", item_id=1, is_complete=True, completed_by="john@example.com")

        # Sign off on the gate
        update_pre_uat_gate("UAT-PROP-12345678", sign_off=True, signed_by="jane@example.com")
    """
    logger.info(f"update_pre_uat_gate() called - cycle={cycle_id}, item={item_id}, sign_off={sign_off}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Verify cycle exists
        cursor = conn.execute("SELECT * FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        cycle = cursor.fetchone()
        if not cycle:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        if sign_off:
            # Check if all required items are complete
            cursor = conn.execute("""
                SELECT COUNT(*) as pending FROM pre_uat_gate_items
                WHERE cycle_id = ? AND is_required = 1 AND is_complete = 0
            """, (cycle_id,))
            pending = cursor.fetchone()['pending']

            if pending > 0:
                return f"Error: Cannot sign off - {pending} required gate item(s) still pending"

            if not signed_by:
                return "Error: signed_by is required when sign_off=True"

            # Sign off on the gate
            conn.execute("""
                UPDATE uat_cycles SET
                    pre_uat_gate_passed = 1,
                    pre_uat_gate_signed_by = ?,
                    pre_uat_gate_signed_date = DATE('now'),
                    pre_uat_gate_notes = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE cycle_id = ?
            """, (signed_by, notes, cycle_id))

            conn.commit()

            return f"""✓ Pre-UAT Gate SIGNED OFF!

Cycle: {cycle_id}
  {cycle['name']}

Signed by: {signed_by}
Date: {date.today().isoformat()}

The UAT cycle is now cleared to proceed to testing.
"""

        elif item_id is not None:
            # Update specific item
            if is_complete is None:
                return "Error: is_complete is required when updating an item"

            completed_date = date.today().isoformat() if is_complete else None

            conn.execute("""
                UPDATE pre_uat_gate_items SET
                    is_complete = ?,
                    completed_by = ?,
                    completed_date = ?,
                    notes = COALESCE(?, notes)
                WHERE item_id = ? AND cycle_id = ?
            """, (1 if is_complete else 0, completed_by, completed_date, notes, item_id, cycle_id))

            conn.commit()

            return f"✓ Gate item {item_id} {'completed' if is_complete else 'marked incomplete'}"

        elif category:
            # Update all items in category
            if is_complete is None:
                return "Error: is_complete is required when updating a category"

            completed_date = date.today().isoformat() if is_complete else None

            cursor = conn.execute("""
                UPDATE pre_uat_gate_items SET
                    is_complete = ?,
                    completed_by = ?,
                    completed_date = ?
                WHERE cycle_id = ? AND category = ?
            """, (1 if is_complete else 0, completed_by, completed_date, cycle_id, category))

            return f"✓ {cursor.rowcount} gate items in '{category}' {'completed' if is_complete else 'marked incomplete'}"

        else:
            # Show gate status
            cursor = conn.execute("""
                SELECT * FROM pre_uat_gate_items
                WHERE cycle_id = ?
                ORDER BY category, sequence
            """, (cycle_id,))
            items = cursor.fetchall()

            result = f"""Pre-UAT Gate Status for {cycle_id}
{'=' * 40}

"""
            current_category = None
            for item in items:
                if item['category'] != current_category:
                    current_category = item['category']
                    result += f"\n{current_category.upper().replace('_', ' ')}:\n"

                icon = '✓' if item['is_complete'] else ('*' if item['is_required'] else '○')
                result += f"  [{icon}] {item['item_text']}"
                if item['is_complete'] and item['completed_by']:
                    result += f" (by {item['completed_by']})"
                result += "\n"

            result += f"\nGate Passed: {'Yes' if cycle['pre_uat_gate_passed'] else 'No'}\n"
            if cycle['pre_uat_gate_signed_by']:
                result += f"Signed by: {cycle['pre_uat_gate_signed_by']} on {cycle['pre_uat_gate_signed_date']}\n"

            return result

    except sqlite3.Error as e:
        logger.error(f"update_pre_uat_gate() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"update_pre_uat_gate() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_cycle_dashboard(cycle_id: str) -> str:
    """
    Get a comprehensive dashboard view of UAT cycle progress.

    Args:
        cycle_id: UAT cycle ID

    Returns:
        Dashboard with test progress, tester workload, and key metrics

    Example:
        get_cycle_dashboard("UAT-PROP-12345678")
    """
    logger.info(f"get_cycle_dashboard() called - cycle={cycle_id}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Use the v_uat_cycle_summary view
        cursor = conn.execute("""
            SELECT * FROM v_uat_cycle_summary WHERE cycle_id = ?
        """, (cycle_id,))
        summary = cursor.fetchone()

        if not summary:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        summary = dict(summary)

        # Get tester progress from view
        cursor = conn.execute("""
            SELECT * FROM v_uat_tester_progress WHERE cycle_id = ?
            ORDER BY completion_pct DESC
        """, (cycle_id,))
        tester_progress = cursor.fetchall()

        # Get tests needing retest
        cursor = conn.execute("""
            SELECT * FROM v_retest_queue WHERE cycle_id = ?
        """, (cycle_id,))
        retest_queue = cursor.fetchall()

        # Build dashboard
        result = f"""
╔══════════════════════════════════════════════════════════════╗
║  UAT CYCLE DASHBOARD                                          ║
╠══════════════════════════════════════════════════════════════╣
║  {summary['name'][:55]:<55} ║
║  {summary['program_prefix']} | {summary['uat_type']:<20} | Status: {summary['status']:<12} ║
╚══════════════════════════════════════════════════════════════╝
"""
        if summary['days_to_launch']:
            if summary['days_to_launch'] > 0:
                result += f"⏰ {summary['days_to_launch']} days until target launch\n"
            elif summary['days_to_launch'] == 0:
                result += "⚠️  TARGET LAUNCH DATE IS TODAY\n"
            else:
                result += f"🚨 {abs(summary['days_to_launch'])} days PAST target launch date\n"

        # Test Progress
        total = summary['total_tests'] or 0
        passed = summary['passed'] or 0
        failed = summary['failed'] or 0
        blocked = summary['blocked'] or 0
        not_run = summary['not_run'] or 0

        result += f"""
TEST PROGRESS
─────────────────────────────────────────
Total Tests: {total}

"""
        if total > 0:
            pass_pct = round(100 * passed / total)
            fail_pct = round(100 * failed / total)
            blocked_pct = round(100 * blocked / total)
            not_run_pct = round(100 * not_run / total)

            # Progress bar
            bar_width = 40
            pass_bar = int(bar_width * passed / total)
            fail_bar = int(bar_width * failed / total)
            blocked_bar = int(bar_width * blocked / total)
            not_run_bar = bar_width - pass_bar - fail_bar - blocked_bar

            result += f"[{'█' * pass_bar}{'░' * fail_bar}{'▒' * blocked_bar}{'·' * not_run_bar}]\n\n"

            result += f"  ✅ Passed:   {passed:>4} ({pass_pct}%)\n"
            result += f"  ❌ Failed:   {failed:>4} ({fail_pct}%)\n"
            result += f"  🚧 Blocked:  {blocked:>4} ({blocked_pct}%)\n"
            result += f"  ⚪ Not Run:  {not_run:>4} ({not_run_pct}%)\n"

        # Tester Progress
        if tester_progress:
            result += f"""
TESTER PROGRESS
─────────────────────────────────────────
"""
            for tp in tester_progress:
                tp = dict(tp)
                pct = tp['completion_pct'] or 0
                bar = '█' * int(pct / 5) + '·' * (20 - int(pct / 5))
                result += f"  {tp['assigned_to'][:20]:<20} [{bar}] {pct:>3}%\n"
                result += f"    {tp['passed'] or 0}✓ {tp['failed'] or 0}✗ {tp['not_run'] or 0}○\n"

        # Retest Queue
        if retest_queue:
            result += f"""
RETEST QUEUE ({len(retest_queue)} tests)
─────────────────────────────────────────
"""
            for rt in retest_queue[:5]:
                rt = dict(rt)
                result += f"  • {rt['test_id']}: {rt['title'][:40]}...\n"
                result += f"    Status: {rt['retest_status']} | Defect: {rt['defect_id'] or 'N/A'}\n"
            if len(retest_queue) > 5:
                result += f"  ... and {len(retest_queue) - 5} more\n"

        # Gate Status
        result += f"""
PRE-UAT GATE
─────────────────────────────────────────
  Gate Passed: {'✓ Yes' if summary.get('pre_uat_gate_passed') else '○ No'}
"""

        return result

    except sqlite3.Error as e:
        logger.error(f"get_cycle_dashboard() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"get_cycle_dashboard() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_tester_workload(
    cycle_id: str = None,
    tester: str = None
) -> str:
    """
    Get tester workload and progress across UAT cycles.

    Args:
        cycle_id: Filter by specific cycle
        tester: Filter by specific tester

    Returns:
        Tester workload summary

    Example:
        get_tester_workload("UAT-PROP-12345678")
        get_tester_workload(tester="john.doe@example.com")
    """
    logger.info(f"get_tester_workload() called - cycle={cycle_id}, tester={tester}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        query = "SELECT * FROM v_uat_tester_progress WHERE 1=1"
        params = []

        if cycle_id:
            query += " AND cycle_id = ?"
            params.append(cycle_id)
        if tester:
            query += " AND assigned_to = ?"
            params.append(tester)

        query += " ORDER BY cycle_name, completion_pct DESC"

        cursor = conn.execute(query, params)
        results = cursor.fetchall()

        if not results:
            filters = []
            if cycle_id:
                filters.append(f"cycle={cycle_id}")
            if tester:
                filters.append(f"tester={tester}")
            filter_str = f" (filters: {', '.join(filters)})" if filters else ""
            return f"No tester workload data found{filter_str}."

        result = "TESTER WORKLOAD REPORT\n"
        result += "=" * 60 + "\n\n"

        current_cycle = None
        for row in results:
            row = dict(row)

            if row['cycle_name'] != current_cycle:
                current_cycle = row['cycle_name']
                result += f"\n{current_cycle} ({row['cycle_id']})\n"
                result += "-" * 50 + "\n"

            pct = row['completion_pct'] or 0
            bar = '█' * int(pct / 5) + '·' * (20 - int(pct / 5))

            result += f"\n{row['assigned_to']}\n"
            result += f"  Progress: [{bar}] {pct}%\n"
            result += f"  Tests: {row['total_tests']} total\n"
            result += f"    ✅ Passed: {row['passed'] or 0}\n"
            result += f"    ❌ Failed: {row['failed'] or 0}\n"
            result += f"    🚧 Blocked: {row['blocked'] or 0}\n"
            result += f"    ⚪ Not Run: {row['not_run'] or 0}\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"get_tester_workload() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"get_tester_workload() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def record_go_nogo_decision(
    cycle_id: str,
    decision: str,
    signed_by: str,
    notes: str = None
) -> str:
    """
    Record the Go/No-Go decision for a UAT cycle.

    Args:
        cycle_id: UAT cycle ID
        decision: Decision - 'go', 'conditional_go', or 'no_go'
        signed_by: Who made the decision
        notes: Decision notes/conditions

    Returns:
        Decision confirmation

    Example:
        record_go_nogo_decision("UAT-PROP-12345678", "go", "jane.smith@example.com",
                               notes="All critical tests passed")
    """
    logger.info(f"record_go_nogo_decision() called - cycle={cycle_id}, decision={decision}")

    valid_decisions = ['go', 'conditional_go', 'no_go']
    if decision not in valid_decisions:
        return f"Error: Invalid decision '{decision}'. Valid: {', '.join(valid_decisions)}"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get cycle
        cursor = conn.execute("SELECT * FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        cycle = cursor.fetchone()
        if not cycle:
            return f"Error: UAT cycle not found with ID '{cycle_id}'"

        # Check if gate passed (recommended but not required)
        warning = ""
        if not cycle['pre_uat_gate_passed']:
            warning = "\n⚠️  Warning: Pre-UAT gate has not been signed off!\n"

        # Record decision
        conn.execute("""
            UPDATE uat_cycles SET
                go_nogo_decision = ?,
                go_nogo_signed_by = ?,
                go_nogo_signed_date = DATE('now'),
                go_nogo_notes = ?,
                status = CASE WHEN ? = 'go' THEN 'complete' ELSE status END,
                updated_at = CURRENT_TIMESTAMP
            WHERE cycle_id = ?
        """, (decision, signed_by, notes, decision, cycle_id))

        # Log to audit
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, changed_date, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            'uat_cycle', cycle_id, 'Go/No-Go Decision', 'go_nogo_decision',
            None, decision, f'MCP:record_go_nogo_decision:{signed_by}',
            now, notes or f'Decision: {decision}'
        ))

        conn.commit()

        decision_icon = {'go': '✅', 'conditional_go': '⚠️', 'no_go': '❌'}.get(decision, '•')
        decision_text = {'go': 'GO', 'conditional_go': 'CONDITIONAL GO', 'no_go': 'NO-GO'}.get(decision)

        result = f"""
╔══════════════════════════════════════════════════════════════╗
║  {decision_icon} GO/NO-GO DECISION RECORDED                              ║
╠══════════════════════════════════════════════════════════════╣
║  Decision: {decision_text:<48} ║
║  Cycle: {cycle_id:<50} ║
║  Signed by: {signed_by:<46} ║
║  Date: {date.today().isoformat():<51} ║
╚══════════════════════════════════════════════════════════════╝
{warning}"""
        if notes:
            result += f"\nNotes:\n  {notes}\n"

        if decision == 'go':
            result += "\n✅ UAT cycle marked as COMPLETE. Ready for production launch!\n"
        elif decision == 'conditional_go':
            result += "\n⚠️  Conditional approval - review notes for required conditions.\n"
        else:
            result += "\n❌ Launch blocked. Address issues and re-evaluate.\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"record_go_nogo_decision() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"record_go_nogo_decision() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


# ============================================================
# ONBOARDING PROJECT MANAGEMENT TOOLS
# ============================================================

# Standard milestones for new onboarding projects
# Each has: milestone_type, milestone_name, sequence_order, auto_verify_type
STANDARD_MILESTONES = [
    ("QUESTIONNAIRE", "Questionnaire Received", 1, None),
    ("KICKOFF", "Kickoff Meeting", 2, None),
    ("CONFIGURATION", "Configuration Complete", 3, "CONFIG_EXISTS"),
    ("EXTRACT_VALIDATED", "Extract Validated", 4, "EXTRACT_VALID"),
    ("USERS", "Users Provisioned", 5, "USERS_EXIST"),
    ("TRAINING", "Training Complete", 6, "TRAINING_COMPLETE"),
    ("UAT", "UAT Passed", 7, "TESTS_PASSING"),
    ("GO_LIVE", "Go-Live", 8, None),
]


@mcp.tool()
def create_onboarding_project(
    program_prefix: str,
    clinic_name: str,
    target_launch_date: str,
    propel_lead: str,
    client_contact_name: str = None,
    client_contact_email: str = None,
    notes: str = None
) -> str:
    """
    Create a new onboarding project with auto-generated milestones.

    Creates a project to track clinic onboarding from intake through launch.
    Automatically creates the 8 standard milestones with target dates based
    on the launch date.

    Args:
        program_prefix: Program prefix (e.g., "P4M", "Px4M")
        clinic_name: Name of the clinic being onboarded
        target_launch_date: Target launch date (YYYY-MM-DD)
        propel_lead: Propel project lead email
        client_contact_name: Primary client contact name
        client_contact_email: Primary client contact email
        notes: Project notes

    Returns:
        Project creation confirmation with project_id and milestone timeline

    Example:
        create_onboarding_project(
            program_prefix="P4M",
            clinic_name="Portland Cancer Institute",
            target_launch_date="2025-03-15",
            propel_lead="glen.lewis@propelhealth.com",
            client_contact_name="Dr. Smith",
            client_contact_email="smith@portland.clinic"
        )
    """
    import sqlite3
    import hashlib

    logger.info(f"create_onboarding_project() called - program={program_prefix}, clinic={clinic_name}")

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Validate program exists
        cursor = conn.execute(
            "SELECT program_id, name FROM programs WHERE prefix = ?",
            (program_prefix,)
        )
        program = cursor.fetchone()
        if not program:
            return f"Error: Program not found with prefix '{program_prefix}'"

        # Parse target date
        try:
            target_date = datetime.strptime(target_launch_date, "%Y-%m-%d").date()
        except ValueError:
            return f"Error: Invalid date format '{target_launch_date}'. Use YYYY-MM-DD."

        # Generate project_id: ONB-<CLINIC_CODE>-<YYYYMM>
        clinic_code = ''.join(c for c in clinic_name.upper() if c.isalpha())[:4]
        date_code = target_date.strftime("%Y%m")
        hash_suffix = hashlib.md5(f"{clinic_name}{datetime.now().isoformat()}".encode()).hexdigest()[:4].upper()
        project_id = f"ONB-{clinic_code}-{date_code}-{hash_suffix}"

        # Create project
        conn.execute("""
            INSERT INTO onboarding_projects (
                project_id, program_id, project_name, clinic_name,
                status, target_launch_date,
                client_contact_name, client_contact_email, propel_lead,
                notes, created_by
            ) VALUES (?, ?, ?, ?, 'INTAKE', ?, ?, ?, ?, ?, 'MCP:create_onboarding_project')
        """, (
            project_id, program['program_id'],
            f"{clinic_name} Onboarding", clinic_name,
            target_launch_date,
            client_contact_name, client_contact_email, propel_lead,
            notes
        ))

        # Calculate milestone target dates (work backwards from launch)
        # Rough schedule: 8 milestones over ~8-12 weeks
        days_until_launch = (target_date - date.today()).days
        if days_until_launch < 30:
            days_until_launch = 60  # Minimum 60 days for planning

        # Distribute milestones across timeline
        milestone_intervals = [0.05, 0.10, 0.30, 0.45, 0.60, 0.75, 0.90, 1.0]

        # Create standard milestones
        for i, (m_type, m_name, seq, auto_verify) in enumerate(STANDARD_MILESTONES):
            # Calculate target date for this milestone
            days_to_milestone = int(days_until_launch * milestone_intervals[i])
            m_target_date = date.today() + timedelta(days=days_to_milestone)

            conn.execute("""
                INSERT INTO onboarding_milestones (
                    project_id, milestone_type, milestone_name, sequence_order,
                    status, target_date, auto_verify_type
                ) VALUES (?, ?, ?, ?, 'NOT_STARTED', ?, ?)
            """, (project_id, m_type, m_name, seq, m_target_date.isoformat(), auto_verify))

        # Log to audit
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, new_value,
                changed_by, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            'onboarding_project', project_id, 'CREATE',
            f'{{"clinic": "{clinic_name}", "program": "{program_prefix}", "target": "{target_launch_date}"}}',
            f'MCP:create_onboarding_project:{propel_lead}',
            f'New onboarding project for {clinic_name}'
        ))

        conn.commit()

        # Build response
        result = f"""
╔══════════════════════════════════════════════════════════════╗
║  ✅ ONBOARDING PROJECT CREATED                               ║
╠══════════════════════════════════════════════════════════════╣
║  Project ID: {project_id:<45} ║
║  Clinic: {clinic_name:<49} ║
║  Program: {program['name']:<48} ║
║  Target Launch: {target_launch_date:<41} ║
║  Propel Lead: {propel_lead:<44} ║
╚══════════════════════════════════════════════════════════════╝

MILESTONE TIMELINE:
"""
        # Show milestone schedule
        cursor = conn.execute("""
            SELECT milestone_name, target_date, status
            FROM onboarding_milestones
            WHERE project_id = ?
            ORDER BY sequence_order
        """, (project_id,))

        for m in cursor.fetchall():
            m_dict = dict(m)
            result += f"  {m_dict['sequence_order'] if 'sequence_order' in m_dict else '•'}. {m_dict['milestone_name']:<30} Target: {m_dict['target_date']}\n"

        result += f"""
Next Steps:
  • get_onboarding_project("{project_id}") - View full project details
  • update_milestone("{project_id}", "QUESTIONNAIRE", "COMPLETE") - Mark questionnaire received
  • add_onboarding_dependency("{project_id}", "EPIC_EXTRACT", "Build ticket #12345") - Add dependency
"""
        return result

    except sqlite3.IntegrityError as e:
        return f"Error: Database integrity error - {str(e)}"
    except Exception as e:
        logger.error(f"create_onboarding_project() error: {e}", exc_info=True)
        return f"Error creating onboarding project: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_onboarding_project(project_id: str) -> str:
    """
    Get detailed status of an onboarding project.

    Shows project info, milestone progress, and pending dependencies.

    Args:
        project_id: Project ID (e.g., "ONB-PORT-202503-A1B2")

    Returns:
        Project details with milestone timeline and dependencies

    Example:
        get_onboarding_project("ONB-PORT-202503-A1B2")
    """
    import sqlite3

    logger.info(f"get_onboarding_project() called - project={project_id}")

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get project
        cursor = conn.execute("""
            SELECT op.*, p.name as program_name, p.prefix as program_prefix
            FROM onboarding_projects op
            JOIN programs p ON op.program_id = p.program_id
            WHERE op.project_id = ?
        """, (project_id,))
        project = cursor.fetchone()

        if not project:
            return f"Error: Onboarding project not found: '{project_id}'"

        project = dict(project)

        # Get milestones
        cursor = conn.execute("""
            SELECT * FROM onboarding_milestones
            WHERE project_id = ?
            ORDER BY sequence_order
        """, (project_id,))
        milestones = [dict(m) for m in cursor.fetchall()]

        # Get dependencies
        cursor = conn.execute("""
            SELECT * FROM onboarding_dependencies
            WHERE project_id = ? AND status != 'RESOLVED'
            ORDER BY due_date
        """, (project_id,))
        dependencies = [dict(d) for d in cursor.fetchall()]

        # Calculate progress
        total_milestones = len(milestones)
        completed = sum(1 for m in milestones if m['status'] == 'COMPLETE')
        progress_pct = int((completed / total_milestones) * 100) if total_milestones > 0 else 0

        # Status indicators
        status_icons = {
            'INTAKE': '📋',
            'IN_PROGRESS': '🔄',
            'UAT_READY': '🧪',
            'LAUNCHED': '🚀',
            'ON_HOLD': '⏸️'
        }

        milestone_icons = {
            'NOT_STARTED': '⬜',
            'IN_PROGRESS': '🔵',
            'COMPLETE': '✅',
            'BLOCKED': '🔴'
        }

        result = f"""
╔══════════════════════════════════════════════════════════════╗
║  {status_icons.get(project['status'], '•')} ONBOARDING PROJECT: {project['clinic_name']:<34} ║
╠══════════════════════════════════════════════════════════════╣
║  Project ID: {project_id:<45} ║
║  Program: {project['program_name']} [{project['program_prefix']}]{' ' * (40 - len(project['program_name']) - len(project['program_prefix']))} ║
║  Status: {project['status']:<49} ║
║  Target Launch: {project['target_launch_date'] or 'TBD':<41} ║
║  Progress: [{('█' * (progress_pct // 5)) + ('·' * (20 - progress_pct // 5))}] {progress_pct}%{' ' * (3 - len(str(progress_pct)))} ║
╚══════════════════════════════════════════════════════════════╝

CONTACTS:
  Propel Lead: {project['propel_lead'] or 'Not assigned'}
  Client Contact: {project['client_contact_name'] or 'Not specified'} ({project['client_contact_email'] or 'no email'})

MILESTONE PROGRESS:
"""
        for m in milestones:
            icon = milestone_icons.get(m['status'], '•')
            completion = f" - Completed {m['actual_completion_date']}" if m['status'] == 'COMPLETE' else f" - Target: {m['target_date']}"
            result += f"  {icon} {m['sequence_order']}. {m['milestone_name']:<28}{completion}\n"
            if m['status'] == 'BLOCKED' and m['blocker_reason']:
                result += f"      ⚠️  Blocked: {m['blocker_reason'][:50]}\n"

        if dependencies:
            result += f"\nPENDING DEPENDENCIES ({len(dependencies)}):\n"
            for d in dependencies:
                status_icon = '🔴' if d['status'] == 'PENDING' else '🟡'
                due = f" (Due: {d['due_date']})" if d['due_date'] else ""
                result += f"  {status_icon} {d['dependency_type']}: {d['description'][:40]}{due}\n"
                if d['owner']:
                    result += f"      Owner: {d['owner']}\n"

        if project['notes']:
            result += f"\nNOTES:\n  {project['notes'][:200]}\n"

        return result

    except Exception as e:
        logger.error(f"get_onboarding_project() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def list_onboarding_projects(
    program_prefix: str = None,
    status: str = None,
    propel_lead: str = None
) -> str:
    """
    List onboarding projects with optional filters.

    Args:
        program_prefix: Filter by program (e.g., "P4M")
        status: Filter by status (INTAKE, IN_PROGRESS, UAT_READY, LAUNCHED, ON_HOLD)
        propel_lead: Filter by Propel lead email

    Returns:
        List of matching projects with progress summary

    Example:
        list_onboarding_projects(program_prefix="P4M")
        list_onboarding_projects(status="IN_PROGRESS")
    """
    import sqlite3

    logger.info(f"list_onboarding_projects() called - program={program_prefix}, status={status}")

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Build query
        query = """
            SELECT op.*, p.name as program_name, p.prefix as program_prefix,
                   (SELECT COUNT(*) FROM onboarding_milestones om
                    WHERE om.project_id = op.project_id AND om.status = 'COMPLETE') as completed_milestones,
                   (SELECT COUNT(*) FROM onboarding_milestones om
                    WHERE om.project_id = op.project_id) as total_milestones,
                   (SELECT COUNT(*) FROM onboarding_dependencies od
                    WHERE od.project_id = op.project_id AND od.status != 'RESOLVED') as pending_dependencies
            FROM onboarding_projects op
            JOIN programs p ON op.program_id = p.program_id
            WHERE 1=1
        """
        params = []

        if program_prefix:
            query += " AND p.prefix = ?"
            params.append(program_prefix)
        if status:
            query += " AND op.status = ?"
            params.append(status)
        if propel_lead:
            query += " AND op.propel_lead LIKE ?"
            params.append(f"%{propel_lead}%")

        query += " ORDER BY op.target_launch_date, op.clinic_name"

        cursor = conn.execute(query, params)
        projects = [dict(p) for p in cursor.fetchall()]

        if not projects:
            filters = []
            if program_prefix:
                filters.append(f"program={program_prefix}")
            if status:
                filters.append(f"status={status}")
            if propel_lead:
                filters.append(f"lead={propel_lead}")
            filter_str = f" (filters: {', '.join(filters)})" if filters else ""
            return f"No onboarding projects found{filter_str}."

        status_icons = {
            'INTAKE': '📋',
            'IN_PROGRESS': '🔄',
            'UAT_READY': '🧪',
            'LAUNCHED': '🚀',
            'ON_HOLD': '⏸️'
        }

        result = "ONBOARDING PROJECTS\n"
        result += "=" * 70 + "\n\n"

        for p in projects:
            icon = status_icons.get(p['status'], '•')
            progress = int((p['completed_milestones'] / p['total_milestones']) * 100) if p['total_milestones'] > 0 else 0
            bar = '█' * (progress // 10) + '·' * (10 - progress // 10)

            result += f"{icon} {p['clinic_name']}\n"
            result += f"   ID: {p['project_id']}\n"
            result += f"   Program: {p['program_name']} [{p['program_prefix']}]\n"
            result += f"   Status: {p['status']} | Target: {p['target_launch_date'] or 'TBD'}\n"
            result += f"   Progress: [{bar}] {progress}% ({p['completed_milestones']}/{p['total_milestones']} milestones)\n"
            if p['pending_dependencies'] > 0:
                result += f"   ⚠️  {p['pending_dependencies']} pending dependencies\n"
            result += "\n"

        result += f"Total: {len(projects)} projects\n"

        return result

    except Exception as e:
        logger.error(f"list_onboarding_projects() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def update_milestone(
    project_id: str,
    milestone_type: str,
    status: str,
    completed_by: str = None,
    notes: str = None,
    blocker_reason: str = None
) -> str:
    """
    Update the status of an onboarding milestone.

    Args:
        project_id: Project ID
        milestone_type: Milestone type (QUESTIONNAIRE, KICKOFF, CONFIGURATION, EXTRACT_VALIDATED, USERS, TRAINING, UAT, GO_LIVE)
        status: New status (NOT_STARTED, IN_PROGRESS, COMPLETE, BLOCKED)
        completed_by: Who completed it (required if status=COMPLETE)
        notes: Additional notes
        blocker_reason: Reason for blocking (required if status=BLOCKED)

    Returns:
        Update confirmation

    Example:
        update_milestone("ONB-PORT-202503-A1B2", "QUESTIONNAIRE", "COMPLETE", completed_by="glen.lewis@propelhealth.com")
        update_milestone("ONB-PORT-202503-A1B2", "CONFIGURATION", "BLOCKED", blocker_reason="Waiting for EPIC build")
    """
    import sqlite3

    logger.info(f"update_milestone() called - project={project_id}, type={milestone_type}, status={status}")

    valid_types = [m[0] for m in STANDARD_MILESTONES]
    if milestone_type not in valid_types:
        return f"Error: Invalid milestone_type '{milestone_type}'. Valid: {', '.join(valid_types)}"

    valid_statuses = ['NOT_STARTED', 'IN_PROGRESS', 'COMPLETE', 'BLOCKED']
    if status not in valid_statuses:
        return f"Error: Invalid status '{status}'. Valid: {', '.join(valid_statuses)}"

    if status == 'COMPLETE' and not completed_by:
        return "Error: completed_by is required when status is COMPLETE"

    if status == 'BLOCKED' and not blocker_reason:
        return "Error: blocker_reason is required when status is BLOCKED"

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get milestone
        cursor = conn.execute("""
            SELECT om.*, op.clinic_name
            FROM onboarding_milestones om
            JOIN onboarding_projects op ON om.project_id = op.project_id
            WHERE om.project_id = ? AND om.milestone_type = ?
        """, (project_id, milestone_type))
        milestone = cursor.fetchone()

        if not milestone:
            return f"Error: Milestone '{milestone_type}' not found for project '{project_id}'"

        milestone = dict(milestone)
        old_status = milestone['status']

        # Update milestone
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        completion_date = date.today().isoformat() if status == 'COMPLETE' else None

        conn.execute("""
            UPDATE onboarding_milestones SET
                status = ?,
                actual_completion_date = COALESCE(?, actual_completion_date),
                completed_by = COALESCE(?, completed_by),
                notes = COALESCE(?, notes),
                blocker_reason = ?,
                updated_date = CURRENT_TIMESTAMP
            WHERE project_id = ? AND milestone_type = ?
        """, (
            status, completion_date, completed_by, notes,
            blocker_reason if status == 'BLOCKED' else None,
            project_id, milestone_type
        ))

        # Check if we should update project status
        if status == 'COMPLETE':
            # Get all milestones to check progress
            cursor = conn.execute("""
                SELECT milestone_type, status FROM onboarding_milestones
                WHERE project_id = ?
            """, (project_id,))
            all_milestones = {m['milestone_type']: m['status'] for m in cursor.fetchall()}

            # Auto-update project status based on milestone completion
            if all_milestones.get('GO_LIVE') == 'COMPLETE':
                conn.execute("""
                    UPDATE onboarding_projects SET
                        status = 'LAUNCHED',
                        actual_launch_date = DATE('now'),
                        updated_date = CURRENT_TIMESTAMP
                    WHERE project_id = ?
                """, (project_id,))
            elif all_milestones.get('UAT') == 'COMPLETE':
                conn.execute("""
                    UPDATE onboarding_projects SET status = 'UAT_READY', updated_date = CURRENT_TIMESTAMP
                    WHERE project_id = ? AND status != 'LAUNCHED'
                """, (project_id,))
            elif milestone_type == 'QUESTIONNAIRE':
                conn.execute("""
                    UPDATE onboarding_projects SET status = 'IN_PROGRESS', updated_date = CURRENT_TIMESTAMP
                    WHERE project_id = ? AND status = 'INTAKE'
                """, (project_id,))

        # Log to audit
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                old_value, new_value, changed_by, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            'onboarding_milestone', f"{project_id}:{milestone_type}",
            'UPDATE', 'status',
            old_status, status,
            f'MCP:update_milestone:{completed_by or "system"}',
            notes or f'Milestone {milestone_type} changed to {status}'
        ))

        conn.commit()

        status_icons = {
            'NOT_STARTED': '⬜',
            'IN_PROGRESS': '🔵',
            'COMPLETE': '✅',
            'BLOCKED': '🔴'
        }

        result = f"""
{status_icons.get(status, '•')} Milestone Updated!

Project: {milestone['clinic_name']} ({project_id})
Milestone: {milestone['milestone_name']}
Status: {old_status} → {status}
"""
        if status == 'COMPLETE':
            result += f"Completed by: {completed_by}\n"
            result += f"Completion date: {date.today().isoformat()}\n"
        if status == 'BLOCKED':
            result += f"Blocker: {blocker_reason}\n"
        if notes:
            result += f"Notes: {notes}\n"

        # Get updated project status
        cursor = conn.execute("SELECT status FROM onboarding_projects WHERE project_id = ?", (project_id,))
        proj = cursor.fetchone()
        if proj and proj['status'] != 'INTAKE':
            result += f"\nProject status: {proj['status']}\n"

        return result

    except Exception as e:
        logger.error(f"update_milestone() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def add_onboarding_dependency(
    project_id: str,
    dependency_type: str,
    description: str,
    milestone_type: str = None,
    external_reference: str = None,
    external_system: str = None,
    owner: str = None,
    owner_email: str = None,
    due_date: str = None
) -> str:
    """
    Add an external dependency to an onboarding project.

    Args:
        project_id: Project ID
        dependency_type: Type of dependency (EPIC_EXTRACT, AWS_APPROVAL, LEGAL_REVIEW, TRAINING_SCHEDULE, CLIENT_SIGN_OFF, OTHER)
        description: Description of the dependency
        milestone_type: Link to specific milestone (optional)
        external_reference: External ticket/case number
        external_system: External system name (EPIC, Salesforce, Jira, etc.)
        owner: Person responsible for resolving
        owner_email: Owner's email
        due_date: When dependency needs to be resolved (YYYY-MM-DD)

    Returns:
        Dependency creation confirmation

    Example:
        add_onboarding_dependency(
            "ONB-PORT-202503-A1B2",
            "EPIC_EXTRACT",
            "EPIC build ticket for patient extract",
            external_reference="EPIC-12345",
            external_system="EPIC",
            owner="John Doe",
            due_date="2025-02-15"
        )
    """
    import sqlite3

    logger.info(f"add_onboarding_dependency() called - project={project_id}, type={dependency_type}")

    valid_types = ['EPIC_EXTRACT', 'AWS_APPROVAL', 'LEGAL_REVIEW', 'TRAINING_SCHEDULE', 'CLIENT_SIGN_OFF', 'OTHER']
    if dependency_type not in valid_types:
        return f"Error: Invalid dependency_type '{dependency_type}'. Valid: {', '.join(valid_types)}"

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Verify project exists
        cursor = conn.execute(
            "SELECT clinic_name FROM onboarding_projects WHERE project_id = ?",
            (project_id,)
        )
        project = cursor.fetchone()
        if not project:
            return f"Error: Project not found: '{project_id}'"

        # Get milestone_id if specified
        milestone_id = None
        if milestone_type:
            cursor = conn.execute(
                "SELECT milestone_id FROM onboarding_milestones WHERE project_id = ? AND milestone_type = ?",
                (project_id, milestone_type)
            )
            milestone = cursor.fetchone()
            if milestone:
                milestone_id = milestone['milestone_id']

        # Parse due_date
        if due_date:
            try:
                datetime.strptime(due_date, "%Y-%m-%d")
            except ValueError:
                return f"Error: Invalid date format '{due_date}'. Use YYYY-MM-DD."

        # Create dependency
        cursor = conn.execute("""
            INSERT INTO onboarding_dependencies (
                project_id, milestone_id, dependency_type, description,
                external_reference, external_system,
                owner, owner_email, due_date,
                created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'MCP:add_onboarding_dependency')
            RETURNING dependency_id
        """, (
            project_id, milestone_id, dependency_type, description,
            external_reference, external_system,
            owner, owner_email, due_date
        ))
        dep_id = cursor.fetchone()['dependency_id']

        # Log to audit
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, new_value, changed_by, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            'onboarding_dependency', str(dep_id), 'CREATE',
            f'{{"type": "{dependency_type}", "desc": "{description[:50]}"}}',
            'MCP:add_onboarding_dependency',
            f'New {dependency_type} dependency added'
        ))

        conn.commit()

        result = f"""
🔗 Dependency Added!

Project: {project['clinic_name']} ({project_id})
Dependency ID: {dep_id}
Type: {dependency_type}
Description: {description}
"""
        if external_reference:
            result += f"External Ref: {external_reference}"
            if external_system:
                result += f" ({external_system})"
            result += "\n"
        if owner:
            result += f"Owner: {owner}"
            if owner_email:
                result += f" ({owner_email})"
            result += "\n"
        if due_date:
            result += f"Due Date: {due_date}\n"
        if milestone_type:
            result += f"Linked to: {milestone_type} milestone\n"

        result += f"\nNext: resolve_dependency({dep_id}, ...) when resolved\n"

        return result

    except Exception as e:
        logger.error(f"add_onboarding_dependency() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def resolve_dependency(
    dependency_id: int,
    resolved_by: str,
    resolution_notes: str = None
) -> str:
    """
    Mark a dependency as resolved.

    Args:
        dependency_id: Dependency ID
        resolved_by: Who resolved it
        resolution_notes: How it was resolved

    Returns:
        Resolution confirmation

    Example:
        resolve_dependency(123, "glen.lewis@propelhealth.com", "EPIC build complete, extract verified")
    """
    import sqlite3

    logger.info(f"resolve_dependency() called - id={dependency_id}, by={resolved_by}")

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get dependency
        cursor = conn.execute("""
            SELECT od.*, op.clinic_name
            FROM onboarding_dependencies od
            JOIN onboarding_projects op ON od.project_id = op.project_id
            WHERE od.dependency_id = ?
        """, (dependency_id,))
        dep = cursor.fetchone()

        if not dep:
            return f"Error: Dependency not found: {dependency_id}"

        dep = dict(dep)

        if dep['status'] == 'RESOLVED':
            return f"Dependency {dependency_id} is already resolved (on {dep['resolved_date']})"

        # Update dependency
        conn.execute("""
            UPDATE onboarding_dependencies SET
                status = 'RESOLVED',
                resolved_date = DATE('now'),
                resolved_by = ?,
                resolution_notes = ?,
                updated_date = CURRENT_TIMESTAMP
            WHERE dependency_id = ?
        """, (resolved_by, resolution_notes, dependency_id))

        # If linked to a milestone, check if we should unblock it
        if dep['milestone_id']:
            # Check if there are other pending dependencies for this milestone
            cursor = conn.execute("""
                SELECT COUNT(*) as pending FROM onboarding_dependencies
                WHERE milestone_id = ? AND status != 'RESOLVED' AND dependency_id != ?
            """, (dep['milestone_id'], dependency_id))
            pending = cursor.fetchone()['pending']

            if pending == 0:
                # Unblock the milestone if it was blocked
                conn.execute("""
                    UPDATE onboarding_milestones SET
                        status = CASE WHEN status = 'BLOCKED' THEN 'IN_PROGRESS' ELSE status END,
                        blocker_reason = NULL,
                        updated_date = CURRENT_TIMESTAMP
                    WHERE milestone_id = ?
                """, (dep['milestone_id'],))

        # Log to audit
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, old_value, new_value,
                changed_by, change_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            'onboarding_dependency', str(dependency_id), 'RESOLVE',
            dep['status'], 'RESOLVED',
            f'MCP:resolve_dependency:{resolved_by}',
            resolution_notes or 'Dependency resolved'
        ))

        conn.commit()

        result = f"""
✅ Dependency Resolved!

Project: {dep['clinic_name']} ({dep['project_id']})
Dependency: {dep['dependency_type']} - {dep['description'][:50]}
Resolved by: {resolved_by}
Resolved date: {date.today().isoformat()}
"""
        if resolution_notes:
            result += f"Notes: {resolution_notes}\n"

        return result

    except Exception as e:
        logger.error(f"resolve_dependency() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def get_go_live_readiness(project_id: str) -> str:
    """
    Check if a project is ready for go-live launch.

    Reviews all milestones and dependencies to generate a readiness report.
    Runs auto-verification checks where applicable.

    Args:
        project_id: Project ID

    Returns:
        Detailed readiness report with pass/fail status and blockers

    Example:
        get_go_live_readiness("ONB-PORT-202503-A1B2")
    """
    import sqlite3

    logger.info(f"get_go_live_readiness() called - project={project_id}")

    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

        # Get project
        cursor = conn.execute("""
            SELECT op.*, p.name as program_name, p.prefix, c.clinic_id
            FROM onboarding_projects op
            JOIN programs p ON op.program_id = p.program_id
            LEFT JOIN clinics c ON op.clinic_id = c.clinic_id
            WHERE op.project_id = ?
        """, (project_id,))
        project = cursor.fetchone()

        if not project:
            return f"Error: Project not found: '{project_id}'"

        project = dict(project)

        # Get milestones
        cursor = conn.execute("""
            SELECT * FROM onboarding_milestones
            WHERE project_id = ?
            ORDER BY sequence_order
        """, (project_id,))
        milestones = [dict(m) for m in cursor.fetchall()]

        # Get pending dependencies
        cursor = conn.execute("""
            SELECT * FROM onboarding_dependencies
            WHERE project_id = ? AND status != 'RESOLVED'
        """, (project_id,))
        pending_deps = [dict(d) for d in cursor.fetchall()]

        # Run auto-verification checks
        auto_checks = {}
        clinic_id = project.get('clinic_id')

        for m in milestones:
            if m['auto_verify_type'] and clinic_id:
                check_type = m['auto_verify_type']
                passed = False
                details = ""

                if check_type == 'CONFIG_EXISTS':
                    # Check if configs exist for this clinic
                    cursor = conn.execute("""
                        SELECT COUNT(*) as cnt FROM config_values
                        WHERE clinic_id = ?
                    """, (clinic_id,))
                    cnt = cursor.fetchone()['cnt']
                    passed = cnt > 0
                    details = f"{cnt} configurations found"

                elif check_type == 'USERS_EXIST':
                    # Check if users exist with access
                    cursor = conn.execute("""
                        SELECT COUNT(DISTINCT user_id) as cnt FROM user_access
                        WHERE clinic_id = ? AND status = 'Active'
                    """, (clinic_id,))
                    cnt = cursor.fetchone()['cnt']
                    passed = cnt > 0
                    details = f"{cnt} users with access"

                elif check_type == 'TRAINING_COMPLETE':
                    # Check if users have completed training
                    cursor = conn.execute("""
                        SELECT
                            COUNT(DISTINCT ut.user_id) as trained,
                            (SELECT COUNT(DISTINCT ua.user_id) FROM user_access ua
                             WHERE ua.clinic_id = ? AND ua.status = 'Active') as total
                        FROM user_training ut
                        JOIN user_access ua ON ut.user_id = ua.user_id
                        WHERE ua.clinic_id = ? AND ut.status = 'Completed'
                    """, (clinic_id, clinic_id))
                    result = cursor.fetchone()
                    trained, total = result['trained'], result['total']
                    passed = trained >= total and total > 0
                    details = f"{trained}/{total} users trained"

                elif check_type == 'TESTS_PASSING':
                    # Check if there's a UAT cycle with passing tests
                    cursor = conn.execute("""
                        SELECT
                            COUNT(CASE WHEN tc.test_status = 'Pass' THEN 1 END) as passed,
                            COUNT(*) as total
                        FROM uat_cycles uc
                        JOIN uat_test_cases tc ON uc.cycle_id = tc.cycle_id
                        WHERE uc.program_id = ? AND uc.status != 'complete'
                    """, (project['program_id'],))
                    result = cursor.fetchone()
                    passed_tests, total_tests = result['passed'] or 0, result['total'] or 0
                    passed = passed_tests == total_tests and total_tests > 0
                    details = f"{passed_tests}/{total_tests} tests passing"

                auto_checks[m['milestone_type']] = {
                    'passed': passed,
                    'details': details
                }

                # Update auto-verification in database
                conn.execute("""
                    UPDATE onboarding_milestones SET
                        auto_verified_date = CURRENT_TIMESTAMP,
                        auto_verified_result = ?
                    WHERE milestone_id = ?
                """, (passed, m['milestone_id']))

        conn.commit()

        # Build readiness report
        all_milestones_complete = all(m['status'] == 'COMPLETE' for m in milestones if m['milestone_type'] != 'GO_LIVE')
        no_pending_deps = len(pending_deps) == 0
        ready = all_milestones_complete and no_pending_deps

        result = f"""
╔══════════════════════════════════════════════════════════════╗
║  🚀 GO-LIVE READINESS REPORT                                 ║
╠══════════════════════════════════════════════════════════════╣
║  Project: {project['clinic_name']:<47} ║
║  Program: {project['program_name']:<48} ║
║  Target Launch: {project['target_launch_date'] or 'TBD':<41} ║
╚══════════════════════════════════════════════════════════════╝

OVERALL STATUS: {'✅ READY FOR LAUNCH' if ready else '❌ NOT READY'}

MILESTONE CHECKLIST:
"""
        blockers = []
        for m in milestones:
            if m['milestone_type'] == 'GO_LIVE':
                continue  # Skip GO_LIVE milestone itself

            is_complete = m['status'] == 'COMPLETE'
            icon = '✅' if is_complete else '❌'

            auto_info = ""
            if m['milestone_type'] in auto_checks:
                check = auto_checks[m['milestone_type']]
                auto_icon = '✓' if check['passed'] else '✗'
                auto_info = f" [Auto: {auto_icon} {check['details']}]"

            result += f"  {icon} {m['milestone_name']:<30}{auto_info}\n"

            if not is_complete:
                blockers.append(f"- {m['milestone_name']}: {m['status']}")
                if m['blocker_reason']:
                    blockers.append(f"  Blocker: {m['blocker_reason']}")

        result += f"\nDEPENDENCY CHECK:\n"
        if pending_deps:
            result += f"  ❌ {len(pending_deps)} pending dependencies:\n"
            for d in pending_deps:
                result += f"     • {d['dependency_type']}: {d['description'][:40]}\n"
                blockers.append(f"- Dependency: {d['dependency_type']} - {d['description'][:30]}")
        else:
            result += "  ✅ All dependencies resolved\n"

        if blockers:
            result += f"\n⚠️  BLOCKERS ({len(blockers)}):\n"
            for b in blockers:
                result += f"  {b}\n"

        if ready:
            result += f"""
✅ All prerequisites met!

Next step:
  update_milestone("{project_id}", "GO_LIVE", "COMPLETE", completed_by="your.email@propelhealth.com")
"""
        else:
            result += f"""
❌ Address blockers before launching.

Commands:
  • update_milestone("{project_id}", "<TYPE>", "COMPLETE", ...) - Mark milestone complete
  • resolve_dependency(<id>, ...) - Resolve dependency
"""

        return result

    except Exception as e:
        logger.error(f"get_go_live_readiness() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


# ============================================================
# UAT TOOLKIT INTEGRATION
# ============================================================
# These tools integrate with the uat_toolkit for generating
# HTML trackers and importing results from Formspree emails.

UAT_TOOLKIT_PATH = os.path.expanduser("~/projects/uat_toolkit")


@mcp.tool()
def setup_uat_cycle_with_testers(
    program_prefix: str,
    cycle_name: str,
    uat_type: str,
    target_launch_date: str,
    testers: str,
    overlap_percent: int = 20,
    formspree_id: str = None,
    clinical_pm: str = None,
    clinical_pm_email: str = None,
    description: str = None
) -> str:
    """
    Create a UAT cycle AND assign testers with configurable overlap in one step.

    This is the main setup tool for new UAT cycles. It:
    1. Creates the UAT cycle record
    2. Assigns tests to testers based on even split
    3. Creates overlap assignments for cross-checking

    Args:
        program_prefix: Program prefix (e.g., "ONB", "NCCN", "GRX")
        cycle_name: Display name (e.g., "ONB Questionnaire v1")
        uat_type: Type of UAT - 'feature', 'rule_validation', or 'regression'
        target_launch_date: Target date (YYYY-MM-DD)
        testers: JSON array of tester objects with name and email
                 Example: '[{"name": "Glen Lewis", "email": "glen@..."}, {"name": "Kim", "email": "kim@..."}]'
        overlap_percent: Percentage of tests both testers should execute (default 20%)
        formspree_id: Formspree form ID for email submissions
        clinical_pm: Clinical PM name
        clinical_pm_email: Clinical PM email
        description: Optional description

    Returns:
        Summary of created cycle and assignments

    Example:
        setup_uat_cycle_with_testers(
            program_prefix="ONB",
            cycle_name="ONB Questionnaire v1",
            uat_type="feature",
            target_launch_date="2025-05-02",
            testers='[{"name": "Glen Lewis", "email": "glen.lewis@propelhealth.com"}, {"name": "Kim Childers", "email": "kim.childers@providence.org"}]',
            overlap_percent=20,
            formspree_id="mqeekjjz"
        )
    """
    import math

    logger.info(f"setup_uat_cycle_with_testers() called - program={program_prefix}, name={cycle_name}")

    # Validate uat_type
    valid_types = ['feature', 'rule_validation', 'regression']
    if uat_type not in valid_types:
        return f"Error: Invalid uat_type '{uat_type}'. Valid types: {', '.join(valid_types)}"

    # Parse testers JSON
    try:
        testers_list = json.loads(testers)
    except json.JSONDecodeError as e:
        return f"Error: Invalid testers JSON: {str(e)}"

    if not testers_list:
        return "Error: At least one tester is required"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Resolve program
        cursor = conn.execute(
            "SELECT program_id, name, prefix FROM programs WHERE UPPER(prefix) = UPPER(?)",
            (program_prefix,)
        )
        program = cursor.fetchone()
        if not program:
            return f"Error: Program not found with prefix '{program_prefix}'"

        # Generate cycle_id
        cycle_id = f"UAT-{program['prefix']}-{str(uuid.uuid4())[:8].upper()}"

        # Insert cycle
        conn.execute("""
            INSERT INTO uat_cycles (
                cycle_id, program_id, name, description, uat_type,
                target_launch_date, clinical_pm, clinical_pm_email,
                formspree_id, status, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'planning', 'MCP:setup_uat_cycle_with_testers')
        """, (
            cycle_id, program['program_id'], cycle_name, description, uat_type,
            target_launch_date, clinical_pm, clinical_pm_email, formspree_id
        ))

        # Get all test cases for this program (matching test_id prefix)
        cursor = conn.execute("""
            SELECT test_id, workflow_section, workflow_order
            FROM uat_test_cases
            WHERE test_id LIKE ?
            AND uat_cycle_id IS NULL
            ORDER BY workflow_section, workflow_order, test_id
        """, (f"{program['prefix'].upper()}-%",))

        all_tests = cursor.fetchall()
        total_tests = len(all_tests)

        if total_tests == 0:
            # Try looking for tests without cycle assigned
            cursor = conn.execute("""
                SELECT test_id, workflow_section, workflow_order
                FROM uat_test_cases
                WHERE test_id LIKE ?
                ORDER BY workflow_section, workflow_order, test_id
            """, (f"{program['prefix'].upper()}-%",))
            all_tests = cursor.fetchall()
            total_tests = len(all_tests)

            if total_tests == 0:
                conn.rollback()
                return f"Error: No test cases found for program prefix '{program_prefix}'"

        # Calculate overlap
        overlap_count = max(1, math.ceil(total_tests * (overlap_percent / 100)))
        num_testers = len(testers_list)

        # Assign cycle to tests and distribute among testers
        if num_testers == 1:
            # Single tester gets all tests
            tester = testers_list[0]
            for test in all_tests:
                conn.execute("""
                    UPDATE uat_test_cases
                    SET uat_cycle_id = ?, assigned_to = ?, assignment_type = 'primary'
                    WHERE test_id = ?
                """, (cycle_id, tester['email'], test['test_id']))

        elif num_testers == 2:
            # Two testers: split with overlap
            tester1 = testers_list[0]
            tester2 = testers_list[1]

            # Overlap tests come from the middle (both get them)
            mid = total_tests // 2
            overlap_start = mid - overlap_count // 2
            overlap_end = overlap_start + overlap_count

            for i, test in enumerate(all_tests):
                test_id = test['test_id']

                # Link test to cycle
                conn.execute("""
                    UPDATE uat_test_cases SET uat_cycle_id = ? WHERE test_id = ?
                """, (cycle_id, test_id))

                # Assign primary tester
                if i < mid:
                    conn.execute("""
                        UPDATE uat_test_cases
                        SET assigned_to = ?, assignment_type = 'primary'
                        WHERE test_id = ?
                    """, (tester1['email'], test_id))
                else:
                    conn.execute("""
                        UPDATE uat_test_cases
                        SET assigned_to = ?, assignment_type = 'primary'
                        WHERE test_id = ?
                    """, (tester2['email'], test_id))

        else:
            # More than 2 testers: distribute evenly
            tests_per_tester = total_tests // num_testers

            for i, test in enumerate(all_tests):
                test_id = test['test_id']
                tester_idx = min(i // tests_per_tester, num_testers - 1)
                tester = testers_list[tester_idx]

                conn.execute("""
                    UPDATE uat_test_cases
                    SET uat_cycle_id = ?, assigned_to = ?, assignment_type = 'primary'
                    WHERE test_id = ?
                """, (cycle_id, tester['email'], test_id))

        # Log to audit
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                new_value, changed_by, change_reason
            ) VALUES (?, ?, 'CREATE', 'cycle', ?, 'system', ?)
        """, (
            'uat_cycles', cycle_id,
            json.dumps({"testers": num_testers, "tests": total_tests, "overlap": overlap_count}),
            f"Created UAT cycle with testers via MCP tool"
        ))

        conn.commit()

        logger.info(f"setup_uat_cycle_with_testers() SUCCESS - created {cycle_id}")

        # Build summary
        summary_lines = [
            f"✅ UAT Cycle Created: {cycle_id}",
            f"",
            f"📋 **{cycle_name}**",
            f"   Program: {program['name']} [{program['prefix']}]",
            f"   Type: {uat_type}",
            f"   Target: {target_launch_date or 'TBD'}",
            f"   Formspree: {formspree_id or 'Not configured'}",
            f"",
            f"📊 **Test Distribution:**",
            f"   Total tests: {total_tests}",
            f"   Overlap: {overlap_count} tests ({overlap_percent}%)",
            f""
        ]

        # Get tester counts
        for tester in testers_list:
            cursor = conn.execute("""
                SELECT COUNT(*) as cnt FROM uat_test_cases
                WHERE uat_cycle_id = ? AND assigned_to = ?
            """, (cycle_id, tester['email']))
            count = cursor.fetchone()['cnt']
            summary_lines.append(f"   {tester['name']}: {count} tests")

        summary_lines.extend([
            f"",
            f"🔜 **Next Steps:**",
            f"   1. Run: generate_uat_trackers(cycle_id='{cycle_id}')",
            f"   2. Share tracker URLs with testers"
        ])

        return "\n".join(summary_lines)

    except sqlite3.Error as e:
        logger.error(f"setup_uat_cycle_with_testers() database error: {e}")
        if conn:
            conn.rollback()
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"setup_uat_cycle_with_testers() error: {e}", exc_info=True)
        if conn:
            conn.rollback()
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def assign_uat_testers(
    cycle_id: str,
    testers: str,
    sections: str = None,
    overlap_percent: int = 20
) -> str:
    """
    Add or modify tester assignments for an existing UAT cycle.

    Args:
        cycle_id: UAT cycle ID (e.g., "UAT-ONB-12345678")
        testers: JSON array of tester objects with name and email
                 Example: '[{"name": "Sarah", "email": "sarah@..."}]'
        sections: Optional JSON array of section codes to assign
                  Example: '["P4M", "GRX"]' - only assign tests in these sections
        overlap_percent: Percentage of tests for cross-checking (default 20%)

    Returns:
        Assignment summary

    Example:
        assign_uat_testers(
            cycle_id="UAT-ONB-12345678",
            testers='[{"name": "Sarah Johnson", "email": "sarah.johnson@providence.org"}]',
            sections='["GRX"]'
        )
    """
    logger.info(f"assign_uat_testers() called - cycle={cycle_id}")

    # Parse testers JSON
    try:
        testers_list = json.loads(testers)
    except json.JSONDecodeError as e:
        return f"Error: Invalid testers JSON: {str(e)}"

    # Parse sections if provided
    section_filter = None
    if sections:
        try:
            section_filter = json.loads(sections)
        except json.JSONDecodeError as e:
            return f"Error: Invalid sections JSON: {str(e)}"

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Verify cycle exists
        cursor = conn.execute("SELECT * FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        cycle = cursor.fetchone()
        if not cycle:
            return f"Error: UAT cycle not found: {cycle_id}"

        # Get unassigned tests (or tests in specified sections)
        if section_filter:
            placeholders = ','.join('?' * len(section_filter))
            cursor = conn.execute(f"""
                SELECT test_id, workflow_section FROM uat_test_cases
                WHERE uat_cycle_id = ?
                AND workflow_section IN ({placeholders})
                ORDER BY workflow_section, workflow_order
            """, [cycle_id] + section_filter)
        else:
            cursor = conn.execute("""
                SELECT test_id, workflow_section FROM uat_test_cases
                WHERE uat_cycle_id = ? AND (assigned_to IS NULL OR assigned_to = '')
                ORDER BY workflow_section, workflow_order
            """, (cycle_id,))

        tests = cursor.fetchall()
        if not tests:
            return f"No unassigned tests found for cycle {cycle_id}" + (
                f" in sections {section_filter}" if section_filter else ""
            )

        # Distribute tests among new testers
        total_tests = len(tests)
        num_testers = len(testers_list)
        tests_per_tester = total_tests // num_testers

        assigned_counts = {t['email']: 0 for t in testers_list}

        for i, test in enumerate(tests):
            tester_idx = min(i // tests_per_tester, num_testers - 1)
            tester = testers_list[tester_idx]

            conn.execute("""
                UPDATE uat_test_cases
                SET assigned_to = ?, assignment_type = 'primary'
                WHERE test_id = ?
            """, (tester['email'], test['test_id']))
            assigned_counts[tester['email']] += 1

        conn.commit()

        # Build summary
        summary = [f"✅ Assigned {total_tests} tests to {num_testers} tester(s)", ""]
        for tester in testers_list:
            count = assigned_counts[tester['email']]
            summary.append(f"   {tester['name']}: {count} tests")

        return "\n".join(summary)

    except sqlite3.Error as e:
        logger.error(f"assign_uat_testers() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"assign_uat_testers() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def generate_uat_trackers(
    cycle_id: str,
    push_to_github: bool = True
) -> str:
    """
    Generate personalized UAT tracker HTML files for each tester and optionally push to GitHub.

    This runs the tracker generator script which:
    1. Queries test assignments from database
    2. Creates one HTML file per tester with their assigned tests
    3. Creates a tester selection index page
    4. Updates the main landing page
    5. Optionally commits and pushes to GitHub

    Args:
        cycle_id: UAT cycle ID (e.g., "UAT-ONB-12345678")
        push_to_github: Whether to commit and push changes (default True)

    Returns:
        List of generated files and tester URLs

    Example:
        generate_uat_trackers(cycle_id="UAT-ONB-12345678")
    """
    import subprocess

    logger.info(f"generate_uat_trackers() called - cycle={cycle_id}")

    script_path = os.path.join(UAT_TOOLKIT_PATH, "scripts", "generate_tester_trackers.py")

    if not os.path.exists(script_path):
        return f"Error: Generator script not found: {script_path}"

    # Get formspree_id from cycle
    conn = None
    formspree_id = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.execute("SELECT formspree_id FROM uat_cycles WHERE cycle_id = ?", (cycle_id,))
        row = cursor.fetchone()
        if row:
            formspree_id = row['formspree_id']
    finally:
        if conn:
            conn.close()

    # Run generator script
    cmd = ["python", script_path, cycle_id]
    if formspree_id:
        cmd.append(formspree_id)

    try:
        result = subprocess.run(
            cmd,
            cwd=UAT_TOOLKIT_PATH,
            capture_output=True,
            text=True,
            timeout=60
        )

        output = result.stdout

        if result.returncode != 0:
            logger.error(f"generate_uat_trackers() script failed: {result.stderr}")
            return f"Error: Generator failed:\n{result.stderr}"

        # Push to GitHub if requested
        if push_to_github:
            # Add docs folder
            subprocess.run(
                ["git", "add", "docs/"],
                cwd=UAT_TOOLKIT_PATH,
                capture_output=True
            )

            # Commit
            commit_result = subprocess.run(
                ["git", "commit", "-m", f"Generated trackers for {cycle_id}\n\nCo-Authored-By: Claude <noreply@anthropic.com>"],
                cwd=UAT_TOOLKIT_PATH,
                capture_output=True,
                text=True
            )

            if "nothing to commit" in commit_result.stdout + commit_result.stderr:
                output += "\n\n📝 No changes to commit (files unchanged)"
            else:
                # Push
                push_result = subprocess.run(
                    ["git", "push", "origin", "main"],
                    cwd=UAT_TOOLKIT_PATH,
                    capture_output=True,
                    text=True
                )

                if push_result.returncode == 0:
                    output += "\n\n✅ Pushed to GitHub. URLs will be live in ~1 minute."
                else:
                    output += f"\n\n⚠️ Push failed: {push_result.stderr}"

        logger.info(f"generate_uat_trackers() completed")
        return output

    except subprocess.TimeoutExpired:
        logger.error(f"generate_uat_trackers() timed out")
        return "Error: Generator timed out after 60 seconds"
    except Exception as e:
        logger.error(f"generate_uat_trackers() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def import_uat_results_json(
    json_data: str,
    partial: bool = True
) -> str:
    """
    Import UAT results from pasted JSON (from Formspree email).

    Paste the JSON payload from your Formspree email directly into this tool.
    It will update test results in the database.

    Args:
        json_data: The JSON string from Formspree email (paste the entire JSON)
        partial: If True (default), only update executed tests.
                 If False, treat as final submission and update all tests.

    Returns:
        Import summary with updated test counts

    Example:
        import_uat_results_json(json_data='{"tester": "Kim Childers", "synced_at": "2025-01-13T22:00:00Z", "results": [...]}')

    Usage:
        Just paste the JSON from your email:

        import_uat_results_json(json_data='''
        {
            "_subject": "UAT Progress: ONB v1 - Kim Childers (45%)",
            "tester": "Kim Childers",
            "synced_at": "2025-01-13T22:00:00.000Z",
            "progress": {...},
            "results": [...]
        }
        ''')
    """
    logger.info(f"import_uat_results_json() called")

    # Parse JSON
    try:
        data = json.loads(json_data)
    except json.JSONDecodeError as e:
        return f"Error: Invalid JSON: {str(e)}\n\nMake sure you pasted the complete JSON from the email."

    tester = data.get('tester', 'Unknown')
    tester_email = data.get('tester_email', '')
    sync_type = data.get('sync_type', 'manual')
    synced_at = data.get('synced_at', data.get('submitted_at', datetime.now().isoformat()))
    progress = data.get('progress', {})
    results = data.get('results', [])

    if not results:
        return "Error: No results found in JSON"

    # Auto-detect partial mode from sync_type
    if sync_type in ('auto_open', 'auto_10pm', 'manual'):
        partial = True

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        updated = 0
        skipped = 0
        not_found = []

        for result in results:
            test_id = result.get('test_id')
            status = result.get('status', result.get('test_status', 'Not Run'))
            notes = result.get('notes', '')
            tested_date = result.get('tested_date', synced_at)

            if not test_id:
                continue

            # Skip unexecuted tests in partial sync
            if partial and status in ('Not Run', 'Not_Run', ''):
                skipped += 1
                continue

            # Update test case
            cursor = conn.execute("""
                UPDATE uat_test_cases
                SET test_status = ?,
                    tested_by = ?,
                    tested_date = ?,
                    execution_notes = CASE
                        WHEN ? != '' AND (execution_notes IS NULL OR execution_notes = '')
                        THEN ?
                        WHEN ? != ''
                        THEN execution_notes || char(10) || '[' || ? || '] ' || ?
                        ELSE execution_notes
                    END,
                    updated_date = CURRENT_TIMESTAMP
                WHERE test_id = ?
            """, (
                status, tester, tested_date,
                notes, notes,  # For empty case
                notes, synced_at[:10], notes,  # For append case
                test_id
            ))

            if cursor.rowcount > 0:
                updated += 1
            else:
                not_found.append(test_id)

        # Log import to audit history
        conn.execute("""
            INSERT INTO audit_history (
                record_type, record_id, action, field_changed,
                new_value, changed_by, change_reason
            ) VALUES (?, ?, 'IMPORT', 'test_status', ?, ?, ?)
        """, (
            'uat_test_cases',
            f"batch-{len(results)}",
            json.dumps({"updated": updated, "skipped": skipped, "sync_type": sync_type}),
            tester,
            f"Imported from Formspree {sync_type} sync"
        ))

        conn.commit()

        logger.info(f"import_uat_results_json() SUCCESS - updated {updated} tests")

        # Build summary
        lines = [
            f"✅ **Import Complete**",
            f"",
            f"📥 **Source:** {sync_type} sync from {tester}",
            f"   Synced at: {synced_at[:19]}",
            f"",
            f"📊 **Progress reported:**"
        ]

        if progress:
            lines.append(f"   {progress.get('percent_complete', '?')}% complete ({progress.get('executed', '?')}/{progress.get('total', '?')})")
            lines.append(f"   ✅ {progress.get('passed', 0)} Pass | ❌ {progress.get('failed', 0)} Fail | 🚧 {progress.get('blocked', 0)} Blocked")

        lines.extend([
            f"",
            f"📝 **Import results:**",
            f"   Updated: {updated} tests",
            f"   Skipped: {skipped} (not executed yet)",
        ])

        if not_found:
            lines.append(f"   ⚠️ Not found: {len(not_found)} ({', '.join(not_found[:3])}{'...' if len(not_found) > 3 else ''})")

        lines.extend([
            f"",
            f"🔜 **Next:** Run `regenerate_uat_dashboard('{data.get('cycle_id', 'CYCLE_ID')}')` to update the dashboard"
        ])

        return "\n".join(lines)

    except sqlite3.Error as e:
        logger.error(f"import_uat_results_json() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"import_uat_results_json() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


@mcp.tool()
def regenerate_uat_dashboard(
    cycle_id: str,
    push_to_github: bool = True
) -> str:
    """
    Regenerate the UAT dashboard HTML with current data from database.

    Call this after importing results to update the dashboard.

    Args:
        cycle_id: UAT cycle ID (e.g., "UAT-ONB-12345678")
        push_to_github: Whether to commit and push changes (default True)

    Returns:
        Confirmation with dashboard URL

    Example:
        regenerate_uat_dashboard(cycle_id="UAT-ONB-12345678")
    """
    import subprocess

    logger.info(f"regenerate_uat_dashboard() called - cycle={cycle_id}")

    script_path = os.path.join(UAT_TOOLKIT_PATH, "scripts", "generate_dashboard.py")

    if not os.path.exists(script_path):
        return f"Error: Dashboard generator not found: {script_path}"

    try:
        result = subprocess.run(
            ["python", script_path, cycle_id],
            cwd=UAT_TOOLKIT_PATH,
            capture_output=True,
            text=True,
            timeout=30
        )

        if result.returncode != 0:
            logger.error(f"regenerate_uat_dashboard() script failed: {result.stderr}")
            return f"Error: Generator failed:\n{result.stderr}"

        output = result.stdout

        if push_to_github:
            # Add docs folder
            subprocess.run(
                ["git", "add", "docs/"],
                cwd=UAT_TOOLKIT_PATH,
                capture_output=True
            )

            # Commit
            commit_result = subprocess.run(
                ["git", "commit", "-m", f"Dashboard update for {cycle_id}\n\nCo-Authored-By: Claude <noreply@anthropic.com>"],
                cwd=UAT_TOOLKIT_PATH,
                capture_output=True,
                text=True
            )

            if "nothing to commit" in commit_result.stdout + commit_result.stderr:
                output += "\n\n📝 No changes to commit (dashboard unchanged)"
            else:
                # Push
                push_result = subprocess.run(
                    ["git", "push", "origin", "main"],
                    cwd=UAT_TOOLKIT_PATH,
                    capture_output=True,
                    text=True
                )

                if push_result.returncode == 0:
                    output += "\n\n✅ Pushed to GitHub"
                else:
                    output += f"\n\n⚠️ Push failed: {push_result.stderr}"

        # Add dashboard URL
        cycle_slug = cycle_id.lower().replace('_', '-')
        output += f"\n\n🔗 Dashboard: https://glewis05.github.io/uat_toolkit/dashboard-{cycle_slug}.html"

        logger.info(f"regenerate_uat_dashboard() completed")
        return output

    except subprocess.TimeoutExpired:
        logger.error(f"regenerate_uat_dashboard() timed out")
        return "Error: Generator timed out after 30 seconds"
    except Exception as e:
        logger.error(f"regenerate_uat_dashboard() error: {e}", exc_info=True)
        return f"Error: {str(e)}"


@mcp.tool()
def get_uat_progress(
    cycle_id: str = None,
    program_prefix: str = None
) -> str:
    """
    Get quick UAT progress summary for chat display.

    Simpler than get_cycle_dashboard - designed for quick status checks.

    Args:
        cycle_id: Specific cycle ID (e.g., "UAT-ONB-12345678")
        program_prefix: Program prefix to find most recent cycle (e.g., "ONB")

    Returns:
        Formatted progress summary

    Example:
        get_uat_progress(cycle_id="UAT-ONB-12345678")
        get_uat_progress(program_prefix="ONB")
    """
    logger.info(f"get_uat_progress() called - cycle={cycle_id}, program={program_prefix}")

    conn = None
    try:
        conn = sqlite3.connect(REQ_DB_PATH)
        conn.row_factory = sqlite3.Row

        # Find cycle
        if cycle_id:
            cursor = conn.execute("""
                SELECT c.*, p.prefix, p.name as program_name
                FROM uat_cycles c
                JOIN programs p ON c.program_id = p.program_id
                WHERE c.cycle_id = ?
            """, (cycle_id,))
        elif program_prefix:
            cursor = conn.execute("""
                SELECT c.*, p.prefix, p.name as program_name
                FROM uat_cycles c
                JOIN programs p ON c.program_id = p.program_id
                WHERE UPPER(p.prefix) = UPPER(?)
                ORDER BY c.created_date DESC LIMIT 1
            """, (program_prefix,))
        else:
            cursor = conn.execute("""
                SELECT c.*, p.prefix, p.name as program_name
                FROM uat_cycles c
                JOIN programs p ON c.program_id = p.program_id
                ORDER BY c.created_date DESC LIMIT 1
            """)

        cycle = cursor.fetchone()
        if not cycle:
            return "Error: No UAT cycle found"

        cycle = dict(cycle)
        cycle_id = cycle['cycle_id']

        # Get tester progress
        cursor = conn.execute("""
            SELECT
                assigned_to,
                COUNT(*) as total,
                SUM(CASE WHEN test_status NOT IN ('Not Run', '') THEN 1 ELSE 0 END) as executed,
                SUM(CASE WHEN test_status = 'Pass' THEN 1 ELSE 0 END) as passed,
                SUM(CASE WHEN test_status = 'Fail' THEN 1 ELSE 0 END) as failed,
                SUM(CASE WHEN test_status = 'Blocked' THEN 1 ELSE 0 END) as blocked,
                MAX(tested_date) as last_activity
            FROM uat_test_cases
            WHERE uat_cycle_id = ?
            AND assigned_to IS NOT NULL
            GROUP BY assigned_to
        """, (cycle_id,))

        testers = cursor.fetchall()

        # Calculate overall
        total_tests = sum(t['total'] for t in testers)
        total_executed = sum(t['executed'] or 0 for t in testers)
        total_passed = sum(t['passed'] or 0 for t in testers)
        total_failed = sum(t['failed'] or 0 for t in testers)
        total_blocked = sum(t['blocked'] or 0 for t in testers)
        overall_pct = round((total_executed / total_tests) * 100) if total_tests > 0 else 0

        # Build output
        lines = [
            f"📊 **{cycle['name']}** ({cycle_id})",
            f"   Target: {cycle['target_launch_date'] or 'TBD'}",
            f"",
            f"## Overall Progress: {overall_pct}%",
            f"   {total_executed}/{total_tests} tests executed",
            f"   ✅ {total_passed} Pass | ❌ {total_failed} Fail | 🚧 {total_blocked} Blocked",
            f"",
            f"## Tester Progress:"
        ]

        for t in testers:
            t = dict(t)
            email = t['assigned_to']
            name = email.split('@')[0].replace('.', ' ').title() if email else 'Unassigned'
            total = t['total']
            executed = t['executed'] or 0
            pct = round((executed / total) * 100) if total > 0 else 0

            bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
            lines.append(f"   **{name}**: {bar} {pct}% ({executed}/{total})")
            lines.append(f"      ✅{t['passed'] or 0} ❌{t['failed'] or 0} 🚧{t['blocked'] or 0}")

        return "\n".join(lines)

    except sqlite3.Error as e:
        logger.error(f"get_uat_progress() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"get_uat_progress() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


# ============================================================
# REQUIREMENTS DASHBOARD
# ============================================================
# Generate the requirements dashboard HTML from database state.
# Produces a standalone HTML file with all stories, test counts,
# filtering by program/priority, and pushes to GitHub Pages.
# ============================================================


def escape_html(text: str) -> str:
    """
    PURPOSE:
        Escape HTML special characters for safe embedding.

    R EQUIVALENT:
        htmltools::htmlEscape()
    """
    if not text:
        return ""
    return (text
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;"))


# =============================================================================
# DASHBOARD TEMPLATE VERSION
# =============================================================================
# Frozen template structure - do not change layout without updating version
DASHBOARD_TEMPLATE_VERSION = "1.0"
DASHBOARD_TEMPLATE_DATE = "2025-01-14"


def generate_full_html_template(
    total_stories: int,
    total_tests: int,
    program_cards_html: str,
    program_filters_html: str,
    categories_html: str,
    priority_counts: dict,
    roadmap_counts: dict,
    today: str
) -> str:
    """
    PURPOSE:
        Generate the complete HTML template with all CSS and JS.
        This creates a standalone requirements dashboard page.

    TEMPLATE VERSION:
        v1.0 (2025-01-14) - Frozen layout approved by Glen Lewis
        See: templates/dashboard_template_v1.0.html for baseline reference

    PARAMETERS:
        total_stories (int): Total number of user stories
        total_tests (int): Total number of test cases
        program_cards_html (str): HTML for program summary cards
        program_filters_html (str): HTML for program filter buttons
        categories_html (str): HTML for all category sections with stories
        priority_counts (dict): Count of stories by priority
        roadmap_counts (dict): Count of stories by roadmap target quarter
        today (str): Formatted date string for "Last Updated"

    RETURNS:
        str: Complete HTML document as a string

    WHY THIS APPROACH:
        Single-file HTML with embedded CSS/JS for easy GitHub Pages hosting.
        No build step required - just commit and push.
    """

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Providence GenomicsNow | Requirements Dashboard</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
    <style>
        :root {{
            --color-bg: #f8fafb;
            --color-surface: #ffffff;
            --color-border: #e2e8f0;
            --color-border-light: #f1f5f9;
            --color-text: #1e293b;
            --color-text-secondary: #64748b;
            --color-text-muted: #94a3b8;
            --color-primary: #0f766e;
            --color-primary-light: #ccfbf1;
            --color-primary-dark: #0d9488;
            --color-accent: #0ea5e9;
            --color-accent-light: #e0f2fe;
            --color-must-have: #dc2626;
            --color-must-have-bg: #fee2e2;
            --color-should-have: #059669;
            --color-should-have-bg: #d1fae5;
            --color-could-have: #d97706;
            --color-could-have-bg: #fef3c7;
            --color-draft: #6366f1;
            --color-draft-bg: #e0e7ff;
            --color-p4m: #0f766e;
            --color-p4m-light: #ccfbf1;
            --color-plat: #7c3aed;
            --color-plat-light: #ede9fe;
            --shadow-sm: 0 1px 2px rgba(0,0,0,0.04);
            --shadow-md: 0 4px 6px -1px rgba(0,0,0,0.05), 0 2px 4px -2px rgba(0,0,0,0.05);
            --shadow-lg: 0 10px 15px -3px rgba(0,0,0,0.05), 0 4px 6px -4px rgba(0,0,0,0.05);
            --radius-sm: 6px;
            --radius-md: 10px;
            --radius-lg: 16px;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif;
            background: var(--color-bg);
            color: var(--color-text);
            line-height: 1.6;
            font-size: 15px;
        }}
        .header {{
            background: linear-gradient(135deg, #134e4a 0%, #1e3a5f 100%);
            color: white;
            padding: 2.5rem 2rem;
            position: relative;
            overflow: hidden;
        }}
        .header::before {{
            content: '';
            position: absolute;
            top: -50%;
            right: -10%;
            width: 500px;
            height: 500px;
            background: radial-gradient(circle, rgba(255,255,255,0.06) 0%, transparent 70%);
            border-radius: 50%;
        }}
        .header-content {{
            max-width: 1400px;
            margin: 0 auto;
            position: relative;
            z-index: 1;
        }}
        .header-badge {{
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
            background: rgba(255,255,255,0.12);
            padding: 0.4rem 0.9rem;
            border-radius: 100px;
            font-size: 0.8rem;
            font-weight: 500;
            margin-bottom: 0.75rem;
        }}
        .header h1 {{
            font-size: 2rem;
            font-weight: 700;
            margin-bottom: 0.4rem;
        }}
        .header p {{
            opacity: 0.85;
            font-size: 1rem;
        }}
        .header-meta {{
            display: flex;
            gap: 2rem;
            margin-top: 1.25rem;
            flex-wrap: wrap;
        }}
        .header-meta-item {{
            display: flex;
            flex-direction: column;
        }}
        .header-meta-label {{
            font-size: 0.7rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            opacity: 0.6;
        }}
        .header-meta-value {{
            font-weight: 600;
            font-size: 0.95rem;
        }}
        .main {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 2rem;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
            gap: 1rem;
            margin-bottom: 2rem;
        }}
        .summary-card {{
            background: var(--color-surface);
            border: 1px solid var(--color-border);
            border-radius: var(--radius-md);
            padding: 1rem;
            box-shadow: var(--shadow-sm);
        }}
        .summary-card-label {{
            font-size: 0.75rem;
            color: var(--color-text-secondary);
        }}
        .summary-card-value {{
            font-size: 1.5rem;
            font-weight: 700;
            color: var(--color-primary);
        }}
        .summary-card-sub {{
            font-size: 0.7rem;
            color: var(--color-text-muted);
        }}
        .filters {{
            display: flex;
            gap: 0.75rem;
            margin-bottom: 1.5rem;
            flex-wrap: wrap;
            align-items: center;
        }}
        .filter-group {{
            display: flex;
            gap: 0.5rem;
        }}
        .filter-btn {{
            padding: 0.5rem 1rem;
            border: 1px solid var(--color-border);
            background: var(--color-surface);
            border-radius: 100px;
            font-family: inherit;
            font-size: 0.85rem;
            font-weight: 500;
            color: var(--color-text-secondary);
            cursor: pointer;
            transition: all 0.15s ease;
        }}
        .filter-btn:hover {{
            border-color: var(--color-primary);
            color: var(--color-primary);
        }}
        .filter-btn.active {{
            background: var(--color-primary);
            border-color: var(--color-primary);
            color: white;
        }}
        .filter-btn.p4m.active {{ background: var(--color-p4m); border-color: var(--color-p4m); }}
        .filter-btn.plat.active {{ background: var(--color-plat); border-color: var(--color-plat); }}
        .search-box {{
            flex: 1;
            min-width: 200px;
            max-width: 300px;
            position: relative;
        }}
        .search-box input {{
            width: 100%;
            padding: 0.5rem 1rem 0.5rem 2.5rem;
            border: 1px solid var(--color-border);
            border-radius: 100px;
            font-family: inherit;
            font-size: 0.85rem;
        }}
        .search-box input:focus {{
            outline: none;
            border-color: var(--color-primary);
        }}
        .search-box svg {{
            position: absolute;
            left: 0.85rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--color-text-muted);
        }}
        .category-section {{ margin-bottom: 2rem; }}
        .category-header {{
            display: flex;
            align-items: center;
            gap: 0.75rem;
            margin-bottom: 1rem;
            padding-bottom: 0.5rem;
            border-bottom: 2px solid var(--color-border);
        }}
        .category-icon {{
            width: 32px;
            height: 32px;
            border-radius: var(--radius-sm);
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 0.9rem;
        }}
        .category-icon.analytics {{ background: #dbeafe; }}
        .category-icon.audit {{ background: #fce7f3; }}
        .category-icon.config {{ background: #fef3c7; }}
        .category-icon.dash {{ background: #d1fae5; }}
        .category-icon.referral {{ background: #ccfbf1; }}
        .category-icon.role {{ background: #fae8ff; }}
        .category-title {{ font-size: 1.1rem; font-weight: 600; }}
        .category-count {{ font-size: 0.8rem; color: var(--color-text-muted); margin-left: auto; }}
        .stories-grid {{ display: grid; gap: 1rem; }}
        .story-card {{
            background: var(--color-surface);
            border: 1px solid var(--color-border);
            border-radius: var(--radius-lg);
            box-shadow: var(--shadow-sm);
            overflow: hidden;
        }}
        .story-card:hover {{ box-shadow: var(--shadow-md); }}
        .story-card.hidden {{ display: none; }}
        .story-header {{
            padding: 1.25rem;
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            gap: 1rem;
            cursor: pointer;
        }}
        .story-header:hover {{ background: var(--color-bg); }}
        .story-id {{
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.75rem;
            font-weight: 500;
            padding: 0.2rem 0.5rem;
            border-radius: var(--radius-sm);
            margin-bottom: 0.4rem;
            display: inline-block;
        }}
        .story-id.p4m {{ background: var(--color-p4m-light); color: var(--color-p4m); }}
        .story-id.plat {{ background: var(--color-plat-light); color: var(--color-plat); }}
        .story-title {{ font-size: 1.05rem; font-weight: 600; margin-bottom: 0.25rem; }}
        .story-meta {{ display: flex; gap: 0.5rem; flex-wrap: wrap; }}
        .badge {{
            font-size: 0.65rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.03em;
            padding: 0.25rem 0.5rem;
            border-radius: 100px;
        }}
        .badge-must-have {{ background: var(--color-must-have-bg); color: var(--color-must-have); }}
        .badge-should-have {{ background: var(--color-should-have-bg); color: var(--color-should-have); }}
        .badge-could-have {{ background: var(--color-could-have-bg); color: var(--color-could-have); }}
        .badge-status {{ background: var(--color-draft-bg); color: var(--color-draft); }}
        /* Roadmap badges - blue/teal color scheme to differentiate from priority (red) and status (gray) */
        .badge-roadmap {{ background: #e0f2fe; color: #0369a1; font-weight: 600; }}
        /* 2025 quarters - lighter teal tones */
        .badge-q1-2025 {{ background: #ccfbf1; color: #0d9488; }}
        .badge-q2-2025 {{ background: #cffafe; color: #0891b2; }}
        .badge-q3-2025 {{ background: #a5f3fc; color: #0e7490; }}
        .badge-q4-2025 {{ background: #99f6e4; color: #0f766e; }}
        /* 2026 quarters - blue tones */
        .badge-q1-2026 {{ background: #dbeafe; color: #1e40af; }}
        .badge-q2-2026 {{ background: #c7d2fe; color: #4338ca; }}
        .badge-q3-2026 {{ background: #e0e7ff; color: #4f46e5; }}
        .badge-q4-2026 {{ background: #ede9fe; color: #6d28d9; }}
        /* Future and other */
        .badge-2027-plus {{ background: #fae8ff; color: #a21caf; }}
        .badge-backlog {{ background: #f1f5f9; color: #64748b; }}
        .badge-unscheduled {{ background: #f1f5f9; color: #94a3b8; }}
        .story-toggle {{ color: var(--color-text-muted); transition: transform 0.2s; }}
        .story-card.open .story-toggle {{ transform: rotate(180deg); }}
        .story-body {{
            display: none;
            padding: 0 1.25rem 1.25rem;
            border-top: 1px solid var(--color-border-light);
        }}
        .story-card.open .story-body {{ display: block; }}
        .story-section {{ margin-top: 1rem; }}
        .story-section-title {{
            font-size: 0.7rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--color-text-muted);
            margin-bottom: 0.5rem;
            display: flex;
            align-items: center;
            gap: 0.4rem;
        }}
        .story-section-title::before {{
            content: '';
            width: 3px;
            height: 12px;
            background: var(--color-primary);
            border-radius: 2px;
        }}
        .user-story {{
            font-size: 0.95rem;
            background: linear-gradient(135deg, var(--color-primary-light) 0%, #f0fdfa 100%);
            padding: 0.9rem 1rem;
            border-radius: var(--radius-md);
            border-left: 3px solid var(--color-primary);
            font-style: italic;
        }}
        .acceptance-criteria {{ list-style: none; counter-reset: criteria; }}
        .acceptance-criteria li {{
            counter-increment: criteria;
            display: flex;
            gap: 0.6rem;
            padding: 0.5rem 0;
            border-bottom: 1px solid var(--color-border-light);
            font-size: 0.9rem;
        }}
        .acceptance-criteria li:last-child {{ border-bottom: none; }}
        .acceptance-criteria li::before {{
            content: counter(criteria);
            min-width: 20px;
            height: 20px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: var(--color-accent-light);
            color: var(--color-accent);
            border-radius: 50%;
            font-size: 0.7rem;
            font-weight: 600;
        }}
        .test-count {{
            font-size: 0.8rem;
            color: var(--color-text-secondary);
            background: var(--color-bg);
            padding: 0.5rem 0.75rem;
            border-radius: var(--radius-sm);
            margin-top: 1rem;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}
        /* Status Timeline Styles */
        .status-timeline {{
            display: flex;
            align-items: flex-start;
            gap: 0;
            padding: 1rem 0.5rem;
            background: var(--color-bg);
            border-radius: var(--radius-md);
            overflow-x: auto;
        }}
        .timeline-stage {{
            display: flex;
            flex-direction: column;
            align-items: center;
            min-width: 70px;
            text-align: center;
        }}
        .timeline-dot {{
            width: 12px;
            height: 12px;
            border-radius: 50%;
            background: var(--color-border);
            border: 2px solid var(--color-border);
            margin-bottom: 0.4rem;
        }}
        .timeline-stage.completed .timeline-dot {{
            background: var(--color-primary);
            border-color: var(--color-primary);
        }}
        .timeline-stage.current .timeline-dot {{
            background: #fff;
            border-color: var(--color-primary);
            box-shadow: 0 0 0 3px var(--color-primary-light);
        }}
        .timeline-label {{
            font-size: 0.65rem;
            font-weight: 600;
            color: var(--color-text-muted);
            text-transform: uppercase;
            letter-spacing: 0.02em;
        }}
        .timeline-stage.completed .timeline-label,
        .timeline-stage.current .timeline-label {{
            color: var(--color-text);
        }}
        .timeline-date {{
            font-size: 0.7rem;
            color: var(--color-text-muted);
            font-family: 'JetBrains Mono', monospace;
        }}
        .timeline-stage.completed .timeline-date,
        .timeline-stage.current .timeline-date {{
            color: var(--color-primary);
            font-weight: 500;
        }}
        .timeline-connector {{
            flex: 1;
            height: 2px;
            background: var(--color-border);
            margin-top: 5px;
            min-width: 20px;
        }}
        /* Section Title */
        .section-title {{
            font-size: 13px;
            font-weight: 600;
            color: var(--color-text-secondary);
            margin-bottom: 0.5rem;
        }}
        /* Roadmap Summary Section */
        .roadmap-summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(100px, 1fr));
            gap: 0.75rem;
            margin-top: 1rem;
            margin-bottom: 1.75rem;
            padding: 1rem;
            background: var(--color-surface);
            border: 1px solid var(--color-border);
            border-radius: var(--radius-md);
        }}
        .roadmap-item {{
            text-align: center;
            padding: 0.5rem;
        }}
        .roadmap-item-label {{
            font-size: 0.7rem;
            font-weight: 600;
            color: var(--color-text-secondary);
        }}
        .roadmap-item-value {{
            font-size: 1.25rem;
            font-weight: 700;
            color: var(--color-text);
        }}
        .footer {{
            text-align: center;
            padding: 2rem;
            color: var(--color-text-muted);
            font-size: 0.85rem;
            border-top: 1px solid var(--color-border);
            margin-top: 2rem;
        }}
        .footer a {{ color: var(--color-primary); text-decoration: none; }}
        .attribution {{
            position: fixed;
            bottom: 12px;
            right: 16px;
            font-size: 10px;
            color: #999;
        }}
        @media (max-width: 768px) {{
            .header {{ padding: 1.5rem 1rem; }}
            .header h1 {{ font-size: 1.5rem; }}
            .main {{ padding: 1rem; }}
            .summary-grid {{ grid-template-columns: repeat(2, 1fr); }}
        }}
        @media print {{
            .story-body {{ display: block !important; }}
            .filters {{ display: none; }}
        }}
    </style>
</head>
<body>
    <header class="header">
        <div class="header-content">
            <div class="header-badge">
                <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                    <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path>
                    <polyline points="14 2 14 8 20 8"></polyline>
                </svg>
                Requirements Dashboard
            </div>
            <h1>Providence GenomicsNow Requirements</h1>
            <p>User stories and acceptance criteria for all programs.</p>
            <div class="header-meta">
                <div class="header-meta-item">
                    <span class="header-meta-label">Total Stories</span>
                    <span class="header-meta-value">{total_stories}</span>
                </div>
                <div class="header-meta-item">
                    <span class="header-meta-label">Test Cases</span>
                    <span class="header-meta-value">{total_tests}</span>
                </div>
                <div class="header-meta-item">
                    <span class="header-meta-label">Last Updated</span>
                    <span class="header-meta-value">{today}</span>
                </div>
            </div>
        </div>
    </header>

    <main class="main">
        <div class="section-title">Programs</div>
        <div class="summary-grid">
            {program_cards_html}
            <div class="summary-card">
                <div class="summary-card-label">Must Have</div>
                <div class="summary-card-value">{priority_counts.get("Must Have", 0)}</div>
                <div class="summary-card-sub">Critical priority</div>
            </div>
            <div class="summary-card">
                <div class="summary-card-label">Should Have</div>
                <div class="summary-card-value">{priority_counts.get("Should Have", 0)}</div>
                <div class="summary-card-sub">High priority</div>
            </div>
            <div class="summary-card">
                <div class="summary-card-label">Could Have</div>
                <div class="summary-card-value">{priority_counts.get("Could Have", 0)}</div>
                <div class="summary-card-sub">Future scope</div>
            </div>
        </div>

        <!-- Roadmap Distribution Summary -->
        <div class="roadmap-summary">
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q1 '25</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q1 2025", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q2 '25</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q2 2025", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q3 '25</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q3 2025", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q4 '25</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q4 2025", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q1 '26</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q1 2026", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q2 '26</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q2 2026", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q3 '26</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q3 2026", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Q4 '26</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Q4 2026", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">2027+</div>
                <div class="roadmap-item-value">{roadmap_counts.get("2027+", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Backlog</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Backlog", 0)}</div>
            </div>
            <div class="roadmap-item">
                <div class="roadmap-item-label">Unscheduled</div>
                <div class="roadmap-item-value">{roadmap_counts.get("Unscheduled", 0)}</div>
            </div>
        </div>

        <div class="filters">
            <div class="filter-group">
                {program_filters_html}
            </div>
            <div class="filter-group">
                <button class="filter-btn active" data-priority="all">All Priorities</button>
                <button class="filter-btn" data-priority="must-have">Must Have</button>
                <button class="filter-btn" data-priority="should-have">Should Have</button>
                <button class="filter-btn" data-priority="could-have">Could Have</button>
            </div>
            <div class="filter-group">
                <button class="filter-btn active" data-roadmap="all">All</button>
                <button class="filter-btn" data-roadmap="q1-2025">Q1 '25</button>
                <button class="filter-btn" data-roadmap="q2-2025">Q2 '25</button>
                <button class="filter-btn" data-roadmap="q3-2025">Q3 '25</button>
                <button class="filter-btn" data-roadmap="q4-2025">Q4 '25</button>
                <button class="filter-btn" data-roadmap="q1-2026">Q1 '26</button>
                <button class="filter-btn" data-roadmap="q2-2026">Q2 '26</button>
                <button class="filter-btn" data-roadmap="q3-2026">Q3 '26</button>
                <button class="filter-btn" data-roadmap="q4-2026">Q4 '26</button>
                <button class="filter-btn" data-roadmap="2027-plus">2027+</button>
                <button class="filter-btn" data-roadmap="backlog">Backlog</button>
                <button class="filter-btn" data-roadmap="unscheduled">Unscheduled</button>
            </div>
            <div class="search-box">
                <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                    <circle cx="11" cy="11" r="8"></circle>
                    <path d="m21 21-4.35-4.35"></path>
                </svg>
                <input type="text" placeholder="Search stories..." id="searchInput">
            </div>
        </div>

        {categories_html}
    </main>

    <footer class="footer">
        <p>Providence GenomicsNow Requirements Dashboard</p>
        <p style="margin-top: 0.5rem;">Requirements Toolkit v{DASHBOARD_TEMPLATE_VERSION} | Designed by Glen Lewis</p>
        <p style="margin-top: 0.25rem; font-size: 0.75rem;">Generated {today}</p>
    </footer>

    <script>
        function toggleStory(header) {{
            header.closest('.story-card').classList.toggle('open');
        }}

        document.querySelectorAll('.filter-btn[data-filter]').forEach(btn => {{
            btn.addEventListener('click', () => {{
                document.querySelectorAll('.filter-btn[data-filter]').forEach(b => b.classList.remove('active'));
                btn.classList.add('active');
                filterStories();
            }});
        }});

        document.querySelectorAll('.filter-btn[data-priority]').forEach(btn => {{
            btn.addEventListener('click', () => {{
                document.querySelectorAll('.filter-btn[data-priority]').forEach(b => b.classList.remove('active'));
                btn.classList.add('active');
                filterStories();
            }});
        }});

        document.querySelectorAll('.filter-btn[data-roadmap]').forEach(btn => {{
            btn.addEventListener('click', () => {{
                document.querySelectorAll('.filter-btn[data-roadmap]').forEach(b => b.classList.remove('active'));
                btn.classList.add('active');
                filterStories();
            }});
        }});

        document.getElementById('searchInput').addEventListener('input', filterStories);

        function filterStories() {{
            const programFilter = document.querySelector('.filter-btn[data-filter].active')?.dataset.filter || 'all';
            const priorityFilter = document.querySelector('.filter-btn[data-priority].active')?.dataset.priority || 'all';
            const roadmapFilter = document.querySelector('.filter-btn[data-roadmap].active')?.dataset.roadmap || 'all';
            const searchTerm = document.getElementById('searchInput').value.toLowerCase();

            document.querySelectorAll('.story-card').forEach(card => {{
                const program = card.dataset.program;
                const priority = card.dataset.priority;
                const roadmap = card.dataset.roadmap;
                const searchText = card.dataset.search + ' ' + card.querySelector('.story-title').textContent.toLowerCase();

                const matchesProgram = programFilter === 'all' || program === programFilter;
                const matchesPriority = priorityFilter === 'all' || priority === priorityFilter;
                const matchesRoadmap = roadmapFilter === 'all' || roadmap === roadmapFilter;
                const matchesSearch = !searchTerm || searchText.includes(searchTerm);

                card.classList.toggle('hidden', !(matchesProgram && matchesPriority && matchesRoadmap && matchesSearch));
            }});

            document.querySelectorAll('.category-section').forEach(section => {{
                const visibleStories = section.querySelectorAll('.story-card:not(.hidden)').length;
                section.querySelector('.category-count').textContent = visibleStories + ' stor' + (visibleStories === 1 ? 'y' : 'ies');
                section.style.display = visibleStories === 0 ? 'none' : 'block';
            }});
        }}
    </script>
</body>
</html>'''


def validate_dashboard_structure(html_content: str) -> list:
    """
    PURPOSE:
        Validate that generated dashboard HTML contains all required structural
        elements. This catches accidental breaks during template updates.

    R EQUIVALENT:
        Like using rvest to check for expected HTML nodes in a document.

    PARAMETERS:
        html_content (str): The generated HTML content to validate

    RETURNS:
        list: List of missing elements (empty if all present)

    WHY THIS APPROACH:
        Simple string checks are fast and catch major structural breaks.
        More sophisticated DOM parsing would be overkill for this use case.
    """
    # Required structural elements for v1.0 template
    # Each tuple: (element_name, search_string)
    required_elements = [
        ("Header section", 'class="header"'),
        ("Programs section", 'class="summary-card"'),
        ("Roadmap summary", 'class="roadmap-summary"'),
        ("Filter buttons", 'class="filters"'),
        ("Program filter buttons", 'data-filter='),
        ("Priority filter buttons", 'data-priority='),
        ("Roadmap filter buttons", 'data-roadmap='),
        ("Search input", 'id="searchInput"'),
        ("Category sections", 'class="category-section"'),
        ("Story cards", 'class="story-card"'),
        ("Footer", 'class="footer"'),
        ("Version attribution", f'Requirements Toolkit v{DASHBOARD_TEMPLATE_VERSION}'),
        ("Filter JavaScript", 'function filterStories()'),
        ("Toggle JavaScript", 'function toggleStory('),
    ]

    missing = []
    for element_name, search_string in required_elements:
        if search_string not in html_content:
            missing.append(element_name)

    return missing


@mcp.tool()
async def generate_requirements_dashboard(
    output_path: str = None,
    push_to_github: bool = True,
    repo_path: str = None
) -> str:
    """
    PURPOSE:
        Generate the requirements dashboard HTML from current database state.
        Queries stories, test counts, and generates a filterable HTML dashboard.

    R EQUIVALENT:
        Like using DBI to query, then htmltools to build HTML, then git2r to push.

    PARAMETERS:
        output_path (str): Where to write the HTML file.
                           Default: ~/projects/requirements_toolkit/docs/index.html
        push_to_github (bool): If True, commits and pushes changes to GitHub
        repo_path (str): Path to the requirements_toolkit repo.
                         Default: ~/projects/requirements_toolkit

    RETURNS:
        str: Confirmation message with file path and URL

    EXAMPLE:
        generate_requirements_dashboard()  # Generate and push
        generate_requirements_dashboard(push_to_github=False)  # Generate only

    WHY THIS APPROACH:
        Single-file HTML with embedded CSS/JS for easy GitHub Pages hosting.
        No build step required - just generate and push.
    """
    import re
    import sqlite3
    import subprocess
    from pathlib import Path

    # Set default paths
    if repo_path is None:
        repo_path = Path.home() / "projects" / "requirements_toolkit"
    else:
        repo_path = Path(repo_path).expanduser()

    if output_path is None:
        output_path = repo_path / "docs" / "index.html"
    else:
        output_path = Path(output_path).expanduser()

    # Ensure docs directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # =========================================================================
    # CONNECT TO DATABASE
    # =========================================================================
    # Use the shared database via symlink
    db_path = Path.home() / "projects" / "data" / "client_product_database.db"

    if not db_path.exists():
        return f"Error: Database not found at {db_path}"

    conn = None
    try:
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row

        # =====================================================================
        # QUERY DATA FROM DATABASE
        # =====================================================================

        # Get all stories with program info and roadmap/status audit fields
        # Includes roadmap_target for annual planning and status date fields for timeline
        stories_query = """
            SELECT s.story_id, s.title, s.user_story, s.acceptance_criteria,
                   s.status, s.priority, s.category, s.version,
                   s.roadmap_target, s.draft_date, s.internal_review_date,
                   s.client_review_date, s.approved_date, s.needs_discussion_date,
                   p.prefix, p.name as program_name
            FROM user_stories s
            JOIN programs p ON s.program_id = p.program_id
            ORDER BY p.prefix, s.category, s.story_id
        """

        cursor = conn.execute(stories_query)
        stories = cursor.fetchall()

        if not stories:
            return "No stories found in database. Add stories first."

        # Get test counts per story
        test_counts_query = """
            SELECT story_id, COUNT(*) as test_count
            FROM uat_test_cases
            GROUP BY story_id
        """

        cursor = conn.execute(test_counts_query)
        test_counts = {row['story_id']: row['test_count'] for row in cursor.fetchall()}

        # Get summary stats
        total_stories = len(stories)
        total_tests_query = "SELECT COUNT(*) as cnt FROM uat_test_cases"
        total_tests_result = conn.execute(total_tests_query).fetchone()
        total_tests = total_tests_result['cnt'] if total_tests_result else 0

        # Count by priority
        priority_counts = {"Must Have": 0, "Should Have": 0, "Could Have": 0, "Won't Have": 0}
        for story in stories:
            priority = story['priority']
            if priority in priority_counts:
                priority_counts[priority] += 1

        # Count by roadmap target (annual planning timeline)
        # Valid values in order: Q1-Q4 2025, Q1-Q4 2026, 2027+, Backlog, NULL (Unscheduled)
        roadmap_counts = {
            "Q1 2025": 0,
            "Q2 2025": 0,
            "Q3 2025": 0,
            "Q4 2025": 0,
            "Q1 2026": 0,
            "Q2 2026": 0,
            "Q3 2026": 0,
            "Q4 2026": 0,
            "2027+": 0,
            "Backlog": 0,
            "Unscheduled": 0
        }
        for story in stories:
            roadmap = story['roadmap_target']
            if roadmap is None or roadmap == '':
                roadmap_counts["Unscheduled"] += 1
            elif roadmap in roadmap_counts:
                roadmap_counts[roadmap] += 1
            else:
                # Handle any unexpected values by putting them in Unscheduled
                roadmap_counts["Unscheduled"] += 1

        # Count by program
        program_counts = {}
        program_test_counts = {}
        for story in stories:
            prefix = story['prefix']
            if prefix not in program_counts:
                program_counts[prefix] = 0
                program_test_counts[prefix] = 0
            program_counts[prefix] += 1
            program_test_counts[prefix] += test_counts.get(story['story_id'], 0)

        # =====================================================================
        # GROUP STORIES BY CATEGORY
        # =====================================================================

        categories = {}
        for story in stories:
            category = story['category'] or "UNCATEGORIZED"
            if category not in categories:
                categories[category] = []
            categories[category].append(story)

        # Category display info
        # Includes all known categories from database with display names and icons
        category_info = {
            # P4M categories
            "REFERRAL": {"icon": "📋", "title": "Referral Workflow", "css": "referral"},
            "ANALYTICS": {"icon": "📊", "title": "Analytics", "css": "analytics"},
            "DASH": {"icon": "🖥️", "title": "Dashboard", "css": "dash"},
            "CONFIG": {"icon": "⚙️", "title": "Configuration", "css": "config"},
            "AUDIT": {"icon": "🔒", "title": "Compliance (Part 11)", "css": "audit"},
            "RECORD": {"icon": "🔒", "title": "Compliance (Part 11)", "css": "audit"},
            "ROLE": {"icon": "👤", "title": "Roles & Permissions", "css": "role"},
            # ONB categories
            "ADMIN": {"icon": "⚙️", "title": "Administration", "css": "config"},
            "EXPORT": {"icon": "📤", "title": "Export & Download", "css": "analytics"},
            "FORM": {"icon": "📝", "title": "Form Management", "css": "referral"},
            "GENE": {"icon": "🧬", "title": "Genetic Testing", "css": "dash"},
            "SAVE": {"icon": "💾", "title": "Save & Draft", "css": "config"},
            "STEP": {"icon": "📋", "title": "Workflow Steps", "css": "referral"},
            "UI": {"icon": "🎨", "title": "User Interface", "css": "dash"},
            # DIS/TEST2 categories
            "RECRUIT": {"icon": "👥", "title": "Patient Recruitment", "css": "referral"},
            # PLAT categories
            "RPT": {"icon": "📑", "title": "Reports", "css": "analytics"},
            # Fallback
            "UNCATEGORIZED": {"icon": "📁", "title": "Other", "css": "other"},
        }

        # Merge AUDIT and RECORD into single compliance category
        if "AUDIT" in categories and "RECORD" in categories:
            categories["COMPLIANCE"] = categories.pop("AUDIT", []) + categories.pop("RECORD", [])
            category_info["COMPLIANCE"] = {"icon": "🔒", "title": "Compliance (Part 11)", "css": "audit"}
        elif "AUDIT" in categories:
            categories["COMPLIANCE"] = categories.pop("AUDIT")
            category_info["COMPLIANCE"] = category_info["AUDIT"]
        elif "RECORD" in categories:
            categories["COMPLIANCE"] = categories.pop("RECORD")
            category_info["COMPLIANCE"] = category_info["RECORD"]

        # =====================================================================
        # GENERATE HTML
        # =====================================================================

        today = datetime.now().strftime("%B %d, %Y")

        # Build program summary cards
        program_cards_html = ""
        for prefix, count in program_counts.items():
            program_cards_html += f"""
                <div class="summary-card">
                    <div class="summary-card-label">{prefix}</div>
                    <div class="summary-card-value">{count}</div>
                    <div class="summary-card-sub">{program_test_counts[prefix]} test cases</div>
                </div>
            """

        # Build program filter buttons
        program_filters_html = '<button class="filter-btn active" data-filter="all">All Programs</button>'
        for prefix, count in program_counts.items():
            program_filters_html += f'<button class="filter-btn {prefix.lower()}" data-filter="{prefix.lower()}">{prefix} ({count})</button>'

        # Build category sections
        categories_html = ""

        # Define category order (preferred display order)
        # Categories not in this list will be added at the end alphabetically
        preferred_order = [
            "REFERRAL", "ANALYTICS", "DASH", "CONFIG", "COMPLIANCE", "ROLE",
            "RECRUIT", "STEP", "FORM", "ADMIN", "GENE", "SAVE", "EXPORT", "UI",
            "RPT", "UNCATEGORIZED"
        ]

        # Build complete category order: preferred categories first, then any others
        category_order = [cat for cat in preferred_order if cat in categories]
        for cat in sorted(categories.keys()):
            if cat not in category_order:
                # Insert before UNCATEGORIZED if present, else append
                if "UNCATEGORIZED" in category_order:
                    idx = category_order.index("UNCATEGORIZED")
                    category_order.insert(idx, cat)
                else:
                    category_order.append(cat)

        for cat_key in category_order:
            if cat_key not in categories:
                continue

            cat_stories = categories[cat_key]
            info = category_info.get(cat_key, category_info["UNCATEGORIZED"])

            stories_html = ""
            for story in cat_stories:
                story_id = story['story_id']
                title = story['title']
                user_story = story['user_story']
                acceptance_criteria = story['acceptance_criteria']
                status = story['status']
                priority = story['priority']
                version = story['version']
                prefix = story['prefix']
                roadmap_target = story['roadmap_target']

                # Get status timeline dates
                draft_date = story['draft_date']
                internal_review_date = story['internal_review_date']
                client_review_date = story['client_review_date']
                approved_date = story['approved_date']
                needs_discussion_date = story['needs_discussion_date']

                # Parse acceptance criteria (handle numbered list or newline separated)
                ac_items = []
                if acceptance_criteria:
                    # Remove any leading numbers and clean up
                    lines = acceptance_criteria.strip().split('\n')
                    for line in lines:
                        line = line.strip()
                        if line and not line.isdigit():
                            # Remove leading numbers like "1." or "1)"
                            line = re.sub(r'^\d+[\.\)]\s*', '', line)
                            if line:
                                ac_items.append(line)

                ac_html = ""
                for item in ac_items:
                    ac_html += f"<li>{escape_html(item)}</li>"

                test_count = test_counts.get(story_id, 0)

                # Priority badge class
                priority_class = priority.lower().replace(" ", "-") if priority else "should-have"

                # Roadmap badge (short format for display)
                # Convert full format (Q1 2025) to abbreviated (Q1 '25)
                roadmap_display = "Unscheduled"
                roadmap_class = "unscheduled"
                if roadmap_target:
                    # Q1 2025 -> Q1 '25, Q1 2026 -> Q1 '26, 2027+ -> 2027+, Backlog -> Backlog
                    roadmap_display = roadmap_target.replace(" 2025", " '25").replace(" 2026", " '26")
                    roadmap_class = roadmap_target.lower().replace(" ", "-").replace("+", "-plus")

                # Build status timeline HTML
                # Format dates to short form (M/D/YY e.g., 1/10/26)
                def format_short_date(date_str):
                    if not date_str:
                        return None
                    try:
                        # Parse YYYY-MM-DD HH:MM:SS format
                        dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
                        # Format as M/D/YY (e.g., 1/10/26)
                        return dt.strftime('%-m/%-d/%y')
                    except:
                        return date_str[:8] if date_str else None

                # Timeline stages in order
                timeline_stages = [
                    ("Draft", draft_date, status == "Draft"),
                    ("Internal Review", internal_review_date, status == "Internal Review"),
                    ("Client Review", client_review_date, status == "Pending Client Review"),
                    ("Approved", approved_date, status == "Approved"),
                ]

                # Add Needs Discussion if it has a date (shows it was in that state at some point)
                if needs_discussion_date:
                    timeline_stages.append(("Discussion", needs_discussion_date, status == "Needs Discussion"))

                timeline_html = '<div class="status-timeline">'
                for i, (stage_name, stage_date, is_current) in enumerate(timeline_stages):
                    formatted_date = format_short_date(stage_date)
                    has_date = formatted_date is not None
                    stage_class = "completed" if has_date else "pending"
                    if is_current:
                        stage_class = "current"

                    timeline_html += f'''
                        <div class="timeline-stage {stage_class}">
                            <div class="timeline-dot"></div>
                            <div class="timeline-label">{stage_name}</div>
                            <div class="timeline-date">{formatted_date if formatted_date else '-'}</div>
                        </div>
                    '''
                    # Add connector between stages (except after last)
                    if i < len(timeline_stages) - 1:
                        timeline_html += '<div class="timeline-connector"></div>'

                timeline_html += '</div>'

                # Roadmap filter data attribute (must match filter button data-roadmap values)
                roadmap_filter = roadmap_target.lower().replace(" ", "-").replace("+", "-plus") if roadmap_target else "unscheduled"

                stories_html += f"""
                    <article class="story-card" data-program="{prefix.lower()}" data-priority="{priority_class}" data-roadmap="{roadmap_filter}" data-search="{title.lower()} {user_story.lower() if user_story else ''}">
                        <div class="story-header" onclick="toggleStory(this)">
                            <div>
                                <span class="story-id {prefix.lower()}">{story_id}</span>
                                <h3 class="story-title">{escape_html(title)}</h3>
                                <div class="story-meta">
                                    <span class="badge badge-{priority_class}">{priority}</span>
                                    <span class="badge badge-roadmap badge-{roadmap_class}">{roadmap_display}</span>
                                    <span class="badge badge-status">{status} v{version}</span>
                                </div>
                            </div>
                            <svg class="story-toggle" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="6 9 12 15 18 9"></polyline></svg>
                        </div>
                        <div class="story-body">
                            <div class="story-section">
                                <h4 class="story-section-title">Status Timeline</h4>
                                {timeline_html}
                            </div>
                            <div class="story-section">
                                <h4 class="story-section-title">User Story</h4>
                                <p class="user-story">{escape_html(user_story) if user_story else 'No user story defined.'}</p>
                            </div>
                            <div class="story-section">
                                <h4 class="story-section-title">Acceptance Criteria</h4>
                                <ol class="acceptance-criteria">
                                    {ac_html if ac_html else '<li>No acceptance criteria defined.</li>'}
                                </ol>
                            </div>
                            <div class="test-count">
                                <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 11l3 3L22 4"></path><path d="M21 12v7a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11"></path></svg>
                                {test_count} test case{'s' if test_count != 1 else ''}
                            </div>
                        </div>
                    </article>
                """

            categories_html += f"""
                <section class="category-section" data-category="{cat_key.lower()}">
                    <div class="category-header">
                        <div class="category-icon {info['css']}">{info['icon']}</div>
                        <h2 class="category-title">{info['title']}</h2>
                        <span class="category-count">{len(cat_stories)} stor{'y' if len(cat_stories) == 1 else 'ies'}</span>
                    </div>
                    <div class="stories-grid">
                        {stories_html}
                    </div>
                </section>
            """

        # Build the full HTML
        html_content = generate_full_html_template(
            total_stories=total_stories,
            total_tests=total_tests,
            program_cards_html=program_cards_html,
            program_filters_html=program_filters_html,
            categories_html=categories_html,
            priority_counts=priority_counts,
            roadmap_counts=roadmap_counts,
            today=today
        )

        # =====================================================================
        # WRITE FILE
        # =====================================================================

        # =====================================================================
        # VALIDATE STRUCTURE (catches accidental breaks)
        # =====================================================================
        missing_elements = validate_dashboard_structure(html_content)
        if missing_elements:
            logger.warning(f"Dashboard validation: missing elements: {missing_elements}")

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        result = f"✅ Dashboard generated: {output_path}\n"
        result += f"   Stories: {total_stories} | Tests: {total_tests}\n"

        # Report validation warnings (but don't fail - still generate the file)
        if missing_elements:
            result += f"   ⚠️ Structure warning: missing {', '.join(missing_elements)}\n"

        # =====================================================================
        # PUSH TO GITHUB (optional)
        # =====================================================================

        if push_to_github:
            try:
                original_dir = os.getcwd()
                os.chdir(repo_path)

                # Pull latest
                subprocess.run(["git", "pull"], check=True, capture_output=True)

                # Add and commit
                subprocess.run(["git", "add", "docs/"], check=True, capture_output=True)

                commit_msg = f"Update requirements dashboard - {total_stories} stories, {total_tests} tests"
                commit_result = subprocess.run(
                    ["git", "commit", "-m", commit_msg],
                    capture_output=True,
                    text=True
                )

                if "nothing to commit" in commit_result.stdout or "nothing to commit" in commit_result.stderr:
                    result += "   No changes to commit.\n"
                else:
                    # Push
                    subprocess.run(["git", "push"], check=True, capture_output=True)
                    result += "   ✅ Pushed to GitHub\n"
                    result += "   🔗 https://glewis05.github.io/requirements_toolkit/\n"

                os.chdir(original_dir)

            except subprocess.CalledProcessError as e:
                result += f"   ⚠️ Git error: {e}\n"
            except Exception as e:
                result += f"   ⚠️ Error pushing to GitHub: {e}\n"

        return result

    except sqlite3.Error as e:
        logger.error(f"generate_requirements_dashboard() database error: {e}")
        return f"Database error: {str(e)}"
    except Exception as e:
        logger.error(f"generate_requirements_dashboard() error: {e}", exc_info=True)
        return f"Error: {str(e)}"
    finally:
        if conn:
            conn.close()


# ============================================================
# RUN SERVER
# ============================================================
if __name__ == "__main__":
    mcp.run()
